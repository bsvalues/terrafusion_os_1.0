"""
COPYRIGHT 2018 ESRI

TRADE SECRETS: ESRI PROPRIETARY AND CONFIDENTIAL
Unpublished material - all rights reserved under the
Copyright Laws of the United States.

For additional information, contact:
Environmental Systems Research Institute, Inc.
Attn: Contracts Dept
380 New York Street
Redlands, California, USA 92373

email: contracts@esri.com


---------------------------------------------------------------------------
Source Name:   TopographicGeneralizationUtilities.py
Version:       ArcGIS 2.3
Author:        Environmental Systems Research Institute Inc.
Description:   Contains functions commonly used for generalization.
---------------------------------------------------------------------------
"""

import os
import arcpy
import openpyxl
import time
import sys

# String that folds casing during comparison
#
class CaseStr(str):

    def __new__(cls, key, ignoreChars=None):
        newKey = key
        if ignoreChars:
            for c in ignoreChars:
                newKey = newKey.replace(c, '')
        return str.__new__(cls, newKey)

    def __init__(self, key, ignoreChars=None):
        newKey = key
        if ignoreChars:
            for c in ignoreChars:
                newKey = newKey.replace(c, '')
        str.__init__(newKey)

    def __hash__(self):
        return hash(str.lower(self))

    def __eq__(self, other):
        return str.lower(self) == other.lower()

    def __ne__(self, other):
        return not self.__eq__(other)

    def __lt__(self, other):
        return str.lower(self) < other.lower()

    def startswith(self, other):
        return self.lower().startswith(other.lower())

    def endswith(self, other):
        return self.lower().endswith(other.lower())


# Set with case-insensitive values
#
class CaseSet(set):
    def __init__(self, data=[], ignoreChars=None):
        super().__init__()

        if ignoreChars and not isinstance(ignoreChars, (list,str)):
            raise Exception('invalid type {}. Expected list'.format(type(ignoreChars)))
        self.__ignoreChars = ignoreChars

        if data:
            if isinstance(data, list):
                for val in data:
                    self.add(val)
            else:
                raise Exception('invalid type {}. Expected list'.format(type(data)))


    def norm(self, val):
        return CaseStr(val, self.__ignoreChars) if isinstance(val, str) else val

    def __contains__(self, value):
        return super().__contains__(self.norm(value))

    def add(self, value):
        return super().add(self.norm(value))

    def remove(self, value):
        return super().remove(self.norm(value))


# Dictionary with case-insensitive keys
#
class CaseDict(dict):

    def __init__(self, data=None, ignoreChars=None):

        # ignoreChars allows dictionary keys like "MyKey" and "My Key" to be treated equal
        if ignoreChars and not isinstance(ignoreChars, (list, str)):
            raise Exception('invalid type {}. Expected list'.format(type(ignoreChars)))

        self.__ignoreChars = ignoreChars

        super(CaseDict, self).__init__()
        if data:
            if isinstance(data, list):
                for item in data:
                    self.__setitem__(item, item)
            elif isinstance(data, dict):
                for key, val in data.items():
                    self.__setitem__(key, val)
            else:
                raise Exception('Invalid type {}'.format(type(data)))

    def norm(self, val):
        return CaseStr(val, self.__ignoreChars) if isinstance(val, str) else val

    def __contains__(self, key):
        return super(CaseDict, self).__contains__(self.norm(key))

    def __setitem__(self, key, value):
        return super(CaseDict, self).__setitem__(self.norm(key), value)

    def __getitem__(self, key):
        return super(CaseDict, self).__getitem__(self.norm(key))

    def __delitem__(self, key):
        return super(CaseDict, self).__delitem__(self.norm(key))

    def keys(self):
        return [self.norm(key) for key in super(CaseDict, self).keys()]

    def values(self):
        return [self.norm(value) for value in super(CaseDict, self).values()]

    def get(self, key, default_value = None):
        return self[key] if self.__contains__(key) else default_value


class ToolException(Exception):
    def __init__(self, message_id, arg1='', arg2=''):
        msg_text = arcpy.GetIDMessage(message_id, "Unexpected error")
        if arg1:
            msg_text = msg_text.replace('%1', arg1)
            msg_text = msg_text.replace('%s', arg1)
            msg_text = msg_text.replace('%d', arg1)
        if arg2:
            msg_text = msg_text.replace('%2', arg2)
        super().__init__(msg_text)
        self.id = message_id
        self.args += (arg1,arg2)


class Logger:
    def __init__(self):
        self.DEBUG_MODE = True
        self.reset()

    def message(self, msg):
        log_time = time.strftime("%M:%S", time.gmtime(time.time() - self.__start_time))
        delta_time = time.gmtime(time.time() - self.__prev_time )
        if self.DEBUG_MODE:
            print(f"{log_time} INFO \t {msg} (prev:{delta_time.tm_min} mins {delta_time.tm_sec} secs)")
        else:
            arcpy.AddMessage(f"{log_time} INFO \t {msg} (prev:{delta_time.tm_min} mins {delta_time.tm_sec} secs)")
        self.__prev_time = time.time()

    def error(self, msg):
        log_time = time.strftime("%M:%S", time.gmtime(time.time() - self.__start_time))
        delta_time = time.gmtime(time.time() - self.__prev_time )
        if self.DEBUG_MODE:
            print(f"{log_time} ERROR\t {msg} (prev:{delta_time.tm_min} mins {delta_time.tm_sec} secs)")
        else:
            arcpy.AddMessage(f"{log_time} ERROR\t {msg} (prev:{delta_time.tm_min} mins {delta_time.tm_sec} secs)")

        self.__prev_time = time.time()

    def reset(self):
        self.__start_time = time.time()
        self.__prev_time = time.time()


class WorkspaceLock:
    def __init__(self, workspace):
        self.original_workspace = arcpy.env.workspace
        arcpy.env.workspace = workspace

    def release(self):
        arcpy.env.workspace = self.original_workspace


def is_value_empty(value):
    return value in ('#', ' ', '') or not value


def CheckField(fc_class, field_name, field_type="Long"):
    """
    Checks to see if a field existins in a feature class, if not existing
    the field is added
    """

    field_list = [field.name.upper() for field in arcpy.ListFields(fc_class)]
    if field_name.upper() not in field_list:
        arcpy.AddField_management(fc_class, field_name, field_type)

    return


def CheckFieldName(fc_class, field_name):
    """
    Checks to see if a field exists in a feature class
    the field is added
    """
    if not fc_class or not field_name:
        return None

    field_dict = CaseDict([item.name for item in arcpy.ListFields(fc_class)])
    if field_name in field_dict:
        return field_dict[field_name]

    return None


def CheckFields(fc_class, field_dict):
    """
    Checks to see if a field exists in a feature class, if not existing
    the field is added
    """
    field_names = CaseDict([item.name for item in arcpy.ListFields(fc_class)])
    for field_name, field_type in field_dict.items():
        if field_name not in field_names:
            arcpy.AddField_management(fc_class, field_name, field_type)

    return


def GetRuleFields(sheet, field_type):
    """Get fields for a rule in generalization spreadsheet"""

    if sheet is None:
        raise Exception('sheet parameter is None')
    if not isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet):
        raise Exception('Invalid type {}. Expecting type openpyxl.worksheet.worksheet.Worksheet'.format(type(sheet)))

    fields = []

    common_columns = ['THEME', 'FIELDUSE', 'FIELD', 'VALUE']
    column_dict = CheckColumns(sheet, common_columns)

    for row_idx in range(1, sheet.max_row + 1):
        f_use = sheet.cell(row_idx, column_dict['FIELDUSE']).value
        if f_use is not None and f_use.upper() == field_type.upper():
            field = sheet.cell(row_idx, column_dict['FIELD']).value
            if field is not None and field not in fields:
                fields.append(field)

    return fields


def GetFields(layer, ignore_null=False):
    """
    Checks to see if a field exists in a feature class, if not existing
    the field is added
    """

    fields = arcpy.ListFields(layer)
    upper_fields = {}

    for field in fields:
        name = field.name
        if not field.required:
            field_type = field.type
            if field_type not in ('Blob', 'Raster'):
                upper_fields[name.upper()] = name

        if ignore_null:
            values = [row[0] for row in arcpy.da.SearchCursor(layer, name)]
            uni_values = list(set(values))
            if len(uni_values) == 1:
                if not uni_values or uni_values == '' or uni_values is None:
                    upper_fields.pop(name.upper())

    return upper_fields


def GetFeatureClasses(workspace, fc_names=[]):

    """
    Gets all the feature classes in a workspace.  Includes feature datasets.
    Returns a dictionary of feature class names (upper case) and full path
    to the feature class.
    Can optionally limit to a subset of feature classes using the fc_names list.
    """

    fc_dict = CaseDict()
    input_fc_set = CaseSet(fc_names)  # for case-insensitive search

    walk = arcpy.da.Walk(workspace, datatype="FeatureClass")
    for dirpath, dirnames, filenames in walk:
        for filename in filenames:
            if input_fc_set:
                if filename in input_fc_set:
                    fc_class_path = os.path.join(dirpath, filename)
                    fc_dict[filename.upper()] = fc_class_path
            else:
                fc_class_path = os.path.join(dirpath, filename)
                fc_dict[filename.upper()] = fc_class_path

    return fc_dict


def GetWorkspacePath(input_table):
    """
    Return the Geodatabase path from the input table or feature class.
    """

    workspace = os.path.dirname(input_table)
    if [any(ext) for ext in ('.gdb', '.mdb', '.sde') if ext in os.path.splitext(workspace)]:
        return workspace
    else:
        return os.path.dirname(workspace)


def GetThemeFeatureClassPath(theme_workspace, theme, fc_name, geometry_type):
    """
    Generates the full path of a theme feature class based on its shape type
    """

    type_suffix = ''
    if geometry_type.upper() == 'POLYGON' or geometry_type.upper() == 'POLY':
        type_suffix = "A"
    elif geometry_type.upper() == 'LINE' or geometry_type.upper() == 'POLYLINE':
        type_suffix = "L"
    elif geometry_type.upper() == 'POINT' or geometry_type.upper() == 'PT':
        type_suffix = "P"

    theme_fc_name = '{}_{}_{}'.format(theme, fc_name, type_suffix)
    return os.path.join(theme_workspace, theme_fc_name)


def MakeWhereClause(inFC, field_name, value, operator='='):
    whereclause = None
    try:
        # determine field type
        fields = arcpy.ListFields(inFC)
        field_type = None

        for field in fields:
            if field.name.upper() == field_name.upper():
                field_type = field.type

        if field_type:
            if field_type in ('Double', 'Integer', 'OID', 'Single', 'SmallInteger'):
                whereclause = "{0} {1} {2}".format(field_name, operator, value)
            elif field_type in ('Date', 'GlobalID', 'Guid', 'String'):
                whereclause = "{0} {1} '{2}'".format(field_name, operator, value)
            else:
                arcpy.AddIDMessage('WARNING', 90293, field_name, field_type)
        else:
            raise ToolException(90293, field_name, "<unknown>")
    finally:
        return whereclause


def MakeInWhereclause(inFC, field_name, in_list):

    whereClause = None
    try:
        if len(in_list) >= 1:
            # determine field type
            fields = arcpy.ListFields(inFC)
            field_type = None
            for field in fields:
                if field.name.upper() == field_name.upper():
                    field_type = field.type

            if field_type:
                csv = ""
                if field_type in ('Double', 'Integer', 'OID', 'Single', 'SmallInteger'):
                    for i in in_list:
                        csv += "{0},".format(i)
                    # Remove trailing comma
                    csv = csv[:-1]
                elif field_type in ('Date', 'GlobalID', 'OID', 'Guid', 'String'):
                    for i in in_list:
                        csv += "'{0}',".format(i)
                    # Remove trailing comma
                    csv = csv[:-1]
                else:
                    arcpy.AddIDMessage('WARNING', 90293, field_name, field_type)

                if not csv == "":
                    whereClause = '{0} IN ({1})'.format(arcpy.AddFieldDelimiters(inFC, field_name), csv)
            else:
                arcpy.AddIDMessage('INFORMATIVE', 413, field_name, inFC)
        else:
            arcpy.AddIDMessage('WARNING', 10245, field_name )  # No feature is selected in %1.
    finally:
        return whereClause


def GetMinimumAreaProjected(fc, min_val, area_field, min_units):

    min_proj = 0
    min_val = float(min_val)
    order = 'ORDER BY '+ area_field

    #search the feature class, sorted by area
    with arcpy.da.SearchCursor(fc, ['SHAPE@', area_field], sql_clause=(None, order)) as cursor:
        for row in cursor:
            # if the area is less than the min value, keep searching
            proj_area = row[0].getArea('GEODESIC', min_units.upper())
            if proj_area <= min_val and proj_area > min_proj:
                min_proj = row[0].area
            # if the value is greater than the minimum value, break
            # at this point we have find the largest value that is less than the
            # minimum value
            elif proj_area > min_val:
                return min_proj
    return min_proj


def GetMinimumLengthProjected(fc, min_val, len_field, min_units):

    min_proj = 0
    min_val = float(min_val)
    order = 'ORDER BY '+ len_field

    #search the feature class, sorted by area
    with arcpy.da.SearchCursor(fc, ['SHAPE@', len_field], sql_clause=(None, order)) as cursor:
        for row in cursor:
            # if the area is less than the min value, keep searching
            proj_area = row[0].getLength('GEODESIC', min_units.upper())
            if proj_area <= min_val and proj_area > min_proj:
                min_proj = row[0].length
            # if the value is greater than the minimum value, break
            # at this point we have find the largest value that is less than the
            # minimum value
            elif proj_area > min_val:
                return min_proj
    return min_proj


def CreateMinimumAreaQuery(fc, minimum_area, operation='<='):
    """
    Returns a SQL query to return features in a polygon feature class smaller
    than the specified minimum area
    """

    query = ''
    supported_units = ('ACRES', 'ARES', 'HECTARES', 'SQUAREFEET', 'SQUAREKILOMETERS', 'SQUAREMETERS', 'SQUAREMILES', 'SQUAREYARDS')

    # Determine if geographic coordinate system
    desc = arcpy.Describe(fc)
    spatial_ref = desc.spatialReference
    if spatial_ref.name == spatial_ref.GCS.name:
        sr_type = 'GCS'
    else:
        sr_type = 'PROJ'

    min_val, min_units = minimum_area.split(' ')

    arcpy.AddMessage("{} : {}".format(min_val, min_units))
    area_field = desc.areaFieldName

    if sr_type == 'GCS' and min_units.upper() == 'DECIMALDEGREES':
        query = '{} < {}'.format(area_field, min_val)
    elif min_units.upper() in supported_units:
        min_proj = GetMinimumAreaProjected(fc, min_val, area_field, min_units)
        query = '{} {} {}'.format(area_field, operation, min_proj)
    else:
        raise ToolException(817)  # Invalid linear unit type.
    return query


def CreateMinimumLengthQuery(fc, minimum_len, operation='<='):
    """
    Returns a SQL query to return features in a polygon feature class smaller
    than the specified minimum area
    """
    query = ''
    supported_units = ('FEET', 'KILOMETERS', 'METERS', 'MILES', 'NAUTICALMILES', 'YARDS')

    # Determine if geographic coordinate system
    desc = arcpy.Describe(fc)
    spatial_ref = desc.spatialReference

    if desc.shapeType in ('Polygon', 'Polyline'):
        if spatial_ref.name == spatial_ref.GCS.name:
            sr_type = 'GCS'
        else:
            sr_type = 'PROJ'

        min_val, min_units = minimum_len.split(' ')
        len_field = desc.lengthFieldName

        if sr_type == 'GCS' and min_units.upper() == 'DECIMALDEGREES':
            query = MakeWhereClause(fc, len_field, min_val, operation)
        elif min_units.upper() in supported_units:
            min_proj = GetMinimumLengthProjected(fc, min_val, len_field, min_units)
            query = MakeWhereClause(fc, len_field, min_proj, operation)
        else:
            raise ToolException(817)  # Invalid linear unit type.
    else:
        raise ToolException(160746)  # Geometry type not supported.
    return query



def CheckSheets(rules_book, sheet_name, sheet_index=-1):
    """
    Checks to ensure that a sheet with the specified name exists in the
    choosen generalization rules spreadsheet
    """

    if not isinstance(rules_book, openpyxl.Workbook):
        arcpy.AddMessage((arcpy.GetIDMessage(241)+' (rules_book)'))
        return None

    sheet_names = rules_book.sheetnames

    existing_name = None
    for name in sheet_names:
        if name.upper() == sheet_name.upper():
            existing_name = name
            break

    if existing_name:
        sheet = rules_book[existing_name]
        return sheet
    else:
        if sheet_index != -1:
            sheet = rules_book.worksheets[sheet_index]
            if not sheet:
                raise ToolException(90294, sheet_name)
            elif sheet.title.lower() != sheet_name:
                raise ToolException(90294, sheet_name)
            return sheet
        else:
            raise ToolException(90294, sheet_name)


def CheckColumns(sheet, column_names=[]):
    """
    Returns a dictionary of column name and column index.  If a column with the
    specified name does not exists, returns error and exits.
    """

    if not isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet):
        raise Exception('Invalid type {}. Expecting type openpyxl.worksheet.worksheet.Worksheet'.format(type(sheet)))

    # Ignore case and spaces in column names
    dict_column_indices = CaseDict(None, [' '])
    dict_column_names = CaseDict(column_names, [' '])

    # Search the top row of the sheet for specific titles
    title_cells = []
    for cell in sheet[1]:
        if cell.value:
            title_cells.append(cell.value)
    for index, title in enumerate(title_cells):
        if title in dict_column_names:
            dict_column_indices[title] = index + 1

    # if we did not find a match for every column name
    if len(column_names) != len(dict_column_indices):
        # determine which columns are missing
        missing_columns = list(CaseSet(dict_column_names.keys()) - CaseSet(dict_column_indices.keys()))
        raise ToolException(90295, str(missing_columns))
    else:
        return dict_column_indices


def GetColumns(sheet, theme):
    """
    Returns a dictionary of column name and column index.  If a column with the
    specified name does not exists, returns error.
    """

    title_cells = []
    for cell in sheet[1]:
        if cell.value:
            title_cells.append(cell.value)
    column_dict = CaseDict()
    field_dict = CaseDict()

    # Search the top row of the sheet for specific titles
    for index, col_name in enumerate(title_cells):
        t_name = col_name.upper()

        # columns for a theme with start with the theme name followed by :
        if t_name.startswith(theme.upper()):
            theme_name, t_name = t_name.split(':')
            t_name.replace(' ', '')
            t_name.replace(';', '_')

            # Ignore case and spaces in column names
            field_name = '{}_{}'.format(theme, t_name.replace(' ', ''))
            column_dict[col_name] = index + 1
            field_dict[field_name] = col_name

    return column_dict, field_dict


def query_excel_rows(excel_file, sheet_name, column_list, theme_list = [], as_dict = True):
    """
    Retrieves cell values from a generalization rules spreadsheet in table format

    :param excel_file : Path to the excel file that contains generalization rules
    :param sheet_name : Name or index of the sheet
    :param column_list: Column names in string or list format, where first column name must be suffixed with '@OID'
                        if it's not the first one in the list. If column list is set to '*', the first row is assumed
                        to contain the column names and it will return all non-theme fields, i.e. column names without
                        a theme prefix separated with a colon.
    :param theme_list : list of theme names used to identify theme columns.
    :param as_dict    : whether the results will be generated as a dictionary or a list. If it's requested as a list,
                        the first item will contain the column names.
    :return           : List of row values that are indexed by index or column names (case insensitive)
    """
    skipChars = '- '

    if isinstance(column_list,str):
        column_list = column_list.split(',')
    columns = [CaseStr(col.strip()) for col in column_list]

    if isinstance(theme_list,str):
        theme_list = theme_list.split(',')
    themes = [CaseStr(name.strip()) for name in theme_list]

    if not columns:
        return []

    # Get sheet from workbook
    workbook = None
    workbook = OpenWorkbook(excel_file)
##    arcpy.AddMessage(workbook)

    sheet = None
    if isinstance(sheet_name, str):
        sheet_names = workbook.sheetnames
        for name in sheet_names:
            if name.strip().upper() == sheet_name.strip().upper():
                sheet = workbook[name]
                break
    elif isinstance(sheet_name,int):
        sheet_index = int(sheet_name)
        if sheet_index >= 1 or sheet_index < len(workbook._sheets):
            sheet = workbook.worksheets[sheet_index]

    if not sheet:
        # Sheet %1 does not exist within the Generalization Rules file.
        raise ToolException(90294, sheet_name)

    # Find row index for column names. If all fields are requested, assume it is the first row
    start_index = None
    if len(columns) == 1 and columns[0] == '*':
        start_index = 1
        columns = []
        for col_index in range(1, sheet.max_column + 1):
            val = sheet.cell(1, col_index).value
            if val is not None:
                col_name = CaseStr(str(val).strip(), skipChars)
                if ':' not in col_name and col_name:
                    columns.append(col_name)
    else:
        # Find first column name and index
        col_index, first_col_name = next((x for x in enumerate(columns) if '@oid' in x[1].lower()), (-1, columns[0]))
        if col_index != -1:
            first_col_name = first_col_name[:-4]
            columns[col_index] = CaseStr(first_col_name, skipChars)

        # Find column indices
        for row_index in range(1, sheet.max_row + 1):
            val = sheet.cell(row_index, 1).value
            if val is not None:
                col_name = str(val).strip()
                if col_name.upper() == first_col_name.upper():
                    start_index = row_index
                    break

    if start_index == None:
        # If your header row does not exist, you must enter columnNames.
        raise ToolException(100165)

    # Retrieve column infos
    theme_columns = []
    all_column_indices = CaseDict(ignoreChars=skipChars)
    for col_index in range(1, sheet.max_column + 1):
        row_column_name = sheet.cell(start_index, col_index).value
        if row_column_name is not None:
            column_name = CaseStr(str(row_column_name).strip(), skipChars)
            if ':' in column_name:
                theme_name = column_name.split(':')[0]
                if theme_name in themes:
                    theme_columns.append(column_name)
##                    print(column_name)
            all_column_indices[ column_name ] = col_index

    columns += theme_columns
    column_indices = []
    missing_fields = ''
    for col_name in columns:
        if col_name not in all_column_indices:
            missing_fields += f"{'' if not missing_fields else ', '}{col_name}"
        column_indices.append( all_column_indices[col_name] if col_name in all_column_indices else -1)

    if missing_fields:
        # Column: %1 does not exist in worksheet: %2.
        raise ToolException(180130, missing_fields, os.path.basename(excel_file))

##    print(columns)

    # Retrieve data
    results = []
    if not as_dict:
        results.append(columns) # add column names to list

    for row_index in range(start_index + 1, sheet.max_row + 1):

        if not str(sheet.cell(row_index, 1).value).strip():
            continue # skip blank rows

        vals = CaseDict(ignoreChars=skipChars) if as_dict else []
        for col_index, col_name in enumerate(columns):
            cell_value = sheet.cell(row_index, column_indices[col_index]).value
            # don't check for null values, want to add all values even if null
            if as_dict:
                vals[col_name] = cell_value
            else:
                vals.append(cell_value)
        results.append(vals)

    return results


def OpenWorkbook(excel_file):
    try:
        if str(excel_file).endswith(".xls"):
            arcpy.AddIDMessage("ERROR", 10158, excel_file)  # Unable to open file %1.
            sys.exit(1)
        else:
            rule_book = openpyxl.load_workbook(excel_file)
    except Exception as e:
        arcpy.AddIDMessage("ERROR", 10158, excel_file)  # Unable to open file %1.
        sys.exit(1)

    if not isinstance(rule_book, openpyxl.Workbook):
        arcpy.AddMessage((arcpy.GetIDMessage(241)+' (rules_book)'))
        return None
    return rule_book
