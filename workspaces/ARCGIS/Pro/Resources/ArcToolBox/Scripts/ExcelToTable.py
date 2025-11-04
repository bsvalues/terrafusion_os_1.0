"""
Tool name: Excel To Table
Source: ExcelToTable.py
Author: ESRI

Convert a Microsoft Excel (xls or xlsx) file to an gdb, dbf or INFO table.

"""

import arcpy
import os
import sys
from datetime import datetime
import xlrd
import re
import locale
import openpyxl
import arcgisscripting
import ExcelToTableX
from functools import cached_property


INTEGER_MAX = 2**31
BIG_INTEGER_MAX = 2**53

FIELD_TYPE_MAP = {1: 'Text',
                  2: 'Double',
                  3: 'Date',
                  4: 'Long',
                  'MGRS': 'MGRS',
                  'USNG': 'USNG',
                  'bigint': "BigInteger",
                  'dateonly': 'DateOnly',
                  'timeonly': 'TimeOnly',
                  }


class TableInfo:
    """ Class containing information about the output table"""
    def __init__(self, out_table=None):
        self.is_dbf = False
        self.is_sde = False
        self.ext = None
        self.supportsBigInteger = True
        self.supportsTimeOnly = True
        self.supportsDateOnly = True
        self.supportsTimestampOffset = True
        self.out_path = ""
        self._out_table_name = ""

        if out_table:
            self._out_table_name = os.path.split(out_table)[1]
            self.out_path = os.path.split(out_table)[0]
            try:
                self.out_path_desc = arcpy.Describe(self.out_path)

                self.is_dbf = (self.out_path_desc.workspaceFactoryProgID == ""
                                    and os.path.isdir(self.out_path))

                self.is_sde = self.out_path_desc.workspaceFactoryProgID == 'esriDataSourcesGDB.SdeWorkspaceFactory'

                self.ext = os.path.splitext(self._out_table_name)[1].lower()
                self.supportsBigInteger = getattr(self.out_path_desc, 'supportsBigInteger', False)
                self.supportsTimeOnly = getattr(self.out_path_desc, 'supportsTimeOnly', False)
                self.supportsDateOnly = getattr(self.out_path_desc, 'supportsDateOnly', False)
                self.supportsTimestampOffset = getattr(self.out_path_desc, 'supportsTimestampOffset', False)
            except Exception as e:
                print(e)

    @cached_property
    def out_table_name(self):
        """ produce a useable name for the out table"""

        out_table_name = self._out_table_name
        if self.is_sde:
            try:
                # pluck 'table' from 'db.owner.table'
                out_table_name = arcpy.ParseTableName(out_table_name, self.out_path).split(", ")[2]
            except:
                pass

        if self.is_dbf:
            if not self.ext in [".dbf",]:
                out_table_name += ".dbf"
        else:
            out_table_name = arcpy.ValidateTableName(out_table_name, self.out_path)

        return out_table_name

class Field(object):
    """ Class to hold properties and behavior of the output fields """

    def __init__(self, name, tableInfo, field_names, col_num=None, wb_datemode=0):
        """ Validate name of field based on output table format as well
              as existing/used field names
        """
        self.length = None
        self._cell_type = None
        self.col_num = col_num
        self.alias = name if name else None
        self.has_errors = False
        self.tableInfo = tableInfo
        self.wb_datemode = wb_datemode
        self.original_name = name   # if name else None

        name = arcpy.ValidateFieldName(name, self.tableInfo.out_path)
        name = arcpy.ValidateFieldName(name, "in_memory")

        if not name:
            try:
                name = 'COL_' + xlrd.colname(col_num)
            except IndexError:
                name = 'COL_'

        if self.tableInfo.is_dbf:
            # dbf can't take non-alpha character to begin field name
            if not name[0].isalpha():
                name = 'f' + name
            name = name[:10]

        names = field_names + ['objectid']

        # generate a unique name (using "_1", "_2", etc...)
        i = 0
        orig_name = name
        while names.count(name.lower()) != 0:
            i += 1
            name = f"{orig_name[:8]}_{i}"
        self.name = name

    def __repr__(self):
        return f'<Field.name={self.name}, alias={self.alias}, ftype={self.ftype}, length={self.length}>'

    def set_field_type(self, sheet, eval_count, check_for_MGRS=False,
                       cell_range=None, field_names_row=None):
        """ Set the output field type based on first eval_count records """

        if eval_count is None or eval_count == 0:
            eval_bottom = cell_range['bottom']
        else:
            eval_bottom = min(cell_range['top'] + eval_count, sheet.nrows)

        # Get list of all types in column (don't count blank 0,5,6)
        cell_types = sheet.col_types(colx=self.col_num,
                                     start_rowx=cell_range.get("top"),
                                     end_rowx=eval_bottom)

        if (field_names_row >= cell_range.get("top") and
            field_names_row <= cell_range.get("bottom")):

            # remove one instance of the field_name_row's type (always a string)
            field_name_row_type = sheet.col_types(colx=self.col_num,
                                                  start_rowx=field_names_row,
                                                  end_rowx=field_names_row + 1)

            # remove a single matching value from the list
            cell_types.remove(field_name_row_type[0])

        # capture if errors exist in the field
        self.has_errors = (5 in cell_types)

        # remove nulls
        cell_types = set(cell_types) - {0, 5, 6}

        if len(cell_types) == 1:
            # if all eval cells are of the same type use that
            # if all are Double, check if cells are styled like integer
            cell_type = cell_types.pop()
            if cell_type == 2:
                try:
                    max_value = 0
                    min_value = 0
                    for row in range(cell_range.get("top"), eval_bottom):
                        if row == field_names_row:
                            continue
                        value = sheet.cell(row, self.col_num).value
                        if value:
                            assert ((value % 1) == 0.0)  # no remainder so Integer
                            max_value = max([value, max_value])
                            min_value = min([value, min_value])

                    if (max_value < INTEGER_MAX) and (min_value >= INTEGER_MAX*-1):
                        cell_type = 4
                    elif (max_value <= BIG_INTEGER_MAX) and (min_value >= BIG_INTEGER_MAX*-1) \
                                                       and self.tableInfo.supportsBigInteger:
                        cell_type = 'bigint'
                    else:
                        cell_type = 2
                except:
                    pass

            if cell_type == 3 and (self.tableInfo.supportsTimeOnly and self.tableInfo.supportsDateOnly):
                # determine if Date is actually TimeOnly or DateOnly

                dateOnly = True
                timeOnly = True
                for row in range(cell_range.get("top"), eval_bottom):
                    if row == field_names_row:
                        continue
                    value = sheet.cell(row, self.col_num).value
                    if value:
                        # turn the Date value from numeric to a tuple of
                        value_as_tuple = xlrd.xldate_as_tuple(value, self.wb_datemode)
                        if value_as_tuple[3:] != (0, 0, 0):
                            dateOnly = False
                        if value_as_tuple[:3] != (0, 0, 0):
                            timeOnly = False

                        if (dateOnly is False) and (timeOnly is False):
                            # no need to continue checking values
                            break

                if dateOnly:
                    cell_type = "dateonly"
                elif timeOnly:
                    cell_type = "timeonly"

        else:
            cell_type = 1

        # if text in a gdb, set field's length based on longest value
        if cell_type == 1  and (not self.tableInfo.is_dbf):
            maxlen = 255

            cLines = 1      # Counter for all values
            cMGRSLines = 1  # Counter for fields that contain a MGRS value

            for row in range(cell_range["top"], eval_bottom):
                cLines +=1
                value = sheet.cell(row, self.col_num).value

                if value and isinstance(value, (str)):
                    maxlen = max(maxlen, len(value) + 10)

                try:
                    if check_for_MGRS and arcgisscripting._sharing._TextFileFieldHelper().IsLatLong(value):
                            cMGRSLines += 1
                except:
                    pass

            # set max len encountered in cells
            self.length = maxlen

            if check_for_MGRS and (cMGRSLines == cLines):
                # all cells examined were MGRS, so set field type to MGRS
                # Note: We have no way to distinguish MGRS from USNG so we
                #       check the column heading

                self.length = 16  # 15 + 1 characters
                if 'USNG' in self.name.upper():
                    cell_type = 'USNG'
                else:
                    cell_type = 'MGRS'

        self._cell_type = cell_type

    @property
    def ftype(self):
        return FIELD_TYPE_MAP[self._cell_type]

    def validate_value(self, value):
        """ Validate the value against the output field """
        if self.ftype == 'Date' and value:
            value = xlrd.xldate_as_tuple(value, self.wb_datemode)

            if value[:3] == (0, 0, 0):
                # time only is no good for datetime.datetime() use
                value = (1899, 12, 30, value[3], value[4], value[5])
            value = datetime(*value)

        # Non-gdb do not support None/Null (except with Date)
        if self.tableInfo.is_dbf:
            if value in [None, '']:
                if self.ftype == 'Text':
                    value = ""
                elif self.ftype == 'Date':
                    value = None
                else:
                    value = 0

        # Cannot set '' into a numeric or date field in a gdb
        elif value == '' and self.ftype in ['Double', 'Long', 'BigInteger',
                                            'Date', 'DateOnly', 'TimeOnly',]:
            value = None

        # Return the validated value
        return value


def get_sheet_names(in_excel):
    """ 
    Returns a list of sheet names for the selected excel file. 
    *This gets called during validation*
    """
    workbook = xlrd.open_workbook(in_excel)
    return [sheet.name for sheet in workbook.sheets()]

def open_excel_table(in_excel, sheet_name):
    """ Open the excel file, return sheet and workbook. """

    try:
        workbook = xlrd.open_workbook(in_excel)
        worksheet = workbook.sheet_by_name(sheet_name)
        return worksheet, workbook

    except Exception as err:
        arcpy.AddError(err)
        raise

def validate_fields(in_excel, out_table, sheet_name, rows=100):
    """ Validates field names, eliminating duplicate names and invalid
        characters. This is only used in the script tool's Validation.
    """
    # Code works, but xlrd.open_workbook can be quite slow (20 sec), don't want
    #  to pay this cost in updateParameter as it's called in tool dialog each
    #  time a parameter value is changed.
    return []

    # workbook = xlrd.open_workbook(in_excel)
    # sheet, workbook = open_excel_table(in_excel, sheet_name)
    # out_path, out_table_name = os.path.split(out_table)
    # is_dbf = arcpy.Describe(out_path).workspaceType == "FileSystem"
    # return [f.name for f in gen_out_fields(workbook,
    #                                        sheet,
    #                                        out_path,
    #                                        is_dbf,
    #                                        rows,
    #                                        check_for_MGRS)]
    # Note: check_for_MGRS is only true it this is called from ExcelAnalyze
    #            from within the server context


def gen_out_fields(sheet, tableInfo=TableInfo(), eval_count=None,
                   check_for_MGRS=False, cell_range=None, field_names_row=0,
                   validate_workspace=None, wb_datemode=0):
    """ Generate the list of output field names based on inputs

        wb_datemode: the xlrd.workbook's datemode
    """

    if tableInfo is None:
        tableInfo = TableInfo()
        if validate_workspace:
            # here for legacy code from ExcelAnalyze.py
            tableInfo.out_path = validate_workspace

    if sheet.nrows == 0:
        return []

    if cell_range is None:
        cell_range = {'left': 0,
                      'right': sheet.ncols,
                      'top': 0,
                      'bottom': sheet.nrows
                      }

    # Set the Field types based on values
    checkNextFieldForMGRS = check_for_MGRS
    out_fields = []
    # Generate the list of output fields
    for i, f in enumerate(sheet.row_values(field_names_row,
                                           cell_range.get("left"),
                                           cell_range.get("right", None))):
        field = Field("" if field_names_row == -1 else f , tableInfo,
                         [i.name.lower() for i in out_fields], i + cell_range.get("left"),
                         wb_datemode=wb_datemode)

        field.set_field_type(sheet, eval_count, checkNextFieldForMGRS,
                             cell_range, field_names_row)

        if field.ftype == 'MGRS' or field.ftype == 'USNG':
            # Once we have found one MGRS column, we stop to look for them
            checkNextFieldForMGRS = False
        out_fields.append(field)

    return out_fields

def col_letter_to_int(s):
    """ converts column header to an integer
        A=0, Z=26, AA=27, CV=100, ...

        Not case sensitive, a=A=0, b=B=1
    """

    s = list(s.upper())
    s.reverse()
    val2 = ((ord(s[2]) - 64) * 26 * 26) if len(s) > 2 else 0
    val1 = ((ord(s[1]) - 64) * 26) if len(s) > 1 else 0
    val0 = ord(s[0]) - 64

    # all all of them then substract 0 (index base)
    return val2 + val1 + val0 - 1

def unpack_cell_range(cell_range):
    # validate then turn cell_range to left, right, top, bottom

    if re.match("^[A-Za-z]{1,3}[0-9]{1,7}:[A-Za-z]{1,3}[0-9]{1,7}$",
                  cell_range) is None:
        raise AssertionError

    # pull the 4 components out of the cell_range
    left = col_letter_to_int(re.search("^[A-Za-z]{1,3}", cell_range)[0])
    right = col_letter_to_int(re.search(r"(?<=\:)[A-Za-z]{1,3}",
                                        cell_range)[0])
    top = int(re.search("[0-9]{1,7}(?=:)", cell_range)[0])
    bottom = int(re.search("[0-9]{1,7}$", cell_range)[0])

    # make sure ranges are valid
    assert left <= right
    assert top <= bottom

    return left, right + 1, top - 1, bottom

def excel_to_table(in_excel, out_table, sheet_name=None, field_names_row=1,
                   cell_range=None, eval_count=None):
    """ Convert an excel sheet to a gdb table, dbf, or info table

        :param in_excel: path to an xls or xlsx file
        :param out_table: path to gdb table or dbf to be created
        :param sheet_name: sheet from input file to be converted - if blank sheet[0] is used
        :param cell_range: cell range to convert (A5:B10) - if blank use all
        :param eval_count: number of input rows to eval to determine field data type - if blank use all
        :param field_names_row: row for field names
        :return: None
    """
    if sheet_name in [None, '', '#']:
        sheet_name = get_sheet_names(in_excel)[0]

    if arcpy.Exists(out_table):
        if not arcpy.env.overwriteOutput:
            arcpy.AddIDMessage("ERROR", 258, out_table)
            if __name__ == '__main__':
                print('ERROR ' + arcpy.GetIDMessage(258))
                sys.exit(1)
            else:
                raise arcpy.ExecuteError(arcpy.GetIDMessage(258))

    sheet, workbook = open_excel_table(in_excel, sheet_name)

    if cell_range in [None, '', "#"]:
        cell_range = {'left': 0,
                      'right': sheet.ncols,
                      'top': 0,
                      'bottom': sheet.nrows
                      }
    else:
        try:
            left, right, top, bottom = unpack_cell_range(cell_range)
        except Exception as e:
            # todo: proper error msg
            arcpy.AddIDMessage("ERROR", 2926, "cell_range", str(cell_range))
            if __name__ == '__main__':
                sys.exit(1)
            else:
                msg = arcpy.GetIDMessage(2926).replace("%1", "cell_range"
                                             ).replace("%2", str(cell_range))
                raise arcpy.ExecuteError(msg)

        # turn cell_range into a dictionary for easier use later
        cell_range = {'left': left,
                      'right': min(right, sheet.ncols),
                      'top': top,
                      'bottom': min(bottom, sheet.nrows) }

        # throw error if the right or bottom of cell_range is outside the data
        if (cell_range['left'] >= cell_range['right'] or
            cell_range['top'] > cell_range['bottom'] or
            cell_range['top'] < 0):

            # the cell_range is below or to the left of the data range
            arcpy.AddIDMessage("ERROR", 3301)
            if __name__ == '__main__':
                sys.exit(1)
            else:
                raise arcpy.ExecuteError(arcpy.GetIDMessage(3301))

    tableInfo = TableInfo(out_table)

    # Note: If the check_for_MGRS flag would be set to True, then we would
    # get some field type that we cannot create. Only set this to True to
    # get for evaluating fields, such as MGRS (text) fields

    # user enters row id (1 based in from excel) but xlrd uses index (0 based)
    field_names_row -=1

    out_fields = gen_out_fields(sheet, tableInfo, eval_count,
                                check_for_MGRS=False, cell_range=cell_range,
                                field_names_row=field_names_row,
                                wb_datemode=workbook.datemode)

    if tableInfo.ext == ".csv":
        coll_sep = "," if locale.localeconv()['decimal_point'] == "." else";"
        with open(out_table, 'w', encoding="utf-8") as _outf:
            _outf.write("\ufeff" + str(coll_sep).join([i.name for i in out_fields]) + "\n")

    else:
        # For performance reasons, add the fields to an in_memory table
        tmp_table = os.path.join('in_memory', 'tmp_exceltotable_template')
        if arcpy.Exists(tmp_table):
            arcpy.Delete_management(tmp_table)

        arcpy.CreateTable_management(*os.path.split(tmp_table))

        # Add the fields that were validates using the previous function
        arcpy.AddFields_management(tmp_table,
                                   [(f.name, f.ftype, f.alias, f.length)
                                               for f in out_fields])

        # Now create the actual output table
        out_table = arcpy.CreateTable_management(tableInfo.out_path,
                                                 tableInfo.out_table_name,
                                                 template=tmp_table)[0]

        arcpy.Delete_management(tmp_table)

        # Output info table has trouble with OBJECTID field
        if tableInfo.is_dbf and (tableInfo.ext != '.dbf'):
            arcpy.DeleteField_management(out_table, 'OBJECTID')

    arcpy.SetParameterAsText(1, out_table)

    # If the sheet has no rows, warn and exit
    if sheet.nrows < 2:
        arcpy.AddIDMessage('WARNING', 117)
        return

    # Loop through each row and insert values into the output table
    with arcpy.da.InsertCursor(out_table, [f.name for f in out_fields]) as cursor:
        for rowid in range(cell_range['top'], cell_range['bottom']):
            if rowid == field_names_row: continue
            row = sheet.row_values(rowid, cell_range['left'], cell_range['right'])

            for colid, field in enumerate(out_fields):
                value = row[colid]

                #  cells with value=42 and ctype=5 contain #N/A
                if field.has_errors and value == 42 and  \
                        (sheet.cell(rowid, colid).ctype == 5):
                    value = None

                row[colid] = field.validate_value(value)

            try:
                cursor.insertRow(row)
            except Exception as e:
                print(e)


if __name__ == "__main__":
    # Get the parameters from the script tool
    kwargs = {"in_excel": arcpy.GetParameterAsText(0),
              "out_table": arcpy.GetParameterAsText(1),
              "sheet_name": arcpy.GetParameterAsText(2),
              "field_names_row": arcpy.GetParameter(3),
              "cell_range": arcpy.GetParameterAsText(4), }

    if os.path.splitext(kwargs["in_excel"])[1].upper() == ".XLSX":
        ExcelToTableX.excel_to_table(**kwargs)
    else:
        excel_to_table(**kwargs)
