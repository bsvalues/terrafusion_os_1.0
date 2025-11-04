"""
Tool name: Excel To Table
Source: ExcelToTableX.py
Author: ESRI

Convert a Microsoft Excel xlsx file to an gdb, csv or dbftable.

ExcelToTable vs ExcelToTableX
 - ExcelToTable.py uses xlrd to convert xls files
 - ExcelToTableX.py uses openpyxl to convert xlsx files

"""

import arcpy
import os
import sys
import datetime
import re
import locale
import openpyxl
import arcgisscripting
from functools import cached_property
fieldHelper = arcgisscripting._sharing._TextFileFieldHelper()


INTEGER_MAX = 2**31
BIG_INTEGER_MAX = 2**53

FIELD_TYPE_MAP = {'str': 'Text',
                'float': 'Double',
                'datetime': 'Date',
                'time': 'TimeOnly',
                'dateonly': 'DateOnly',
                'int': 'Long',
                'bigint': 'BigInteger',
                'bool': 'Short',
                'NoneType': "none",
                'USNG': 'USNG',
                'MGRS': 'MGRS'
                }

class TableInfo:
    """ Class containing information about the output table"""
    def __init__(self, out_table=None):
        self.is_dbf = False
        self.is_sde = False
        self.ext = None
        self.supportsBigInteger = True
        self.supportsTimeOnly = True
        self.supportsDateOnly = True
        self.supportsTimestampOffset = True
        self.check_for_MGRS = False
        self.out_path = ""
        self._out_table_name = ""

        if out_table:
            self._out_table_name = os.path.split(out_table)[1]
            self.out_path = os.path.split(out_table)[0]
            try:
                self.out_path_desc = arcpy.Describe(self.out_path)

                self.is_dbf = (self.out_path_desc.workspaceFactoryProgID == ""
                            and os.path.isdir(self.out_path))

                self.is_sde = self.out_path_desc.workspaceFactoryProgID == 'esriDataSourcesGDB.SdeWorkspaceFactory'

                self.ext = os.path.splitext(self._out_table_name)[1].lower()
                self.supportsBigInteger = getattr(self.out_path_desc, 'supportsBigInteger', False)
                self.supportsTimeOnly = getattr(self.out_path_desc, 'supportsTimeOnly', False)
                self.supportsDateOnly = getattr(self.out_path_desc, 'supportsDateOnly', False)
                self.supportsTimestampOffset = getattr(self.out_path_desc, 'supportsTimestampOffset', False)
            except Exception as e:
                print(e)

    @cached_property
    def out_table_name(self):

        out_table_name = self._out_table_name
        if self.is_sde:
            try:
                # pluck 'table' from 'db.owner.table'
                out_table_name = arcpy.ParseTableName(out_table_name, self.out_path).split(", ")[2]
            except:
                pass

        if self.is_dbf:
            if not self.ext in [".dbf",]:
                out_table_name += ".dbf"
        else:
            out_table_name = arcpy.ValidateTableName(out_table_name, self.out_path)

        return out_table_name


class Field:
    """ Class to hold properties and behavior of the output fields """

    def __init__(self, name, tableInfo, field_names=[], col_num=None,
                    check_for_MGRS=False):
        """ Validate name of field based on output table format as well
              as existing/used field names
              tableInfo can be passed as TableInfo()
        """
        self._ftype = None
        self._dtype = set()
        self._maxlen = None
        self._is_lat_long = set()
        self.col_num = col_num
        self.tableInfo = tableInfo
        self.original_name = name   # if name else None

        if name and type(name) is str:
            name = openpyxl.utils.escape.unescape(name)

        self.alias = name if name else None

        name = arcpy.ValidateFieldName("" if name is None else name, self.tableInfo.out_path)
        name = arcpy.ValidateFieldName(name, "in_memory")

        if not name:
            try:
                name = 'COL_' + openpyxl.utils.get_column_letter(col_num)
            except IndexError:
                name = 'COL_'

        if self.tableInfo.is_dbf:
            # dbf can't take non-alpha character to begin field name
            if not name[0].isalpha():
                name = 'f' + name
            name = name[:10]

        names = field_names + ['objectid']

        # generate a unique name (using "_1", "_2", etc...)
        i = 0
        orig_name = name
        while names.count(name.lower()) != 0:
            i += 1
            name = f"{orig_name[:8]}_{i}"
        self.name = name

    @cached_property
    def ftype(self):
        """ determine output field type based on cell info (_dtype) and other stuff """

        cell_types = self._dtype - {'NoneType'}
        if len(cell_types) == 0:
            dtype = 'str'

        elif len(cell_types) == 1:
            # if all cells are same type use that
            dtype = list(cell_types)[0]

        elif (len(cell_types - {'bigint', 'int'}) == 0) and \
                self.tableInfo.supportsBigInteger:
            # if mix of integer and biginteger, go with biginteger
            dtype = 'bigint'

        elif len(cell_types - {'float', 'int', 'bigint'}) == 0:
            # if mix of float and integer, go with float
            dtype = 'float'

        elif len(cell_types - {'dateonly', 'datetime', 'time'}) == 0:
            dtype = 'datetime'

        else:
            # if all cells are mix of types use text
            dtype = 'str'

        if dtype == 'dateonly' and not self.tableInfo.supportsDateOnly:
            # if text in a gdb, set field's length based on longest value
            dtype = 'datetime'

        elif dtype == 'time' and not self.tableInfo.supportsTimeOnly:
            dtype = 'datetime'

        elif dtype == 'str':
            self._maxlen  = max((self._maxlen if self._maxlen else 0) + 10, 255)

            if all(self._is_lat_long) and self.tableInfo.check_for_MGRS:
                self._maxlen = 16    # 15 + 1 characters
                self.tableInfo.check_for_MGRS = False # stop checking for MGRS
                if 'USNG' in self.name.upper():
                    dtype = 'USNG'
                else:
                    dtype = 'MGRS'

        return FIELD_TYPE_MAP[dtype]

    @property
    def length(self):
        return self._maxlen

    @property
    def dtype(self):
        """ the cell's data types """
        return self._dtype

    def collect_cell_info(self, cell):
        """ collect all the collumn's cell info into this class' member variable
             - self._dtype : set(cell.value.__class__.__name__)
             - self._maxlen : the longest str
             - self._is_lat_long : set(fieldHelper.IsLatLong(value))
        """

        value = cell.value
        data_type = value.__class__.__name__

        # check if we can refine the type for later mapping
        if data_type == "int":
            if (value < INTEGER_MAX) and (value >= INTEGER_MAX*-1):
                data_type = 'int'
            elif (value <= BIG_INTEGER_MAX) and (value >= BIG_INTEGER_MAX*-1):
                data_type = 'bigint'
            else:
                # float maybe inadequate for these values
                data_type = 'float'

        # treat #N/A same as None
        elif value == "#N/A":
            data_type = "NoneType"

        # at Pro 3.2 we have a dateonly type for datetime with 0 for hour, minute and second
        elif (data_type == "datetime") and (set([value.hour, value.minute, value.second]) == {0}):
            data_type = "dateonly" # dateOnly

        if data_type == "str":
            # if str, track the length
            self._maxlen = max(self._maxlen if self._maxlen else 0, len(value))
            # string could be MGRS
            if self.tableInfo.check_for_MGRS:
                if not (False in self._is_lat_long):
                    self._is_lat_long.add(fieldHelper.IsLatLong(value))

        # add this cell's data_type to the self._dtype set
        self._dtype.add(data_type)

    def __repr__(self):
        return f'<Field.name={self.name}, alias={self.alias}, ftype={self.ftype}, length={self.length}>'

    def validate_value(self, value):
        """ Validate the value against the output field type """
        # Non-gdb output do not support None/Null (except with Date)
        if self.tableInfo.is_dbf:
            if value in [None, '', "#N/A"]:
                if self.ftype == 'Text':
                    value = ""
                elif self.ftype == 'Date':
                    value = None
                else:
                    value = 0
        else: # gdb
            if value == "#N/A":
                value = None  #value of #N/A is always None
            elif self.ftype in ['Double', 'Date', 'Long'] and value in ['', ]:
                # Cannot set '' into a numeric or date field in a gdb
                value = None
            elif self.ftype == "Text":
                if value == None:
                    value = ""
                elif isinstance(value, (datetime.date, datetime.time, datetime.datetime)):
                    value = str(value)
                elif isinstance(value, str):
                    # deal with xlsx escape charcters
                    value = openpyxl.utils.escape.unescape(value)
        return value

def get_sheet_names(in_excel):
    """ 
    Returns sheet names and named ranges (if present)
        
        Returns:
            sheets_nr: list of sheet names and named ranges
            nr_dict: dictionary of named ranges
            wb.sheetnames: just the sheet names
        *This function gets called during validation*
    """
    wb = openpyxl.load_workbook(in_excel, read_only=True, data_only=True, keep_links=False)
    # get sheet names
    sheets_nr = wb.sheetnames
    nr_dict = {}

    # fetch sheet scoped named ranges
    for ws in wb.worksheets:
        for dn in ws.defined_names.values():
            nr = ws.defined_names.get(dn.name, dn.localSheetId)
            if nr.type == "RANGE":
                split_nr = nr.value.split('!')
                # only add named ranges with a single range
                if len(split_nr) <= 2:
                    # do not add named ranges with only a single cell
                    if split_nr[1].find(':') > 0:
                        sheets_nr.append(dn.name)
                        nr_dict.update({dn.name: dn.value})
    # fetch global scoped named ranges
    for dn in wb.defined_names.values():
        nr = wb.defined_names.get(dn.name, dn.localSheetId)
        if nr.type == "RANGE":
            split_nr = nr.value.split('!')
            # only add named ranges with a single range
            if len(nr.value.split('!')) <= 2:
                # do not add named ranges with only a single cell
                if split_nr[1].find(':') > 0:
                    sheets_nr.append(dn.name)
                    nr_dict.update({dn.name: dn.value})

    return sheets_nr, nr_dict, wb.sheetnames

def get_named_ranges(in_excel, sheet_name):
    """ 
    Determines cell range of named range and the sheet it belongs to.

    Args:
        in_excel (string): input excel file
        sheet_name (string): input sheet or named range

    Returns:
        sheet_name (string): sheet that named range belongs to
        cell_range (string): cell range of input named range
        None (None): returned if named range not found
    """
    # get named ranges
    nr_dict = get_sheet_names(in_excel)[1]
    # get the value associated with the named range
    nr_value = nr_dict.get(sheet_name)
    cell_range = ""

    if nr_value is not None:
        split_nr = nr_value.split('!')
        if len(split_nr) <= 2:
            if split_nr[1].find(':') > 0:
                sheet_name = split_nr[0]
                # remove single quotes (') from sheet name
                if sheet_name[0] == "'" and sheet_name[-1] == "'":
                    sheet_name = sheet_name[1:-1]
                cell_range = split_nr[1].replace('$', '')
            else:
                # change to appropriate error message
                arcpy.AddIDMessage("ERROR", 3958)
                # named range with single cell not allowed
                raise arcpy.ExecuteError(arcpy.GetIDMessage(3958))
            
        # else:
            # possible implementation in the future that supports
            # multiple ranges
            
    # theoretically this will never happen because users are not allowed
    # to enter sheets/named ranges that don't exist
    # else:
        # throw error message - can't find named range      
    return sheet_name, cell_range 

def open_excel_table(in_excel, sheet_name):
    """ Open the excel file, return sheet and workbook. """

    try:
        workbook = openpyxl.load_workbook(in_excel, read_only=True, data_only=True, keep_links=False)
        worksheet = workbook[sheet_name]
        
        # Read-only mode relies on applications and libraries that created the
        # file providing correct information about the worksheets, specifically
        # the used part of it, known as the dimensions. Some applications set
        # this incorrectly. Resetting will correct this.
        if worksheet.max_column == worksheet.max_row == 1:
            worksheet.reset_dimensions()
        return worksheet, workbook

    except Exception as err:
        arcpy.AddError(err)
        raise

def validate_fields(in_excel, out_table, sheet_name, rows=100):
    """ Validates field names, eliminating duplicate names and invalid
        characters. This is only used in the script tool's Validation.
    """
    # Code works, but open_workbook can be quite slow (20 sec), don't want
    #  to pay this cost in updateParameter as it's called in tool dialog each
    #  time a parameter value is changed.
    return []

def gen_out_fields(sheet, tableInfo=None, eval_count=None,
                   check_for_MGRS=False, cell_range=None,
                   field_names_row=1, validate_workspace=None):
    """ Generate the list of output field names based on inputs """

    if tableInfo is None:
        tableInfo = TableInfo()
        tableInfo.check_for_MGRS = check_for_MGRS
        if validate_workspace:
            # here for legacy code from ExcelAnalyzeX.py
            tableInfo.out_path = validate_workspace

    if cell_range is None:
        cell_range = gen_cell_range(None, sheet)

    if sheet.max_row == 0:
        return []

    # Set the Field types based on values
    out_fields = []

    # Generate the list of output fields
    # need loop since can be independent of  data range
    for row in sheet.iter_rows(max(field_names_row, 1),
                               max(field_names_row, 1),
                               cell_range.get("left"),
                               cell_range.get("right")):

        for i, value in enumerate([j.value for j in row]):
            name = "" if field_names_row == 0 else value
            out_fields.append(Field(name, tableInfo,
                                    [i.name.lower() for i in out_fields],
                                    i + (cell_range.get("left") or 1)) )

    # Get list of all types in column
    rowid = None
    for row_counter, row in enumerate(
            sheet.iter_rows(cell_range.get("top"), cell_range.get("bottom"),
                            cell_range['left'], cell_range['right'])):

        if eval_count is not None:
            if row_counter >= eval_count:
                break

        if field_names_row is not None:
            # not all cells have row property, so loop until find one
            for cell in row:
                rowid = getattr(cell, 'row', None)
                if rowid:
                    break

        for i, cell in enumerate(row):
            if rowid == field_names_row:
                field_names_row = None # set to none, so stop looking for it
                break # break out of the cells loop

            # track all the column cell's data types
            out_fields[i].collect_cell_info(cell)

    return out_fields

def col_letter_to_int(s):
    """ converts column header to an integer
        A=0, Z=26, AA=27, CV=100, ...

        Not case sensitive, a=A=0, b=B=1
    """
    s = list(s.upper())
    s.reverse()
    val2 = ((ord(s[2]) - 64) * 26 * 26) if len(s) > 2 else 0
    val1 = ((ord(s[1]) - 64) * 26) if len(s) > 1 else 0
    val0 = ord(s[0]) - 64

    # all all of them then substract 0 (index base)
    return val2 + val1 + val0 - 1

def unpack_cell_range(cell_range):
    """ insure cell_range is valid and turn it into
        left, right, top, bottom

        So B10:E50 becomes left: 2, right: 5, top: 10, bottom: 50
    """

    if re.match("^[A-Za-z]{1,3}[0-9]{1,7}:[A-Za-z]{1,3}[0-9]{1,7}$",
                  cell_range) is None:
        raise AssertionError

    # pull the 4 components out of the cell_range
    left = col_letter_to_int(re.search("^[A-Za-z]{1,3}", cell_range)[0])
    right = col_letter_to_int(re.search(r"(?<=\:)[A-Za-z]{1,3}", cell_range)[0])

    top = int(re.search("[0-9]{1,7}(?=:)", cell_range)[0])
    bottom = int(re.search("[0-9]{1,7}$", cell_range)[0])

    # make sure ranges are valid
    assert left <= right
    assert top <= bottom

    return left +1, right +1, top, bottom

def gen_cell_range(cell_range, sheet):
    """ create a dict of the left, right, top, bottom of the cell_range """

    if cell_range in [None, '', "#"]:
        cell_range = {'left': None,
                      'right': None,
                      'top': None,
                      'bottom': None}
    else:
        try:
            left, right, top, bottom = unpack_cell_range(cell_range)
        except Exception as e:

            arcpy.AddIDMessage("ERROR", 2926, "cell_range", str(cell_range))
            if __name__ == '__main__':
                sys.exit(1)
            else:
                msg = arcpy.GetIDMessage(2926).replace("%1", "cell_range"
                                             ).replace("%2", str(cell_range))
                raise arcpy.ExecuteError(msg)

        # turn cell_range into a dictionary for easier use later
        cell_range = {'left': left,
                      'right': min(right, sheet.max_column),
                      'top': top,
                      'bottom': min(bottom, sheet.max_row) }

        # throw error if the right or bottom of cell_range is outside the data
        if ((cell_range['left'] > cell_range['right']) or
             cell_range['top'] > cell_range['bottom'] or
             cell_range['top'] < 0):

            # the cell_range is below or to the left of the data range
            arcpy.AddIDMessage("ERROR", 3301)
            if __name__ == '__main__':
                sys.exit(1)
            else:
                raise arcpy.ExecuteError(arcpy.GetIDMessage(3301))

    return cell_range

def excel_to_table(in_excel, out_table, sheet_name=None, field_names_row=1,
                   cell_range=None, eval_count=None):
    """
    Convert an excel sheet to a gdb table, dbf, or info table

    :param in_excel: path to an xls or xlsx file
    :param out_table: path to gdb table or dbf to be created
    :param sheet_name: sheet from input file to be converted - if blank sheet[0] is used
    :param cell_range: cell range to convert (A5:B10) - if blank use all
    :param eval_count: number of input rows to eval to determine field data type - if blank use all
    :param field_names_row: row for field names
    :return: None
    """
    # get sheets
    sheets = get_sheet_names(in_excel)[2]
    
    # if user didn't provide sheet input, select first sheet
    if sheet_name in [None, '', '#']:
        sheet_name = sheets[0]
    # get cell range and sheet of the NR
    elif sheet_name not in sheets:
        sheet_name, temp_cr = get_named_ranges(in_excel, sheet_name)
        cell_range = temp_cr if temp_cr not in [None, '', '#'] else cell_range
        if cell_range in [None, '', '#']:
            # change to appropriate error message
            arcpy.AddIDMessage("ERROR", 3958)
            # named range with single cell not allowed
            raise arcpy.ExecuteError(arcpy.GetIDMessage(3958))
        # get the min and max rows of the named range
        min_max_temp = re.findall(r'\d+', cell_range)
        min = list(map(int, min_max_temp))[0]
        max = list(map(int, min_max_temp))[1]
        # set to min if input is outside of named range
        if field_names_row > max or field_names_row < min:
            field_names_row = min
    
    if arcpy.Exists(out_table):
        if not arcpy.env.overwriteOutput:
            arcpy.AddIDMessage("ERROR", 258, out_table)
            if __name__ == '__main__':
                print('ERROR ' + arcpy.GetIDMessage(258))
                sys.exit(1)
            else:
                raise arcpy.ExecuteError(arcpy.GetIDMessage(258))

    sheet, workbook = open_excel_table(in_excel, sheet_name)
    cell_range = gen_cell_range(cell_range, sheet)
    #out_path, out_table_name = os.path.split(out_table)

    tableInfo = TableInfo(out_table)

    # Note: If the check_for_MGRS flag wo out_path_ws_type == "FileSysuld be set to True, then we would
    # get some field type that we cannot create. Only set this to True to
    # get for evaluating fields, such as MGRS (text) fields
    out_fields = gen_out_fields(sheet, tableInfo, eval_count,
                                check_for_MGRS=False, cell_range=cell_range,
                                field_names_row=field_names_row,  )

    if tableInfo.ext == ".csv":
        coll_sep = "," if locale.localeconv()['decimal_point'] == "." else";"
        with open(out_table, 'w', encoding="utf-8") as _outf:
            _outf.write("\ufeff" + str(coll_sep).join([i.name for i in out_fields]) + "\n")
    else:
        # For performance reasons, add the fields to an in_memory table
        tmp_table = os.path.join('in_memory', 'tmp_exceltotable_template')
        if arcpy.Exists(tmp_table):
            arcpy.Delete_management(tmp_table)

        arcpy.CreateTable_management(*os.path.split(tmp_table))

        # Add the fields that were validates using the previous function
        arcpy.AddFields_management(tmp_table,
                                   [(f.name, f.ftype, f.alias, f.length)
                                               for f in out_fields])

        # Now create the actual output table
        out_table = arcpy.CreateTable_management(tableInfo.out_path,
                                                 tableInfo.out_table_name,
                                                 template=tmp_table)[0]

        arcpy.Delete_management(tmp_table)

        # Output info table has trouble with OBJECTID field
        if tableInfo.is_dbf and (tableInfo.ext != '.dbf'):
            arcpy.DeleteField_management(out_table, 'OBJECTID')

    arcpy.SetParameterAsText(1, out_table)

    # If the sheet has no rows, warn and exit
    #if sheet.max_row < 2:
    #    arcpy.AddIDMessage('WARNING', 117)
    #    return

    # Loop through each row and insert values into the output table
    with arcpy.da.InsertCursor(out_table, [f.name for f in out_fields]) as cursor:
        field_names_row_id = field_names_row - (cell_range['top'] or 1)
        blank_rows = []

        for rowid, rowvals in enumerate(sheet.iter_rows(cell_range['top'],
                                                        cell_range['bottom'],
                                                        cell_range['left'],
                                                        cell_range['right'],
                                                        values_only=True)):

            if rowid == field_names_row_id:
                continue

            row = list(rowvals)

            for colid, field in enumerate(out_fields):
                try:
                    row[colid] = field.validate_value(row[colid], )
                except IndexError:
                    row.append(field.validate_value(None))

            # collect blank rows
            if not any(row):
                blank_rows.append(row)
                continue

            # if blank rows were not at end, then add them to the output
            if blank_rows:
                for row2 in blank_rows:
                    oid = cursor.insertRow(row2)
                blank_rows = []

            # add row to the output table
            oid = cursor.insertRow(row)