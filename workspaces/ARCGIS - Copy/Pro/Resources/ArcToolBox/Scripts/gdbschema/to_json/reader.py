import abc
from typing import TYPE_CHECKING, Type, overload, Literal, Optional

from .._logging import get_logger

if TYPE_CHECKING:
    from ..conversion.helper import Base
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.worksheet.table import Table
    from openpyxl.cell.cell import Cell

logger = get_logger(__name__)

__all__ = [
    "ExcelWorkbook",
    "ExcelWorksheet",
    "ExcelTable",
]


class Excel:
    def __repr__(self):
        return f"<{self.__class__.__name__} {self.name!r}>"

    @property
    @abc.abstractmethod
    def name(self) -> str:
        """The name of the object"""


class ExcelTable(Excel):
    def __init__(self, sheet: "ExcelWorksheet", table: "Table"):
        self.ws = sheet
        self.t = table

    @property
    def name(self) -> str:
        return self.t.name

    @overload
    def data(self, key_value: Literal[False] = False) -> tuple[dict, ...]:
        ...

    @overload
    def data(self, key_value: Literal[True]) -> dict:
        ...

    @staticmethod
    def _extract_value(cell: "Cell"):
        """Extracts value from cell"""
        val = cell.value

        if hyper := cell.hyperlink:
            # openpyxl reads empty cells with hyperlinks as having data == hyperlink.
            if hyper.location == val:
                return

        return val

    def data(self, key_value: bool = False):
        """The data in the table"""
        header, *data = (tuple(map(self._extract_value, row)) for row in self.ws.get_cells(self.t.ref))
        # Excel doesn't support header-only tables, so if there is a single row, the table might really be empty.
        if len(data) == 1:
            if set(data[0]) == {None}:
                data = []
        if key_value:
            return {k: v for k, v in data}
        return tuple(dict(zip(header, row)) for row in data)


class ExcelWorksheet(Excel):
    def __init__(self, book: "ExcelWorkbook", sheet: "Worksheet"):
        self.workbook: ExcelWorkbook = book
        self.ws = sheet
        self.tables: list[ExcelTable] = [ExcelTable(self, table) for table in self.ws.tables.values()]

    @property
    def name(self) -> str:
        return self.ws.title

    def get_cells(self, ref: str) -> tuple[tuple["Cell", ...], ...]:
        return self.ws[ref]

    def get_table(
        self,
        *table_types: Type["Base"],
        properties: bool = False,
        extension_properties: bool = False,
        controller_membership: bool = False,
    ) -> Optional[ExcelTable]:
        if properties:
            prefix = "properties_"
            if table_types:
                prefix = f"{table_types[0].class_name_lower()}_{prefix}"
        elif extension_properties:
            prefix = "extensionproperties_"
        elif controller_membership:
            prefix = "controllermembership_"
        else:
            prefix = tuple(f"{t.class_name_lower()}_" for t in table_types)
        for table in self.tables:
            if table.name.casefold().startswith(prefix):
                return table


class ExcelWorkbook:
    def __init__(self, file):
        self.file = file
        self.wb = self._load_workbook(file)
        self.sheets: list[ExcelWorksheet] = [ExcelWorksheet(self, ws) for ws in self.wb]
        self.name_lookup: dict[str, ExcelWorksheet] = {s.name.casefold(): s for s in self.sheets}

    @staticmethod
    def _load_workbook(path: str) -> "Workbook":
        import io
        import openpyxl
        import warnings

        # Load the file into memory, so we don't hold file locks and performance is optimal.
        with open(path, "rb") as f:
            stream = io.BytesIO(f.read())

        # Suppress DataValidation warnings from openpyxl.
        with logger.timing(message="Loading Workbook"), warnings.catch_warnings(action="ignore", category=UserWarning):
            return openpyxl.load_workbook(stream, data_only=True, keep_vba=False)

    def get_sheets_by_type(self, *sheet_types: Type["Base"]) -> list[ExcelWorksheet]:
        import re

        prefix = "|".join(s.SHEET_PREFIX for s in sheet_types)
        pat = re.compile(rf"({prefix})\d+_", re.IGNORECASE)
        return [sheet for sheet in self.sheets if pat.match(sheet.name) is not None]
