"""
COPYRIGHT 2018 ESRI

TRADE SECRETS: ESRI PROPRIETARY AND CONFIDENTIAL
Unpublished material - all rights reserved under the
Copyright Laws of the United States.

For additional information, contact:
Environmental Systems Research Institute, Inc.
Attn: Contracts Dept
380 New York Street
Redlands, California, USA 92373

email: contracts@esri.com

---------------------------------------------------------------------------
Source Name:   ImportGeneralizationData.py
Version:       ArcGIS 2.3
Author:        Environmental Systems Research Institute Inc.
Description:   Imports production data into a theme geodatabase.
---------------------------------------------------------------------------
"""

import openpyxl
import arcpy
import os
import sys
import TopographicGeneralizationUtilities as utils
import DefenseUtilities as defense
from TopographicGeneralizationUtilities import CaseStr, CaseDict, CaseSet, ToolException

cstr = utils.CaseStr

def get_tolerance_row(sheet, name):
    """
    Returns a row in the Tolerances sheet given the name of the tolerance
    """
    if not isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet):
        raise Exception('Invalid type {}. Expecting type openpyxl.worksheet.worksheet.Worksheet'.format(type(sheet)))

    tolerance_row = None
    # make sure the sheet has the exepcted columns
    common_columns = ['TOLERANCE NAME', 'VALUE', 'UNITS', 'TOLERANCE TYPE', 'OPERATION', 'MODELS']
    column_indices = utils.CheckColumns(sheet, common_columns)

    for row_index in range(1, sheet.max_row):
        tolname = sheet.cell(row_index, column_indices['TOLERANCE NAME']).value
        if tolname != "" or tolname is not None:
            if tolname.lower() == name.lower():
                tolerance_row = utils.CaseDict()
                tolerance_row['TOLERANCE NAME'] = tolname
                tolerance_row['Value'] = sheet.cell(row_index, column_indices['Value']).value
                tolerance_row['Units'] = sheet.cell(row_index, column_indices['Units']).value
                tolerance_row['TOLERANCE TYPE'] = sheet.cell(row_index, column_indices['TOLERANCE TYPE']).value
                tolerance_row['OPERATION'] = sheet.cell(row_index, column_indices['OPERATION']).value
                tolerance_row['MODELS'] = sheet.cell(row_index, column_indices['MODELS']).value
                break

    return tolerance_row

def check_license():
    # Check out licenses
    if defense.licenselevel() == 'Basic' or defense.licenselevel() == 'None':
        raise defense.LicenseException("This tool requires at least the Standard License to run this tool.")
    if defense.licenselevel == 'Server':
        if arcpy.CheckExtension('defense') != 'Available' or arcpy.CheckExtension('foundation') != 'Available':
            defense.LicenseException('Tool requires either Defense or Foundation extension to run.')
    else:
        if 'Available' == arcpy.CheckExtension('defense'):
            defense.checkoutextensions(['defense'])
        elif 'Available' == arcpy.CheckExtension('Foundation'):
            defense.checkoutextensions(['Foundation'])
        else:
            raise defense.LicenseException('Tool requires either defense or Foundation extension to run.')

def populate_theme_fields(input_values, field_mapping):
    """
    Copies the values from a source field to target fields.
    :param row_values: values stored as case-insensitive dictionary
    :param field_mapping: mapping of fields
    :return: updated values as a dictionary
    """
    new_values = CaseDict(input_values, ':_- ')
    for key, target_columns in field_mapping.items():
        for column_name in target_columns:
            new_values[column_name] = new_values.get(key)
    return new_values

def import_simplify_data(input_workspace, output_workspace, rule_file, theme):
    # Get sheet from workbook
    rule_book = utils.OpenWorkbook(rule_file)

    rule_sheet = utils.CheckSheets(rule_book, 'Rules', 0)
    tol_sheet = utils.CheckSheets(rule_book, "Tolerances", 1)

    # make sure the sheet has the exepcted columns
    common_columns = ['GEOMETRY TYPE','SIMPLIFY: OBJECTCLASS',
    'SIMPLIFY: SIMPLIFY TOLERANCE', 'SIMPLIFY: SMOOTH TOLERANCE']
    column_indices = utils.CheckColumns(rule_sheet, common_columns)

    #keep track of the simplify table records
    simplify_records = utils.CaseDict()
    simplify_count = 0
    tolerances = utils.CaseSet()

    # for each row - if the row is part of Simplify theme, create a record to be inserted into table
    for row_index in range(1, rule_sheet.max_row):

        simplify_type = rule_sheet.cell(row_index, column_indices['SIMPLIFY: OBJECTCLASS']).value
        if simplify_type is not None:
            if simplify_type.lower() == "individual" or simplify_type.lower() == "shared":
                  operation_type = None
                  geometry_type = rule_sheet.cell(row_index, column_indices['GEOMETRY TYPE']).value
                  tolerance_name = rule_sheet.cell(row_index, column_indices['SIMPLIFY: SIMPLIFY TOLERANCE']).value
                  if tolerance_name != "" and tolerance_name is not None and tolerance_name not in tolerances:
                    operation_type = "Simplify"
                    tolrow = get_tolerance_row(tol_sheet, tolerance_name)
                    if tolrow is not None:
                      tolrow['PROCESSINGTYPE'] = simplify_type
                      tolrow['GEOMETRYTYPE'] = geometry_type
                      tolrow['OPERATIONTYPE'] = operation_type
                      tolrow['TOLERANCETEXT'] = '{} {}'.format(tolrow['VALUE'], tolrow['UNITS'])
                      simplify_records[simplify_count] = tolrow
                      tolerances.add(tolerance_name)
                      simplify_count += 1

                  tolerance_name = rule_sheet.cell(row_index, column_indices['SIMPLIFY: SMOOTH TOLERANCE']).value
                  if tolerance_name != "" and tolerance_name is not None and tolerance_name not in tolerances:
                    operation_type = "Smooth"
                    tolrow = get_tolerance_row(tol_sheet, tolerance_name)
                    if tolrow != None:
                      tolrow['PROCESSINGTYPE'] = simplify_type
                      tolrow['GEOMETRYTYPE'] = geometry_type
                      tolrow['OPERATIONTYPE'] = operation_type
                      tolrow['TOLERANCETEXT'] = '{} {}'.format(tolrow['VALUE'], tolrow['UNITS'])
                      simplify_records[simplify_count] = tolrow
                      tolerances.add(tolerance_name)
                      simplify_count += 1


    #create Simplify_Tolerances table
    table_name = "Simplify_Tolerances"
    tolerance_fields = [('ToleranceName', 'TEXT'), ('ProcessingType','TEXT'), ('GeometryType','TEXT'),
        ('Operation','TEXT'), ('OperationType','TEXT'), ('ToleranceValue','FLOAT'), ('ToleranceUnits', 'TEXT'),
        ('Tolerance', 'TEXT')]
    tolerance_table = os.path.join(output_workspace, table_name)
    if arcpy.Exists(tolerance_table):
        #remove duplicates
        tol_names = ""
        for i in range(simplify_count):
            if tol_names == "":
                tol_names = "'{}'".format(simplify_records[i]['Tolerance Name'])
            else:
                tol_names += ", '{}'".format(simplify_records[i]['Tolerance Name'])

        where_clause = ""
        if tol_names == "":
            #if no tolerances for theme delete all records
            oidField = arcpy.Describe(tolerance_table).OIDFieldName
            where_claue = "{} >= 0".format(oidField)
        else:
            where_clause = 'ToleranceName IN ({})'.format(tol_names)

        table_view = arcpy.MakeTableView_management(tolerance_table, "ToleranceView", where_clause)
        if table_view != None:
            arcpy.DeleteRows_management(table_view)
    else:
        arcpy.AddIDMessage('INFORMATIVE', 86180)  # Creating output table...
        result = arcpy.CreateTable_management(output_workspace, table_name)
        tolerance_table = result[0]
        for field in tolerance_fields:
            utils.CheckField(tolerance_table, field[0], field[1])

    #add records into Simplify_Tolerances table
    arcpy.AddIDMessage('INFORMATIVE', 86174)  # Writing output...

    simplify_list = []
    for i in range(simplify_count):
        simplify_list.append(simplify_records[i])

    #sort by tolerance value
    simplify_list = sorted(simplify_list, key=lambda simplify_rec: simplify_rec['Value'])

    fields = []
    for field in tolerance_fields:
        fields.append(field[0])

    with arcpy.da.InsertCursor(tolerance_table, fields) as cursor:
        for rec in simplify_list:
            cursor.insertRow((rec['TOLERANCE NAME'],rec['PROCESSINGTYPE'],rec['GEOMETRYTYPE'],
            rec['OPERATION'],rec['OPERATIONTYPE'],rec['VALUE'],rec['UNITS'], rec['TOLERANCETEXT']))


def add_theme_fields(feature_class, theme_fields):
    table_fields = [field.name.upper() for field in arcpy.ListFields(feature_class)]
    missing_fields = [name for name in theme_fields if name.upper() not in table_fields]
    new_fields = [[name, 'TEXT', name, 255] for name in missing_fields if not utils.is_value_empty(name)]
    arcpy.AddFields_management(feature_class, new_fields)

def import_generalization_data(input_workspace, output_workspace, rule_file, theme):

    rule_file_name = os.path.splitext(os.path.basename(rule_file))[0]

##    arcpy.AddIDMessage('INFORMATIVE', 86071, rule_file_name)  # %s is loading...

    # get feature classes from input workspace
    input_fc_dict = utils.GetFeatureClasses(input_workspace)

    spatial_ref = arcpy.Describe(input_fc_dict.values()[0]).spatialReference


    ''' --------------------------------------------------
    Get Rule info from spreadsheet for theme
    --------------------------------------------------'''
    # Get sheet from workbook
    common_columns = ['OBJECTCLASS', 'SUBTYPENUMBER', 'SUBTYPEFIELD', 'SUBSET QUERY', 'CONVERT OBJECTCLASS',
                      'THEMES', 'GEOMETRY TYPE', 'CONVERT SUBTYPE', 'CONVERT MAPPING']
    try:
        themeRules = utils.query_excel_rows(rule_file, 'Rules', common_columns, theme)

    except ToolException as e:
        arcpy.AddIDMessage("ERROR", e.id, e.args[1], e.args[2])
        sys.exit(1)
    except Exception as e:
        arcpy.AddIDMessage("ERROR", 10158, rule_file_name)  # Unable to open file %1.
        sys.exit(1)

    if not themeRules:
        arcpy.AddIDMessage("ERROR", 2719) # No attribute rules were found.
        sys.exit(1)

    # Populate theme fields
    #
    theme_table_field = f'{theme}:ObjectClass'
    theme_fcname_field = f'{theme.upper()}_FCNAME'

    all_column_names = themeRules[0].keys()
    if theme_table_field not in all_column_names:
        arcpy.AddIDMessage('ERROR', 413, theme_table_field, rule_file_name)
        sys.exit(1)

    theme_fields = [column_name.replace(':','_') for column_name in all_column_names if cstr(column_name).startswith(f'{theme}:')]
    theme_fields += [ theme_fcname_field, 'CONVERT_OBJECTCLASS', 'CONVERT_SUBTYPE', 'CONVERT_MAPPING']

    # determine the field that contains the theme feature class name
    query_field = None
    for field_name in theme_fields:
        if cstr(field_name,':_ ') == cstr(theme_table_field,':_ '):
            query_field = field_name



    # mapping of internal theme fields to rule file columns
    theme_field_mapping = { 'CONVERT OBJECTCLASS': ['CONVERT_OBJECTCLASS'],
                            'CONVERT SUBTYPE'    : ['CONVERT_SUBTYPE'],
                            'CONVERT MAPPING'    : ['CONVERT_MAPPING'],
                            'OBJECTCLASS'        : [theme_fcname_field]}



    copy_fc_dict = CaseDict()

    out_fc_dict = CaseDict()
    out_type_dict = CaseDict()

    clean_fc_set = CaseSet()


    ''' --------------------------------------------------
    Get Theme feature classes info from spreadsheet
    --------------------------------------------------'''

    # try to get info including new updated by theme column
    common_columns = ['THEME', 'OBJECTCLASS', 'FEATURE TYPE', 'SCHEMA CLASSES', 'UPDATED BY THEME']
    try:
        themeObjectClasses = utils.query_excel_rows(rule_file, 'ThemeObjectClasses', common_columns)
    except:
        try:
            # if column doesn't exist, get the rest of the info and assumen value is yes
            common_columns = ['THEME', 'OBJECTCLASS', 'FEATURE TYPE', 'SCHEMA CLASSES']
            themeObjectClasses = utils.query_excel_rows(rule_file, 'ThemeObjectClasses', common_columns)

        except ToolException as e:
            arcpy.AddIDMessage("ERROR", e.id, e.args[1], e.args[2])
            sys.exit(1)
        except Exception as e:
            arcpy.AddIDMessage("ERROR", 10158, rule_file_name)  # Unable to open file %1.
            sys.exit(1)

    theme_info_dict = {}
    for themeClass in themeObjectClasses:
        fc_theme = themeClass['THEME']

        # Check if object class belongs to this theme
        if fc_theme.upper() != theme.upper():
            continue

        fc_name = themeClass['OBJECTCLASS']
        fc_name = fc_name.strip()
        geometry_type = themeClass['FEATURE TYPE']

        out_fc_name = None
        create_type = None

        # determine the output feature class name for the row
        if geometry_type.upper() == 'POLYGON' or geometry_type.upper() == 'POLY':
            out_fc_name = f"{theme}_{fc_name}_A"
        elif geometry_type.upper() == 'LINE' or geometry_type.upper() == 'POLYLINE':
            out_fc_name = f"{theme}_{fc_name}_L"
        elif geometry_type.upper() == 'POINT' or geometry_type.upper() == 'PT':
            out_fc_name = f"{theme}_{fc_name}_P"

        if out_fc_name:
            theme_info_dict[out_fc_name.upper()] = themeClass




    # for each row - if the row is part of the theme, get the feature class and add and populate the required fields
    #
    for rule in themeRules:
##        arcpy.AddMessage('Rule {}'.format(rule))
        # calculate theme field values
        rule = populate_theme_fields(rule, theme_field_mapping)

        # find the theme feature class name
        input_fc_name = rule['OBJECTCLASS']
        if input_fc_name is not None:
            input_fc_name = input_fc_name.strip()
            # Issue 2998 - cast to upper to match foreign characters
            input_fc_name = input_fc_name.upper()
        theme_type = rule[theme_table_field]

        theme_obj = rule[theme_table_field]

        if utils.is_value_empty(theme_type):
            continue # no theme feature class specified for input feature class



        if input_fc_name not in input_fc_dict:
##            arcpy.AddWarning('Cannot find {} in {}'.format(input_fc_name, input_workspace))
            arcpy.AddIDMessage('WARNING', 413, input_fc_name, input_workspace)  # %1 not found in %2
            continue # feature class not found in input workspace

        input_fc_path = input_fc_dict[input_fc_name]

        geometry_type = arcpy.Describe(input_fc_path).shapeType


##        arcpy.AddMessage('input_fc_name {}'.format(input_fc_name))
##        arcpy.AddMessage('copy_fc_dict {}'.format(copy_fc_dict))

        # --------------------------------------------------
        # Create a temporary fc by exploding and add theme fields
        # --------------------------------------------------
        if input_fc_name not in copy_fc_dict:

##            arcpy.AddMessage('Create Temp Feature')
            new_fc_path = f'in_memory\\{input_fc_name}'

            # if the feature class is multipoint, write to disk
            if geometry_type.upper() == 'MULTIPOINT':
                scratch = arcpy.env.scratchGDB
                new_fc_path  = f'{scratch}\\{input_fc_name}_temp'
                clean_fc_set.add(new_fc_path )

            # explode features to single part and add theme fields
            arcpy.RepairGeometry_management(input_fc_path)
            input_fc_path = arcpy.MultipartToSinglepart_management(input_fc_path, new_fc_path )
            arcpy.RepairGeometry_management(input_fc_path)

            input_fc_path = new_fc_path
            copy_fc_path = input_fc_path
            add_theme_fields(copy_fc_path, theme_fields)
            copy_fc_dict[input_fc_name] = copy_fc_path

        # if the fields have been added, get the path to the fc with the fields
        else:
            copy_fc_path = copy_fc_dict[input_fc_name]


        # --------------------------------------------------
        # Determine target Theme FC
        # --------------------------------------------------

        # determine the geometry type to generate the name of the output
        # feature class - <theme>_<theme object class column value>_<A/L/P>
        out_fc_type = theme_type.replace(' ', '')


        out_fc_name = None
        if geometry_type.upper() == 'POLYGON' or geometry_type.upper() == 'POLY':
            out_fc_name = f'{theme}_{out_fc_type}_A'
        elif geometry_type.upper() == 'LINE' or geometry_type.upper() == 'POLYLINE':
            out_fc_name = f'{theme}_{out_fc_type}_L'
        elif geometry_type.upper() == 'POINT':
            out_fc_name = f'{theme}_{out_fc_type}_P'
        elif geometry_type.upper() == 'MULTIPOINT':
            out_fc_name = f'{theme}_{out_fc_type}_P'

        if not out_fc_name:
            arcpy.AddIDMessage('WARNING', 294, input_fc_name)
            continue

        # create a dictionary out_fc_dict
        # keys: The theme feature class name
        # values: list of all input feature classes that will need to be merged to create the feature class
        out_fc_name = out_fc_name.upper()
        if out_fc_name in out_fc_dict:
            out_fc_dict[out_fc_name].add(copy_fc_path)
        else:
            out_fc_dict[out_fc_name] = CaseSet([copy_fc_path])
            out_type_dict[out_fc_name] = theme_type




        # --------------------------------------------------
        # Determine subset of features to apply rule to
        # and update the theme fields with value from rule
        # --------------------------------------------------

        # determine the subset of features to apply the rule
        subtype_field = rule['SUBTYPEFIELD']
        subtype_value = rule['SUBTYPENUMBER']
        subset_query  = rule['SUBSET QUERY']

        where_clause = ''
        if subtype_field and subtype_value is not None:
            where_clause = f"{subtype_field} = {int(subtype_value)}"

        # Combine subtype and subset queries. If no subtypes, use only the subset query.
        if subset_query:
            where_clause = f"{where_clause} AND ({subset_query})" if where_clause else subset_query


        # populate the spreadsheet value on the features in the input feature class that
        # meets the subset query
        try:
            update_count = 0


            with arcpy.da.UpdateCursor(copy_fc_path, '*', where_clause) as cursor:
                # get the list of field for the fc
                fields_list = cursor.fields

                # update each record with the theme information
                for row in cursor:
                    # Update theme fields using rule values
                    for index, field_name in enumerate(fields_list):
                        if field_name in rule:
                            val = rule[field_name]
                            if val is None:
                                val = ''
                            row[index] = val

                    cursor.updateRow(row)
                    update_count += 1

        except Exception as e:
            arcpy.AddIDMessage('ERROR', 729, where_clause, copy_fc_path) # %1: Value is not valid. %2
            sys.exit(1)


        if update_count >= 1:
            input_fc_name = os.path.basename(input_fc_path) + ('' if not where_clause else f" ({where_clause})")
            arcpy.AddIDMessage('INFORMATIVE', 86089, input_fc_name)  # ...Importing entities from %s

    # --------------------------------------------------
    # Extra logic for simplify theme
    # --------------------------------------------------
    if theme.lower() == "simplify":
        import_simplify_data(input_workspace, output_workspace, rule_file, theme)

    ''' --------------------------------------------------
    Delete any rows that do not have the theme: Object Class populated
    --------------------------------------------------'''



    copy_fc_list = copy_fc_dict.values()

##    arcpy.AddMessage('query field {}'.format(query_field))

    del_cnt = 0
    row_cnt = 0

    for copy_fc in copy_fc_list:
        with arcpy.da.UpdateCursor(copy_fc, [query_field]) as cursor:
            for row in cursor:
                row_cnt += 1
                if row[0] is None or row[0] in ['', ' ']:
                    cursor.deleteRow()
                    del_cnt += 1

##    arcpy.AddMessage('start {} deleted {}'.format(row_cnt, del_cnt))





##    arcpy.AddMessage('theme info {}'.format(theme_info_dict))

    # --------------------------------------------------
    # Loop through each theme feature class
    # and create feature class
    # --------------------------------------------------

    created_fc_list = CaseSet()
    if len(out_fc_dict) >= 1:
        for out_fc_name, copy_fc_list in out_fc_dict.items():

            updated = 'YES'
            geo_type = None
##            arcpy.AddMessage(out_fc_name)

##            for theme_fc_name in theme_info_dict.keys():

            if out_fc_name.upper().strip() in theme_info_dict:
##                arcpy.AddMessage('in theme dict')
                themeClass = theme_info_dict[out_fc_name.upper().strip()]
                if 'UPDATED BY THEME' in themeClass:
                    updated = themeClass['UPDATED BY THEME']
                else:
                    updated = 'YES'
                geo_type = themeClass['FEATURETYPE'].upper()

##            arcpy.AddMessage('updated by theme: {}'.format(updated))

            # ---------------------------
            # select only the features from the input fc where the theme: objectclass is the target theme fc
            # -----------------------------
##            theme_oc = out_fc_name.split('_')[1]
##            theme_oc = theme_oc.title()
            theme_oc = out_type_dict[out_fc_name]
            where = f"{query_field} = \'{theme_oc}\'"

            merge_lyrs = []
            copy_fc_list = list(set(copy_fc_list))
            for copy_fc in copy_fc_list:
                lyr = arcpy.SelectLayerByAttribute_management(copy_fc, 'NEW_SELECTION', where)
                merge_lyrs.append(lyr)
##                arcpy.AddMessage('{} - {} Selected cnt {}'.format(copy_fc, where, arcpy.GetCount_management(lyr)[0]))
            spatial_ref = arcpy.Describe(copy_fc_list[0]).spatialReference

            out_fc_path = os.path.join(output_workspace, out_fc_name)

            if updated.upper() == 'YES':
                # ---------------------------
                # If the theme fc is updated in the models, create by merging inputs to get all attributes
                # -----------------------------

                arcpy.AddIDMessage('INFORMATIVE', 86078, out_fc_name, output_workspace)  # Adding %1 to %2...
                merge_fc = arcpy.Merge_management(merge_lyrs, out_fc_path)
                created_fc_list.add(out_fc_name)
            else:

                # ---------------------------
                # If the theme fc is not updated in the models, create create empty feature class and append to only have theme attributes
                # -----------------------------
                arcpy.AddIDMessage('INFORMATIVE', 86078, out_fc_name, output_workspace)  # Adding %1 to %2...

                if geo_type == 'LINE':
                    geo_type = 'POLYLINE'
                arcpy.CreateFeatureclass_management(output_workspace, out_fc_name, geo_type, spatial_reference=spatial_ref)
                add_theme_fields(out_fc_path, theme_fields)
                arcpy.Append_management(merge_lyrs, out_fc_path, 'NO_TEST')
                created_fc_list.add(out_fc_name)


    ''' --------------------------------------------------
    Create remaining Theme feature classes

    these are fcs with no rules define and will get schema from themeobjectclasses tab
    --------------------------------------------------'''
    for out_fc_name, themeClass in theme_info_dict.items():

        # if the output feature class was not created by merging features, create an empty feature class using
        # the feature class in the Schema Classes column as a template for the attributes.
        if out_fc_name and out_fc_name not in created_fc_list:

            out_fc_name = out_fc_name.upper()
            arcpy.AddIDMessage('INFORMATIVE', 86078, out_fc_name, output_workspace)  # Adding %1 to %2...

            schema_fcs = themeClass['SCHEMA CLASSES']
            schema_list = CaseSet([fc.strip() for fc in schema_fcs.split(';')])

            schema_paths = ''
            for schema_fc in schema_list:
                if schema_fc in input_fc_dict:
                    input_fc_path = input_fc_dict[schema_fc]
                    schema_paths += f"{'; ' if schema_paths else ''}{input_fc_path}"
                    spatial_ref = arcpy.Describe(input_fc_path).spatialReference

            out_fc_path = os.path.join(output_workspace, out_fc_name)
            if arcpy.Exists(out_fc_path):
                arcpy.Delete_management(out_fc_path)

            if 'UPDATED BY THEME' in themeClass:
                updated = themeClass['UPDATED BY THEME']
            else:
                updated = 'YES'

            if updated.upper() == 'YES':

                # ---------------------------
                # If the theme fc is updated in the models, create by feature class using schema fc as template
                # -----------------------------
##                arcpy.AddMessage('update {} - add with schema fields'.format(updated))
                new_feature_class = arcpy.CreateFeatureclass_management(output_workspace, out_fc_name, create_type, schema_paths, spatial_reference=spatial_ref)
            else:

                # ---------------------------
                # If the theme fc is not updated in the models, create by feature class with no template
                # -----------------------------
##                arcpy.AddMessage('update {} - add with no schema fields'.format(updated))
                geo_type = themeClass['FEATURETYPE'].upper()
                if geo_type == 'LINE':
                    geo_type = 'POLYLINE'

                new_feature_class = arcpy.CreateFeatureclass_management(output_workspace, out_fc_name, geo_type, spatial_reference=spatial_ref)

            # ---------------------------
            # Add theme fields to feature class
            # -----------------------------

            add_theme_fields(new_feature_class, theme_fields)

    # Clean up feature classes and theme fields
    #
    arcpy.AddIDMessage('INFORMATIVE', 86386) # Cleaning up

##    # Delete the theme fields that were added to the input feature classes
##    if not use_memory_workspace and theme_fields:
##        for input_fc_path in copy_fc_dict.values():
##            arcpy.DeleteField_management(input_fc_path, theme_fields)
##
    # Cleanup feature classes in scratch workspace
    for temp_fc in clean_fc_set:
        if arcpy.Exists(temp_fc):
            arcpy.Delete_management(temp_fc)

    return


def main():

    input_workspace  = arcpy.GetParameterAsText(0)
    output_workspace = arcpy.GetParameterAsText(1)
    rule_file        = arcpy.GetParameterAsText(2)
    theme            = arcpy.GetParameterAsText(3)   # Theme types: SOE, TRANS, STRUCTURE, HYDRO, GENERAL, ELEVATION

##    try:
    check_license()

    import_generalization_data(input_workspace, output_workspace, rule_file, theme)

##    except Exception as e:
##        arcpy.AddIDMessage('ERROR', 86195, str(e))
##        sys.exit(1)

if __name__ == '__main__':
    main()


