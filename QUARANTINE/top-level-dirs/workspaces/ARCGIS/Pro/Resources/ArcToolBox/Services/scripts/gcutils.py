#__author__ = 'vict7669'
from tkinter import N
import arcpy
import json
import os
import sys
import time
import csv
from queue import *
import concurrent.futures
import time
import gc
import zipfile
import time
import codecs
import xlrd
import stat
import copy
import arcgisscripting
import logging
import re
import shutil
import ssl
import traceback
import socket
import requests
from shutil import copyfile
import openpyxl
# import httplib
# import pandas
# import signal
# import threading
# from threading import Thread
# import io
# from collections import OrderedDict
# from urlparse import urlparse
# from operator import itemgetter
# import numpy


LOG_LEVEL = 'USER'
#LOG_LEVEL = 'INFO'


def log_error_call_stack():
    '''Adds the callstack from the exception as GP messages'''
    msgs = traceback.format_exception(*sys.exc_info())[1:]
    for msg in msgs:
        arcpy.AddMessage(msg.strip())


def give_portion_of_recordset(recordset,start,end):
    ''' Helper function used in rematch logic
        its sole purpose is to take in a recordset,
        a start and an end and return a smaller
        recordset'''
    recordset_json = json.loads(recordset)
    list_of_records = recordset_json['records']
    portion = []
    for i in range(start,end):
        portion.append(list_of_records[i])

    portion_final = {"records":portion}

    return portion_final


def convert_to_int_if_possible(value):
    ''' This function helps when parsing
        REST inputs and outputs '''
    try:
        returned_int = int(value)
        return returned_int
    except ValueError:
        return value


def rounded_convert_to_int_if_possible(value):
    """ Same as before, but rounded nicely"""
    try:
        returned_int = round(float(value),6)
        return returned_int
    except Exception:
        return value


def convert_list_to_ascii(list):
    new_ascii_list = []
    for unicode_val in list:
        new_ascii_list.append(unicode_val.encode('ascii','ignore'))
    return new_ascii_list


def token_check(url):
    # If both service description and geocodeAddresses calls don't give a 499 error,
    # then service is completely unsecure. No need to get server token in that case
    service_secure = False
    if not url.startswith("http"):
        return "error"
    else:
        try:
            #arcpy.AddMessage(url)
            r = requests.post(url, data={"f":"json"}, verify=False)
            #arcpy.AddMessage(r)
            json_response = r.json()
            #arcpy.AddMessage(json_response)
            r.raise_for_status()
        except requests.exceptions.HTTPError as err:
            return "error"
        except Exception as err:
            return "error"

        if "error" in json_response:
            error_code = json_response["error"]["code"]
            message = json_response["error"]["message"]
            if error_code == 499 or "permissions" in message:
                service_secure = True
            else:
                return "error"
        else:
            try:
                recordset = {
                    "records": []
                    }
                values = {
                        "addresses": json.dumps(recordset),
                        "category": "",
                        "sourceCountry": "",
                        "outSR": "",
                        "f": "json"
                    }
                req = requests.post(url + "/geocodeAddresses", data=values, verify=False, timeout=1000)
                ga_json_response = req.json()
                #arcpy.AddMessage(ga_json_response)
                req.raise_for_status()
            except requests.exceptions.HTTPError as err:
                #arcpy.AddMessage(err)
                return "error"

            if "error" in ga_json_response:
                error_code = ga_json_response["error"]["code"]
                if error_code == 499:
                    service_secure = True
                else:
                    #arcpy.AddMessage(ga_json_response)
                    return "error"
        return service_secure


def token_check_big_data(url):
    service_secure = False
    try:
        r = requests.post(url, data={'f':'json'}, verify=False, timeout=60)
        table_info_response = r.json()
        r.raise_for_status()
    except requests.exceptions.HTTPError as err:
        return "error"
    if "error" in table_info_response:
        if table_info_response["error"]["code"] == 499:
            service_secure = True
    return service_secure


def unicode_csv_reader(utf8_data, dialect=csv.excel, **kwargs):
    csv_reader = csv.reader(utf8_data, dialect=dialect, **kwargs)
    for row in csv_reader:
        #arcpy.AddMessage('msg')
        ## Strings are automatically unicode in Python3
        yield [cell for cell in row]


class Logger():
    ''' This class is used to provide debug information
        to the user. Global LOG_LEVEL should be set to
        either USER or INFO, USER displays only gp
        messages that an end-user should see, INFO is
        all debugging checkpoints
    '''
    def __init__(self,logger_level):
        self.level = logger_level

    def log(self,message,message_level):
        if message_level == 'ERROR':
            arcpy.AddError(str(message))
        if message_level == 'INFO' and self.level == 'INFO':
            arcpy.AddMessage(str(message))
        if message_level == 'USER':
            arcpy.AddMessage(str(message))
        if message_level == 'USER_ERROR':
            arcpy.AddError(str(message))
        return


# Helper class to deal with reading feature services
class LayerInfo(object):
    '''Store describe object and other useful information about feature layers'''

    def __init__(self, dataset):
        #If the dataset is None return LayerInfo with empty properties
        self.count = 0
        self.describe = None
        self.name = ""
        self.layer = None
        self.isLayer = True
        if dataset:
            #if dataset is a hosted layer, then we already have layer, layername and count
            #and we describe the layer derived from the hosted layer
            if hasattr(dataset, "count") and not isinstance(dataset, str):
                self.name = dataset.layername
                self.count = dataset.count
                self.layer = dataset.name
                if self.count:
                    self.describe = arcpy.Describe(self.layer)
            else:
                if arcpy.Exists(dataset):
                    self.count = int(arcpy.management.GetCount(dataset).getOutput(0))
                    self.describe = arcpy.Describe(dataset)
                    if self.describe.dataType == "FeatureLayer":
                        self.layer = dataset
                        self.name = self.describe.namestring
                    elif self.describe.dataType == "Table":
                        self.name = self.describe.name
                        self.isLayer = False
                    else:
                        self.name = self.describe.name
                        arcpy.management.MakeFeatureLayer(dataset, self.name)
                        self.layer = self.name


class Analyze():
    """The analyze class gets called twice
        1. By the analyze tool
        2. By the batch geocode tool
        They have different constructors (dictionary passed-in inputs)
        based on which tool is calling the class"""

    ERROR_CODES = {
        100162: "Analyze Geocode Input failed.",
        100163: "Your text qualifier is invalid.",
        100164: "You must enter either inputTable, inputFileItem or columnNames. All three cannot be empty.",
        100165: "If your header row does not exist, you must enter columnNames.",
        100166: "If your fileType is table, then you must enter an inputTable.",
        100167: "CSV Collection is not a supported input type. Please input a CSV or Microsoft Excel spreadsheet.",
        100168: "Your filename or extension is too long.",
        100169: "Your zip file cannot be opened.",
        100170: "There is no file of the specified format in your zip file.",
        100171: "Invalid locator service URL: {service_url}.",
        100172: "Invalid credentials or invalid locator service URL: {service_url}. Make sure proxy items are shared with Everyone.",
        100174: "The geocoding service must be on a federated server.",
        100175: "Invalid token.",
        100176: "Your file must be UTF-8 encoded. Please resubmit your file.",
        100177: "Your Excel file is corrupted and cannot be opened.",
        100178: "Your file's first row is empty. Please edit your file and resubmit.",
        100179: "You must define column widths if your file is fixed width.",
        100180: "Your column widths must be integers separated by commas like so: 5,7,13.",
        100181: "The file's header row could not be parsed or is empty.",
        100182: "Field {field} could map to 2 possible alias fields {field_one} and {field_two}.",
        100202: "Could not make a REST request from the server machine verify that the server machine can connect to the internet.",
        100203: "The geocodeServiceURL entered should be https instead of http.",
        100204: "The inputTable has over 100,000 records and is not supported",
        100205: "Your column delimiter is invalid.",
        100230: "Your file type is not supported. Supported file types are csv and xlsx.",
        100237: "This input table is invalid. Please select another table."
    }

    def __init__(self,**kwargs):
        kwargs = dict(**kwargs)
        # Error reporting

        if kwargs['mode'] == 'analyze':
            # For debugging purposes...
            lg = Logger(LOG_LEVEL)
            self.logger = lg
            if 'error_func' in kwargs:
                self.errorFunc = kwargs['error_func']
            else:
                self.errorFunc = arcpy.AddError
            if 'warning_func' in kwargs:
                self.errorFunc = kwargs['warning_func']
            else:
                self.warningFunc = arcpy.AddWarning
            if 'debug' in kwargs:
                self.DEBUG = kwargs['debug']
            else:
                self.DEBUG = False

            # This file is not necessarily a zip! We'll check though!
            self.in_file = kwargs['in_file_zip']
            self.service_url = kwargs['service_url']
            self.in_file_type = kwargs['file_type']
            self.portal_file_type = kwargs['portal_file_type']

            self.big_data = kwargs['big_data']

            if self.big_data:
                self.in_table_big_data = kwargs['table_input']
                self.in_table = None
            else:
                self.in_table = LayerInfo(kwargs['table_input'])
                self.in_table_big_data = None

            self.column_delimiter_orig = kwargs['delimiter']

            self.token = kwargs['token']
            self.referer = kwargs['referer']            

            supported_file_types = ['csv','xlsx','xls','table','gdb','txt','EGDB_table']

            if self.in_file_type not in supported_file_types:
                msg_code = 100230
                msg = self.ERROR_CODES[msg_code]
                self.addError(msg_code, msg)

            delim_dict = {'SPACE': ' ','TAB': '\t','COMMA': ',','PIPE': '|','SEMICOLON': ';'}
            if self.in_file_type == 'csv' or self.in_file[-3:] == 'csv':
                if self.column_delimiter_orig == '':
                    # If none entered, we assume the default of ,
                    self.column_delimiter = ','
                else:
                    try:
                        self.column_delimiter = delim_dict[self.column_delimiter_orig]
                    except KeyError:
                        msg_code = 100205
                        msg = self.ERROR_CODES[msg_code]
                        self.addError(msg_code, msg)
            else:
                self.column_delimiter = ''

            self.locale = kwargs['locale']

            qualifier = kwargs['qualifier']
            self.text_qualifier_orig = qualifier
            qualifier_dict = {'SINGLE_QUOTE': '\'','DOUBLE_QUOTE': '\"'}
            if qualifier != '':
                try:
                    self.text_qualifier = qualifier_dict[qualifier]
                except KeyError:
                    msg_code = 100163
                    msg = self.ERROR_CODES[msg_code]
                    self.addError(msg_code, msg)
            else:
                self.text_qualifier = ''

            if kwargs['header_row_exists'] == 'false' or kwargs['header_row_exists'] == '':
                self.header_row_exists = False
            else:
                self.header_row_exists = True
            if kwargs['fixed_width_file'] == 'false' or kwargs['fixed_width_file'] == '':
                self.fixed_width = False
            else:
                self.fixed_width = True

            self.col_widths = kwargs['widths_of_columns']
            self.chars_per_row = kwargs['chars_per_row']
            self.column_names = kwargs['col_names']

            # To be filled in by self.execute()
            self.unzipped_file = ''
            self.available_service_fields = []
            self.http_referrer = ''
            self.output_json = ''
            self.unzipped_folder = ''

            if not self.in_file and not self.column_names and not self.in_table and not self.in_table_big_data:
                msg_code = 100164
                msg = self.ERROR_CODES[msg_code]
                self.addError(msg_code, msg)

            if not self.header_row_exists and self.column_names == '':
                msg_code = 100165
                msg = self.ERROR_CODES[msg_code]
                self.addError(msg_code, msg)


            self.execute_no_field_mapping()
        else:
            self.in_file = kwargs['in_file_zip']
            self.in_file_type = kwargs['file_type']
            self.in_table_path = kwargs['in_table']
            self.in_table = LayerInfo(kwargs['in_table'])
            self.service_url = kwargs['service_url']
            self.column_delimiter = kwargs['delimiter']
            self.text_qualifier = kwargs['qualifier']

            self.token = kwargs['token']
            self.referer = kwargs['referer']

            if 'error_func' in kwargs:
                self.errorFunc = kwargs['error_func']
            else:
                self.errorFunc = arcpy.AddError
            if 'warning_func' in kwargs:
                self.errorFunc = kwargs['warning_func']
            else:
                self.warningFunc = arcpy.AddWarning
            if 'debug' in kwargs:
                self.DEBUG = kwargs['debug']
            else:
                self.DEBUG = False


            supported_file_types = ['csv','xlsx','xls','table','gdb','txt', 'EGDB_table']

            if self.in_file_type not in supported_file_types:
                msg_code = 100230
                msg = self.ERROR_CODES[msg_code]
                self.addError(msg_code, msg)

            # self.column_delimiter_orig = self.column_delimiter
            delim_dict = {'SPACE': ' ','TAB': '\t','COMMA': ',','PIPE': '|','SEMICOLON': ';'}
            if self.in_file_type == 'csv' or self.in_file[-3:] == 'csv':
                if self.column_delimiter == '':
                    # If none entered, we assume the default of ,
                    self.column_delimiter = ','
                else:
                    orig = self.column_delimiter
                    try:
                        self.column_delimiter = delim_dict[orig]
                    except KeyError:
                        msg_code = 100205
                        msg = self.ERROR_CODES[msg_code]
                        self.addError(msg_code, msg)

            qualifier = self.text_qualifier
            qualifier_dict = {'SINGLE_QUOTE': '\'','DOUBLE_QUOTE': '\"'}
            if qualifier != '':
                try:
                    self.text_qualifier = qualifier_dict[qualifier]
                except KeyError:
                    msg_code = 100163
                    msg = self.ERROR_CODES[msg_code]
                    self.addError(msg_code, msg)
            else:
                self.text_qualifier = ''

            self.portal_file_type = kwargs['portal_file_type']
            if kwargs['header_row_exists'] == 'false':
                self.header_row_exists = False
            else:
                self.header_row_exists = True
            if kwargs['fixed_width_file'] == 'false' or kwargs['fixed_width_file'] == '':
                self.fixed_width = False
            else:
                self.fixed_width = True
            self.col_widths = kwargs['widths_of_columns']
            self.chars_per_row = kwargs['chars_per_row']
            self.column_names = kwargs['col_names']
            # if zipfile.is_zipfile(self.in_file):
            #     self.is_zip = True
            # else:
            #     self.is_zip = False
            # Extra parameters
            #if self.in_file_type == 'csv':
            #    if self.column_delimiter == 'tab':
            #        self.column_delimiter = '\t'
            #    elif self.column_delimiter == 'space':
            #        self.column_delimiter = ' '

            # For debugging purposes...
            lg = Logger(LOG_LEVEL)
            self.logger = lg

            # To be filled in by self.execute()
            self.portal_self = kwargs['portal_self']
            self.unzipped_file = ''
            self.available_service_fields = []
            self.http_referrer = ''
            self.output_json = ''
            self.folder_to_remove = ''
            self.execute_fields_already_mapped()

    def addError(self, msg_code, msg, msg_params=None, is_warning=False):
        '''Adds the error message and raises arcpy.ExecuteError exception. self.errorFunc function is used to add the
        error message. if error function is arcpy.AddError only add the msg.'''

        error_func_name = self.errorFunc.__name__
        error_func = self.errorFunc
        if is_warning:
            error_func_name = self.warningFunc.__name__
            error_func = self.warningFunc

        if hasattr(arcpy, error_func_name):
            error_func(msg.replace("$", ""))
        else:
            error_func(msg_code, msg, msg_params, is_warning)

        raise arcpy.ExecuteError

    def execute_no_field_mapping(self):
        ''' Method gets called when Analyze is called from the Analyze
            GP script. It generates a field mapping and returns a
            JSON of relevant file information and the new generated
            field mapping. It is up to the user to take this JSON
            and change the field mapping as they deem fit before
            submiting the JSON (Analyze_obj) to the batchGeocode GP script'''
        was_originally_table = False
        #arcpy.AddMessage(self.big_data)
        if self.big_data:
            # try:
            #     r = requests.post(self.in_table_big_data, data={'f':'json', 'referer':self.referer}, verify=False, timeout=60)
            #     table_info_response = r.json()
            #     r.raise_for_status()
            # except requests.exceptions.HTTPError as err:
            #     msg_code = 100237
            #     msg = self.ERROR_CODES[msg_code]
            #     self.addError(msg_code, msg)
            # if "error" in table_info_response:
            #     if table_info_response["error"]["code"] == 499:
            #         #arcpy.AddMessage('gets to the passing the token version')
            #         #arcpy.AddMessage('token is {}'.format(self.token))
            #         #retry REST call with a token
            #         try:
            #             r = requests.post(self.in_table_big_data, data={'f':'json', 'token':self.token, 'referer':self.referer}, verify=False, timeout=60)
            #             table_info_response = r.json()
            #             r.raise_for_status()
            #         except requests.exceptions.HTTPError as err:
            #             msg_code = 100237
            #             msg = self.ERROR_CODES[msg_code]
            #             self.addError(msg_code, msg)
            #     else:
            #         msg_code = 100237
            #         msg = self.ERROR_CODES[msg_code]
            #         self.addError(msg_code, msg)
            #arcpy.AddMessage(table_info_response)
            desc = arcpy.Describe(self.in_table_big_data)
            header_row = [field.name for field in desc.fields]
            #arcpy.AddMessage('row is {}'.format(header_row))
            # table_fields = table_info_response["fields"]
            # header_row = [field["name"] for field in table_fields]           
            
        elif self.column_names:
            # Then we don't need to crack open the file!!!
            header_row = self.column_names.split(',')
            ### Don't need lines below since Python3 strings are unicode
            ### by default
            # ascii_header_row = []
            # for col_header in header_row:
            #     ascii_header_row.append(col_header.encode('ascii','ignore'))
            # header_row = ascii_header_row
        elif self.in_file_type == 'table' and not self.big_data:
            if not self.in_table:
                msg_code = 100166
                msg = self.ERROR_CODES[msg_code]
                self.addError(msg_code, msg)

            was_originally_table = True
            # Run the arcpy conversion and turn it into a csv... process as normal
            scratch_folder = arcpy.env.scratchFolder
            # Generate the field mapping to add an ObjectID Field because TableToTable
            # does not do it quite correctly
            intermediate_csv = os.path.join(scratch_folder, 'intermediate.csv')
            # self.oid_field = self.in_table.describe.oidFieldName
            try:
                self.list_of_fields = [fld for fld in self.in_table.describe.fields]
            except:
                msg_code = 100166
                msg = self.ERROR_CODES[msg_code]
                self.addError(msg_code, msg)
            # fieldmappings = arcpy.FieldMappings()
            header_row = []
            for field in self.list_of_fields:
                # don't copy over shape!
                if field.name != 'shape':
                    header_row.append(field.name)

            #arcpy.AddMessage('Header row for the table is {}'.format(header_row))

            # self.in_file = self.in_table.layer
        else:
            header_row = []
            # Open up the zip file
            if self.portal_file_type == "CSV Collection":
                # try:
                #     zip_ref = zipfile.ZipFile(self.in_file, 'r')
                # except Exception as e:
                msg_code = 100167
                msg = self.ERROR_CODES[msg_code]
                self.addError(msg_code, msg)
            else:
                # then it is a regular unzipped csv/xlsx file
                in_file = self.in_file


        # Let's check if the service is on a federated server. If it is, then we need
        # to inherit the username and password in a special way

        index_of_REST = self.service_url.lower().find('rest')
        if index_of_REST == -1:
            self.logger.log('Your service URL did not have a valid schema page ' + \
                 'please check your service URL: Your geocode service must be ' + \
                               'through REST, your geocodeServiceURL ' + \
                               'does not have REST or rest in it.','USER_ERROR')



        self.logger.log('token is {}'.format(self.token),'INFO')

        possible_service_fields = []

        # Populate the available fields from the service dynamically
        # Get the fields that are available to be mapped to by the service
        # This way, we can check the mapping provided by the user and make
        # sure the fields they are trying to map to exist for the service
        # that they are using

        # This line below gets rid of the 'geocodeAddresses' part of the URL
        self.ssl_issues = False
        almost_schema_page = self.service_url[:-17]
        try:
            r = requests.post(almost_schema_page, data={'f':'json', 'referer':self.referer}, verify=True, timeout=20)
            self.json_info_response = r.json()
            r.raise_for_status()
        except requests.exceptions.HTTPError as err:
            msg_dict = dict(service_url=self.service_url)
            msg_code = 100171
            msg = self.ERROR_CODES[msg_code].format(**msg_dict)
            self.addError(msg_code, msg, msg_dict)
        except requests.exceptions.SSLError as err:
            self.ssl_issues = True
        # values = ([('token',self.token),('f','json')])
        if self.ssl_issues:
            try:
                r = requests.post(almost_schema_page, data={'f':'json', 'referer':self.referer}, verify=False, timeout=20)
                self.json_info_response = r.json()
                r.raise_for_status()
            except requests.exceptions.HTTPError as err:
                msg_dict = dict(service_url=self.service_url)
                msg_code = 100171
                msg = self.ERROR_CODES[msg_code].format(**msg_dict)
                self.addError(msg_code, msg, msg_dict)

        # Now that we have the URL for the info page, let's discover
        # the schema page for the service using the info on the infopage
        # self.service_info = json_info_response


        # Now we are going to generate a field mapping. The general approach
        # is take in the header fields provided in the file and try to match them
        # to address fields available from the service via string comparison. If
        # there are recognizedNames in the locale provided, then we will check
        # against those possible aliases as well


        try:
            #arcpy.AddMessage("self.json_info_response without the token is: {}".format(self.json_info_response))
            addr_fields = self.json_info_response['addressFields']
            for addr in addr_fields:
                # Add each field to possible_service_fields
                possible_service_fields.append(addr['name'])
        except KeyError as e:
            # If the service returns a 499, then we need a token
            if self.json_info_response['error']['code'] == 499 or self.json_info_response['error']['code'] == 403:
                try:
                    #arcpy.AddMessage("making a req to {0} with {1} token".format(almost_schema_page,self.token))
                    r = requests.post(almost_schema_page, data={'f':'json','token':self.token, 'referer':self.referer}, verify=False, timeout=20)
                    self.json_info_response = r.json()
                    r.raise_for_status()
                except requests.exceptions.HTTPError as err:
                    msg_dict = dict(service_url=self.service_url)
                    msg_code = 100171
                    msg = self.ERROR_CODES[msg_code].format(**msg_dict)
                    self.addError(msg_code, msg, msg_dict)

                try:
                    #arcpy.AddMessage("self.json_info_response with the token is: {}".format(self.json_info_response))
                    addr_fields = self.json_info_response['addressFields']
                    for addr in addr_fields:
                        # Add each field to possible_service_fields
                        possible_service_fields.append(addr['name'])

                except KeyError as e:
                    msg_dict = dict(service_url=self.service_url)
                    msg_code = 100172
                    msg = self.ERROR_CODES[msg_code].format(**msg_dict)
                    self.addError(msg_code, msg, msg_dict)

            else:
                msg_dict = dict(service_url=self.service_url)
                msg_code = 100172
                msg = self.ERROR_CODES[msg_code].format(**msg_dict)
                self.addError(msg_code, msg, msg_dict)

            # if json_info_response['error']['code'] == 498:


        try:
            singleLine_field = self.json_info_response['singleLineAddressField']['name']
            possible_service_fields.append(singleLine_field)
            possible_service_fields.append('OBJECTID')
        except KeyError:
            msg_dict = dict(service_url=self.service_url)
            msg_code = 100171
            msg = self.ERROR_CODES[msg_code].format(**msg_dict)
            self.addError(msg_code, msg, msg_dict)

        # The two try-except statements above populate possible_service_fields
        # which is an important parameter to pass on.
        # self.logger.log('POSSIBLE_ARCGIS_FIELDS ARE {}'.\
        #                 format(possible_service_fields),'INFO')

        self.available_service_fields = possible_service_fields

        # Now we diverge based on whether or not we have a
        # recognizedNames section in the service description
        try:
            # Just pick the first one to check if
            # recognized_names exists in the service
            dummy_var = addr_fields[0]['recognizedNames']
            recognized_names_exists = True
            recognized_names_dict = {}
            if self.locale == '':
                # If no locale was specified, try en!
                self.locale = 'en'
        except KeyError:
            # There are no recognized_names. We will proceed accordingly
            recognized_names_exists = False
            pass
        if not header_row:
            # Then we need to get the header row

            # Now lets get the header row from the file
            # IMPORTANT TO NOTICE: self.in_file is a zip file whereas in_file
            # is a csv/txt/xls file
            if self.in_file_type == 'csv':
                self.prelim_csv_file = in_file
                self.csv_file_handle = open(self.prelim_csv_file, encoding='utf-8')
                if self.column_delimiter == '' or self.column_delimiter == ',':
                    self.prelim_csvReader = unicode_csv_reader(self.csv_file_handle)
                else:
                    self.prelim_csvReader = unicode_csv_reader(self.csv_file_handle,delimiter=self.column_delimiter)
                try:
                    header_row = []
                    almost_header_row = next(self.prelim_csvReader)
                    # arcpy.AddMessage(almost_header_row)
                    for elem in almost_header_row:
                        header_row.append(elem.strip())
                    # parse_success = True
                    #arcpy.AddMessage(header_row)
                except Exception as e:
                    #arcpy.AddMessage("exception is {}".format(e))
                    self.csv_file_handle.close()
                    msg_code = 100176
                    msg = self.ERROR_CODES[msg_code]
                    self.addError(msg_code, msg)


                # new_header_row = header_row


                # if self.text_qualifier == '\'':
                #     new_header_row_to_use = []
                #     for value in new_header_row:
                #         new_header_row_to_use.extend(re.findall("'.*?'",value))
                #     new_header_row = new_header_row_to_use
                #     header_row_final = []
                #     for val in new_header_row:
                #         header_row_final.append(val.strip('\''))
                #     new_header_row = header_row_final

                # Python's CSV reader might just be able to handle double quotes by default??
                # still investigating honestly.


                # The next 3 lines remove a \xef\xbb\xbf from the start
                # of a csv if necessary. Usually \xef\xbb\xbf are markers
                # that the csv is unicode-encoded
                try:
                    if header_row[0].startswith(str(codecs.BOM_UTF8)):
                        beginning_of_file = header_row[0]
                        header_row[0] = beginning_of_file[3:]
                except UnicodeDecodeError:
                    pass

                self.csv_file_handle.close()
                # header_row = new_header_row

            elif in_file[-4:] == 'xlsx' or self.in_file_type == 'xlsx'\
                    or in_file[-4:] == '.xls':
                header_row = []
                try:
                    #arcpy.AddMessage("in_file is {}".format(in_file))
                    ### it looks like when it downloads it on server the file extension is being dropped...
                    if not in_file.endswith(".xlsx"):
                        try:
                            ### try it as an xlsx file
                            shutil.copy(in_file, in_file + ".xlsx")
                            workbook = openpyxl.load_workbook(in_file + ".xlsx")
                            self.in_worksheet = workbook.active
                            first_row = self.in_worksheet[1]
                        except:
                            shutil.copy(in_file, in_file + ".xls")
                            workbook = xlrd.open_workbook(in_file + ".xls")
                            self.in_worksheet = workbook.sheet_by_index(0)
                            first_row = self.in_worksheet.row(0)
                        
                    else:
                        workbook = openpyxl.load_workbook(in_file)
                        self.in_worksheet = workbook.active
                except Exception as e:
                    msg_code = 100177
                    msg = self.ERROR_CODES[msg_code]
                    self.addError(msg_code, msg)

                for cell in first_row:
                    header_row.append(cell.value.strip())
                first_row_values = header_row
                if self.text_qualifier == '\'':
                    new_header_row_to_use = []
                    for value in first_row_values:
                        new_header_row_to_use.extend(re.findall("'.*?'",value))
                    first_row_values = new_header_row_to_use
                    header_row_final = []
                    for val in first_row_values:
                        header_row_final.append(val.strip('\''))
                    first_row_values = header_row_final

                if self.text_qualifier == "\"":
                    new_header_row_to_use = []
                    for value in first_row_values:
                        new_header_row_to_use.extend(re.findall('".*?"',value))
                    first_row_values = new_header_row_to_use
                    header_row_final = []
                    for val in first_row_values:
                        header_row_final.append(val.strip('\"'))
                    first_row_values = header_row_final
                header_row = first_row_values
            elif self.in_file_type == 'txt':

                f = open(in_file)
                raw_first_line = f.readline()

                if self.fixed_width:
                    if not self.col_widths:
                        f.close()
                        msg_code = 100179
                        msg = self.ERROR_CODES[msg_code]
                        self.addError(msg_code, msg)

                    if self.chars_per_row != '':
                        almost_header_row = raw_first_line[0:int(self.chars_per_row)]
                    else:
                        almost_header_row = raw_first_line
                    counter = 0
                    row = []
                    list_to_iter = self.col_widths.split(',')
                    num_of_chars_error = False
                    for num_of_chars in list_to_iter:
                        try:
                            some_int = int(num_of_chars)
                        except ValueError:
                            num_of_chars_error = True
                            break

                        row.append(almost_header_row[counter:counter+int(num_of_chars)].strip(' '))
                        counter = counter + int(num_of_chars)

                    if num_of_chars_error:
                        f.close()
                        msg_code = 100180
                        msg = self.ERROR_CODES[msg_code]
                        self.addError(msg_code, msg)

                    almost_header_row = row
                # Because tabs are the only delimiter we are supporting
                if not self.fixed_width:
                    almost_header_row = raw_first_line.split('\t')
                if self.text_qualifier == '\'':
                    new_header_row_to_use = []
                    for value in almost_header_row:
                        new_header_row_to_use.extend(re.findall("'.*?'",value))
                    first_row_array = new_header_row_to_use
                    header_row_final = []
                    for val in first_row_array:
                        header_row_final.append(val.strip('\''))
                    almost_header_row = header_row_final


                if self.text_qualifier == "\"":
                    new_header_row_to_use = []
                    for value in almost_header_row:
                        new_header_row_to_use.extend(re.findall('".*?"',value))
                    first_row_array = new_header_row_to_use
                    header_row_final = []
                    for val in first_row_array:
                        header_row_final.append(val.strip('\"'))
                    almost_header_row = header_row_final
                header_row = almost_header_row

                f.close()
                if self.col_widths == '':
                    final_header_row = []
                    for val in header_row:
                        final_header_row.append(val.strip('\n'))
                    header_row = final_header_row

        self.logger.log('Header Row is {}'.format(header_row),'INFO')

        # Now that we have both the header row and the available service fields
        # we should perform string comparisons between the two. If
        # recognized_names exists within the service, we'll compare against
        # those as well.

        # header_row_ascii = []
        # for val in header_row:
        #     header_row_ascii.append(val.encode('ascii','ignore'))
        # header_row = header_row_ascii
        field_info = []
        for fld in header_row:
            field_info.append((fld,'TEXT',255))

        header_row_lower = [x.lower() for x in header_row]
        arcgis_fields_lower = [x.lower() for x in self.available_service_fields]
        self.field_mapping = []

        if recognized_names_exists:
            skip_recognized_names = False
            # Do recognized names stuff
            try:
                # addr_fields = json_info_response['addressFields']
                locale_valid = False
                for addr in addr_fields:
                    # # Add each field to possible_service_fields
                    # possible_service_fields.append(addr['name'])
                    recognized_names = addr['recognizedNames']
                    try:
                        recognized_names_to_iter = recognized_names[self.locale]
                        # If we get one valid locale hit with the user's locale,
                        # then do not discard all of the recognized names
                        locale_valid = True
                    except KeyError as e:
                        # The locale entered by the user was bad. Try its
                        # truncated form instead
                        try:
                            recognized_names_to_iter = recognized_names[self.locale[0:-3]]
                        except KeyError as e:
                            if not locale_valid:
                                skip_recognized_names = True
                            else:
                                skip_recognized_names = False

                    if not skip_recognized_names:
                        self.logger.log('Locale worked, working on aliases','INFO')
                        recognized_names_exists = True
                        recognized_names_dict[addr['name']] = set()
                        for name in recognized_names_to_iter:
                            # Create a dictionary where the keys are
                            # the possible service fields, and the values are
                            # a set of possible aliases.
                            recognized_names_dict[addr['name']].add(name.lower())
                if skip_recognized_names:
                    self.logger.log('The locale you entered' +\
                                    ' is not supported by the' +\
                                    ' geocoding service.','USER')
            except Exception as e:
                if self.DEBUG:
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    self.logger.log('{0} {1} {2}'.format(exc_type, fname, exc_tb.tb_lineno),'ERROR')

            # We have done all of the address fields except single line
            # Let's add that one to recognized_names_dict as well!
            if not skip_recognized_names:
                try:
                    singleline_recognized_names = self.json_info_response['singleLineAddressField']['recognizedNames'][self.locale]
                    recognized_names_dict[self.json_info_response['singleLineAddressField']['name']] = set()
                    for name in singleline_recognized_names:
                            # Create a dictionary where the keys are
                            # the possible service fields, and the values are
                            # a set of possible aliases.
                            recognized_names_dict[self.json_info_response\
                                ['singleLineAddressField']['name']].\
                                add(name.lower())
                except KeyError as e:
                    # The locale entered by the user was bad. Try its
                    # truncated form instead
                    try:
                        singleline_recognized_names = self.json_info_response['singleLineAddressField']['recognizedNames']\
                            [self.locale[0:-3]]
                        for name in singleline_recognized_names:
                            # Create a dictionary where the keys are
                            # the possible service fields, and the values are
                            # a set of possible aliases.
                            recognized_names_dict[self.json_info_response['singleLineAddressField']['name']].\
                                add(name.lower())
                    except KeyError as e:
                        pass
                        # They already know.
                        # self.logger.log('The locale you entered' +\
                        #                 ' is not supported by the' +\
                        #                 ' geocoding service.','USER')

                recognized_names_dict_keys = list(recognized_names_dict.keys())


            service_fields_already_taken = set()
            choose_available_service_fields = copy.copy(self.available_service_fields)

            # First pass through does direct string mappings
            field_counter = 0
            for field in header_row_lower:
                # mapped = False
                current_index = header_row_lower.index(field)
                if field in arcgis_fields_lower:
                    service_index = arcgis_fields_lower.index(field)
                    if field not in service_fields_already_taken:
                        self.field_mapping.append([header_row[current_index],\
                                                   self.available_service_fields[service_index]])
                        service_fields_already_taken.add(field)
                    else:
                        # Unable to map this header field
                        self.field_mapping.append([header_row[field_counter], ''])
                else:
                    # we will do these on the second go-around
                    pass

            # Second pass through maps header fields to fields that
            # are still available to be mapped, represented in
            # choose_available_service_fields
            analyze_error_q = Queue(maxsize=2)
            field_counter = 0
            for field in header_row_lower:
                mapped = False
                if field not in arcgis_fields_lower:
                    if skip_recognized_names:
                        self.field_mapping.append([header_row[field_counter],''])
                    else:
                        # Was not able to find a direct string match, try against
                        # aliases
                        analyze_error_q.queue.clear()
                        for key in recognized_names_dict_keys:
                            if field in recognized_names_dict[key]:
                                self.logger.log("field is {0} and alias is {1}".\
                                                format(field,key),'INFO')
                                if key in choose_available_service_fields:
                                    analyze_error_q.put(key)
                                    # Basically, if the header field matches
                                    # recognized names of 2 possible fields
                                    # that are available for mapping from
                                    # the service, then we need to throw an
                                    # error.
                                    if analyze_error_q.full():
                                        msg_params = dict(field=field,field_one=analyze_error_q.get(),field_two=analyze_error_q.get())
                                        msg_code = 100182
                                        msg = self.ERROR_CODES[msg_code].format(**msg_params)
                                        self.addError(msg_code, msg, msg_params=msg_params)

                        # If there is only one possible alias, then map
                        # it correctly
                        if analyze_error_q._qsize() == 1:
                            arcgis_field = analyze_error_q.get()
                            if arcgis_field.lower() in service_fields_already_taken:
                                mapped = False
                            else:
                                self.field_mapping.append([header_row[field_counter].encode('utf-8'),\
                                                       arcgis_field.encode('utf-8')])
                                mapped = True
                        # If there are 0 alias matches, then map to ''
                        if not mapped:
                            self.field_mapping.append([header_row[field_counter],''])

                field_counter += 1

            # If we used recognized names, we have the potential to produce
            # a field mapping that is out-of-order in terms of columns of the
            # original file. Let's put them back together
            # field_mapping_unordered = copy.copy(self.field_mapping)
            # field_mapping_dict = {}
            # for lst in field_mapping_unordered:
            #     field_mapping_dict[lst[0]] = lst[1]
            # field_mapping_ordered = []
            # for field in header_row:
            #     field_mapping_ordered.append([unicode(field,"utf-8"),field_mapping_dict[field.encode('utf-8')]])
            # self.field_mapping = field_mapping_ordered


        else:
            # No recognized names means that we only have to do string
            # comparison between header_row and self.available_service_fields
            # choose_available_service_fields = []
            # choose_available_service_fields = copy.copy(self.available_service_fields)
            service_fields_already_taken = set()
            field_counter = 0
            for field in header_row_lower:
                current_index = header_row_lower.index(field)
                if field in arcgis_fields_lower:
                    service_index = arcgis_fields_lower.index(field)
                    if field not in service_fields_already_taken:
                        self.field_mapping.append([header_row[current_index],self.available_service_fields[service_index]])
                        service_fields_already_taken.add(field)
                    else:
                        # Unable to map this header field
                        self.field_mapping.append([header_row[current_index], ''])
                else:
                    # Unable to map this header field
                    self.field_mapping.append([header_row[current_index],''])

                field_counter += 1

        self.logger.log('generated field mapping is {0}'.\
                        format(self.field_mapping),'INFO')
        if len(self.field_mapping) == 1:
            # Then we should map to singleline
            self.field_mapping = [[header_row[0].encode('utf-8'),singleLine_field.encode('ascii', 'ignore')]]

        new_field_mapping = []
        for this_list in self.field_mapping:
            list_lower = [l.lower() for l in this_list]
            if "oid" in list_lower or "objectid" in list_lower:
                pass
            else:
                new_field_mapping.append(this_list)
        
        final_field_map = []
        #arcpy.AddMessage(new_field_mapping)
        for this_list in new_field_mapping:
            if isinstance(this_list[0],bytes):
                final_field_map.append([this_list[0].decode("utf-8") ,this_list[1].decode("utf-8")])               
            else:
                final_field_map.append(this_list)

        self.field_mapping = final_field_map

        #arcpy.AddMessage(self.field_mapping)

        #arcpy.AddMessage(field_info)

        if was_originally_table:
            self.in_file_type = 'table'

        if self.in_file_type == 'txt':
            json_object_as_dict = dict(header_row_exists=self.header_row_exists,\
                                        column_delimiter=self.column_delimiter_orig,\
                                        # http_referer=self.http_referrer,\
                                        text_qualifier=self.text_qualifier_orig,\
                                        # unzipped_file=self.unzipped_file,\
                                        # available_service_fields=\
                                        #     str(self.available_service_fields),\
                                        field_info=field_info,\
                                        field_mapping=self.field_mapping,\
                                        fixed_width=self.fixed_width,\
                                        width_of_cols=self.col_widths,\
                                        chars_per_row=self.chars_per_row,\
                                        column_names=self.column_names,\
                                        file_type=self.in_file_type)
        else:
            json_object_as_dict = dict(header_row_exists=self.header_row_exists,\
                                        column_delimiter=self.column_delimiter_orig,\
                                        # http_referer=self.http_referrer,\
                                        text_qualifier=self.text_qualifier_orig,\
                                        # unzipped_file=self.unzipped_file,\
                                        # available_service_fields=\
                                        #     str(self.available_service_fields),\
                                        field_info=json.dumps(field_info,ensure_ascii=False),\
                                        field_mapping=json.dumps(self.field_mapping,ensure_ascii=False),\
                                        # fixed_width=self.fixed_width,\
                                        # width_of_cols=self.col_widths,\
                                        # chars_per_row=self.chars_per_row,\
                                        column_names=self.column_names,\
                                        file_type=self.in_file_type,\
                                        singleline_field=singleLine_field)
        self.output_json = json.dumps(json_object_as_dict)
        # self.output_json = self.field_mapping
        # raise SystemExit

        # If we needed to unzip, we need to remove the intermittent folder
        # if self.in_file[-4:] == '.zip':
        #     try:
        #         # os.remove(self.unzipped_folder)
        #         shutil.rmtree(self.unzipped_folder,ignore_errors=True)
        #     except Exception as e:
        #         self.logger.log('Unable to remove intermittent files {}'.format(e)\
        #                         ,'INFO')
        return

    def execute_fields_already_mapped(self):
        ''' This is a preprocessing step that gets called
            from the batchGeocode GP script. It first calls
            this, takes the resultant analyze object and then
            calls the batchGeocode class with an Analyze object
            as one of the inputs (for more on this, see BatchGeocode.py)'''

        arcpy.env.autoCancelling = False


        self.remove_global_id = False

        if self.in_file_type == "EGDB_table":
            self.in_file_type = "table"
            self.EGDB_table = True
        else:
            self.EGDB_table = False

        if self.in_file_type != 'table':
            self.started_as_feature_service = False
            if self.portal_file_type == "CSV Collection":
                # then its a zip file...
                # We don't support CSV Collections!
                msg_code = 100167
                msg = self.ERROR_CODES[msg_code]
                self.addError(msg_code, msg)
            else:
                self.unzipped_file = self.in_file
        else:
            # # We need to store some information for when we later write to a feature class

            self.started_as_feature_service = True

            was_originally_table = True
            # Run the arcpy conversion and turn it into a csv... process as normal
            scratch_folder = arcpy.env.scratchFolder
            # Generate the field mapping to add an ObjectID Field because TableToTable
            # does not do it quite correctly
            intermediate_csv = os.path.join(scratch_folder, 'intermediate.csv')
            try:
                self.oid_field = self.in_table.describe.oidFieldName
            except AttributeError as e:
                if "NoneType" in str(e):
                    arcpy.AddError('The table submitted is empty or corrupt.')

            self.list_of_fields = [fld for fld in self.in_table.describe.fields]
            fieldmappings = arcpy.FieldMappings()
            # fieldmappings.addTable(self.in_file)
            for field in self.list_of_fields:
                # don't add shape field!!!
                if field.name != "shape" or field.name != self.oid_field or field.name != 'globalid':
                    new_field_mapping = arcpy.FieldMap()
                    if self.in_table.isLayer:
                        new_field_mapping.addInputField(self.in_table.layer, field.name)
                    else:
                        new_field_mapping.addInputField(self.in_table_path, field.name)
                    out_field = new_field_mapping.outputField
                    out_field.name = field.name
                    if field.type == "OID":
                        out_field.type = "Double"
                    else:
                        out_field.type = field.type
                    new_field_mapping.outputField = out_field
                    fieldmappings.addFieldMap(new_field_mapping)


            # arcpy.conversion.TableToTable(self.in_table.layer, scratch_folder, 'intermediate.csv',
            #                               field_mapping=fieldmappings)
            # self.unzipped_file = intermediate_csv
            # # reset the in file type to csv now that its been converted
            # self.in_file_type = 'csv'


            # just in case it gets used somewhere...
            if self.in_table.isLayer:
                self.in_file = self.in_table.layer
            else:
                self.in_file = self.in_table_path
            self.list_of_fields = [fld.name for fld in arcpy.ListFields(self.in_file)]


        # Let's check if the service is on a federated server. If it is, then we need
        # to inherit the username and password in a special way
        index_of_REST = self.service_url.lower().find('rest')
        if index_of_REST == -1:
            self.logger.log('Your service URL did not have a valid schema page ' + \
                 'please check your service URL: Your geocode service must be ' + \
                               'through REST, your service_url ' + \
                               'does not have REST or rest in it.','USER_ERROR')




        hopefully_geocode_addresses = self.service_url[-16:].lower()
        if hopefully_geocode_addresses != 'geocodeaddresses':
            msg_dict = dict(service_url=self.service_url)
            msg_code = 100171
            msg = self.ERROR_CODES[msg_code].format(**msg_dict)
            self.addError(msg_code, msg, msg_dict)


        # Need to initialize these, they won't get used though
        # There is logic in geocode_batch that ignores these
        # inputs based on the booleans beow

        self.logger.log('token is {}'.format(self.token),'INFO')

        possible_service_fields = []

        # Populate the available fields from the service dynamically
        # Get the fields that are available to be mapped to by the service
        # This way, we can check the mapping provided by the user and make
        # sure the fields they are trying to map to exist for the service
        # that they are using

        # This line below gets rid of the 'geocodeAddresses' part of the URL
        self.ssl_issues = False
        almost_schema_page = self.service_url[:-17]
        try:
            r = requests.post(almost_schema_page, data={'f':'json', 'referer':self.referer}, verify=True, timeout=20)
            self.json_info_response = r.json()
            r.raise_for_status()
        except requests.exceptions.HTTPError as err:
            msg_dict = dict(service_url=self.service_url)
            msg_code = 100171
            msg = self.ERROR_CODES[msg_code].format(**msg_dict)
            self.addError(msg_code, msg, msg_dict)
        except requests.exceptions.SSLError as err:
            self.ssl_issues = True
        # values = ([('token',self.token),('f','json')])
        if self.ssl_issues:
            try:
                r = requests.post(almost_schema_page, data={'f':'json', 'referer':self.referer}, verify=False, timeout=20)
                self.json_info_response = r.json()
                r.raise_for_status()
            except requests.exceptions.HTTPError as err:
                msg_dict = dict(service_url=self.service_url)
                msg_code = 100171
                msg = self.ERROR_CODES[msg_code].format(**msg_dict)
                self.addError(msg_code, msg, msg_dict)

        possible_service_fields = []

        # Now that we have the URL for the info page, let's discover
        # the schema page for the service using the info on the infopage
        # self.service_info = json_info_response
        try:
            addr_fields = self.json_info_response['addressFields']
            # categories_inner = json_response['categories']['categories']
            for addr in addr_fields:
                possible_service_fields.append(addr['name'])

        except KeyError as e:
            # If the service returns a 499, then the token was invalid
            #self.notSSL = not self.ssl_issues
            if self.json_info_response['error']['code'] == 499 or self.json_info_response['error']['code'] == 403:
                # needs a token
                r = requests.post(almost_schema_page, data={'token': self.token, 'f':'json', 'referer':self.referer}, verify=not self.ssl_issues)
                self.json_info_response = r.json()
                try:
                    addr_fields = self.json_info_response['addressFields']
                    for addr in addr_fields:
                        # Add each field to possible_service_fields
                        possible_service_fields.append(addr['name'])

                except KeyError as e:
                    msg_dict = dict(service_url=self.service_url)
                    msg_code = 100172
                    msg = self.ERROR_CODES[msg_code].format(**msg_dict)
                    self.addError(msg_code, msg, msg_dict)

            else:
                msg_dict = dict(service_url=self.service_url)
                msg_code = 100172
                msg = self.ERROR_CODES[msg_code].format(**msg_dict)
                self.addError(msg_code, msg, msg_dict)

        try:
            singleLine_field = self.json_info_response['singleLineAddressField']['name']
            possible_service_fields.append(singleLine_field)
            possible_service_fields.append('OBJECTID')
        except KeyError:
            msg_dict = dict(service_url=self.service_url)
            msg_code = 100171
            msg = self.ERROR_CODES[msg_code].format(**msg_dict)
            self.addError(msg_code, msg, msg_dict)


        # self.logger.log('POSSIBLE_ARCGIS_FIELDS ARE {}'.\
        #                 format(possible_service_fields),'INFO')

        self.available_service_fields = possible_service_fields
        return


class BatchGeocode():
    """Performs asynchronous batch geocoding by making asynchronous
    REST calls a geocode service. Then, returns a results file"""

    ERROR_CODES = {
        100161: "Batch Geocode failed.",
        100171: "Invalid locator service URL: {service_url}.",
        100172: "Invalid credentials or invalid locator service URL: {service_url}. Make sure proxy items are shared with Everyone.",
        100173: "Your user role doesn't include the create items privilege",
        100176: "Your file has a corrupt or empty header.",
        100177: "Your Excel file is corrupted and cannot be opened.",
        100178: "Your file's first row is empty. Please edit your file and resubmit.",
        100183: "Field mapping entered in an incorrect format.",
        100184: "Incorrect outputFields. Enter outputFields like so: x,y,score,match_addr.",
        100185: "chars_per_row must be an integer.",
        100186: "Incorrect widths_of_cols. Enter widths_of_cols like so: 10,15,6",
        100187: "Your outputName must be blank to get out a Feature Collection.",
        100188: "Your outputName cannot be empty.",
        100189: "Incorrect geocodeServiceURL, verify that the geocoding service is online and has a valid geocodeAddresses REST endpoint.",
        100190: "Check your header_rows_to_skip parameter",
        100191: "Unable to write header: check to see if your headerRowsToSkip is incorrect or if the geocoding service is unavailable.",
        100192: "If your header row exists, you must provide the headerRowsToSkip",
        100193: "Your headerRowsToSkip must be a positive number less than the number of rows in the inputTable or inputFileItem",
        100194: "Results file is being edited by another program",
        100195: "File entered was empty, returning an empty file",
        100196: "Field mapping must be a list of lists",
        100197: "Your geocodingServiceURL may be incorrect, it does not have a maxBatchSize on the info page",
        100198: "Edit your field mapping and resubmit, none of your fields are mapped to valid locator fields",
        100199: "Edit your field mapping and resubmit, your mapping includes a locator field that is not supported",
        100200: "Edit your field mapping and resubmit, your mapping includes an unsupported locator field. Supported locator fields are {supported_fields}",
        100201: "Your field_mapping includes an input field {field} not present in your input file: {supported_fields}.",
        100202: "Could not make a REST request from the server machine verify that the server machine can connect to the internet.",
        100227: "Your header row is not the same length as the number of lists in your field mapping.",
        100228: "Your headerRowsToSkip must be a positive number less than the number of rows in the inputTable or inputFileItem.",
        100229: "Your input file has too many rows. xls only supports up to 65536 rows. Please submit a different outputType.",
        100231: "Your outputName is not formatted correctly.",
        100232: "Your user role doesn't include the geocode privilege.",
        100233: "The number of records in the input table is greater than the maximum records allowed for geocoding.",
        100245: "Invalid expression for {paramName}, malformed JSON.",
        100253: "Empty inputTable and empty inputFileItem. Enter either an inputTable or inputFileItem."
    }

    def __init__(self,analyze_object,source_country, \
                 category,output_fields,out_file_type,\
                 fixed_width,header_row_exists,header_rows_to_skip,qualifier,\
                 widths_of_columns,field_names,chars_per_row,field_mapping,field_info,\
                 output_name,context,concurrent_batches, location_type,\
                 to_pass_errorFunc,to_pass_warningFunc,to_pass_DEBUG,output_location,\
                 table_name):
        # Time parameters
        # Have to initialize this first to properly benchmark
        try:
            self.start_time = time.time()
            self.end_time = None

            # Logs messages
            lg = Logger(LOG_LEVEL)
            self.logger = lg

            # Error reporting
            if to_pass_errorFunc:
                self.errorFunc = to_pass_errorFunc
            else:
                self.errorFunc = arcpy.AddError
            if to_pass_warningFunc:
                self.errorFunc = to_pass_warningFunc
            else:
                self.warningFunc = arcpy.AddWarning
            if to_pass_DEBUG:
                self.DEBUG = to_pass_DEBUG
            else:
                self.DEBUG = False

            # Input parameters
            self.analyze_obj = analyze_object
            self.output_location = output_location
            self.table_name = table_name
            self.source_country = source_country
            self.category = category
            self.location_type = location_type
            if "%" in self.category:
                new_category = self.category.replace("%", "/")
                self.category = new_category
            #try:
            #    field_mapping_good = "field_mapping_final = " + field_mapping
            #    exec(field_mapping_good)
            #except SyntaxError:
            #    msg_code = 100183
            #    msg = self.ERROR_CODES[msg_code]
            #    self.addError(msg_code, msg)
            #except Exception:
            #    field_mapping_final = field_mapping
            try:
                self.field_mapping = json.loads(field_mapping)
            except:
                msg_code = 100183
                msg = self.ERROR_CODES[msg_code]
                self.addError(msg_code, msg)

            
            try:
                self.field_info = json.loads(field_info)
            except:
                msg_code = 100183
                msg = self.ERROR_CODES[msg_code]
                self.addError(msg_code, msg)


            #try:
            #    field_info_good = "field_info_final = " + field_info
            #    exec (field_info_good)
            #except Exception:
            #    field_info_final = field_info
            #except SyntaxError:
            #    msg_code = 100183
            #    msg = self.ERROR_CODES[msg_code]
            #    self.addError(msg_code, msg)


            #self.field_info = field_info_final

            self.spatial_reference = ''
            # reset this later if necessary

            self.concurrent_batches = int(concurrent_batches)

            self.out_file_type = out_file_type

            #self.out_file_type = "csv" ## to debug... let's just try to get csv output

            if self.out_file_type == "append" or self.out_file_type == "fc":
                self.egdb_output_type = self.out_file_type
                self.out_file_type = "EGDB_table"

            propertyDictJSON = self.analyze_obj.portal_self
            # Checking privileges at the UI level using hostedgp instead

            #privileges = propertyDictJSON[u'user'][u'privileges']
            ## propertyDictJSON = json.loads(propertyDict,encoding='utf-8')
            #if self.out_file_type == "csv" or self.out_file_type == 'xls':
            #    if not "portal:user:createItem" in privileges:
            #        # shutil.rmtree(self.unzipped_folder,ignore_errors=True)
            #        msg_code = 100173
            #        msg = self.ERROR_CODES[msg_code]
            #        self.addError(msg_code, msg)


            #if self.out_file_type == "Feature Service":
            #    if not 'premium:user:geocode' in privileges:
            #        # shutil.rmtree(self.unzipped_folder,ignore_errors=True)
            #        msg_code = 100232
            #        msg = self.ERROR_CODES[msg_code]
            #        self.addError(msg_code, msg)

            self.is_agol = False
            self.supports_countries = False
            self.supports_categories = False
            # needs_token = False
            self.max_records = None

            # determine if it is the AGOL world locator
            # 'token': self.token, 
            #try:
            #    r = requests.post(self.analyze_obj.service_url[:-17], data={'f':'json', 'referer':self.analyze_obj.referer}, verify=not self.analyze_obj.ssl_issues, timeout=20)
            #    json_info_response = r.json()
            #    r.raise_for_status()
            #    #values = ([('f','json')])
            #    #encoded_data = urllib.urlencode(values)
            #    #http_header = {'referer':self.analyze_obj.referer}
            #    #info_request = urllib2.Request(self.analyze_obj.service_url[:-17],encoded_data,http_header)
            #    #if self.analyze_obj.ssl_issues:
            #    #    info_response = urllib2.urlopen(info_request,context=ssl._create_unverified_context())
            #    #else:
            #    #    info_response = urllib2.urlopen(info_request)
            #    #json_info_response = json.loads(info_response.read())
            #except requests.exceptions.HTTPError as err:
            #    msg_code = 100171
            #    msg_dict = dict(service_url=self.analyze_obj.service_url)
            #    msg = self.ERROR_CODES[msg_code].format(**msg_dict)
            #    self.addError(msg_code, msg)

            if "countries" in self.analyze_obj.json_info_response:
                self.supports_countries = True

            if "categories" in self.analyze_obj.json_info_response:
                self.supports_categories = True

            try:
                loc_prop = self.analyze_obj.json_info_response['locatorProperties']
            except KeyError:
                msg_code = 100171
                msg_dict = dict(service_url=self.analyze_obj.service_url)
                msg = self.ERROR_CODES[msg_code].format(**msg_dict)
                self.addError(msg_code, msg)

            try:
                self.is_agol = loc_prop['isAGOWorldLocator']
            except KeyError:
                pass

            current_path = os.path.dirname(os.path.realpath(__file__))

            if self.is_agol:
                try:
                    privileges = propertyDictJSON['user']['privileges']
                    if not "premium:user:geocode" in privileges:
                        msg_code = 100232
                        msg = self.ERROR_CODES[msg_code]
                        self.addError(msg_code, msg)
                except Exception as e:
                    msg_code = 100232
                    msg = self.ERROR_CODES[msg_code]
                    self.addError(msg_code, msg)
                
                self.concurrent_batches = 1

            if self.is_agol and os.path.exists(os.path.join(current_path,"batchGeocodingProperties.json")):
                try:
                    with open(os.path.join(current_path,"batchGeocodingProperties.json")) as json_file:
                        thread_info = json.load(json_file)

                        self.concurrent_batches = thread_info["numBatchThreads"]

                        self.max_records = thread_info["maxRecords"]
                except Exception as e:
                    self.concurrent_batches = 1

            self.specific_output_fields = False
            self.orig_output_fields = output_fields
            if output_fields == "*":
                output_fields = ""

            if output_fields == "NONE":
                self.output_fields = []
                self.output_fields.append("Shape X")
                self.output_fields.append("Shape Y")
            elif output_fields != '':
                self.specific_output_fields = True
                try:
                    self.output_fields = output_fields.split(',')
                    shape_x = ["shape x", "Shape x", "Shape X", "shape X"]
                    shape_y = ["shape y", "Shape y", "Shape Y", "shape Y"]
                    if all(x not in self.output_fields for x in shape_x):
                        self.output_fields.append("Shape X")
                    if all(y not in self.output_fields for y in shape_y):
                        self.output_fields.append("Shape Y")
                    if "status" in self.output_fields:
                        status_index = self.output_fields.index("status")
                        self.output_fields.insert(status_index, "Status")
                        self.output_fields.remove("status")
                    if "score" in self.output_fields:
                        score_index = self.output_fields.index("score")
                        self.output_fields.insert(score_index, "Score")
                        self.output_fields.remove("score")
                    if "match_addr" in self.output_fields:
                        matchaddr_index = self.output_fields.index("match_addr")
                        self.output_fields.insert(matchaddr_index, "Match_addr")
                        self.output_fields.remove("match_addr")
                    if "addr_type" in self.output_fields:
                        addrtype_index = self.output_fields.index("addr_type")
                        self.output_fields.insert(addrtype_index, "Addr_type")
                        self.output_fields.remove("addr_type")
                    # if "Match_addr" not in self.output_fields and "match_addr" not in self.output_fields:
                    #     self.output_fields.append("Match_addr")
                except Exception as e:
                    msg_code = 100184
                    msg = self.ERROR_CODES[msg_code]
                    self.addError(msg_code, msg)
            else:
                ### this means it was empty string... which is all out fields.
                self.output_fields = output_fields

            if str(header_row_exists) == 'false' or \
                            str(header_row_exists) == 'False':
                self.header_row_exists = False
            else:
                self.header_row_exists = True

            if str(fixed_width) == 'false' or str(fixed_width) == 'False':
                self.fixed_width = False
            else:
                self.fixed_width = True

            self.header_rows_to_skip = header_rows_to_skip
            self.text_qualifier = qualifier
            if chars_per_row != '':
                try:
                    self.chars_per_row = int(chars_per_row)
                except Exception as e:
                    msg_code = 100185
                    msg = self.ERROR_CODES[msg_code]
                    self.addError(msg_code, msg)
            else:
                self.chars_per_row = chars_per_row
            self.field_names = field_names
            if widths_of_columns != '':
                try:
                    self.chars_per_column = widths_of_columns.split(',')
                except Exception as e:
                    msg_code = 100186
                    msg = self.ERROR_CODES[msg_code]
                    self.addError(msg_code, msg)
            else:
                self.chars_per_column = widths_of_columns

            if self.out_file_type == "EGDB_table":
                if self.egdb_output_type == "append":
                    self.output_location = self.analyze_obj.in_table_path
                self.working_dir = arcpy.env.scratchFolder
            else:
                if self.output_location:
                    path,out_file = os.path.split(self.output_location)
                    self.working_dir = path
                else:
                    self.working_dir = arcpy.env.scratchFolder

            if output_name:
                output_name_json = json.loads(output_name)
                try:
                    self.output_name = output_name_json["serviceProperties"]['name']
                except KeyError:
                    try:
                        self.output_name = output_name_json["itemProperties"]['title']
                    except KeyError:
                        msg_code = 100231
                        msg = self.ERROR_CODES[msg_code]
                        self.addError(msg_code, msg)

                if self.out_file_type == "Feature Collection":
                    msg_code = 100187
                    msg = self.ERROR_CODES[msg_code]
                    self.addError(msg_code, msg)

            else:
                if self.out_file_type != "Feature Collection":
                    msg_code = 100188
                    msg = self.ERROR_CODES[msg_code]
                    self.addError(msg_code, msg)

                else:
                    self.output_name = "outFC"

            # for testing purposes
            # self.output_name = "charles"
            self.out_SR = {}
            if context:
                try:
                    context_json = json.loads(context)
                    if context_json:
                        self.out_SR = context_json['outSR']
                        self.spatial_reference = self.out_SR['wkid']
                except Exception:
                    msg_code = 100245
                    msg_dict = dict(paramName=context)
                    msg = self.ERROR_CODES[msg_code].format(**msg_dict)
                    self.addError(msg_code, msg)
            #arcpy.AddMessage("self.out_SR is {}".format(self.out_SR))
            self.feature_service_output = False

            if self.out_file_type == 'Feature Service' or self.out_file_type == 'Feature Collection' or self.out_file_type == "EGDB_table":
                self.feature_service_output = True
                self.out_file_type = 'csv'

            # This might get reset later...
            self.batch_size = 500

            self.input_data = {}
            self.num_of_batches = 0
            self.remainder = 0
            # Keeps track of where we are in the file
            # if it is fixed width with chars_per_row as
            # not null
            self.total_counter = 0

            # Starts false, main thread reads .env before firing each REST call
            #
            self.is_cancelled = False

            # If it needs a token to call geocodeAddresses or not. Determined earlier in BatchGeocode.py
            if self.analyze_obj.token:
                self.needs_token = True
            else:
                self.needs_token = False

            # Number of empty rows starts at 0. Prevents a .stopIteration() exception later on
            self.num_empty_rows = 0

            # Keeps track if we are in the process of exiting the job or not. If the job succeeds,
            # this will always be false
            self.error_out = False

            # Unable to parse record numbers, gets added as an item resource if the list ends up with
            # elements in it
            self.cannot_parse_records = []
            self.check_item_resources = False

            # Get counts on number of matched, unmatched, and tied records
            self.matched = 0
            self.unmatched = 0
            self.tied = 0

            self.zipped_gdb_output = False
            if out_file_type == "gdb":
                self.zipped_gdb_output = True
                self.feature_service_output = True
                self.out_file_type = 'csv'

            # Data q's
            self.processed_data_q = Queue(maxsize=self.concurrent_batches*2)
            #self.processed_data_q = Queue(maxsize=7)
            self.REST_response_q = Queue(maxsize=self.concurrent_batches*2)
            #self.REST_response_q = Queue(maxsize=7)
            self.num_of_REST_threads = Queue(maxsize=self.concurrent_batches)
            self.error_q = Queue(maxsize=self.concurrent_batches*3)
            self.dropped_records_q = Queue(maxsize=20)
            # self.leftover_records_q = Queue(maxsize=self.batch_size)

        except Exception as e:
            arcpy.AddError(str(e))
            msg_code = 100161
            msg = self.ERROR_CODES[msg_code]
            self.addError(msg_code, msg)
            if not self.analyze_obj.started_as_feature_service:
                self.delete_unwanted_folder()

        self.preprocess()
        #arcpy.AddMessage("done with pre-process")

        self.execute()
        #arcpy.AddMessage("done with execute")

        self.postprocess()

    def addError(self, msg_code, msg, msg_params=None, is_warning=False):
        '''Adds the error message and raises arcpy.ExecuteError exception. self.errorFunc function is used to add the
        error message. if error function is arcpy.AddError only add the msg.'''

        error_func_name = self.errorFunc.__name__
        error_func = self.errorFunc
        if is_warning:
            error_func_name = self.warningFunc.__name__
            error_func = self.warningFunc

        if hasattr(arcpy, error_func_name):
            error_func(msg.replace("$", ""))
        else:
            error_func(msg_code, msg, msg_params, is_warning)

        if not is_warning:
            raise arcpy.ExecuteError

    def delete_unwanted_folder(self):
        # shutil.rmtree(self.analyze_obj.folder_to_remove, ignore_errors=True)
        return

    def geocode_batch_handler(self,batch_num):
        '''Calls geocodeBatch which actually makes the REST calls,
        and queues up the rest responses to be written out.
        REST_response_q has no ordering, just puts responses
        in the queue in the order that they come in, and that
        is the order in which they are written out to the results
        file, self.output.csv/self.output.xls/self.output.txt

        Each thread makes a call to geocode_batch_handler, which then
        makes a call to geocodeBatch, waits for the response, queues the
        the response and then releases the lock on the number of REST
        threads being fired (num_of_REST_threads) since only 4 rest
        threads can fire at once'''
        #arcpy.AddMessage('made it to handler')
        try:
            while self.processed_data_q.empty():
                # If there is no data to be sent, sleep the thread
                #arcpy.AddMessage('still empty')
                time.sleep(0.1)
                # Check again if there is data to be sent after 0.3 seconds
                if not self.error_q.empty():
                    self.logger.log('There is an error in some thread','INFO')
                    dummy_obj = self.num_of_REST_threads.get()
                    # Case in which there was an error in a different REST thread
                    return
                if self.is_cancelled:
                    dummy_obj = self.num_of_REST_threads.get()
                    # IN PROCESS OF DOING THIS THING RN CHECK IF IT WORKED
                    # self.logger.log('The script has been aborted by the user. Returning geocoding results thus far',\
                    #                 'USER_ERROR')
                    return
                if self.error_out:
                    dummy_obj = self.num_of_REST_threads.get()
                    return
            # Otherwise, fire off a REST call
            if self.processed_data_q.empty() == False:
                #arcpy.AddMessage('picked something off the queue and ready to rock')
                last_batch = False
                batch_request = self.processed_data_q.get()
                #arcpy.AddMessage('picked something off the q')
                # If this is the last batch, then set a flag
                if batch_num + 1 == self.num_of_batches:
                    last_batch = True
                self.logger.log('Firing off REST CALL number {}'.format(batch_num),'INFO')
                #arcpy.AddMessage('time to geocode the recordset')
                feature_set = self.geocode_batch(batch_request[1], last_batch)
                #arcpy.AddMessage(feature_set)
                self.logger.log('REST CALL number {} came back'.\
                                format(batch_num),'INFO')
                if not self.error_q.empty():
                    self.logger.log('There is an error in some REST thread','INFO')
                    self.error_out = True
                    # Case in which there was an error in a different REST thread
                    return
                self.REST_response_q.put([batch_request[0], feature_set])
                batch_request = None
                feature_set = None
                # gc.collect()
                self.logger.log('Just put something in the REST response q','INFO')
            self.logger.log('About to release the lock','INFO')
            try:
                dummy_obj = self.num_of_REST_threads.get()
            except arcgisscripting.ExecuteAbort:
                pass
            self.logger.log('Released the lock','INFO')
        except Exception as e:
            if self.DEBUG:
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                self.logger.log('{0} {1} {2}'.format(exc_type, fname, exc_tb.tb_lineno),'ERROR')
            # # This error is fatal, so just release the lock on the num_threads queue
            # so that we can exit the whole system
            self.error_q.put([1,str(e)])
            self.num_of_REST_threads.get()
            return

        return

    def geocode_batch(self, recordset, last_batch):
        ### Makes REST calls to geocodeAddresses
        if not self.out_SR:
            self.out_SR = self.analyze_obj.json_info_response["spatialReference"].get("wkid", 4326)
            #arcpy.AddMessage("outSR is {}".format(outSR))
        try:
        #arcpy.AddMessage("outSR is {}".format(self.out_SR))
            #arcpy.AddMessage(recordset)
            values = {
                    "addresses": json.dumps(recordset,ensure_ascii=False),
                    "outSR": json.dumps(self.out_SR,ensure_ascii=True),
                    "f": "json"
                }
            #arcpy.AddMessage('makes it past json.dumps')
            if self.needs_token:
                #arcpy.AddMessage('sending the token!!!')
                values["token"] = self.analyze_obj.token
            if self.location_type:
                values["locationType"] = self.location_type
            #arcpy.AddMessage("supports country: {}".format(self.supports_countries))
            #arcpy.AddMessage("country is {}".format(self.source_country))
            if self.supports_countries:
                values["sourceCountry"] = self.source_country
            if self.supports_categories:
                values["category"] = self.category

                
            # Turns everything into bytes to be sent
            #arcpy.AddMessage(values)

            # Secured local service case needs an http_header
            http_header = {'referer':self.analyze_obj.referer}

            # uncomment if you want to test the retry logic
            #response = self.retry_REST_request_until_done(values,last_batch)
            #return response
            try:
                response = requests.post(self.analyze_obj.service_url, data=values, headers=http_header, verify=True)
            except requests.exceptions.SSLError:
                self.analyze_obj.ssl_issues = True
                response = requests.post(self.analyze_obj.service_url, data=values, headers=http_header, verify=False)
            # Store the response
            json_rep = response.json()
            #arcpy.AddMessage('response is {}'.format(json_rep))
            self.logger.log('Our rest response is in memory now, ready to return','INFO')

            if "error" in json_rep:
                error = json_rep['error']
                if str(error['code']) == '400':
                    # Unsupported Category
                    error_message = str(error['details'][0])
                    if 'category' in error_message.lower():
                        self.error_q.put([1,str(error_message)])
                        self.num_of_REST_threads.get()
                        return "error"

                # whatever
                if str(error['code']) == '498':
                    # This means that there is an expired token
                    server_request_properties = arcpy.gp._arc_object.serverrequestproperties()
                    s_json = json.loads(server_request_properties)
                    token = s_json['token']
                    self.analyze_obj.token = token
                    self.logger.log('new token is {}'.format(self.analyze_obj.token),'INFO')

                if "Token" in error['message']:
                    try:
                        self.needs_token = True
                        # try it with a token
                        values["token"] = self.analyze_obj.token

                        # Secured local service case needs an http_header
                        http_header = {'referer':self.analyze_obj.referer}
                        # http_header_required = True


                        # Store the response
                        if self.analyze_obj.ssl_issues:
                            response = requests.post(self.analyze_obj.service_url, data=values, headers=http_header, verify=False)
                        else:
                            response = requests.post(self.analyze_obj.service_url, data=values, headers=http_header, verify=True)
                        
                        json_rep = response.json()
                        if "error" in json_rep and str(json_rep["error"]["code"]) == "500":
                            retried_feature_set = self.retry_REST_request_until_done(values,last_batch)
                            if retried_feature_set == str(self.error_message):
                                self.error_q.put([1,str(self.error_message)])
                                self.num_of_REST_threads.get()
                                return "error"
                            else:
                                return retried_feature_set
                        if "error" in json_rep:
                            error = json_rep['error']
                            error_message = str(error['details'][0])
                            self.error_q.put([1,str(error_message)])
                            self.num_of_REST_threads.get()
                            return "error"
                        else:
                            #this means it worked!!!
                            return json_rep
                    except Exception as e:
                        #arcpy.AddMessage('problem at 1804')
                        self.error_q.put([1,str(e)])
                        self.num_of_REST_threads.get()
                        return "error"

                if str(error['code']) == '403':
                    # Permissions issue
                    error_message = str(error['message'])
                    self.error_q.put([1,str(error_message)])
                    self.num_of_REST_threads.get()
                    return "error"
                if str(error['code']) == '500':
                    # this is a timeout. We need to hit retry logic!
                    try:
                        retried_feature_set = self.retry_REST_request_until_done(values,last_batch)
                    except Exception as e:
                        arcpy.AddError(e)
                        if self.DEBUG:
                            exc_type, exc_obj, exc_tb = sys.exc_info()
                            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                            self.logger.log('{0} {1} {2}'.format(exc_type, fname, exc_tb.tb_lineno),'ERROR')
                    if retried_feature_set == str(self.error_message):
                        #arcpy.AddMessage('error in retry logic')
                        self.error_q.put([1,str(self.error_message)])
                        self.num_of_REST_threads.get()
                        return "error"
                    else:
                        #arcpy.AddMessage('retry logic returned successfully')
                        return retried_feature_set
                else:
                    try:
                        error_message = str(error['message'])
                    except KeyError:
                        try:
                            error_message = str(error['details'][0])
                        except KeyError:
                            error_message = "There was an error in geocoding."
                    self.error_q.put([1,str(error_message)])
                    self.num_of_REST_threads.get()
                    return "error"

                # If the token wasn't expired OR invalid, then we need
                # to launch into rematch logic.
            else:
                return json_rep

        except Exception as e:
            if self.DEBUG:
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                self.logger.log('{0} {1} {2}'.format(exc_type, fname, exc_tb.tb_lineno),'ERROR')
            # # This error is fatal, so just release the lock on the num_threads queue
            # so that we can exit the whole system
            self.error_q.put([1,str(e)])
            self.num_of_REST_threads.get()
            return

    def retry_REST_request_until_done(self,values,last_batch):
        ''' The following rematch code contains logic to half
            recordsets and try again until the entirety of the
            recordset is geocoded '''
        # errors_set = set()
        # error_counter = 0
        self.error_message = ""
        try:
            ultimate_full_featureset = []
            #arcpy.AddMessage(values)
            full_recordset = values['addresses']
            # Initialize variables for this function
            partial_batch_begin = 0
            partial_batch_end = 0
            # This variable changes if it is the last batch
            batch_size_for_this_function = 0
            # Set the initial partial batch size correctly
            if last_batch:
                partial_batch_size = self.remainder - 1
                batch_size_for_this_function = self.remainder - 1
            else:
                partial_batch_size = self.batch_size
                batch_size_for_this_function = self.batch_size
            just_failed = True
            done_with_this_batch = False
            #Set the indexes that we want to use to make the partial batch
            while just_failed or (partial_batch_end < batch_size_for_this_function):
                time.sleep(1)
                if just_failed and partial_batch_size == 1:
                    # construct unmatched featureSet of size 1
                    try:
                        featureSet = {
                            "locations": [
                                {
                                "address": "",
                                "score": 0,
                                "attributes": {
                                # "ResultID": 2, #some special formula here
                                # "Loc_name": "",
                                # "Status": "U",
                                # "Score": 0,
                                # "Match_addr": "",
                                # "LongLabel": "",
                                # "ShortLabel": "",
                                # "Addr_type": "",
                                # "Type": "",
                                # "PlaceName": "",
                                # "Place_addr": "",
                                # "Phone": "",
                                # "URL": "",
                                # "Rank": 0,
                                # "AddBldg": "",
                                # "AddNum": "",
                                # "AddNumFrom": "",
                                # "AddNumTo": "",
                                # "AddRange": "",
                                # "Side": "",
                                # "StPreDir": "",
                                # "StPreType": "",
                                # "StName": "",
                                # "StType": "",
                                # "StDir": "",
                                # "BldgType": "",
                                # "BldgName": "",
                                # "LevelType": "",
                                # "LevelName": "",
                                # "UnitType": "",
                                # "UnitName": "",
                                # "SubAddr": "",
                                # "StAddr": "",
                                # "Block": "",
                                # "Sector": "",
                                # "Nbrhd": "",
                                # "District": "",
                                # "City": "",
                                # "MetroArea": "",
                                # "Subregion": "",
                                # "Region": "",
                                # "RegionAbbr": "",
                                # "Territory": "",
                                # "Postal": "",
                                # "PostalExt": "",
                                # "Country": "",
                                # "LangCode": "",
                                # "Distance": 0,
                                # "X": 0,
                                # "Y": 0,
                                # "DisplayX": 0,
                                # "DisplayY": 0,
                                # "Xmin": 0,
                                # "Xmax": 0,
                                # "Ymin": 0,
                                # "Ymax": 0
                                }
                                }
                            ]
                        }
                        for attr in self.nice_ordering:
                            if attr == "Rank" or attr == "Score" or attr == "Rank" or attr == "Distance" or attr == "X" \
                                    or attr == "Y" or attr == "DisplayX" or attr == "DisplayY" or attr == "Xmin" \
                                    or attr == "Xmax" or attr == "Ymin" or attr == "Ymax":
                                featureSet['locations'][0]['attributes'][attr] = 0
                            elif attr == "Status":
                                featureSet['locations'][0]['attributes'][attr] = "U"
                            else:
                                featureSet['locations'][0]['attributes'][attr] = ""
                        # some stuff here to get resultID
                        featureSet['locations'][0]['attributes']["ResultID"] = values[0][1]['records'][0]['attributes']['OBJECTID']
                        ultimate_full_featureset.append(featureSet)
                        just_failed = False
                        if done_with_this_batch:
                            break
                        continue
                    except Exception as e:
                        # stack trace...
                        exc_type, exc_obj, exc_tb = sys.exc_info()
                        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                        self.logger.log('{0} {1} {2}'.format(exc_type, fname, exc_tb.tb_lineno),'ERROR')
                        sys.exit()
                elif just_failed:
                    if partial_batch_size > 1:
                        # Because we just failed, let's try to half the batch size
                        # and try again
                        partial_batch_size = partial_batch_size // 2
                    # if partial_batch_size == 1:
                    #     return self.error_message
                    # Adjust the end, given that we have halved the batch size
                    partial_batch_end = partial_batch_begin + partial_batch_size
                else:
                    # We didn't fail, so let's increase the batch size
                    partial_batch_size = partial_batch_size * 2

                    if partial_batch_size > batch_size_for_this_function:
                        partial_batch_size = batch_size_for_this_function
                    # Move to the next batch
                    partial_batch_begin = partial_batch_end
                    partial_batch_end = partial_batch_end + partial_batch_size

                if partial_batch_end > batch_size_for_this_function:
                    # Make sure the batch_end is reasonable
                    partial_batch_end = batch_size_for_this_function
                    done_with_this_batch = True

                # Create a partial recordset
                partial_recordset = give_portion_of_recordset(full_recordset, \
                                                                partial_batch_begin,\
                                                                partial_batch_end)

                # Make a REST call with the partial recordset
                # Only change addresses, rest of the request stays the same
                values['addresses'] = json.dumps(partial_recordset,ensure_ascii=False)
                # encodedData = urllib.urlencode(values).encode('utf-8')

                http_header = {'referer':self.analyze_obj.http_referrer}
                #try:
                if self.analyze_obj.ssl_issues:
                    response = requests.post(self.analyze_obj.service_url, data=values, headers=http_header, verify=False)
                else:
                    response = requests.post(self.analyze_obj.service_url, data=values, headers=http_header, verify=True)
                featureSet = response.json()
                if "error" in featureSet:
                    error = featureSet['error']
                    if str(error['code']) == '500':
                        # cut in half and retry
                        just_failed = True
                        continue
                    else:
                        self.error_message = str(error["message"])
                        if str(error['code']) == '498':
                            # This means that there is an expired token
                            server_request_properties = arcpy.gp._arc_object.serverrequestproperties()
                            s_json = json.loads(server_request_properties)
                            token = s_json['token']
                            self.analyze_obj.token = token
                            self.logger.log('new token is {}'.format(self.analyze_obj.token),'INFO')
                            values['token'] = token
                            just_failed = True
                            continue
                        else:
                            return str(self.error_message)
                else:
                    # No error was thrown!
                    ultimate_full_featureset.append(featureSet)
                    just_failed = False
                    if done_with_this_batch:
                        break

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            self.logger.log('{0} {1} {2}'.format(exc_type, fname, exc_tb.tb_lineno),'ERROR')
            return e

        # Let's just append to the list of locations from the first
        # featureset we returned. Then, return the featureset

        first_featureset = ultimate_full_featureset[0]

        list_of_records = first_featureset['locations']

        # For each rest response recieved, we tacked it on to
        # ultimate_full_featureset. Now we need to loop through
        # that array and add all of the records from each REST
        # response into one list.
        for i in range(len(ultimate_full_featureset)-1):
            some_featureset = ultimate_full_featureset[i+1]
            some_list_of_records = some_featureset['locations']
            # Append to the list of all locations
            list_of_records.extend(some_list_of_records)

        ultimate_full_featureset[0]['locations'] = list_of_records

        return ultimate_full_featureset[0]

    def create_JSON(self):
        '''Figures out which column is the ID column, creates
        a start and an end point for each batch, and passes
        on fields to recordset_builder. Each file format
        has its own recordset_builder. create_JSON is written
        to be file-agnostic, whereas each recordset_builder
        function is file-specific and contains file-specific parsers'''
        try:
            #arcpy.AddMessage("num_of_batches is {}".format(self.num_of_batches))
            #arcpy.AddMessage("remainder is {}".format(self.remainder))
            # This loop creates the start and end points for each
            # recordset_builder
            #arcpy.AddMessage("inside create_JSON")
            self.iter_rows_created = False
            for current_batch in range(self.num_of_batches):
                # if its the last batch
                if (current_batch == self.num_of_batches-1):
                    if current_batch == 0:
                        # If its also the first batch...
                        start = self.header_rows_to_skip
                        end = self.remainder + 1
                        if self.remainder == 0 and self.row_count > 1:
                            # Special case of 1000 rows exactly
                            end = self.row_count
                    else:
                        #if its the last batch, end at the remainder
                        start = current_batch*self.batch_size + 1
                        end = start + self.remainder

                        if self.remainder == 0:
                            # Case where number of rows is a multiple of
                            # batch_size
                            end = self.batch_size + start
                else:
                    # if its the first batch
                    if current_batch == 0:
                        start = self.header_rows_to_skip
                        end = start + self.batch_size
                    # this means its a middle-batch...
                    else:
                        start = current_batch*self.batch_size + 1
                        end = start + self.batch_size
                self.logger.log('Going to try to create JSON number ' + \
                        str(current_batch),'INFO')
                #arcpy.AddMessage('made it past start/end calculations')
                if self.analyze_obj.in_file_type == 'csv':
                    #arcpy.AddMessage("time to write CSV")
                    #arcpy.AddMessage('start is {}'.format(start))
                    #arcpy.AddMessage('end is {}'.format(end))
                    one_batch_JSON = self.recordset_builder_csv(start, \
                                                            end,self.id_col)
                elif self.analyze_obj.in_file_type == 'xlsx' or self.analyze_obj.in_file_type == 'xls':
                    one_batch_JSON = self.recordset_builder_xls(start,\
                                                            end,self.id_col)
                elif self.analyze_obj.in_file_type == 'txt':
                    one_batch_JSON = self.recordset_builder_txt(start,\
                                                            end, self.id_col)
                # self.logger.log('one batch json is {0}'.format(one_batch_JSON),'INFO')
                #arcpy.AddMessage('donewith recordset_builder')
                while self.processed_data_q.full():
                    # This is where create_JSON is stuck most of the time its thread
                    # is running. So, we check to make sure there are no critical
                    # errors in any other threads
                    # time.sleep(1.5)
                    if not self.error_q.empty():
                        self.logger.log('there is an error somewhere,' + \
                                        'shutting down create_JSON','INFO')
                        break
                    if self.is_cancelled:
                        error = 'The script has been aborted by the user. Returning geocoding results thus far'
                        # # This error is fatal, so just release the lock on the num_threads queue
                        # so that we can exit the whole system
                        self.error_q.put([1,str(error)])
                        self.num_of_REST_threads.get()
                        break
                if not self.error_q.empty():
                    self.error_out = True
                # if self.error_out:
                #     self.leftover_records_q.put([current_batch,one_batch_JSON])
                if not self.error_out:
                    #arcpy.AddMessage('tried to add to the data q')
                    #arcpy.AddMessage("Size of processed_data_q before is {}".format(self.processed_data_q.qsize()))
                    self.processed_data_q.put([current_batch,one_batch_JSON])
                    #arcpy.AddMessage("Size of processed_data_q after {}".format(self.processed_data_q.qsize()))
                    #arcpy.AddMessage('successfully added to data q')
                one_batch_JSON = None

            # We are done processing data, so this thread is finished
            return
        except Exception as e:
            self.logger.log('Exception thrown in create_JSON, exception is ' + \
                             str(e),'ERROR')
            # This error is fatal, so just release the lock on the num_threads queue
            # so that we can exit the whole system
            # We are about to put something in the error_q, which communicates to
            # other threads that the process needs to shut down.
            self.error_q.put([1,str(e)])
            self.num_of_REST_threads.get()
            return

    def recordset_builder_csv(self,start,end,id_col):
        '''Creates the JSON to send to REST from a the input data if the
        input data is in csv format'''
        #arcpy.AddMessage('relevant cols is {}'.format(self.relevent_columns))
        #arcpy.AddMessage('user_arcgis_fields is {}'.format(self.user_arcgis_fields))
        try:
            records = []
            #arcpy.AddMessage(self.user_arcgis_fields)
            #arcpy.AddMessage(self.relevent_columns)
            #arcpy.AddMessage("start is {}".format(start))
            #arcpy.AddMessage("end is {}".format(end))
            for i in range(start,end):
                try:
                    #arcpy.AddMessage("about to get a row")
                    row = next(self.prelim_csvReader)
                    #arcpy.AddMessage("able to get a row")
                    #arcpy.AddMessage("row is {}".format(row))
                except Exception as e:
                    if not self.check_item_resources:
                        arcpy.AddWarning("Unable to parse some records. Check item resources.")
                        self.check_item_resources = True
                    self.cannot_parse_records.append(i)
                    continue
                if len(row) != self.correct_len:
                    self.logger.log('Length 1 is {0}'.format(len(row)),'INFO')
                    self.logger.log('Length 2 is {0}'.format(len(self.all_fields)),'INFO')
                    # If your csv has less columns in a specific row
                    # than the number of header columns, then
                    # skip this row in  converting the batch to JSON
                    if len(row) == 0:
                        # if the row is empty...
                        self.num_empty_rows += 1
                    if self.analyze_obj.started_as_feature_service:
                        new_row = row[1:]
                        row = new_row
                    else:
                        if not self.check_item_resources:
                            arcpy.AddWarning("Unable to parse some records. Check item resources.")
                            self.check_item_resources = True
                        self.cannot_parse_records.append(i)
                        continue

                if self.text_qualifier == '\'':
                    row_to_use = []
                    for value in row:
                        row_to_use.extend(re.findall("'.*?'",value))
                    row = row_to_use
                    row_final = []
                    for val in row:
                        row_final.append(val.strip('\''))
                    row = row_final

                if self.object_id_in_input:
                    if row[id_col] != '':
                        try:
                            if int(row[id_col]) == -1:
                                # if id col exists but is -1, ignore it
                                self.object_id_in_input = False
                        except:
                            # if id col exists but is not an int
                            self.object_id_in_input = False
                            pass
                    if row[id_col] == '':
                        # if id col is empty...
                        self.object_id_in_input = False
                if self.object_id_in_input:
                    self.input_data[row[id_col]] = row
                else:
                    self.input_data[str(i)] = row
                relevant_row = []                                
                for attr_num in range(len(row)):
                    if attr_num in self.relevent_columns:
                        relevant_row.append(row[attr_num])
                #arcpy.AddMessage("relevant row is {}".format(relevant_row))
                if self.object_id_in_input:
                    attributes = {"attributes":{self.user_arcgis_fields[idx]: convert_to_int_if_possible(elem)\
                                        for idx,elem in enumerate(relevant_row)}}
                else:
                    attributes = {"attributes":{self.user_arcgis_fields[idx]: convert_to_int_if_possible(elem)\
                                        for idx,elem in enumerate(relevant_row)}}
                    # Since there is no object ID in the input, assign one
                    attributes["attributes"]["OBJECTID"] = i

                records.append(attributes)
            self.logger.log('We put another batch '+ \
                             'onto the input q, ending at {}'.format(end),'INFO')
            if not self.error_out:
                records_to_send = {"records":records}
                #arcpy.AddMessage(records_to_send)
                return records_to_send
            if self.error_out:
                return
            # if self.error_out:
            #     return errored_out_records
        except Exception as e:
            if isinstance(e,StopIteration):
                msg_code = 100190
                msg = self.ERROR_CODES[msg_code]
                self.addError(msg_code, msg)
            # self.logger.log('Exception in recordset_builder_csv {}'.format(e),'ERROR')
            if self.DEBUG:
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                self.logger.log('{0} {1} {2}'.format(exc_type, fname, exc_tb.tb_lineno),'ERROR')
            # Shut down all threads. If we cannot build recordsets, there
            # is no way the job will finish.
            self.error_q.put([1,str(e)])
            self.num_of_REST_threads.get()
            return "error"

    def recordset_builder_xls(self,start,end,id_col):
        '''Creates the JSON to send to REST from a the input data if the
        input data is in xlsx format'''
        try:
            records = []

            for i in range(start,end):
                # Gets called differently
                try:
                    ### need i + 1 because openpyxl starts indexing at 1 instead of at 0
                    if self.using_xlsx:
                        if not self.iter_rows_created:
                            self.iter_rows = self.in_worksheet.iter_rows()
                            self.iter_rows_created = True
                        row = next(self.iter_rows)
                        #arcpy.AddMessage("got row !!")
                    else:
                        row = self.in_worksheet.row(i)
                        #arcpy.AddMessage("got row !!")
                except Exception as e:
                    if not self.check_item_resources:
                        arcpy.AddWarning("Unable to parse some records. Check item resources.")
                        self.check_item_resources = True
                    self.cannot_parse_records.append(i)
                    continue
                new_row = []
                for entry in row:
                    if entry.value == None:
                        new_row.append("")
                    else:
                        new_row.append(entry.value)

                row = new_row
                #arcpy.AddMessage(row)
                if self.text_qualifier == '\'':
                    row_to_use = []
                    for value in row:
                        row_to_use.extend(re.findall("'.*?'",value))
                    row = row_to_use
                    row_final = []
                    for val in row:
                        row_final.append(val.strip('\''))
                    row = row_final

                if self.text_qualifier == "\"":
                    row_to_use = []
                    for value in row:
                        row_to_use.extend(re.findall('".*?"',value))
                    row = row_to_use
                    row_final = []
                    for val in row:
                        row_final.append(val.strip('\"'))
                    row = row_final
                # Creating a dictionary from the
                # input data so that we may write
                # the input data into the self.output file
                if self.object_id_in_input:
                    if row[id_col] != '':
                        try:
                            if int(row[id_col]) == -1:
                                # if id col exists but is -1, ignore it
                                self.object_id_in_input = False
                        except:
                            # if id col exists but is not an int
                            self.object_id_in_input = False
                            pass
                    if row[id_col] == '':
                        # if id col is empty...
                        self.object_id_in_input = False
                if self.object_id_in_input:
                    self.input_data[str(convert_to_int_if_possible(row[id_col]))] = row
                else:
                    self.input_data[str(i)] = row
                relevant_row = []
                for attr_num in range(len(row)):
                    if attr_num in self.relevent_columns:
                        # don't need this because strings are unicode by default in Python3
                        # if isinstance(row[attr_num],str):
                        #     relevant_row.append(convert_to_int_if_possible\
                        #                             (row[attr_num].\
                        #                              encode('utf-8')))
                        # else:
                        relevant_row.append(convert_to_int_if_possible\
                                                (row[attr_num]))
                #arcpy.AddMessage(relevant_row)
                if len(row) != self.correct_len:
                    # If your xls has less columns in a specific row
                    # than the number of header columns, then
                    # skip this row in  converting the batch to JSON
                    self.cannot_parse_records.append(i)
                    # self.logger.log('Row {} of your input file'.format(i) + \
                    #                 ' is not properly formatted xls','USER')
                    continue
                if self.object_id_in_input:
                    attributes = {"attributes":{self.user_arcgis_fields[j]:\
                                                    convert_to_int_if_possible\
                                                        (relevant_row[j])
                                        for j in range(len(self.relevant_user_fields))}}
                else:
                    attributes = {"attributes":{self.user_arcgis_fields[j]:\
                                                    convert_to_int_if_possible\
                                                        (relevant_row[j])
                                        for j in range(len(self.relevant_user_fields))}}
                    # Since there is no object ID in the input, assign one
                    attributes["attributes"]["OBJECTID"] = i

                records.append(attributes)
                #arcpy.AddMessage("sending records {}".format(records))
            if not self.error_out:
                records = {"records":records}
                #arcpy.AddMessage(records)
                return records
            if self.error_out:
                return
        except Exception as e:
            #self.logger.log('Exception in recordset_builder_xls {}'.format(e),'ERROR')
            if self.DEBUG:
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                self.logger.log('{0} {1} {2}'.\
                                format(exc_type, fname, exc_tb.tb_lineno),'ERROR')
            # Shut down all threads. If we cannot build recordsets, there
            # is no way the job will finish.
            self.error_q.put([1,str(e)])
            self.num_of_REST_threads.get()
            return "error"

    def recordset_builder_txt(self,start,end,id_col):
        '''Creates the JSON to send to REST from a the input data if the
        input data is in txt format'''
        try:
            records = []
            # self.total_counter gets initialized in the constructor
            for i in range(start,end):
                try:
                    string_row = self.f.readline()
                    if self.chars_per_row == '':
                        almost_row = string_row.strip('\n')
                    if self.fixed_width:
                        if self.chars_per_row != '':
                            # Special case where the whole file is in one line
                            records = self.one_row_recordset_builder(string_row,id_col)
                            return records
                        else:
                            almost_row = string_row
                        counter = 0
                        row = []
                        for num_of_chars in self.chars_per_column:
                            row.append(almost_row[counter:counter+int(num_of_chars)].strip(' '))
                            counter += int(num_of_chars)


                    # Get rid of delimiters and new line character
                    # and turn it into an array
                    if not self.fixed_width:
                        # we are only going to support splitting on tabs for
                        # txt files
                        row = almost_row.split('\t')
                except Exception as e:
                    self.cannot_parse_records.append(i)
                    self.logger.log('the problem is reading a row out of range ' +\
                                str(e),'ERROR')
                if self.text_qualifier == '\'':
                    row_to_use = []
                    for value in row:
                        row_to_use.extend(re.findall("'.*?'",value))
                    row = row_to_use
                    row_final = []
                    for val in row:
                        row_final.append(val.strip('\''))
                    row = row_final

                if self.text_qualifier == "\"":
                    row_to_use = []
                    for value in row:
                        row_to_use.extend(re.findall('".*?"',value))
                    row = row_to_use
                    row_final = []
                    for val in row:
                        row_final.append(val.strip('\"'))
                    row = row_final
                # Creating a dictionary from the
                # input data so that we may write
                # the input data into the self.output file
                if self.object_id_in_input:
                    if row[id_col] != '':
                        try:
                            if int(row[id_col]) == -1:
                                # if id col exists but is -1, ignore it
                                self.object_id_in_input = False
                        except:
                            # if id col exists but is not an int
                            self.object_id_in_input = False
                            pass
                    if row[id_col] == '':
                        # if id col is empty...
                        self.object_id_in_input = False
                if self.object_id_in_input:
                    #NOTE: THIS MIGHT NEED TO BE CHANGED TO A REGEX OF THINGS THAT AREN'T INTS
                    self.input_data[str(row[id_col]).strip(' ')] = row
                else:
                    self.input_data[str(i)] = row
                relevant_row = []
                for attr_num in range(len(row)):
                    if attr_num in self.relevent_columns:
                        relevant_row.append(convert_to_int_if_possible\
                                                    (row[attr_num]))

                if len(row) != self.correct_len:
                    #If your csv has less columns in a specific row
                    #than the number of header columns, then
                    #skip this row in  converting the batch to JSON
                    self.cannot_parse_records.append(i)
                    self.logger.log('Row {} of your input file'.format(i) + \
                          ' is not properly formatted txt','USER')
                    continue

                if self.object_id_in_input:
                    attributes = {"attributes":{self.user_arcgis_fields[j]:\
                                                    convert_to_int_if_possible\
                                                        (relevant_row[j])
                                        for j in range(len(self.relevant_user_fields))}}
                else:
                    attributes = {"attributes":{self.user_arcgis_fields[j]:\
                                                    convert_to_int_if_possible\
                                                        (relevant_row[j])
                                        for j in range(len(self.relevant_user_fields))}}
                    # Since there is no object ID in the input, assign one
                    attributes["attributes"]["OBJECTID"] = i

                records.append(attributes)

            self.logger.log('We put another batch onto the input q, ending at ' +\
                        str(end),'INFO')
            if not self.error_out:
                records = {"records":records}
                return records
            if self.error_out:
                return
        except Exception as e:
            # self.logger.log('Exception in recordset_builder_txt {}'.format(e),'ERROR')
            if self.DEBUG:
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                self.logger.log('{0} {1} {2}'.format(exc_type, fname, exc_tb.tb_lineno),'ERROR')
            # Shut down all threads. If we cannot build recordsets, there
            # is no way the job will finish.
            self.error_q.put([1,str(e)])
            self.num_of_REST_threads.get()
            return "error"

    def recordset_builder_fs(self,batch_num):
        try:
            if batch_num == 0:
                first_batch = True
            else:
                first_batch = False

            # determine start and ending points
            if (batch_num == self.num_of_batches-1):
                # if its the last batch
                if batch_num == 0:
                    # If its also the first batch...
                    start = self.header_rows_to_skip
                    end = self.remainder + 1
                    if self.remainder == 0 and self.row_count > 1:
                        # Special case of 1000 rows exactly
                        end = self.row_count
                else:
                    #if its the last batch, end at the remainder
                    start = batch_num*self.batch_size + 1
                    end = start + self.remainder

                    if self.remainder == 0:
                        # Case where number of rows is a multiple of
                        # batch_size
                        end = self.batch_size + start
            else:
                # if its the first batch
                if batch_num == 0:
                    start = self.header_rows_to_skip
                    end = start + self.batch_size
                # this means its a middle-batch...
                else:
                    start = batch_num*self.batch_size + 1
                    end = start + self.batch_size

            # begin building recordsets
            #arcpy.AddMessage("start is {}".format(start))
            #arcpy.AddMessage("end is {}".format(end))
            records = []
            for i in range(start,end):
                try:
                    row = next(self.cursor)
                    if first_batch and i == start:
                        # this is the very first record in the input GDB
                        # figure out if this input GDB has OBJECTIDs that 
                        # start at 1. If it does, we can assume its OID
                        # column is "regular" (starts at 1, increases by 1)
                        if row[self.oid_index] == 1:
                            self.oid_table_starts_at_one = True
                        else:
                            self.oid_table_starts_at_one = False
                    #arcpy.AddMessage("row is {}".format(row))
                except Exception as e:
                    self.cannot_parse_records.append(i)
                    #arcpy.AddWarning("Issue is: {}".format(e))
                    pass

                if len(row) != self.correct_len:
                    self.logger.log('Length 1 is {0}'.format(len(row)),'INFO')
                    self.logger.log('Length 2 is {0}'.format(len(self.all_fields)),'INFO')
                    # If your csv has less columns in a specific row
                    # than the number of header columns, then
                    # skip this row in  converting the batch to JSON
                    self.cannot_parse_records.append(i)
                    if len(row) == 0:
                        # if the row is empty...
                        self.num_empty_rows += 1
                    # if self.analyze_obj.started_as_feature_service:
                    #     new_row = row[1:]
                    #     row = new_row
                    else:
                        self.logger.log('Row {} of your input file'.format(i) + \
                          ' is not properly formatted csv','INFO')
                        continue

                if self.text_qualifier == '\'':
                    row_to_use = []
                    for value in row:
                        row_to_use.extend(re.findall("'.*?'",value))
                    row = row_to_use
                    row_final = []
                    for val in row:
                        row_final.append(val.strip('\''))
                    row = row_final

                # we can say row[self.oid_index] because we know the table has an OID field at position self.oid_index
                if self.oid_table_starts_at_one:
                    self.input_data[str(row[self.oid_index])] = row
                else:
                    self.input_data[str(i)] = row
                relevant_row = []
                for attr_num in range(len(row)):
                    if attr_num in self.relevent_columns:
                        # if isinstance(row[attr_num],str):
                        #     relevant_row.append(row[attr_num].encode('utf-8'))
                        # else:
                        relevant_row.append(row[attr_num])
                        # to_append = row[attr_num]
                        # if "\r\n" in to_append:
                        #     to_append = to_append.strip("\r\n")
                        # relevant_row.append(to_append)
                attributes = {"attributes":{self.user_arcgis_fields[j]:relevant_row[j]\
                                    for j in range(len(self.relevant_user_fields))}}

                usr_ags_flds_lower = [fld.lower() for fld in self.user_arcgis_fields]
                if self.oid_table_starts_at_one:
                    attributes["attributes"]["OBJECTID"] = row[self.oid_index]
                    if not isinstance(row[self.oid_index], int):
                        self.logger.log('ObjectID is not an integer','ERROR')
                        self.error_out = True
                else:
                    attributes["attributes"]["OBJECTID"] = i

                records.append(attributes)
            self.logger.log('We put another batch '+ \
                             'onto the input q','INFO')
            REST_to_send = {"records":records}
            #arcpy.AddMessage("REST TO SEND IS: {}".format(REST_to_send))
            if not self.error_q.empty():
                self.error_out = True
            if not self.error_out:
                self.processed_data_q.put([batch_num,REST_to_send])
            if self.error_out:
                return
            # if self.error_out:
            #     return errored_out_records
        except Exception as e:
            if isinstance(e,StopIteration):
                msg_code = 100190
                msg = self.ERROR_CODES[msg_code]
                self.addError(msg_code, msg)
            # self.logger.log('Exception in recordset_builder_csv {}'.format(e),'ERROR')
            if self.DEBUG:
                arcpy.AddError(e)
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                self.logger.log('{0} {1} {2}'.format(exc_type, fname, exc_tb.tb_lineno),'ERROR')
            # Shut down all threads. If we cannot build recordsets, there
            # is no way the job will finish.
            self.error_q.put([1,str(e)])
            self.num_of_REST_threads.get()
            return "error"

    def one_row_recordset_builder(self,string_row,id_col):

        num_of_rows, leftovers = divmod(len(string_row),self.chars_per_row)
        try:
            records = []
            # self.total_counter gets initialized in the constructor
            for i in range(num_of_rows):
                try:
                    almost_row = string_row[self.total_counter:self.total_counter+int(self.chars_per_row)]
                    self.total_counter += int(self.chars_per_row)

                    counter = 0
                    row = []
                    # list_to_iter = self.chars_per_column.split(',')
                    for num_of_chars in self.chars_per_column:
                        row.append(almost_row[counter:counter+int(num_of_chars)].strip(' '))
                        counter += int(num_of_chars)
                except Exception as e:
                    self.logger.log('the problem is reading a row out of range ' +\
                                str(e),'ERROR')
                if self.text_qualifier == '\'':
                    row_to_use = []
                    for value in row:
                        row_to_use.extend(re.findall("'.*?'",value))
                    row = row_to_use
                    row_final = []
                    for val in row:
                        row_final.append(val.strip('\''))
                    row = row_final

                if self.text_qualifier == "\"":
                    row_to_use = []
                    for value in row:
                        row_to_use.extend(re.findall('".*?"',value))
                    row = row_to_use
                    row_final = []
                    for val in row:
                        row_final.append(val.strip('\"'))
                    row = row_final
                # Creating a dictionary from the
                # input data so that we may write
                # the input data into the self.output file
                if self.object_id_in_input:
                    #NOTE: THIS MIGHT NEED TO BE CHANGED TO A REGEX OF THINGS THAT AREN'T INTS
                    self.input_data[str(row[id_col]).strip(' ')] = row
                else:
                    self.input_data[str(i)] = row
                relevant_row = []
                for attr_num in range(len(row)):
                    if attr_num in self.relevent_columns:
                        relevant_row.append(convert_to_int_if_possible\
                                                    (row[attr_num]))

                if len(row) != len(self.all_fields):
                    #If your csv has less columns in a specific row
                    #than the number of header columns, then
                    #skip this row in  converting the batch to JSON
                    self.logger.log('Row {} of your input file'.format(i) + \
                          ' is not properly formatted txt','USER')
                    continue

                if self.object_id_in_input:
                    attributes = {"attributes":{self.user_arcgis_fields[j]:\
                                                    convert_to_int_if_possible\
                                                        (relevant_row[j])
                                        for j in range(len(self.relevant_user_fields))}}
                else:
                    attributes = {"attributes":{self.user_arcgis_fields[j]:\
                                                    convert_to_int_if_possible\
                                                        (relevant_row[j])
                                        for j in range(len(self.relevant_user_fields))}}
                    # Since there is no object ID in the input, assign one
                    attributes["attributes"]["OBJECTID"] = i

                records.append(attributes)

            records = {"records":records}
            self.logger.log('We put another batch onto the input q, ending at ' +\
                        str(num_of_rows),'INFO')
            return records
        except Exception as e:
            # self.logger.log('Exception in recordset_builder_txt {}'.format(e),'ERROR')
            if self.DEBUG:
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                self.logger.log('{0} {1} {2}'.format(exc_type, fname, exc_tb.tb_lineno),'ERROR')
            # Shut down all threads. If we cannot build recordsets, there
            # is no way the job will finish.
            self.error_q.put([1,str(e)])
            self.num_of_REST_threads.get()
            return "error"

    def create_results_csv(self):
        '''Only gets called if the output_file_format is csv.
            Takes rest responses and writes them to self.output.csv'''
        # Open the file for the first time, write the header row
        #arcpy.AddMessage('inside create results')
        header_has_been_written = False
        #arcpy.AddMessage('inside creating_results future')

        # This is a nice ordering for how to write out the self.output
        # We need to know the service self.output fields
        # Let's query for them
        # This line below gets rid of the 'geocodeAddresses' part of the URL
        
        cand_fields = self.analyze_obj.json_info_response['candidateFields']


        try:
            ctr = 0
            for field in self.output_fields:
                # remove extrenuous spaces
                # we will remove junk input for this later
                if field == "Shape X" or field == "Shape Y":
                    continue
                new_field = field.replace(' ','')
                self.output_fields[ctr] = new_field
                ctr += 1
            self.nice_ordering = []
            # Now we only want to return certain candidate fields
            for field in cand_fields:
                self.nice_ordering.append(field['name'])
            # If user is hitting a local geocoding service, then candidateFields
            # does not contain status even though it will be an output. RETRIEVE
            # THAT OUTPUT!!!
            if "Status" not in self.nice_ordering:
                self.nice_ordering.append("Status")
            try:
                # Remove the geometry if necessary
                self.nice_ordering.remove('Shape')

            except Exception as e:
                # I guess it wasn't necessary!
                pass
            if '*' in self.output_fields:
                self.output_fields.remove('*')
                particular_output_fields = False
            elif self.output_fields == '':
                particular_output_fields = False
            else:
                particular_output_fields = True

            if particular_output_fields:
                output_fields_lower = [x.lower() for x in self.output_fields]
                nice_ordering_lower = [x.lower() for x in self.nice_ordering]
                for requested_field in output_fields_lower:
                    if requested_field not in nice_ordering_lower:
                        if requested_field == "shape x" or requested_field == "shape y":
                            continue
                        output_fields_lower.remove(requested_field)
                        self.logger.log(\
                            'Your requested self.output field ' +\
                            '{} does not exist,'.format(requested_field) +\
                            ' skipping adding it to the self.output','USER')
                # Now that we have removed junk possible self.output...
                if not output_fields_lower:
                    # Then the list is empty!!! set boolean to False
                    particular_output_fields = False
                desired_fields = self.output_fields
                # indexes_to_keep = []
                # for output_field in output_fields_lower:
                #     if output_field in nice_ordering_lower:
                #         counter = 0
                #         for service_field in nice_ordering_lower:
                #             if service_field == output_field:
                #                 indexes_to_keep.append(counter)
                #             counter += 1
                # desired_fields = []
                # for index in indexes_to_keep:
                #     desired_fields.append(self.nice_ordering[index])

        except Exception as e:
            # self.logger.log('Exception in create_results_csv {}'\
                            # .format(e),'ERROR')
            if self.DEBUG:
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                self.logger.log('{0} {1} {2}'.format(exc_type, fname, exc_tb.tb_lineno),'ERROR')
            # Shut down all threads. If we cannot build recordsets, there
            # is no way the job will finish.
            self.error_q.put([1,str(e)])
            self.num_of_REST_threads.get()
            return "error"

        #arcpy.AddMessage('checkpoint 1')
        for batch_num in range(self.num_of_batches):
            while self.REST_response_q.empty():
                # We wait for REST responses. In the meantime,
                # be ready to shut down this thread in case a different
                # thread encounters an error
                
                #arcpy.AddMessage("waiting and chewing mem")
                time.sleep(0.2)
                if not self.error_q.empty():
                    self.logger.log('there is an error somewhere,' + \
                                     'shut down create_results','INFO')
                    self.error_out = True
                    break
                if self.is_cancelled:
                    error = 'The script has been aborted by the user. Returning geocoding results thus far'
                    self.logger.log('The script has been aborted by the user. Returning geocoding results thus far',\
                                    'USER_ERROR')
                    self.error_out = True
                    # # This error is fatal, so just release the lock on the num_threads queue
                    # so that we can exit the whole system
                    self.error_q.put([1,str(error)])
                    # self.num_of_REST_threads.get()
                    break
                if self.error_out:
                    # already has an error
                    break
            if not self.error_q.empty():
                # There is an error somewhere, shut down all threads
                self.error_out = True
                self.logger.log('Shutting down create_results','INFO')
            try:
                #arcpy.AddMessage('about to error out')
                if self.error_out:
                    if particular_output_fields:
                        cntr = 0
                        for field in desired_fields:
                            if "Shape X" == field:
                                self.shapex_field = cntr
                            cntr += 1
                    else:
                        cntr = 0
                        for field in self.nice_ordering:
                            if "Status" == field:
                                self.status_field = cntr
                            cntr += 1
                    if not header_has_been_written:
                        if self.all_fields[0] == 'ID':
                            # Excel won't open csv files that have "ID" as their
                            # first value. This is the work-around
                            self.all_fields[0] = 'id'

                        if not particular_output_fields:
                            all_fields = []
                            for field in self.all_fields:
                                all_fields.append("USER_{}".format(field))

                            mapped_user_fields = []
                            for field in self.user_field_nums:
                                mapped_user_fields.append("IN_{}".format(field[1]))
                        else:
                            all_fields = self.all_fields
                            mapped_user_fields = self.user_field_nums
                        # if self.analyze_obj.started_as_feature_service:
                        #     all_fields_minus_oid.remove('OID')
                        # Write the first line of the csv as the fields
                        if particular_output_fields:
                            self.csvWriter.writerow(desired_fields + all_fields)
                            num_of_out_fields = len(desired_fields)

                        else:
                            self.csvWriter.writerow(self.nice_ordering + mapped_user_fields + all_fields)
                            num_of_out_fields = len(self.nice_ordering)

                        header_has_been_written = True

                        if particular_output_fields:
                            left = num_of_out_fields - self.shapex_field - 1
                            if list(self.input_data.keys()):
                                keys_to_iterate = list(self.input_data.keys())
                                empty_list = [""] * self.shapex_field
                                second_empty_list = [""] * left
                                for key in keys_to_iterate:
                                    # mapped_fields_to_write_still = []
                                    # for num in self.user_field_nums:
                                    #     mapped_fields_to_write_still.append(self.input_data[key][num[0]])
                                    self.csvWriter.writerow(empty_list + [""] + second_empty_list + list(self.input_data[key]))
                                    del self.input_data[key]

                        else:
                            left = num_of_out_fields - self.status_field - 1
                            if list(self.input_data.keys()):
                                keys_to_iterate = list(self.input_data.keys())
                                empty_list = [""] * self.status_field
                                second_empty_list = [""] * left
                                for key in keys_to_iterate:
                                    mapped_fields_to_write_still = []
                                    for num in self.user_field_nums:
                                        mapped_fields_to_write_still.append(self.input_data[key][num[0]])
                                    self.csvWriter.writerow(empty_list + ["U"] + second_empty_list + list(mapped_fields_to_write_still) + list(self.input_data[key]))
                                    del self.input_data[key]
                    else:
                        # tries = 0
                        # while self.leftover_records_q.empty():
                        #     time.sleep(0.5)
                        #     tries += 1
                        #     if tries > 3:
                        #         break
                        if particular_output_fields:
                            num_of_out_fields = len(desired_fields)

                        else:
                            num_of_out_fields = len(self.nice_ordering)

                        if particular_output_fields:
                            left = num_of_out_fields - self.shapex_field - 1
                            if list(self.input_data.keys()):
                                keys_to_iterate = list(self.input_data.keys())
                                empty_list = [""] * self.shapex_field
                                second_empty_list = [""] * left
                                for key in keys_to_iterate:
                                    # mapped_fields_to_write_still = []
                                    # for num in self.user_field_nums:
                                    #     mapped_fields_to_write_still.append(self.input_data[key][num[0]])
                                    self.csvWriter.writerow(empty_list + [""] + second_empty_list + list(self.input_data[key]))
                                    del self.input_data[key]
                        else:
                            left = num_of_out_fields - self.status_field - 1
                            if list(self.input_data.keys()):
                                keys_to_iterate = list(self.input_data.keys())
                                empty_list = [""] * self.status_field
                                second_empty_list = [""] * left
                                for key in keys_to_iterate:
                                    mapped_fields_to_write_still = []
                                    for num in self.user_field_nums:
                                        mapped_fields_to_write_still.append(self.input_data[key][num[0]])
                                    self.csvWriter.writerow(empty_list + ["U"] + second_empty_list + list(mapped_fields_to_write_still) + list(self.input_data[key]))
                                    del self.input_data[key]
                    continue
            except Exception as e:
                arcpy.AddError(e)
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                self.logger.log('{0} {1} {2}'.format(exc_type, fname, exc_tb.tb_lineno),'ERROR')

            self.logger.log('the REST response q is not empty, going to try  ' + \
                             'to write out batch number {}'.format(batch_num),'INFO')
            full_response = self.REST_response_q.get()
            rest_response = full_response[1]
            #arcpy.AddMessage(rest_response)
            #write the header
            if not header_has_been_written:
                self.logger.log('writing headers for the first time','INFO')
                try:
                    # Basically just pick one from the list
                    sample_loc = rest_response['locations'][0]
                    self.spatial_reference = rest_response['spatialReference']['wkid']
                except Exception as e:
                    msg_code = 100191
                    msg = self.ERROR_CODES[msg_code]
                    self.addError(msg_code, msg)
                    header_has_been_written = True
                    continue
                    # Try continue, see if you can still write other
                    # REST responses, maybe this was just a bad response.
                    # The REST response you just skipped will show up
                    # as a DROPPED record in the end
                if self.all_fields[0] == 'ID':
                    # Excel won't open csv files that have "ID" as their
                    # first value. This is the work-around
                    self.all_fields[0] = 'id'

                # Remove the OID Column (artifact of moving from table to csv)
                if not particular_output_fields:
                    all_fields = []
                    for field in self.all_fields:
                        all_fields.append("USER_{}".format(field))
                else:
                    all_fields = self.all_fields

                self.locator_fields = []
                self.locator_aliases = {}

                if self.single_line_only:
                    singleline_field_name = self.analyze_obj.json_info_response['singleLineAddressField']['name']
                    singleline_field_alias = self.analyze_obj.json_info_response['singleLineAddressField']['alias']
                    self.locator_fields.append("IN_{}".format(singleline_field_name))
                    self.locator_aliases[singleline_field_name] = singleline_field_alias

                else:
                    locator_flds = self.analyze_obj.json_info_response['addressFields']
                    # categories_inner = json_response['categories']['categories']
                    for addr in locator_flds:
                        self.locator_fields.append("IN_{}".format(addr['name']))
                        self.locator_aliases[addr['name']] = addr['alias']

                # for field in self.user_field_nums:
                #     mapped_user_fields.append("IN_{}".format(field[1]))



                # if self.analyze_obj.started_as_feature_service:
                #     all_fields_minus_oid.remove('OID')
                # Write the first line of the csv as the fields
                #arcpy.AddMessage('about to hit bad line')
                if particular_output_fields:
                    # if "x" not in desired_fields and "X" not in desired_fields:
                    #     desired_fields.append("X")
                    # if "y" not in desired_fields and "Y" not in desired_fields:
                    #     desired_fields.append("Y")
                    # if "score" not in desired_fields and "Score" not in desired_fields:
                    #     desired_fields.append("Score")
                    # if "status" not in desired_fields and "Status" not in desired_fields:
                    #     desired_fields.append("Status")
                    if "Match_addr" in desired_fields:
                        index_of_match_addr = desired_fields.index("Match_addr")
                        desired_fields.insert(index_of_match_addr, "Match_type")
                    # desired_fields.append("Shape X")
                    # desired_fields.append("Shape Y")
                    #arcpy.AddMessage("getting written are: {}".format(desired_fields + all_fields))
                    self.csvWriter.writerow(desired_fields + all_fields)
                    #arcpy.AddMessage(self.nice_ordering + self.locator_fields + all_fields)
                    # num_of_out_fields = len(desired_fields)

                else:
                    if "x" not in self.nice_ordering and "X" not in self.nice_ordering:
                        self.nice_ordering.append("X")
                    if "y" not in self.nice_ordering and "Y" not in self.nice_ordering:
                        self.nice_ordering.append("Y")
                    if "score" not in self.nice_ordering and "Score" not in self.nice_ordering:
                        self.nice_ordering.append("Score")
                    if "status" not in self.nice_ordering and "Status" not in self.nice_ordering:
                        self.nice_ordering.append("Status")
                    index_of_match_addr = self.nice_ordering.index("Match_addr")
                    self.nice_ordering.insert(index_of_match_addr, "Match_type")
                    #self.nice_ordering.append("Match_type")
                    self.nice_ordering.append("Shape X")
                    self.nice_ordering.append("Shape Y")
                    # arcpy.AddMessage("mapped_user_fields is {}".format(mapped_user_fields))
                    # arcpy.AddMessage("all_fields is {}".format(all_fields))
                    self.csvWriter.writerow(self.nice_ordering + self.locator_fields + all_fields)
                    #arcpy.AddMessage(self.nice_ordering + self.locator_fields + all_fields)
                    # num_of_out_fields = len(self.nice_ordering)
                header_has_been_written = True
                self.logger.log('Header was written successfully','INFO')

            # Create each row to be written to the self.output.csv
            try:
                self.logger.log('going to attempt to write batch ' +\
                                    str(batch_num),'INFO')
                try:
                    dummy_var = rest_response['locations']

                except Exception as e:
                    if self.DEBUG:
                        exc_type, exc_obj, exc_tb = sys.exc_info()
                        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                        self.logger.log('{0} {1} {2}'.format(exc_type, fname, exc_tb.tb_lineno),'ERROR')
                    # self.logger.log('issue with rest response {}'.format(e),'ERROR')
                    self.error_q.put([1,str(e)])
                    self.num_of_REST_threads.get()
                    return
                for location in dummy_var:
                    try:
                        attributes = location['attributes']

                    except Exception as e:
                        # self.logger.log('issue with rest response '.format(e),'ERROR')
                        if self.DEBUG:
                            exc_type, exc_obj, exc_tb = sys.exc_info()
                            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                            self.logger.log('{0} {1} {2}'.format(exc_type, fname, exc_tb.tb_lineno),'ERROR')
                        self.error_q.put([1,str(e)])
                        self.num_of_REST_threads.get()
                        return
                    attr_list = []
                    attr_options = list(attributes.keys())
                    try:
                        if particular_output_fields:
                            for attr in desired_fields:
                                if attr == "Shape X":
                                    if "location" in location:
                                        attr_list.append(location["location"]["x"])
                                    else:
                                        attr_list.append("")
                                elif attr == "Shape Y":
                                    if "location" in location:
                                        attr_list.append(location["location"]["y"])
                                    else:
                                        attr_list.append("")
                                elif attr not in attr_options:
                                    attr_list.append('')
                                elif attr == "Address":
                                    attr_list.append(location["address"])
                                elif attributes[attr] == '':
                                    attr_list.append('')
                                # elif isinstance(attributes[attr],basestring):
                                #     cleaned_string = attributes[attr].encode('utf-8')
                                #     attr_list.append(cleaned_string)
                                else:
                                    attr_list.append(attributes[attr])
                        else:
                            for attr in self.nice_ordering:
                                if attr == "Shape X":
                                    if "location" in location:
                                        attr_list.append(location["location"]["x"])
                                    else:
                                        attr_list.append("")
                                elif attr == "Shape Y":
                                    if "location" in location:
                                        attr_list.append(location["location"]["y"])
                                    else:
                                        attr_list.append("")
                                elif attr not in attr_options:
                                    attr_list.append('')
                                elif attr == "Address":
                                    attr_list.append(location["address"])
                                elif attributes[attr] == '':
                                    attr_list.append('')
                                # elif isinstance(attributes[attr],basestring):
                                #     cleaned_string = attributes[attr].encode('utf-8')
                                #     attr_list.append(cleaned_string)
                                else:
                                    attr_list.append(attributes[attr])

                        #arcpy.AddMessage(attr_list)

                    except Exception as e:
                        if self.DEBUG:
                            exc_type, exc_obj, exc_tb = sys.exc_info()
                            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                            self.logger.log('{0} {1} {2}'.format(exc_type, fname, exc_tb.tb_lineno),'ERROR')

                    # Write to the self.output file
                    try:
                        if str(attributes["ResultID"]) == '-1':
                            self.logger.log('A REST response had a -1 resultID',\
                                        'INFO')
                            continue
                    except Exception as e:
                        self.logger.log('One record does not have a result ID',\
                                    'INFO')

                    # Get the relevant input row from the dictionary
                    # Delete from the input dictionary because we are done
                    # with that row. Tries to keep the number of rows stored
                    # in memory under 8*batch_size
                    try:
                        input_row = self.input_data.get(str(attributes["ResultID"]))
                        #arcpy.AddMessage("got an input row: {}".format(input_row))
                        if input_row is None:
                            self.logger.log('Empty dictionary entry','INFO')
                            if self.dropped_records_q.full():
                                self.error_q.put([1,str('Dropped more than 20 records')])
                                self.num_of_REST_threads.get()
                                return
                            else:
                                self.dropped_records_q.put(str(attributes["ResultID"]))
                            continue
                        del self.input_data[str(attributes["ResultID"])]
                    except Exception as e:
                        # Not a SUPER severe error...
                        self.logger.log('Dictionary error {}'.format(e),'INFO')

                    # mapped_fields_to_write_still = []
                    # for num in self.user_field_nums:
                    #     mapped_fields_to_write_still.append(input_row[num[0]])

                    dict_of_user_fields_to_their_index_number = {}
                    for num,tup in enumerate(self.field_info):
                        dict_of_user_fields_to_their_index_number[tup[0]] = num

                    if not particular_output_fields:
                        locator_fields_to_write = []

                        for fld in self.locator_fields:
                            if fld[3:] in self.user_arcgis_fields:
                                #arcpy.AddWarning('considering field {}'.format(fld[3:]))
                                for elem_num in range(len(self.field_mapping)):
                                    if self.field_mapping[elem_num][1] == fld[3:]:
                                        #arcpy.AddWarning('considering field2 {}'.format(self.field_mapping[elem_num][1]))
                                        # so now we need the self.field_mapping[elem_num][0] value for this row...
                                        locator_fields_to_write.append(input_row[dict_of_user_fields_to_their_index_number[self.field_mapping[elem_num][0]]])
                            else:
                                locator_fields_to_write.append("")
                        try:
                            row_to_write = list(attr_list) + list(locator_fields_to_write) + list(input_row)
                            #arcpy.AddMessage("row to write is {}".format(row_to_write))
                        except Exception as e:
                            self.logger.log('We are having trouble ' + \
                                                'concatenating rows {}'.format(e),'ERROR')
                    else:
                        try:
                            row_to_write = list(attr_list) + list(input_row)
                            #arcpy.AddMessage("row to write is {}".format(row_to_write))
                        except Exception as e:
                            self.logger.log('We are having trouble ' + \
                                                'concatenating rows {}'.format(e),'ERROR')

                    try:
                        self.csvWriter.writerow(row_to_write)

                    except Exception as e:
                        self.logger.log('Unable to write to csv {}'.format(e),'ERROR')

                self.logger.log('Just wrote batch {}'.format(batch_num) + \
                                        ' to self.output.csv','INFO')

                full_response = None
                rest_response = None
                gc.collect()

            except Exception as e:
                # Clearly a bad batch

                if self.DEBUG:
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    self.logger.log('{0} {1} {2}'.format(exc_type, fname, exc_tb.tb_lineno),'ERROR')
                    self.logger.log('There was a problem with batch {0} : {1}'.format(batch_num,e),'ERROR')
                # Pass in hopes of us being able to still write the
                # rest of t he results
                pass
        self.logger.log('DONE writing to {}'.format(self.output_name + '.csv'),'INFO')

    def create_results_xls(self):
        '''Only gets called if the output_file_format is xls.
            Takes rest responses and writes them to self.output.xls'''
        counter = 0
        header_has_been_written = False

        cand_fields = self.analyze_obj.json_info_response['candidateFields']
        try:
            ctr = 0
            for field in self.output_fields:
                # remove extrenuous spaces
                # we will remove junk input for this later
                if field == "Shape X" or field == "Shape Y":
                    continue
                new_field = field.replace(' ','')
                self.output_fields[ctr] = new_field
                ctr += 1
            self.nice_ordering = []
            # Now we only want to return certain candidate fields
            cand_fields = self.analyze_obj.json_info_response['candidateFields']
            for field in cand_fields:
                self.nice_ordering.append(field['name'])
            # If user is hitting a local geocoding service, then candidateFields
            # does not contain status even though it will be an output. RETRIEVE
            # THAT OUTPUT!!!
            if "Status" not in self.nice_ordering:
                self.nice_ordering.append("Status")
            try:
                # Remove the geometry if necessary
                self.nice_ordering.remove('Shape')
            except Exception as e:
                # I guess it wasn't necessary!
                pass
            if '*' in self.output_fields:
                self.output_fields.remove('*')
                particular_output_fields = False
            elif self.output_fields == '':
                particular_output_fields = False

            else:
                particular_output_fields = True

            if particular_output_fields:
                output_fields_lower = [x.lower() for x in self.output_fields]
                nice_ordering_lower = [x.lower() for x in self.nice_ordering]
                for requested_field in output_fields_lower:
                    if requested_field not in nice_ordering_lower:
                        if requested_field == "shape x" or requested_field == "shape y":
                            continue
                        output_fields_lower.remove(requested_field)
                        self.logger.log(\
                            'Your requested self.output field ' +\
                            '{} does not exist,'.format(requested_field) +\
                            ' skipping adding it to the self.output','USER')
                # Now that we have removed junk possible self.output...
                if not output_fields_lower:
                    # Then the list is empty!!! set boolean to False
                    particular_output_fields = False

                desired_fields = self.output_fields
                # indexes_to_keep = []
                # for output_field in output_fields_lower:
                #     if output_field in nice_ordering_lower:
                #         field_counter = 0
                #         for service_field in nice_ordering_lower:
                #             if service_field == output_field:
                #                 indexes_to_keep.append(field_counter)
                #             field_counter += 1
                # desired_fields = []
                # for index in indexes_to_keep:
                #     desired_fields.append(self.nice_ordering[index])
        except Exception as e:
            if self.DEBUG:
                self.logger.log('Exception in create_results_xls {}'\
                                .format(e),'ERROR')
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                self.logger.log('{0} {1} {2}'.\
                                format(exc_type, fname, exc_tb.tb_lineno),'ERROR')
            # Shut down all threads. If we cannot build recordsets, there
            # is no way the job will finish.
            self.error_q.put([1,str(e)])
            self.num_of_REST_threads.get()
            return "error"
        for batch_num in range(self.num_of_batches):
            while self.REST_response_q.empty():
                # We wait for REST responses. In the meantime, be ready to shut down just in case
                time.sleep(0.1)
                if not self.error_q.empty():
                    self.logger.log('there is an error somewhere, ' +\
                                'shut down create_results','INFO')
                    self.error_out = True
                    break
                if self.is_cancelled:
                    error = 'The script has been aborted by the user. Returning geocoding results thus far'
                    self.logger.log(error,'USER_ERROR')
                    self.error_out = True
                    # # This error is fatal, so just release the lock on the num_threads queue
                    # so that we can exit the whole system
                    self.error_q.put([1,str(error)])
                    # self.num_of_REST_threads.get()
                    break
                if self.error_out:
                    # already has an error
                    break
            if not self.error_q.empty():
                # There is an error somewhere, shut down all threads
                self.error_out = True
                self.logger.log('shutting down create_results','INFO')

            if self.error_out:
                if particular_output_fields:
                    cntr = 0
                    for field in desired_fields:
                        if "Shape X" == field:
                            self.shapex_field = cntr
                        cntr += 1
                else:
                    cntr = 0
                    for field in self.nice_ordering:
                        if "Status" == field:
                            self.status_field = cntr
                        cntr += 1
                if not header_has_been_written:
                    if self.all_fields[0] == 'ID':
                        # Excel won't open csv files that have "ID" as their
                        # first value. This is the work-around
                        self.all_fields[0] = 'id'

                    # Remove the OID Column (artifact of moving from table to csv)
                    # all_fields_minus_oid = self.all_fields
                    # if self.analyze_obj.started_as_feature_service:
                    #     all_fields_minus_oid.remove('OID')

                    # Remove the OID Column (artifact of moving from table to csv)

                    if not particular_output_fields:
                        all_fields = []
                        for field in self.all_fields:
                            all_fields.append("USER_{}".format(field))

                        mapped_user_fields = []
                        for field in self.user_field_nums:
                            mapped_user_fields.append("IN_{}".format(field[1]))
                    else:
                        all_fields = self.all_fields
                        mapped_user_fields = self.user_field_nums

                    
                    # Write the first line of the csv as the fields
                    if particular_output_fields:
                        write_row = desired_fields + all_fields
                        for i in range(len(write_row)):
                            self.out_worksheet.write(0,i,write_row[i])
                        counter += 1
                    else:
                        write_row = self.nice_ordering + mapped_user_fields + all_fields
                        for i in range(len(write_row)):
                            self.out_worksheet.write(0,i,write_row[i])
                        counter += 1
                    header_has_been_written = True


                    if particular_output_fields:
                        left = num_of_out_fields - self.shapex_field - 1
                        if list(self.input_data.keys()):
                            keys_to_iterate = list(self.input_data.keys())
                            empty_list = [""] * self.shapex_field
                            second_empty_list = [""] * left
                            for key in keys_to_iterate:
                                record = empty_list + [""] + second_empty_list + self.input_data[key]
                                for i in range(len(record)):
                                    try:
                                        self.out_worksheet.write(counter,i,record[i])
                                    except Exception as e:
                                        self.logger.log('{0} could not write an entry {1}'.format(record[i],e),'USER')
                                del self.input_data[key]
                                counter += 1
                    else:
                        left = num_of_out_fields - self.status_field - 1
                        if list(self.input_data.keys()):
                            keys_to_iterate = list(self.input_data.keys())
                            empty_list = [""] * self.status_field
                            second_empty_list = [""] * left
                            for key in keys_to_iterate:
                                record = empty_list + ["U"] + second_empty_list + mapped_fields_to_write_still + self.input_data[key]
                                for i in range(len(record)):
                                    try:
                                        self.out_worksheet.write(counter,i,record[i])
                                    except Exception as e:
                                        self.logger.log('{0} could not write an entry {1}'.format(record[i],e),'USER')
                                del self.input_data[key]
                                counter += 1
                else:
                    if list(self.input_data.keys()):
                        if particular_output_fields:
                            left = num_of_out_fields - self.shapex_field - 1
                            empty_list = [""] * self.shapex_field
                            second_empty_list = [""] * left
                            keys_to_iterate = list(self.input_data.keys())
                            for key in keys_to_iterate:
                                # mapped_fields_to_write_still = []
                                # for num in self.user_field_nums:
                                #     mapped_fields_to_write_still.append(self.input_data[key][num[0]])
                                record = empty_list + [""] + second_empty_list + self.input_data[key]
                                for i in range(len(record)):
                                    try:
                                        self.out_worksheet.write(counter,i,record[i])
                                    except Exception as e:
                                        self.logger.log('{0} could not write an entry {1}'.format(record[i],e),'USER')
                                del self.input_data[key]
                                counter += 1
                        else:
                            left = num_of_out_fields - self.status_field - 1
                            empty_list = [""] * self.status_field
                            second_empty_list = [""] * left
                            keys_to_iterate = list(self.input_data.keys())
                            for key in keys_to_iterate:
                                mapped_fields_to_write_still = []
                                for num in self.user_field_nums:
                                    mapped_fields_to_write_still.append(self.input_data[key][num[0]])
                                record = empty_list + ["U"] + second_empty_list + mapped_fields_to_write_still + self.input_data[key]
                                for i in range(len(record)):
                                    try:
                                        self.out_worksheet.write(counter,i,record[i])
                                    except Exception as e:
                                        self.logger.log('{0} could not write an entry {1}'.format(record[i],e),'USER')
                                del self.input_data[key]
                                counter += 1
                continue
                    #
                    # for record in some_batch:
                    #     self.csvWriter.writerow(record)
                    #
            self.logger.log('the REST response q is not empty, gonna try to ' + \
                             'write out batch number {}'.format(batch_num),'INFO')
            full_response = self.REST_response_q.get()
            rest_response = full_response[1]

            #write the header
            if header_has_been_written == False:
                self.logger.log('writing headers for the first time','INFO')
                try:
                    # Basically just pick one from the list
                    sample_loc = rest_response['locations'][0]
                    self.spatial_reference = rest_response['spatialReference']['wkid']
                except Exception as e:
                    msg_code = 100191
                    msg = self.ERROR_CODES[msg_code]
                    self.addError(msg_code, msg)
                    header_has_been_written = True
                    continue
                    # We continue here because the whole batch is bad
                    # Continue keyword skips the batch

                if self.all_fields[0] == 'ID':
                    self.all_fields[0] = 'id'
                self.logger.log('we are about to write the header row','INFO')
                # if self.analyze_obj.started_as_feature_service:
                #     self.all_fields.remove('OID')
                if not particular_output_fields:
                    all_fields = []
                    #arcpy.AddMessage(self.all_fields)
                    for field in self.all_fields:
                        all_fields.append("USER_{}".format(field))
                else:
                    all_fields = self.all_fields

                self.locator_fields = []

                if not particular_output_fields:               

                    locator_flds = self.analyze_obj.json_info_response['addressFields']
                    # categories_inner = json_response['categories']['categories']
                    for addr in locator_flds:
                        self.locator_fields.append("IN_{}".format(addr['name']))


                if particular_output_fields:
                    # if "x" not in desired_fields and "X" not in desired_fields:
                    #     desired_fields.append("X")
                    # if "y" not in desired_fields and "Y" not in desired_fields:
                    #     desired_fields.append("Y")
                    # if "score" not in desired_fields and "Score" not in desired_fields:
                    #     desired_fields.append("Score")
                    # if "status" not in desired_fields and "Status" not in desired_fields:
                    #     desired_fields.append("Status")
                    # desired_fields.append("Shape X")
                    # desired_fields.append("Shape Y")
                    write_row = list(desired_fields) + list(all_fields)
                    self.out_worksheet.append(write_row)
                else:
                    #arcpy.AddMessage('checkpoint 3')
                    if "x" not in self.nice_ordering and "X" not in self.nice_ordering:
                        self.nice_ordering.append("X")
                    if "y" not in self.nice_ordering and "Y" not in self.nice_ordering:
                        self.nice_ordering.append("Y")
                    if "score" not in self.nice_ordering and "Score" not in self.nice_ordering:
                        self.nice_ordering.append("Score")
                    if "status" not in self.nice_ordering and "Status" not in self.nice_ordering:
                        self.nice_ordering.append("Status")
                    self.nice_ordering.append("Shape X")
                    self.nice_ordering.append("Shape Y")
                    #arcpy.AddMessage('checkpoint 4')
                    write_row = list(self.nice_ordering) + list(self.locator_fields) + list(all_fields)
                    #arcpy.AddMessage('checkpoint 5')
                    #arcpy.AddMessage(write_row)
                    self.out_worksheet.append(write_row)
                #arcpy.AddMessage('wrote it')
                header_has_been_written = True
                #write the first line of the csv as the fields

            # Create each row to be written to the output xls
            try:
                self.logger.log('going to attempt to write batch ' +\
                            str(batch_num),'INFO')
                try:
                    #arcpy.AddMessage('going to load a REST response')
                    dummy_var = rest_response['locations']
                    #arcpy.AddMessage('loaded a REST response')
                except Exception as e:
                    self.logger.log('issue with rest response {}'.format(e),'ERROR')
                for location in dummy_var:
                    # We need a counter here because xls demands a row
                    # number to write to
                    counter = counter + 1
                    try:
                        attributes = location['attributes']

                    except Exception as e:
                        self.logger.log('issue with rest response {}'.format(e),'ERROR')
                    attr_list = []
                    attr_options = list(attributes.keys())
                    try:
                        if particular_output_fields:
                            for attr in desired_fields:
                                if attr == "Shape X":
                                    if "location" in location:
                                        attr_list.append(location["location"]["x"])
                                    else:
                                        attr_list.append("")
                                elif attr == "Shape Y":
                                    if "location" in location:
                                        attr_list.append(location["location"]["y"])
                                    else:
                                        attr_list.append("")
                                elif attr not in attr_options:
                                    attr_list.append('')
                                elif attr == "Address":
                                    attr_list.append(location["address"])
                                elif attributes[attr] == '':
                                    attr_list.append('')
                                # elif isinstance(attributes[attr],basestring) or \
                                #         isinstance(attributes[attr],int) or \
                                #         isinstance(attributes[attr],float):
                                #     # cleaned_string = attributes[attr].encode('utf-8')
                                #     attr_list.append(attributes[attr])
                                else:
                                    attr_list.append(attributes[attr]) #.decode('utf-8'))
                        else:
                            for attr in self.nice_ordering:
                                if attr == "Shape X":
                                    if "location" in location:
                                        attr_list.append(location["location"]["x"])
                                    else:
                                        attr_list.append("")
                                elif attr == "Shape Y":
                                    if "location" in location:
                                        attr_list.append(location["location"]["y"])
                                    else:
                                        attr_list.append("")
                                elif attr not in attr_options:
                                    attr_list.append('')
                                elif attr == "Address":
                                    attr_list.append(location["address"])
                                elif attributes[attr] == '':
                                    attr_list.append('')
                                # elif isinstance(attributes[attr],basestring) or \
                                #         isinstance(attributes[attr],int) or \
                                #         isinstance(attributes[attr],float):
                                #     # cleaned_string = attributes[attr].encode('utf-8')
                                #     attr_list.append(attributes[attr])
                                else:
                                    attr_list.append(attributes[attr]) #.decode('utf-8'))

                    except Exception as e:
                        self.logger.log('unicode error while trying to write ' +\
                                    str(e),'ERROR')

                    # Write to the self.output file
                    try:
                        if str(attributes["ResultID"]) == '-1':
                            self.logger.log('No result ID','INFO')
                            continue
                    except Exception as e:
                        self.logger.log('there is no result ID!!!!','INFO')
                    # Get the relevant input row from the dictionary
                    # Delete from the input dictionary because we are done
                    # with that row. Tries to keep the number of rows stored
                    # in memory under 8*batch_size
                    try:
                        input_row = self.input_data.get(str(attributes["ResultID"]))
                        if input_row is None:
                            self.dictionary = 'Empty dictionary'
                            self.dropped_records_q.put(str(attributes["ResultID"]))
                            continue
                        if self.dropped_records_q.full():
                            self.error_q.put([1,str('Dropped more than 20 records')])
                            self.num_of_REST_threads.get()
                            return
                        del self.input_data[str(attributes["ResultID"])]
                    except Exception as e:
                        self.logger.log('Dictionary error {}'.format(e),'INFO')

                    # mapped_fields_to_write_still = []
                    # for num in self.user_field_nums:
                    #     mapped_fields_to_write_still.append(input_row[num[0]])
                    if not particular_output_fields:
                        dict_of_user_fields_to_their_index_number = {}
                        for num,tup in enumerate(self.field_info):
                            dict_of_user_fields_to_their_index_number[tup[0]] = num

                        locator_fields_to_write = []

                        for fld in self.locator_fields:
                            if fld[3:] in self.user_arcgis_fields:
                                #arcpy.AddWarning('considering field {}'.format(fld[3:]))
                                for elem_num in range(len(self.field_mapping)):
                                    if self.field_mapping[elem_num][1] == fld[3:]:
                                        #arcpy.AddWarning('considering field2 {}'.format(self.field_mapping[elem_num][1]))
                                        # so now we need the self.field_mapping[elem_num][0] value for this row...
                                        locator_fields_to_write.append(input_row[dict_of_user_fields_to_their_index_number[self.field_mapping[elem_num][0]]])
                            else:
                                locator_fields_to_write.append("")

                        try:
                            row_to_write = attr_list + list(locator_fields_to_write) + list(input_row)
                        except Exception as e:
                            self.logger.log('having trouble concatenating rows','ERROR')
                    
                    else:
                        try:
                            row_to_write = attr_list + list(input_row)
                        except Exception as e:
                            self.logger.log('having trouble concatenating rows','ERROR')

                    # do not need this part for Python3 strings are already unicode
                    # row_to_write_final = []
                    # for entry in row_to_write:
                    #     if isinstance(entry,str):
                    #         row_to_write_final.append(entry.encode('utf-8'))
                    #     else:
                    #         row_to_write_final.append(entry)

                    # get rid of first column... artifact of table to table
                    # if self.analyze_obj.started_as_feature_service:
                    #     negative_one = row_to_write_final.pop(0)
                    try:
                        self.out_worksheet.append(row_to_write)
                    except Exception as e:
                        # try:
                        #     decoded_entry = row_to_write_final[i].decode('windows-1252')
                        #     encoded_entry = decoded_entry.encode('utf-8')
                        #     self.out_worksheet.write(counter,i,encoded_entry)
                        # except Exception as ex:
                        #     try:
                        #         self.out_worksheet.write(counter,i,row_to_write_final[i].encode('ascii','ignore'))
                        #     except Exception as excep:
                        self.logger.log('{0} could not write an entry {1}'.format(row_to_write[i],e),'USER')


                self.logger.log('Just wrote batch {} to output.xls'.format(batch_num),'INFO')

                # Save after writing each batch
                # self.xls_out_workbook.save(os.path.join(self.working_dir,'results.xls'))

            except Exception as e:
                if self.DEBUG:
                    self.logger.log('There was a problem with batch {0}: {1}'.format(batch_num,e),'ERROR')
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    self.logger.log('{0} {1} {2}'.format(exc_type, fname, exc_tb.tb_lineno),'ERROR')
                pass

    def create_results_txt(self):
        '''Only gets called if the output_file_format is txt.
            Takes rest responses and writes them to self.output.txt'''
        counter = 0
        # Open the file for the first time, write the header row
        header_has_been_written = False
        # We need to know the service self.output fields
        # Let's query for them
        # This line below gets rid of the 'geocodeAddresses' part of the URL
        
        #almost_schema_page = self.analyze_obj.service_url[:-17]
        #if self.analyze_obj.token == None:
        #    #don't need a token!!!
        #    self.logger.log('Assuming the service is unsecure','INFO')
        #    schema_page = almost_schema_page + '/?f=pjson'
        #    request = urllib2.Request(schema_page)
        #    # Store the response
        #    if self.analyze_obj.ssl_issues:
        #        response = urllib2.urlopen(request,context=ssl._create_unverified_context())
        #    else:
        #        response = urllib2.urlopen(request)
        #    json_info_response = json.loads(response.read())
        #else:
        #    values = ([('token',self.analyze_obj.token),('f','json')])
        #    schema_page = almost_schema_page + '/?f=pjson'
        #    encoded_data = urllib.urlencode(values)
        #    http_header = {'referer':self.analyze_obj.referer}
        #    info_request = urllib2.Request(schema_page,\
        #                                   encoded_data,http_header)
        #    if self.analyze_obj.ssl_issues:
        #        info_response = urllib2.urlopen(info_request,context=ssl._create_unverified_context())
        #    else:
        #        info_response = urllib2.urlopen(info_request)
        #    json_info_response = json.loads(info_response.read())

        cand_fields = self.analyze_obj.json_info_response['candidateFields']

        try:
            ctr = 0
            for field in self.output_fields:
                # remove extrenuous spaces
                # we will remove junk input for this later
                new_field = field.replace(' ','')
                self.output_fields[ctr] = new_field
                ctr += 1
            self.nice_ordering = []
            # Now we only want to return certain candidate fields
            cand_fields = json_info_response['candidateFields']
            for field in cand_fields:
                self.nice_ordering.append(field['name'])
            # If user is hitting a local geocoding service, then candidateFields
            # does not contain status even though it will be an output. RETRIEVE
            # THAT OUTPUT!!!
            if "Status" not in self.nice_ordering:
                self.nice_ordering.append("Status")
            try:
                # Remove the geometry if necessary
                self.nice_ordering.remove('Shape')
            except Exception as e:
                # I guess it wasn't necessary!
                pass
            if '*' in self.output_fields:
                self.output_fields.remove('*')
                particular_output_fields = False
            elif self.output_fields == '':
                particular_output_fields = False
            else:
                particular_output_fields = True

            if particular_output_fields:
                output_fields_lower = [x.lower() for x in self.output_fields]
                nice_ordering_lower = [x.lower() for x in self.nice_ordering]
                for requested_field in output_fields_lower:
                    if requested_field not in nice_ordering_lower:
                        output_fields_lower.remove(requested_field)
                        self.logger.log(\
                            'Your requested self.output field ' +\
                            '{} does not exist,'.format(requested_field) +\
                            ' skipping adding it to the self.output','USER')
                # Now that we have removed junk possible self.output...
                if not output_fields_lower:
                    # Then the list is empty!!! set boolean to False
                    particular_output_fields = False
                indexes_to_keep = []
                for output_field in output_fields_lower:
                    if output_field in nice_ordering_lower:
                        counter = 0
                        for service_field in nice_ordering_lower:
                            if service_field == output_field:
                                indexes_to_keep.append(counter)
                            counter += 1
                desired_fields = []
                for index in indexes_to_keep:
                    desired_fields.append(self.nice_ordering[index])
        except Exception as e:
            if self.DEBUG:
                self.logger.log('Exception in create_results_csv {}'.format(e),'ERROR')
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                self.logger.log('{0} {1} {2}'.format(exc_type, fname, exc_tb.tb_lineno),'ERROR')
            # Shut down all threads. If we cannot build recordsets, there
            # is no way the job will finish.
            self.error_q.put([1,str(e)])
            self.num_of_REST_threads.get()
            return "error"
        for batch_num in range(self.num_of_batches):
            # self.logger.log(str(round((float(batch_num)/float(self.num_of_batches))*\
            #                 100,1)) + ' percent done with' +\
            #                 ' the pending job','USER')
            while self.REST_response_q.empty():
                # We wait for REST responses. In the meantime, be ready to shut down just in case
                time.sleep(0.1)
                if not self.error_q.empty():
                    self.logger.log('there is an error somewhere,' +\
                                'shut down create_results','INFO')
                    self.error_out = True
                    break
                if self.is_cancelled:
                    error = 'The script has been aborted by the user. Returning geocoding results thus far'
                    self.logger.log('The script has been aborted by the user. Returning geocoding results thus far',\
                                    'USER_ERROR')
                    self.error_out = True
                    # # This error is fatal, so just release the lock on the num_threads queue
                    # so that we can exit the whole system
                    self.error_q.put([1,str(error)])
                    # self.num_of_REST_threads.get()
                    break

                if self.error_out:
                    # already has an error
                    break

            if not self.error_q.empty():
                # There is an error somewhere, shut down all threads
                self.error_out = True
                self.logger.log('shutting down create_results','INFO')

            if self.error_out:
                if particular_output_fields:
                        cntr = 0
                        for field in desired_fields:
                            if "Status" == field:
                                self.status_field = cntr
                            cntr += 1
                else:
                    cntr = 0
                    for field in self.nice_ordering:
                        if "Status" == field:
                            self.status_field = cntr
                        cntr += 1
                if not header_has_been_written:
                    if self.all_fields[0] == 'ID':
                        # Excel won't open csv files that have "ID" as their
                        # first value. This is the work-around
                        self.all_fields[0] = 'id'

                    # Remove the OID Column (artifact of moving from table to csv)
                    all_fields_minus_oid = self.all_fields
                    # if self.analyze_obj.started_as_feature_service:
                    #     all_fields_minus_oid.remove('OID')

                    # Write the first line of the csv as the fields
                    if particular_output_fields:
                        record = all_fields_minus_oid + desired_fields
                        for field in record:
                            self.txt_file_out.write(str(field) + '\t')
                        self.txt_file_out.write('\n')
                        # for field in desired_fields:
                        #     if "Status" == field:
                        #         self.status_field = cntr
                        #     cntr += 1
                    else:
                        record = all_fields_minus_oid + self.nice_ordering
                        for field in record:
                            self.txt_file_out.write(str(field) + '\t')
                        self.txt_file_out.write('\n')
                        # for field in nice_ordering:
                        #     if "Status" == field:
                        #         self.status_field = cntr
                        #     cntr += 1
                    header_has_been_written = True

                    # then write the first batch
                    # tries = 0
                    # while self.leftover_records_q.empty():
                    #     time.sleep(0.5)
                    #     tries += 1
                    #     if tries > 3:
                    #         break
                    # if not self.leftover_records_q.empty():
                    #     some_batch = self.leftover_records_q.get()[1]
                    #     for record in some_batch:
                    #         empty_list = [""] * self.status_field
                    #         record = record + empty_list + ["U"]
                    #         for field in record:
                    #             self.txt_file_out.write(str(field) + '\t')
                    #         self.txt_file_out.write('\n')
                    if list(self.input_data.keys()):
                        keys_to_iterate = list(self.input_data.keys())
                        empty_list = [""] * self.status_field
                        for key in keys_to_iterate:
                            record = self.input_data[key] + empty_list + ["U"]
                            for field in record:
                                self.txt_file_out.write(str(field) + '\t')
                            self.txt_file_out.write('\n')
                            del self.input_data[key]
                else:
                    # tries = 0
                    # while self.leftover_records_q.empty():
                    #     time.sleep(0.5)
                    #     tries += 1
                    #     if tries > 3:
                    #         break
                    # if not self.leftover_records_q.empty():
                    #     some_batch = self.leftover_records_q.get()[1]
                    #     for record in some_batch:
                    #         empty_list = [""] * self.status_field
                    #         record = record + empty_list + ["U"]
                    #         for field in record:
                    #             self.txt_file_out.write(str(field) + '\t')
                    #         self.txt_file_out.write('\n')

                    if list(self.input_data.keys()):
                        keys_to_iterate = list(self.input_data.keys())
                        empty_list = [""] * self.status_field
                        for key in keys_to_iterate:
                            record = self.input_data[key] + empty_list + ["E"]
                            for field in record:
                                self.txt_file_out.write(str(field) + '\t')
                            self.txt_file_out.write('\n')
                            del self.input_data[key]

                continue
            self.logger.log('the REST response q is not empty, gonna try to ' + \
                             'write out batch number {}'.format(batch_num),'INFO')
            full_response = self.REST_response_q.get()
            rest_response = full_response[1]

            #write the header
            if header_has_been_written == False:
                self.logger.log('Writing headers for the first time','INFO')
                try:
                    # Basically just pick one from the list
                    sample_loc = rest_response['locations'][0]
                    self.spatial_reference = rest_response['spatialReference']['wkid']
                except Exception as e:
                    msg_code = 100191
                    header_has_been_written = True
                    msg = self.ERROR_CODES[msg_code]
                    self.addError(msg_code, msg)
                    continue
                    # We continue here because the whole batch is bad
                # fields_array = []
                # for field in sample_loc['attributes']:
                #     fields_array.append(field)
                if self.all_fields[0] == 'ID':
                    self.all_fields[0] = 'id'

                # Remove OID column... artifact from tabletotable
                if self.analyze_obj.started_as_feature_service:
                    self.all_fields.remove('OID')
                if particular_output_fields:
                    header_row = self.all_fields + desired_fields
                else:
                    header_row = self.all_fields + self.nice_ordering
                for field in header_row:
                    # Hard-coded for txt because we are only supporting tabbed
                    self.txt_file_out.write(field + '\t')
                # make sure to write a newline after the header has been written!
                self.txt_file_out.write('\n')
                header_has_been_written = True
                #write the first line of the csv as the fields

            # Create each row to be written to the self.output.csv
            try:
                self.logger.log('going to attempt to write batch ' + \
                                 str(batch_num),'INFO')
                try:
                    dummy_var = rest_response['locations']
                except Exception as e:
                    self.logger.log('issue with rest response {}'.format(e),'ERROR')
                for location in dummy_var:
                    counter = counter + 1
                    try:
                        attributes = location['attributes']

                    except Exception as e:
                        self.logger.log('issue with rest response #2 ' +\
                                    str(e),'ERROR')
                    #print attributes
                    attr_list = []
                    attr_options = list(attributes.keys())
                    try:
                        if particular_output_fields:
                            for attr in desired_fields:
                                # Get rid of any residual unicode characters if there are any
                                if attr not in attr_options:
                                    attr_list.append('')
                                elif attr == "Address":
                                    attr_list.append(location["address"])
                                elif attributes[attr] == '':
                                    attr_list.append('')
                                # elif isinstance(attributes[attr],basestring):
                                #     cleaned_string = attributes[attr].encode('utf-8')
                                #     attr_list.append(cleaned_string)
                                else:
                                    attr_list.append(attributes[attr])
                        else:
                            for attr in self.nice_ordering:
                                # Get rid of any residual unicode characters if there are any
                                if attr not in attr_options:
                                    attr_list.append('')
                                elif attr == "Address":
                                    attr_list.append(location["address"])
                                elif attributes[attr] == '':
                                    attr_list.append('')
                                # elif isinstance(attributes[attr],basestring):
                                #     cleaned_string = attributes[attr].encode('utf-8')
                                #     attr_list.append(cleaned_string)
                                else:
                                    attr_list.append(attributes[attr])
                    except Exception as e:
                        self.logger.log('unicode error while trying to write ' + \
                                         str(e),'ERROR')

                    # Write to the self.output file
                    try:
                        if str(attributes["ResultID"]) == '-1':
                            self.logger.log('No Result ID','INFO')
                            continue
                    except Exception as e:
                        self.logger.log('No Result ID','INFO')
                    # Get the relevant input row from the dictionary
                    # Delete from the input dictionary because we are done
                    # with that row. Tries to keep the number of rows stored
                    # in memory under 8*batch_size
                    # del input_data[str(attributes["ResultID"])]
                    try:
                        input_row = self.input_data.get(str(attributes["ResultID"]))
                        if input_row is None:
                            self.logger.log('empty on record {}'.format(counter),'INFO')
                            self.dropped_records_q.put(str(attributes["ResultID"]))
                            continue
                        if self.dropped_records_q.full():
                            self.error_q.put([1,\
                                              str('Dropped more than 20 records')])
                            self.num_of_REST_threads.get()
                            return

                        del self.input_data[str(attributes["ResultID"])]

                    except Exception as e:
                        self.logger.log('We are having some ' +\
                                    'dictionary struggles {}'.format(e),'INFO')

                    try:
                        row_to_write = input_row + attr_list
                    except Exception as e:
                        self.logger.log('Having trouble concatenating rows {}'.format(e),\
                                        'ERROR')
                    try:
                        # Since we know we're writing txt...
                        row_to_write_final = []
                        for entry in row_to_write:
                            if isinstance(entry,str):
                                row_to_write_final.append(entry.decode('utf-8','ignore'))
                            else:
                                row_to_write_final.append(entry)

                        # Remove -1, artifact of tabletotable
                        # if self.analyze_obj.started_as_feature_service:
                        #     negative_one = row_to_write_final.pop(0)
                        for field in row_to_write_final:
                            self.txt_file_out.write(str(field) + '\t')
                        self.txt_file_out.write('\n')

                    except Exception as e:
                        self.logger.log('unable to write to txt {}'.format(e),'ERROR')

                self.logger.log('Just wrote batch {}'.format(batch_num) + \
                                     ' to self.output.txt','INFO')

                full_response = None
                rest_response = None
                gc.collect()

            except Exception as e:
                # Clearly a bad batch
                if self.DEBUG:
                    self.logger.log('There was a problem with batch {0}: {1}'.format(batch_num,e),'ERROR')
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    self.logger.log('{0} {1} {2}'.format(exc_type, fname, exc_tb.tb_lineno),'ERROR')
                pass

    def convert_csv_to_feature_service(self):
        # crack open csv out file
        #arcpy.AddMessage("the path is {}".format(os.path.join(self.working_dir,self.output_name + '.csv')))
        #arcpy.AddMessage("starting to fs conversion")
        #arcpy.AddMessage("out file type is {}".format(self.out_file_type))
        if self.analyze_obj.EGDB_table:
            ## do stuff
            if self.egdb_output_type == "append":
                if os.path.exists(os.path.join(self.working_dir,self.output_name + '.csv')):
                    with open(os.path.join(self.working_dir,self.output_name + '.csv'),'r', encoding="utf8") as csvfile:
                        csvReader = csv.reader(csvfile)
                        all_col_names = next(csvReader)
                        needs_to_be_added = []
                        user_field_count = 0
                        for field_value in all_col_names:
                            if field_value[:5] != "USER_" and field_value[:3] != "IN_":
                                needs_to_be_added.append(field_value)
                            else:
                                user_field_count += 1
                        #arcpy.AddMessage(needs_to_be_added)

                        if self.specific_output_fields:
                            ### special case where we need to throw away original fields and not append duplicates
                            needs_to_be_added = self.output_fields


                        all_col_names_lower = [fld.lower() for fld in needs_to_be_added]
                        try:
                            match_type_index = all_col_names_lower.index("match_type")
                            match_type_exists = True
                        except ValueError:
                            match_type_exists = False
                        
                        try:
                            score_index = all_col_names_lower.index("score")
                            score_index_exists = True
                        except ValueError:
                            score_index_exists = False
                            score_index = 9999 #something that will never = the index loop

                        try:
                            status_index = all_col_names_lower.index("status")
                            status_index_exists = True
                        except ValueError:
                            status_index_exists = False


                        try:
                            index_of_X = all_col_names_lower.index("x")
                        except ValueError:
                            index_of_X = 9999 #something that will never = the index loop

                        try:
                            index_of_Y = all_col_names_lower.index("y")
                        except ValueError:
                            index_of_Y = 9999 #something that will never = the index loop
                            
                        index_of_ShapeX = all_col_names_lower.index("shape x")
                        #arcpy.AddMessage("index 1 is {}".format(index_of_ShapeX))
                        index_of_ShapeY = all_col_names_lower.index("shape y")
                        #arcpy.AddMessage("index 2 is {}".format(index_of_ShapeY))
                        try:
                            index_of_xmax = all_col_names_lower.index("xmax")
                            index_of_xmin = all_col_names_lower.index("xmin")
                            index_of_ymax = all_col_names_lower.index("ymax")
                            index_of_ymin = all_col_names_lower.index("ymin")
                            extent_exists = True
                        except ValueError:
                            extent_exists = False

                        

                        # if not self.spatial_reference:
                        #     self.spatial_reference = 4326
                        # sr = arcpy.SpatialReference(int(self.spatial_reference))

                        ### make sure none of the fields we are about to add are duplicates. Otherwise no fields will get added
                        orig_fields = arcpy.ListFields(self.output_location)
                        lower_orig_fields = [fld.name.lower() for fld in orig_fields]

                        ### if there are duplicate names that we need to add "output_" in front of the field name
                        fields_to_bulk_add = []
                        for field_name in needs_to_be_added:
                            if field_name.lower() == "x" or field_name.lower() == "y" or ("xmin" == field_name.lower()) or ("xmax" == field_name.lower()) or ("ymin" == field_name.lower()) or ("ymax" == field_name.lower()) or ("score" == field_name.lower()):
                                if field_name.lower() in lower_orig_fields:
                                    fields_to_bulk_add.append(["output_" + field_name,"DOUBLE", "output_" + field_name])
                                else:
                                    fields_to_bulk_add.append([field_name,"DOUBLE",field_name])
                            else:
                                if field_name.lower() in lower_orig_fields:
                                    fields_to_bulk_add.append(["output_" + field_name,"TEXT"])
                                else:
                                    fields_to_bulk_add.append([field_name,"TEXT"])
                        
                        #arcpy.AddMessage("about to bulk add: {}".format(fields_to_bulk_add))

                        only_xy = False

                        if self.orig_output_fields:
                            if self.orig_output_fields == "NONE":
                                ### NONE is a special keyword asking specifically for location-only
                                only_xy = True

                        if only_xy:
                            ### if we only need x and y...
                            fields_to_bulk_add = []
                            x_exists = False
                            y_exists = False
                            ### the following 4 statements handle the case where x and y are already input fields
                            if "x" in lower_orig_fields:
                                fields_to_bulk_add.append(['output_X', 'DOUBLE', 'output_X'])
                                x_exists = True
                            else:
                                fields_to_bulk_add.append(['X', 'DOUBLE', 'X'])
                            if "y" in lower_orig_fields:
                                fields_to_bulk_add.append(['output_Y', 'DOUBLE', 'output_Y'])
                                y_exists = True
                            else:
                                fields_to_bulk_add.append(['Y', 'DOUBLE', 'Y'])

                            arcpy.management.AddFields(self.output_location, fields_to_bulk_add)
                            ### the following 4 statements handle the case where x and y are already input fields
                            if x_exists and y_exists:
                                cursor = arcpy.da.UpdateCursor(self.output_location, ["output_X", "output_Y"])
                            if not x_exists and not y_exists:
                                cursor = arcpy.da.UpdateCursor(self.output_location, ["X", "Y"])
                            if x_exists and not y_exists:
                                cursor = arcpy.da.UpdateCursor(self.output_location, ["output_X", "Y"])
                            if not x_exists and y_exists:
                                cursor = arcpy.da.UpdateCursor(self.output_location, ["X", "output_Y"])
                            for row_num, row in zip(range(self.row_count - self.num_empty_rows - 1), cursor):
                                try:
                                    csv_row = next(csvReader)
                                    #arcpy.AddMessage("the row I read looks like {}".format(csv_row))
                                    if status_index_exists:
                                        if csv_row[status_index] == "M":
                                            self.matched += 1
                                        elif csv_row[status_index] == "T":
                                            self.tied += 1
                                        else:
                                            self.unmatched += 1
                                    
                                    #arcpy.AddMessage("csv file to look at is {}".format(os.path.join(self.working_dir,self.output_name + '.csv')))
                                    point = (csv_row[index_of_ShapeX],csv_row[index_of_ShapeY])
                                    try:
                                        x = float(point[0])
                                    except ValueError:
                                        x = None
                                    try:
                                        y = float(point[1])
                                    except ValueError:
                                        y = None
                                    final_list = [x, y]
                                    try:
                                        cursor.updateRow(final_list)
                                    except RuntimeError as e:
                                        self.logger.log("A row could not be written {}".format(final_list),'INFO')
                                        #arcpy.AddMessage("ERROR IS {}".format(e))
                                        #arcpy.AddWarning("A row could not be written {}".format(final_list))
                                        pass
                                    except TypeError:
                                        self.logger.log('Tried to write row {}'.format(final_list),'INFO')
                                        #arcpy.AddMessage('tried to write row {}'.format(final_list))
                                        pass
                                        #lol
                                    #big_counter += 1
                                except StopIteration:
                                    break

                            del cursor
                        else:

                            #arcpy.AddMessage("gets here")
                            orig_fields = arcpy.ListFields(self.output_location)
                            #arcpy.AddMessage("gets here2")
                            orig_len = len(orig_fields)
                            arcpy.management.AddFields(self.output_location,fields_to_bulk_add)

                            fc_fields = arcpy.ListFields(self.output_location)
                            fc_field_names = [fld.name for fld in fc_fields]

                            fields_to_update = fc_field_names[orig_len:]

                            #arcpy.AddMessage("fields are: {}".format(fc_field_names))
                            #arcpy.AddMessage("opening update cursor on: {}".format(fields_to_update))

                            out_fields_len = len(fields_to_update)

                            cursor = arcpy.da.UpdateCursor(self.output_location, fields_to_update)
                            # arcpy.AddMessage("index_of_X is: {}".format(index_of_X))
                            # arcpy.AddMessage("index_of_Y is: {}".format(index_of_Y))
                            # arcpy.AddMessage("score_index is: {}".format(score_index))
                            # arcpy.AddMessage("index_of_xmax is: {}".format(index_of_xmax))
                            # arcpy.AddMessage("index_of_ymax is: {}".format(index_of_ymax))
                            # arcpy.AddMessage("index_of_xmin is: {}".format(index_of_xmin))
                            # arcpy.AddMessage("index_of_ymin is: {}".format(index_of_ymin))
                            # arcpy.AddMessage("index_of_ShapeY is: {}".format(index_of_ShapeY))
                            # arcpy.AddMessage("index_of_ShapeX is: {}".format(index_of_ShapeX))

                            for row_num, row in zip(range(self.row_count - self.num_empty_rows - 1), cursor):
                                try:
                                    csv_row = next(csvReader)
                                    row_to_add = []
                                    #arcpy.AddMessage("the row I read looks like {}".format(csv_row))
                                    output_row = csv_row[:out_fields_len]
                                    #arcpy.AddMessage("output rows is: {}".format(output_row))
                                    #arcpy.AddMessage("csv file to look at is {}".format(os.path.join(self.working_dir,self.output_name + '.csv')))
                                    point = (output_row[index_of_ShapeX],output_row[index_of_ShapeY])
                                    #arcpy.AddMessage(point)
                                    #del csv_row[index_of_ShapeX]
                                    #del csv_row[index_of_ShapeY]
                                    for i, field in enumerate(output_row):
                                        if len(field) > 255:
                                            field = field[:255]
                                        else:
                                            if extent_exists:
                                                if i == index_of_X or i == index_of_Y or i == score_index or i == index_of_xmax or i == index_of_ymax or i == index_of_xmin or i == index_of_ymin:
                                                    try:
                                                        new_field = float(field)
                                                    except ValueError:
                                                        new_field = None
                                                    row_to_add.append(new_field)
                                                else:
                                                    row_to_add.append(field)
                                            else:
                                                if i == index_of_X or i == index_of_Y or i == score_index:
                                                    try:
                                                        new_field = float(field)
                                                    except ValueError:
                                                        new_field = None
                                                    row_to_add.append(new_field)
                                                else:
                                                    row_to_add.append(field)


                                    #csv_row_cleaned = [rounded_convert_to_int_if_possible(val) for val in csv_row]
                                    if status_index_exists:
                                        if csv_row[status_index] == "M":
                                            self.matched += 1
                                            if match_type_exists:
                                                row_to_add[match_type_index] = "A"
                                        elif csv_row[status_index] == "T":
                                            self.tied += 1
                                            if match_type_exists:
                                                row_to_add[match_type_index] = "A"
                                        else:                            
                                            self.unmatched += 1
                                    try:
                                        cursor.updateRow(row_to_add)
                                    except RuntimeError as e:
                                        #arcpy.AddMessage("ERROR IS {}".format(e))
                                        #arcpy.AddWarning("A row could not be written {}".format(row_to_add))
                                        self.logger.log("A row could not be written {}".format(final_list),'INFO')
                                        pass
                                    except TypeError as e:
                                        self.logger.log('Tried to write row {}'.format(final_list),'INFO')
                                        #arcpy.AddMessage('tried to write row {}'.format(row_to_add))
                                        #arcpy.AddMessage("error is {}".format(e))
                                        pass
                                        #lol
                                    #big_counter += 1
                                except StopIteration:
                                    break

                            del cursor

                else:
                    pass
                
            if self.egdb_output_type == "fc":
                ### will implement this at 2.8 once I get this from the GP team
                ### https://devtopia.esri.com/ArcGISPro/geoprocessing/issues/3839
                pass

        else:

            if os.path.exists(os.path.join(self.working_dir,self.output_name + '.csv')):
                with open(os.path.join(self.working_dir,self.output_name + '.csv'),'r', encoding="utf8") as csvfile:
                    csvReader = csv.reader(csvfile)

                    # get the header row
                    # self.output.write(os.path.join(self.working_dir,self.output_name + '.csv'),self.output_name + '.csv')

                    all_col_names = next(csvReader)
                    #arcpy.AddMessage('column names are: {}'.format(all_col_names))
                    all_col_names_lower = [fld.lower() for fld in all_col_names]

                    reserved_words = ["status", "score", "match_type", "match_addr", "addr_type"]
                    bad_dict = {}
                    for word in reserved_words:
                        count = 0
                        bad_idx = 0
                        for idx, col_name in enumerate(all_col_names_lower):
                            if word == col_name:
                                count += 1
                                bad_idx = idx
                            
                        if count > 1:
                            bad_dict[word] = bad_idx

                    if bad_dict:
                        ### this means we have a reserved word in the input and need to replace it with word_1
                        # arcpy.AddWarning("Your input table has a reserved word. The field has been updated with a _1 appended.")
                        ### Alexei will translate this warning into a Pro GP Warning. This is the format we decided on for him to
                        ### fill out the warning
                        for key in list(bad_dict.keys()):
                            idx_to_replace = bad_dict[key]
                            orig_value = all_col_names[idx_to_replace]
                            arcpy.AddWarning("003344::{0}::{1}".format(orig_value, orig_value + "_1"))
                            all_col_names[idx_to_replace] = orig_value + "_1"






                    try:
                        match_type_index = all_col_names_lower.index("match_type")
                        match_type_exists = True
                    except ValueError:
                        match_type_exists = False
                    
                    try:
                        score_index = all_col_names_lower.index("score")
                        score_index_exists = True
                    except ValueError:
                        score_index_exists = False
                        score_index = 9999 #something that will never = the index loop

                    try:
                        status_index = all_col_names_lower.index("status")
                        status_index_exists = True
                    except ValueError:
                        status_index_exists = False


                    try:
                        index_of_X = all_col_names_lower.index("x")
                        x_exists = True
                    except ValueError:
                        index_of_X = 9999 #something that will never = the index loop
                        x_exists = False

                    try:
                        index_of_Y = all_col_names_lower.index("y")
                        y_exists = True
                    except ValueError:
                        index_of_Y = 9999 #something that will never = the index loop
                        y_exists = False
                    index_of_ShapeX = all_col_names_lower.index("shape x")
                    #arcpy.AddMessage("index 1 is {}".format(index_of_ShapeX))
                    index_of_ShapeY = all_col_names_lower.index("shape y")
                    #arcpy.AddMessage("index 2 is {}".format(index_of_ShapeY))
                    try:
                        index_of_xmax = all_col_names_lower.index("xmax")
                        index_of_xmin = all_col_names_lower.index("xmin")
                        index_of_ymax = all_col_names_lower.index("ymax")
                        index_of_ymin = all_col_names_lower.index("ymin")
                        extent_exists = True
                    except ValueError:
                        extent_exists = False

                    if not self.spatial_reference:
                        self.spatial_reference = 4326
                    sr = arcpy.SpatialReference(int(self.spatial_reference))

                    
                    # create the GDB feature class
                    scratch_folder = arcpy.env.scratchFolder
                    scratch_gdb = os.path.join(scratch_folder,'scratch.gdb')
                    if self.zipped_gdb_output:
                        temp_output_name = self.output_name.replace(" ","_")
                        self.out_gdb = scratch_gdb
                    else:
                        temp_output_name = "Geocoding_Result"
                    self.output_feature_class = os.path.join(scratch_gdb, temp_output_name)
                    #arcpy.AddMessage("spatial reference is {}".format(self.spatial_reference))
                    if self.sixty_four_bit_OID:
                        arcpy.management.CreateFeatureclass(scratch_gdb, temp_output_name, "POINT",spatial_reference=sr, oid_type="64_BIT")
                    else:
                        arcpy.management.CreateFeatureclass(scratch_gdb, temp_output_name, "POINT",spatial_reference=sr)

                    # don't add shape x and shape y as columns, add them as shape fields instead
                    del all_col_names[index_of_ShapeX]
                    del all_col_names[index_of_ShapeX]

                    if self.analyze_obj.in_file_type == 'table':
                        aliases_lower = [x.lower() for x in self.aliases]

                    fields_to_bulk_add = []

                    new_all_col_names = []

                    #arcpy.AddMessage("OLD FIELD NAMES WERE: {}".format(all_col_names))

                    # chop off pesky trailing underscores that could be a problem
                    for fld_name in all_col_names:
                        #new_fld_name = arcpy.ValidateFieldName(fld_name, self.output_feature_class)
                        if fld_name[-1] == "_":
                            new_all_col_names.append(fld_name[:-1])
                        else:
                            new_all_col_names.append(fld_name)

                    
                    for field_name in new_all_col_names:
                        if field_name.lower() == 'objectid':
                            fields_to_bulk_add.append(["ObjectID_orig","TEXT","ObjectID_orig"])
                            #arcpy.management.AddField(self.output_feature_class, 'ObjectID_orig', "TEXT",255)
                            continue
                        if field_name.lower() == 'oid':
                            fields_to_bulk_add.append(["OID_orig","TEXT","OID_orig"])
                            #arcpy.management.AddField(self.output_feature_class, 'OID_orig', "TEXT",255)
                            continue
                        if field_name.lower() == "x" or field_name.lower() == "y" or ("xmin" == field_name.lower()) or ("xmax" == field_name.lower()) or ("ymin" == field_name.lower()) or ("ymax" == field_name.lower()) or ("score" == field_name.lower()):
                            #arcpy.management.AddField(self.output_feature_class, field_name, "DOUBLE", field_alias=field_name)
                            fields_to_bulk_add.append([field_name,"DOUBLE",field_name])
                        else:
                            if field_name[:3] == "IN_":
                                #arcpy.management.AddField(self.output_feature_class, field_name, "TEXT",field_alias=self.locator_aliases[field_name[3:]])
                                fields_to_bulk_add.append([field_name,"TEXT",self.locator_aliases[field_name[3:]]])
                            else:
                                if self.analyze_obj.in_file_type == 'table':
                                    # maintain any aliases the user already had
                                    if field_name[:5] == "USER_":
                                        if field_name[5:].lower() in aliases_lower:
                                            index_of_alias = aliases_lower.index(field_name[5:].lower())
                                            fields_to_bulk_add.append([field_name,"TEXT",self.aliases[index_of_alias]])
                                            #arcpy.management.AddField(self.output_feature_class, field_name, "TEXT",field_alias=self.aliases[index_of_alias])
                                        else:
                                            fields_to_bulk_add.append([field_name,"TEXT"])
                                    else:
                                        fields_to_bulk_add.append([field_name,"TEXT"])
                                        #arcpy.management.AddField(self.output_feature_class, field_name, "TEXT")
                                else:
                                    fields_to_bulk_add.append([field_name,"TEXT"])
                                    #arcpy.management.AddField(self.output_feature_class, field_name, "TEXT")
                    #arcpy.AddMessage("adding the following fields: {}".format)
                    
                    arcpy.management.AddFields(self.output_feature_class,fields_to_bulk_add)
                    #arcpy.AddMessage("field added! {}".format(field_name))
                    # turn it into a feature class
                    #arcpy.AddMessage('done adding fields')

                    fc_fields = arcpy.ListFields(self.output_feature_class)
                    fc_field_names = [fld.name for fld in fc_fields]

                    #arcpy.AddMessage("fields are: {}".format(fc_field_names))

                    # now we call an insert cursor on all the fields in the Feature Class MINUS 
                    # ObjectID and Shape which will always be the first 2 fields. So we do
                    # fc_field_names[2:]

                    cursor = arcpy.da.InsertCursor(self.output_feature_class, fc_field_names[2:] + ['SHAPE@XY'])

                    big_counter = 1
                    wrote_warning = False
                    if self.analyze_obj.in_file_type == 'table':
                        # adjustment for the very end since row_count from arcpy gives one less
                        self.row_count += 1
                    for row_num in range(self.row_count - self.num_empty_rows - 1):
                        try:
                            matched = True
                            # row_count - 1 because we already have the header
                            ### needs more debugging -> is there a csv output being created at all?
                            ### is sql getting read in at all? check what's being sent in JSON


                            csv_row = next(csvReader)
                            #arcpy.AddMessage("the row I read looks like {}".format(csv_row))
                            #arcpy.AddMessage("csv file to look at is {}".format(os.path.join(self.working_dir,self.output_name + '.csv')))
                            point = (csv_row[index_of_ShapeX],csv_row[index_of_ShapeY])
                            #arcpy.AddMessage(point)
                            #del csv_row[index_of_ShapeX]
                            #del csv_row[index_of_ShapeY]
                            final_row = []
                            for i, field in enumerate(csv_row):
                                if len(field) > 255:
                                    field = field[:255]
                                else:
                                    if extent_exists:
                                        if i == index_of_X or i == index_of_Y or i == index_of_xmax or i == index_of_xmin or i == index_of_ymax or i == index_of_ymin or i == score_index:
                                            final_row.append(rounded_convert_to_int_if_possible(field))
                                        else:
                                            final_row.append(field)
                                    else:
                                        if i == index_of_X or i == index_of_Y or i == score_index:
                                            final_row.append(rounded_convert_to_int_if_possible(field))
                                        else:
                                            final_row.append(field)

                            #csv_row_cleaned = [rounded_convert_to_int_if_possible(val) for val in csv_row]
                            if status_index_exists:
                                if final_row[status_index] == "M":
                                    self.matched += 1
                                    if match_type_exists:
                                        final_row[match_type_index] = "A"
                                    x = float(point[0])
                                    y = float(point[1])
                                    point_to_append = [[x,y]]
                                    if x_exists:
                                        if final_row[index_of_X] == "":
                                            final_row[index_of_X] = None
                                    if y_exists:
                                        if final_row[index_of_Y] == "":
                                            final_row[index_of_Y] = None
                                    #arcpy.AddMessage(point_to_append)
                                    final_list = final_row + point_to_append
                                elif final_row[status_index] == "T":
                                    self.tied += 1
                                    if match_type_exists:
                                        final_row[match_type_index] = "A"
                                    x = float(point[0])
                                    y = float(point[1])
                                    point_to_append = [[x,y]]
                                    if x_exists:
                                        if final_row[index_of_X] == "":
                                            final_row[index_of_X] = None
                                    if y_exists:
                                        if final_row[index_of_Y] == "":
                                            final_row[index_of_Y] = None
                                    #arcpy.AddMessage(point_to_append)
                                    final_list = final_row + point_to_append
                                else:
                                    if extent_exists:
                                        final_row[index_of_xmax] = None
                                        final_row[index_of_xmin] = None
                                        final_row[index_of_ymax] = None
                                        final_row[index_of_ymin] = None                                
                                    self.unmatched += 1
                                    point_to_append = [[None,None]]
                                    if score_index_exists:
                                        final_row[score_index] = None
                                    if x_exists:
                                        final_row[index_of_X] = None
                                    if y_exists:
                                        final_row[index_of_Y] = None
                                    final_list = final_row + point_to_append
                            #arcpy.AddMessage("pre removal {}".format(final_list))
                            else:
                                if point[0] == "" and point[1] == "":
                                    point_to_append = [[None,None]]
                                else:
                                    x = float(point[0])
                                    y = float(point[1])
                                    point_to_append = [[x,y]]
                                final_list = final_row + point_to_append 
                            del final_list[index_of_ShapeX]
                            del final_list[index_of_ShapeX]
                            #arcpy.AddMessage("post removal {}".format(final_list))
                            try:
                                cursor.insertRow(final_list)
                            except RuntimeError as e:
                                #arcpy.AddMessage("ERROR IS {}".format(e))
                                #arcpy.AddWarning("A row could not be written {}".format(final_list))
                                self.logger.log("A row could not be written {}".format(final_list),'INFO')
                                pass
                            except TypeError:
                                #arcpy.AddMessage('tried to write row {}'.format(final_list))
                                self.logger.log('Tried to write row {}'.format(final_list),'INFO')
                                pass
                                #lol
                            #big_counter += 1
                        except StopIteration:
                            break

                    del cursor

            else:
                pass

    def preprocess(self):
        # NOTE: There is tool validation code in the script tool
        # that detects whether the service is secure or unsecure
        # and whether the service has category and source country
        # enabled. Here, we just taken in these as more inputs

        if not self.analyze_obj.EGDB_table:
            # if not writing to an EGDB, write to the scratch folder
            self.output_location = arcpy.env.scratchFolder
            folder_name, file_name = os.path.split(self.output_location)

        # The tool validation also throws an error immediately
        # if the file is not a zip file
        excel_zero = False
        if self.analyze_obj.in_file_type == 'csv':
            self.row_count = sum(1 for line in open(self.analyze_obj.unzipped_file, encoding="utf8"))
        elif self.analyze_obj.in_file_type == 'xlsx' or self.analyze_obj.in_file_type == 'xls':
            try:
                if not self.analyze_obj.unzipped_file.endswith(".xlsx"):
                    shutil.copy(self.analyze_obj.unzipped_file, self.analyze_obj.unzipped_file + ".xlsx")
                    #openpyxl.Workbook()
                    try:
                        workbook = openpyxl.load_workbook(self.analyze_obj.unzipped_file + ".xlsx", read_only=True)
                        self.using_xlsx = True
                    except Exception as e:
                        ### this means we have a .xls file! use xlrd instead
                        shutil.copy(self.analyze_obj.unzipped_file, self.analyze_obj.unzipped_file + ".xls")
                        #  workbook = xlrd.open_workbook(in_file + ".xls")
                        #     self.in_worksheet = workbook.sheet_by_index(0)
                        #     first_row = self.in_worksheet.row(0)
                        workbook = xlrd.open_workbook(self.analyze_obj.unzipped_file + ".xls")
                        self.using_xlsx = False

                else:
                    workbook = openpyxl.load_workbook(self.analyze_obj.unzipped_file, read_only=True)
                    self.using_xlsx = True
            except Exception as e:
                if not self.analyze_obj.started_as_feature_service:
                    self.delete_unwanted_folder()
                msg_code = 100177
                msg = self.ERROR_CODES[msg_code]
                self.addError(msg_code, msg)


            try:
                if self.table_name:
                    #arcpy.AddMessage("TABLE NAME IS {}".format(self.table_name))
                    if "/" in self.table_name or "\\" in self.table_name:
                        ### entire path was passed instead of just sheet name
                        if self.using_xlsx:
                            self.in_worksheet = workbook[os.path.basename(self.table_name)]
                            self.row_count = self.in_worksheet.max_row
                        else:
                            self.in_worksheet = workbook.sheet_by_name(os.path.basename(self.table_name))
                            self.row_count = self.in_worksheet.nrows
                    else:
                        if self.using_xlsx:
                            self.in_worksheet = workbook[self.table_name]
                            self.row_count = self.in_worksheet.max_row
                            #arcpy.AddMessage("DONE GETTING max count1")
                        else:
                            self.in_worksheet = workbook.sheet_by_name(self.table_name)
                            self.row_count = self.in_worksheet.nrows
                            #arcpy.AddMessage("DONE GETTING max count2")
                else:
                    if self.using_xlsx:
                        self.in_worksheet = workbook.active
                        self.row_count = self.in_worksheet.max_row
                    else:
                        self.in_worksheet = workbook.sheet_by_index(0)
                        self.row_count = self.in_worksheet.nrows
            except Exception as e:
                msg_code = 100178
                msg = self.ERROR_CODES[msg_code]
                self.addError(msg_code, msg)

            
        elif self.analyze_obj.in_file_type == 'txt':
            # reader = pandas.read_table(in_file,sep=delimiter,iterator=True)
            self.row_count = sum(1 for line in open(self.analyze_obj.unzipped_file))

        else:
            if self.analyze_obj.in_table.isLayer:
                result = arcpy.GetCount_management(self.analyze_obj.in_table.layer)
                self.row_count = int(result.getOutput(0)) + 1
            else:
                self.row_count = self.analyze_obj.in_table.count + 1
        #arcpy.AddMessage("row count is {}".format(self.row_count))
        if self.max_records:
            if self.row_count > self.max_records:
                msg_dict = dict(max=self.max_records)
                msg_code = 100233
                msg = self.ERROR_CODES[msg_code].format(**msg_dict)
                self.addError(msg_code, msg)

        try:
            int_header_rows = int(self.header_rows_to_skip)

        except ValueError:
            if not self.analyze_obj.started_as_feature_service:
                if not self.analyze_obj.started_as_feature_service:
                    self.delete_unwanted_folder()
            msg_code = 100178
            msg = self.ERROR_CODES[msg_code]
            self.addError(msg_code, msg)


        if self.header_row_exists and self.header_rows_to_skip == '':
            if not self.analyze_obj.started_as_feature_service:
                self.delete_unwanted_folder()
            msg_code = 100192
            msg = self.ERROR_CODES[msg_code]
            self.addError(msg_code, msg)


        if int(self.row_count) >= 65536 and self.out_file_type == 'xls':
            msg_code = 100229
            msg = self.ERROR_CODES[msg_code]
            self.addError(msg_code, msg)


        if self.row_count == 0:
            msg_code = 100178
            msg = self.ERROR_CODES[msg_code]
            self.addError(msg_code, msg)


        if int_header_rows < 0:
            if not self.analyze_obj.started_as_feature_service:
                self.delete_unwanted_folder()
            msg_code = 100193
            msg = self.ERROR_CODES[msg_code]
            self.addError(msg_code, msg)


        if int_header_rows > self.row_count:
            if not self.analyze_obj.started_as_feature_service:
                self.delete_unwanted_folder()
            msg_code = 100228
            msg = self.ERROR_CODES[msg_code]
            self.addError(msg_code, msg)


        self.header_rows_to_skip = int_header_rows

        output_zip_folder = os.path.join(self.output_location,self.output_name + '.zip')
        # self.output = zipfile.ZipFile(output_zip_folder,'w', zipfile.ZIP_DEFLATED)
        self.output = ''
        if self.row_count == 1 and self.header_row_exists or self.row_count == 0:
            if self.analyze_obj.in_file_type == 'table':
                pass
            else:
                arcpy.AddError('Your file or table has no records. Please include records to be geocoded and resubmit.')
                raise arcpy.ExecuteError

        if excel_zero:
            # This means your file was empty
            # We need to return an empty file to the user
            try:
                if self.out_file_type == 'csv':
                    self.csv_out_file = open(os.path.join(self.working_dir,self.output_name + '.csv'),'w',encoding="utf8", newline='')
                elif self.out_file_type == 'xls':
                    self.xls_out_workbook = openpyxl.Workbook()
                    self.out_worksheet = self.xls_out_workbook.active
                    #out_worksheet = self.xls_out_workbook.create_sheet('Sheet1')
                    self.xls_out_workbook.save(os.path.join(self.working_dir,self.output_name + '.xlsx'))
                else:
                    self.txt_file_out = open(os.path.join(self.working_dir,self.output_name + '.txt'),'a')
            except Exception as e:
                shutil.rmtree(self.output_location,ignore_errors=True)
                if not self.analyze_obj.started_as_feature_service:
                    self.delete_unwanted_folder()
                msg_code = 100179
                msg = self.ERROR_CODES[msg_code]
                self.addError(msg_code, msg)

            # Benchmark how long the whole process took
            self.end_time = time.time()

            self.logger.log('About to write report','INFO')

            # Write the report file
            # with open(os.path.join(self.working_dir,'report.txt'), 'w') as report_file:
            #     time_taken = self.end_time-self.start_time
            #     report_file.write('The job took {}'.format(round(time_taken,1)) + \
            #                       ' seconds to finish' + '\n')
            #     report_file.write('Your file was empty or corrupt')

            if self.out_file_type == 'csv':
                self.csv_out_file.close()
            if self.out_file_type == 'xls':
                pass
            if self.out_file_type == 'txt':
                self.txt_file_out.close()

            # Write to self.output file
            if self.out_file_type == 'csv':
                self.output = os.path.join(self.working_dir,self.output_name + '.csv')
            elif self.out_file_type == 'xls':
                self.output = os.path.join(self.working_dir,self.output_name + '.xls')
            elif self.out_file_type == 'txt':
                self.output = os.path.join(self.working_dir,self.output_name + '.txt')
            # self.output.write(os.path.join(self.working_dir,'report.txt'),'report.txt')



            self.logger.log('finished writing to zip file','INFO')
            # Remove the files from the workspace now that they've
            # been written into the zip file
            # try:
            #     if self.out_file_type == 'csv':
            #         os.remove(os.path.join(self.working_dir,self.output_name + '.csv'))
            #     elif self.out_file_type == 'xls':
            #         os.remove(os.path.join(self.working_dir,self.output_name + '.xls'))
            #     elif self.out_file_type == 'txt':
            #         os.remove(os.path.join(self.working_dir,self.output_name + '.txt'))
            #     # os.remove(os.path.join(self.working_dir,'report.txt'))
            #     if not self.analyze_obj.started_as_feature_service:
            #         self.delete_unwanted_folder()
            #     #Why ins't folder to remove the right thing? Is that the issue??
            #     self.logger.log('Was able to remove intermittent file','INFO')
            #
            # except Exception as e:
            #     # Not the end of the world since the only thing
            #     # returned out to the user is the zip file
            #     self.logger.log('Could not remove intermittent files {}'.format(e),'INFO')
            # Done with post-processing for an empty file
            msg_code = 100195
            msg = self.ERROR_CODES[msg_code]
            self.addError(msg_code, msg)




        self.logger.log("self.row_count is {}".format(self.row_count),'INFO')

        # Now that we have the row count, we need to find the max batch size supported
        # by the geocoding service
        self.batch_size = self.analyze_obj.json_info_response['locatorProperties']['MaxBatchSize']

        self.num_of_batches, self.remainder = divmod(self.row_count - self.header_rows_to_skip, self.batch_size)
        if self.remainder > 0:
            self.num_of_batches += 1

        self.logger.log('Number of batches is {}'.format(self.num_of_batches),'INFO')
        self.logger.log('Remainder is {}'.format(self.remainder),'INFO')
        # num_of_batches is now the number of total batches to be sent



        for map in self.field_mapping:
            if len(map) != 2:
                shutil.rmtree(self.output_location,ignore_errors=True)
                os.rmdir(self.output_location)
                if not self.analyze_obj.started_as_feature_service:
                    self.delete_unwanted_folder()
                msg_code = 100196
                msg = self.ERROR_CODES[msg_code]
                self.addError(msg_code, msg)

        # make sure that SOMETHING is mapped
        valid_field_mapping = False
        for one_list in self.field_mapping:
            if one_list[1] != '':
                valid_field_mapping = True
        if not valid_field_mapping:
            shutil.rmtree(self.output_location,ignore_errors=True)
            if not self.analyze_obj.started_as_feature_service:
                self.delete_unwanted_folder()
            msg_code = 100198
            msg = self.ERROR_CODES[msg_code]
            self.addError(msg_code, msg)

        # field_mapping_good is now the acutal field mapping
        # exec is used to get the list of lists that the user
        # actually meant instead of reading it in as a string

        self.relevent_columns = []
        # This bool tracks if their input data has an ObjectID column.
        # If it doesn't, then we need to assign each record one before
        # we send the record off in a rest request. This assignment happens
        # in recordset_builder. We need each record to have an associated ID
        # so that we can tie the input data to the self.output data so that all
        # of the rows are correct. This is important for Asynchronous REST
        # calls because results do not come back in the same order, batch
        # or record order
        self.object_id_in_input = False

        # Note: relevant_columns are the column numbers mapped to arcgis fields
        # the other columns are ignored by create_JSON
        self.relevant_user_fields = []
        self.user_arcgis_fields = []
        self.user_field_nums = []
        self.all_fields = []


        ## Add this error code in at 10.8
        # # Check for duplicate field names and immediately throw an error
        # # if the users have a duplicate field name in their table
        # fields_set = set()
        # for tup in self.field_info:
        #     if tup[0] not in fields_set:
        #         fields_set.add(tup[0])
        #     else:
        #         arcpy.AddWarning("Your input table has a duplicate field name. Please resubmit your input table with no duplicate field names.")
        #         raise arcpy.ExecuteError

        # Adjust the field mapping to be in the correct order
        correct_ordering = []
        for tuple in self.field_info:
            correct_ordering.append(tuple[0])

        ordered_field_map = []
        for fld in correct_ordering:
            written = False
            for fld_list in self.field_mapping:
                if fld_list[0] == fld:
                    ordered_field_map.append(fld_list)
                    written = True
            if not written:
                ordered_field_map.append([fld,''])
        self.field_mapping = ordered_field_map
        #arcpy.AddMessage('new fm is {}'.format(self.field_mapping))
        # Populate some arrays from the user-provided field
        # mapping that we will use later in the program.
        for num, field in enumerate(self.field_mapping):
            if field[1] == '':
                pass
            else:
                self.user_arcgis_fields.append(field[1])
                self.user_field_nums.append([num,field[0]])
                #arcpy.AddMessage(self.user_arcgis_fields)

        #arcpy.AddMessage(self.user_arcgis_fields)
        single_line_field_name = self.analyze_obj.json_info_response['singleLineAddressField']['name']

        self.single_line_only = True

        for mapped_fld in self.user_arcgis_fields:
            if mapped_fld != "" and mapped_fld != single_line_field_name:
                self.single_line_only = False

        #arcpy.AddMessage("to parse: {}".format(self.user_field_nums))
        # Make sure the ArcGIS fields that the user
        # is mapping to are relevant, actual fields
        for field in self.user_arcgis_fields:
            if field not in self.analyze_obj.available_service_fields:
                service_fields_lower = [x.lower() for x in self.analyze_obj.available_service_fields]
                found_it = False
                for some_field in service_fields_lower:
                    if field.lower() == some_field:
                        found_it = True
                        break
                if not found_it:
                    try:
                        # new_list = convert_list_to_ascii()
                        shutil.rmtree(self.output_location,ignore_errors=True)
                        if not self.analyze_obj.started_as_feature_service:
                            self.delete_unwanted_folder()
                        msg_dict = dict(supported_fields=self.analyze_obj.available_service_fields)
                        msg_code = 100200
                        msg = self.ERROR_CODES[msg_code].format(**msg_dict)
                        self.addError(msg_code, msg, msg_dict)

                    except Exception as e:
                        # If for some reason you weren't able to get available service fields...
                        shutil.rmtree(self.output_location,ignore_errors=True)
                        if not self.analyze_obj.started_as_feature_service:
                            self.delete_unwanted_folder()
                        msg_code = 100199
                        msg = self.ERROR_CODES[msg_code]
                        self.addError(msg_code, msg)

        # Make sure the user-defined fields
        # in the field mapping provided by the user
        # are actually in the header of their file
        try:
            self.sixty_four_bit_OID = False
            # Everything below tries to parse the header row
            if self.analyze_obj.in_file_type == 'csv':
                self.prelim_csv_file = self.analyze_obj.unzipped_file
                if self.analyze_obj.column_delimiter == '':
                    # If its a csv and they did not provide a column delimiter
                    # assume the file is comma-delimited!
                    self.analyze_obj.column_delimiter = ','
                #self.analyze_obj.column_delimiter = self.analyze_obj.column_delimiter.encode('ascii','ignore')
                #self.prelim_csvReader = csv.reader(self.prelim_csv_file,delimiter=self.analyze_obj.column_delimiter)
                self.csv_file_handle = open(self.prelim_csv_file, encoding="utf8")
                #arcpy.AddMessage('delimiter is {}'.format(self.analyze_obj.column_delimiter))
                self.prelim_csvReader = unicode_csv_reader(self.csv_file_handle,delimiter=str(self.analyze_obj.column_delimiter))
                if self.header_row_exists:
                    # parsed_success = False
                    try:
                        header_row = []
                        almost_header_row = next(self.prelim_csvReader)
                        for h in almost_header_row:
                            header_row.append(h.strip())
                    except:
                        self.csv_file_handle.close()
                        shutil.rmtree(self.output_location,ignore_errors=True)
                        if not self.analyze_obj.started_as_feature_service:
                            self.delete_unwanted_folder()
                        msg_code = 100176
                        msg = self.ERROR_CODES[msg_code]
                        self.addError(msg_code, msg)

                    if not header_row:
                        shutil.rmtree(self.output_location,ignore_errors=True)
                        if not self.analyze_obj.started_as_feature_service:
                            self.delete_unwanted_folder()
                        self.csv_file_handle.close()
                        msg_code = 100181
                        msg = self.ERROR_CODES[msg_code]
                        self.addError(msg_code, msg)



                    new_header_row = header_row

                    if self.text_qualifier == '\'':
                        new_header_row_to_use = []
                        for value in new_header_row:
                            new_header_row_to_use.extend(re.findall("'.*?'",value))
                        new_header_row = new_header_row_to_use
                        header_row_final = []
                        for val in new_header_row:
                            header_row_final.append(val.strip('\''))
                        new_header_row = header_row_final

                    # The next 3 lines remove a \xef\xbb\xbf from the start
                    # of a csv if necessary. Usually \xef\xbb\xbf are markers
                    # that the csv is unicode-encoded
                    #arcpy.AddMessage("header row is {}".format(new_header_row))
                    #arcpy.AddMessage("checking if it start with {}".format(str(codecs.BOM_UTF8)))
                    if new_header_row[0].startswith(str(codecs.BOM_UTF8)):
                        beginning_of_file = new_header_row[0]
                        new_header_row[0] = beginning_of_file[3:]

                    # account for another UTF-8 BOM encoding
                    if new_header_row[0].startswith(u'\ufeff'):
                        beginning_of_file = new_header_row[0]
                        new_header_row[0] = beginning_of_file.replace(u'\ufeff','')
                        #arcpy.AddMessage("new header row!! Replaced BOM character")
                    #arcpy.AddMessage(new_header_row)
                    header_row_lower = [f.lower() for f in new_header_row]
                    #arcpy.AddMessage(header_row_lower)
                    #arcpy.AddMessage(self.all_fields)
                    field_mapping_claim = [fld[1] for fld in self.user_field_nums]
                    #arcpy.AddMessage(field_mapping_claim)
                    for field in field_mapping_claim:
                        if field.lower() not in header_row_lower:
                            shutil.rmtree(self.output_location,ignore_errors=True)
                            if not self.analyze_obj.started_as_feature_service:
                                self.delete_unwanted_folder()
                            #no_unicode = convert_list_to_ascii(header_row)
                            msg_dict = dict(field=field, supported_fields=header_row)
                            msg_code = 100201
                            msg = self.ERROR_CODES[msg_code].format(**msg_dict)
                            self.addError(msg_code, msg, msg_dict)


                else:
                    header_row = []
                    almost_header_row = self.field_names.split(',')
                    for h in almost_header_row:
                        header_row.append(h.strip())
                    # If there is no header row, then set header_rows_to_skip to be 0
                    self.header_rows_to_skip = 0
                    new_header_row = header_row
                    header_row_lower = [f.lower() for f in new_header_row]
                    if new_header_row[0].startswith(str(codecs.BOM_UTF8)):
                        beginning_of_file = new_header_row[0]
                        new_header_row[0] = beginning_of_file[3:]
                    for field in self.all_fields:
                        if field.lower() not in header_row_lower:
                            shutil.rmtree(self.output_location,ignore_errors=True)
                            if not self.analyze_obj.started_as_feature_service:
                                self.delete_unwanted_folder()
                            no_unicode = convert_list_to_ascii(new_header_row)
                            msg_dict = dict(field=field, supported_fields=no_unicode)
                            msg_code = 100201
                            msg = self.ERROR_CODES[msg_code].format(**msg_dict)
                            self.addError(msg_code, msg, msg_dict)


                for i in range(self.header_rows_to_skip - 1):
                    # Burn through empty lines
                    dummy = next(self.prelim_csvReader)

                #if not self.analyze_obj.started_as_feature_service:
                #    if len(header_row) != len(self.field_mapping):
                #        msg_code = 100227
                #        msg = self.ERROR_CODES[msg_code]
                #        self.addError(msg_code, msg)

                # Ignore input fields that the user chose not to map

                header_row_lower = [elem.lower() for elem in new_header_row]
                self.correct_len = len(new_header_row)

                len_of_field_info = len(self.field_info)

                new_field_info = []

                if self.correct_len != len_of_field_info:
                    #arcpy.AddMessage('GETS INSIDE CRUCIAL LOOP')
                    for idx, tup in enumerate(self.field_info,start=1):
                        if idx <= self.correct_len:
                            new_field_info.append(tup)
                    self.field_info = new_field_info
                
                self.id_col = None
                self.object_id_in_input = False
                counter = 0
                for tuple in self.field_info:
                    self.all_fields.append(tuple[0])
                for field in self.field_mapping:
                    #self.all_fields.append(field[0])
                    if field[1] != '':
                        self.relevant_user_fields.append(field[0])
                        if field[0].lower() in header_row_lower:
                            for fld in new_header_row:
                                if field[0].lower() == fld.lower():
                                    self.relevent_columns.append(counter)
                                    if field[1].lower() == 'objectid' or field[1].lower() == 'oid':
                                        self.id_col = counter
                                        self.object_id_in_input = True
                                    if field[1] == 'SingleLine':
                                        singleLine_field_position = counter
                                
                        else:
                            shutil.rmtree(self.output_location,ignore_errors=True)
                            if not self.analyze_obj.started_as_feature_service:
                                self.delete_unwanted_folder()
                            no_unicode = convert_list_to_ascii(new_header_row)
                            msg_dict = dict(field=field[0], supported_fields=no_unicode)
                            msg_code = 100201
                            msg = self.ERROR_CODES[msg_code].format(**msg_dict)
                            self.addError(msg_code, msg, msg_dict)
                    counter += 1
                    #arcpy.AddMessage(self.relevent_columns)
                            
                lower_arcgis_fields = [fld.lower() for fld in self.user_arcgis_fields]

                # if 'SingleLine' in self.user_arcgis_fields:
                #     len_of_relevant = len(self.relevent_columns) - 1
                #     if "objectid" in lower_arcgis_fields:
                #         len_of_relevant -= 1
                #     if len_of_relevant > 0:
                #         # If there a sufficient amount of multiline fields,
                #         # then drop the singleLine field and use the multiline
                #         # fields
                #         self.relevent_columns.remove(singleLine_field_position)
                #         idx_to_remove = self.user_arcgis_fields.index('SingleLine')
                #         self.user_arcgis_fields.remove('SingleLine')
                #         del self.user_field_nums[idx_to_remove]
                #         # Notice that we weren't precise here, thats because
                #         # this array never actually gets used again. only its
                #         # length gets used! So, we just delete one at random.
                #         del self.relevant_user_fields[0]
                #         self.logger.log('SingleLine removed','INFO')



            elif self.analyze_obj.in_file_type == 'xlsx' or self.analyze_obj.in_file_type == 'xls':
                if self.header_row_exists:
                    first_row_values = []
                    if self.using_xlsx:
                        first_row = self.in_worksheet[1]
                    else:
                        first_row = self.in_worksheet.row(0)
                    for cell in first_row:
                        first_row_values.append(cell.value.strip())
                    if not first_row_values:
                        shutil.rmtree(self.output_location,ignore_errors=True)
                        if not self.analyze_obj.started_as_feature_service:
                            self.delete_unwanted_folder()
                        msg_code = 100181
                        msg = self.ERROR_CODES[msg_code]
                        self.addError(msg_code, msg)

                    if self.text_qualifier == '\'':
                        new_header_row_to_use = []
                        for value in first_row_values:
                            new_header_row_to_use.extend(re.findall("'.*?'",value))
                        first_row_values = new_header_row_to_use
                        header_row_final = []
                        for val in first_row_values:
                            header_row_final.append(val.strip('\''))
                        first_row_values = header_row_final

                    if self.text_qualifier == "\"":
                        new_header_row_to_use = []
                        for value in first_row_values:
                            new_header_row_to_use.extend(re.findall('".*?"',value))
                        first_row_values = new_header_row_to_use
                        header_row_final = []
                        for val in first_row_values:
                            header_row_final.append(val.strip('\"'))
                        first_row_values = header_row_final
                    
                    header_row_lower = [f.lower() for f in first_row_values]
                    for field in self.all_fields:
                        if field.lower() not in header_row_lower:
                            shutil.rmtree(self.output_location,ignore_errors=True)
                            if not self.analyze_obj.started_as_feature_service:
                                self.delete_unwanted_folder()
                            new_list = convert_list_to_ascii(first_row_values)
                            msg_dict = dict(field=field, supported_fields=new_list)
                            msg_code = 100201
                            msg = self.ERROR_CODES[msg_code].format(**msg_dict)
                            self.addError(msg_code, msg, msg_dict)

                else:
                    almost_first_row_values = self.field_names.split(',')
                    for val in almost_first_row_values:
                        first_row_values.append(val.strip())
                    # If there is no header row, then set header_rows_to_skip to be 0
                    header_row_lower = [f.lower() for f in first_row_values]
                    self.header_rows_to_skip = 0
                    for field in self.all_fields:
                        if field.lower() not in header_row_lower:
                            shutil.rmtree(self.output_location,ignore_errors=True)
                            if not self.analyze_obj.started_as_feature_service:
                                self.delete_unwanted_folder()
                            new_list = convert_list_to_ascii(first_row_values)
                            msg_dict = dict(field=field, supported_fields=new_list)
                            msg_code = 100201
                            msg = self.ERROR_CODES[msg_code].format(**msg_dict)
                            self.addError(msg_code, msg, msg_dict)


                #if len(first_row_values) != len(self.field_mapping):
                #    msg_code = 100227
                #    msg = self.ERROR_CODES[msg_code]
                #    self.addError(msg_code, msg)
                header_row_lower = [elem.lower() for elem in first_row_values]
                self.correct_len = len(first_row_values)

                len_of_field_info = len(self.field_info)

                #arcpy.AddMessage("len of field info is {}".format(len_of_field_info))

                ### Note: the excel driver is buggy in Pro. It adds an extra
                ### "ObjectID" field to the end of every excel file right now.
                ### If the length is 1 more than necessary, and the last field
                ### contains "ObjectID", ignore that last field and keep going
                ### forward.

                #arcpy.AddMessage("old field info is: {}".format(self.field_info))
                new_field_info = []

                if self.correct_len != len_of_field_info:
                    if self.correct_len + 1 == len_of_field_info:
                        #arcpy.AddMessage('GETS INSIDE CRUCIAL LOOP')
                        for idx, tup in enumerate(self.field_info,start=1):
                            if idx <= self.correct_len:
                                new_field_info.append(tup)
                        self.field_info = new_field_info

                #arcpy.AddMessage("new field info is: {}".format(self.field_info))
                
                self.id_col = None
                self.object_id_in_input = False
                for tuple in self.field_info:
                    self.all_fields.append(tuple[0])
                for field in self.field_mapping:
                    counter = 0
                    #self.all_fields.append(field[0])
                    if field[1] != '':
                        self.relevant_user_fields.append(field[0])
                        if field[0].lower() in header_row_lower:
                            for fld in first_row_values:
                                if field[0].lower() == fld.lower():
                                    self.relevent_columns.append(counter)
                                    if field[1].lower() == 'objectid' or field[1].lower() == 'oid':
                                        self.id_col = counter
                                        self.object_id_in_input = True
                                    if field[1] == 'SingleLine':
                                        singleLine_field_position = counter
                                counter += 1
                        else:
                            shutil.rmtree(self.output_location,ignore_errors=True)
                            if not self.analyze_obj.started_as_feature_service:
                                self.delete_unwanted_folder()
                            no_unicode = convert_list_to_ascii(new_header_row)
                            msg_dict = dict(field=field[0], supported_fields=no_unicode)
                            msg_code = 100201
                            msg = self.ERROR_CODES[msg_code].format(**msg_dict)
                            self.addError(msg_code, msg, msg_dict)
                            
                lower_arcgis_fields = [fld.lower() for fld in self.user_arcgis_fields]

                # if 'SingleLine' in self.user_arcgis_fields:
                #     len_of_relevant = len(self.relevent_columns) - 1
                #     if "objectid" in lower_arcgis_fields:
                #         len_of_relevant -= 1
                #     if len_of_relevant > 0:
                #         # If there a sufficient amount of multiline fields,
                #         # then drop the singleLine field and use the multiline
                #         # fields
                #         self.relevent_columns.remove(singleLine_field_position)
                #         idx_to_remove = self.user_arcgis_fields.index('SingleLine')
                #         self.user_arcgis_fields.remove('SingleLine')
                #         del self.user_field_nums[idx_to_remove]
                #         # Notice that we weren't precise here, thats because
                #         # this array never actually gets used again. only its
                #         # length gets used! So, we just delete one at random.
                #         del self.relevant_user_fields[0]
                #         self.logger.log('SingleLine removed','INFO')


            elif self.analyze_obj.in_file_type == 'txt':
                self.f = open(self.analyze_obj.unzipped_file)
                if self.header_row_exists:
                    raw_first_line = self.f.readline()
                    if self.fixed_width:
                        if self.chars_per_row != '':
                            almost_row = raw_first_line[0:int(self.chars_per_row)]
                            self.total_counter += int(self.chars_per_row)
                        else:
                            almost_row = raw_first_line
                        counter = 0
                        row = []
                        for num_of_chars in self.chars_per_column:
                            row.append(almost_row[counter:counter+int(num_of_chars)].strip(' '))
                            counter = counter + int(num_of_chars)
                        first_row_array = row
                    # Because tabs are the only delimiter we are supporting
                    if not self.fixed_width:
                        first_row_array = raw_first_line.split('\t')
                    if self.text_qualifier == '\'':
                        new_header_row_to_use = []
                        for value in first_row_array:
                            # Use regex to parse for text qualifier
                            new_header_row_to_use.extend(re.findall("'.*?'",value))
                        first_row_array = new_header_row_to_use
                        header_row_final = []
                        for val in first_row_array:
                            header_row_final.append(val.strip('\''))
                        first_row_array = header_row_final

                    if self.text_qualifier == "\"":
                        new_header_row_to_use = []
                        for value in first_row_array:
                            # Use regex to parse for text qualifier
                            new_header_row_to_use.extend(re.findall('".*?"',value))
                        first_row_array = new_header_row_to_use
                        header_row_final = []
                        for val in first_row_array:
                            header_row_final.append(val.strip('\"'))
                        first_row_array = header_row_final

                    if self.chars_per_row == '':
                        final_header_row = []
                        for val in first_row_array:
                            final_header_row.append(val.strip('\n'))
                        first_row_array = final_header_row

                    if not first_row_array:
                        shutil.rmtree(self.output_location,ignore_errors=True)
                        if not self.analyze_obj.started_as_feature_service:
                            self.delete_unwanted_folder()
                        msg_code = 100181
                        msg = self.ERROR_CODES[msg_code]
                        self.addError(msg_code, msg)

                    header_row_lower = [f.lower() for f in first_row_array]
                    for field in self.all_fields:
                        if field.lower() not in header_row_lower:
                            shutil.rmtree(self.output_location,ignore_errors=True)
                            if not self.analyze_obj.started_as_feature_service:
                                self.delete_unwanted_folder()
                            new_list = convert_list_to_ascii(first_row_array)
                            msg_dict = dict(field=field, supported_fields=new_list)
                            msg_code = 100201
                            msg = self.ERROR_CODES[msg_code].format(**msg_dict)
                            self.addError(msg_code, msg, msg_dict)

                else:
                    first_row_array = self.field_names.split(',')
                    # If there is no header row, then set header_rows_to_skip to be 0
                    self.header_rows_to_skip = 0
                    header_row_lower = [f.lower() for f in first_row_array]
                    for field in self.all_fields:
                        if field.lower() not in header_row_lower:
                            shutil.rmtree(self.output_location,ignore_errors=True)
                            if not self.analyze_obj.started_as_feature_service:
                                self.delete_unwanted_folder()
                            new_list = convert_list_to_ascii(first_row_array)
                            msg_dict = dict(field=field, supported_fields=new_list)
                            msg_code = 100201
                            msg = self.ERROR_CODES[msg_code].format(**msg_dict)
                            self.addError(msg_code, msg, msg_dict)


                for i in range(self.header_rows_to_skip - 1):
                    #literally burn through empty lines
                    dummy = self.f.readline()

                #if len(first_row_array) != len(self.field_mapping):
                #    msg_code = 100227
                #    msg = self.ERROR_CODES[msg_code]
                #    self.addError(msg_code, msg)

                header_row_lower = [elem.lower() for elem in first_row_array]
                self.correct_len = len(first_row_array)
                
                self.id_col = None
                self.object_id_in_input = False
                for tuple in self.field_info:
                    self.all_fields.append(tuple[0])
                for field in self.field_mapping:
                    counter = 0
                    #self.all_fields.append(field[0])
                    if field[1] != '':
                        self.relevant_user_fields.append(field[0])
                        if field[0].lower() in header_row_lower:
                            for fld in first_row_values:
                                if field[0].lower() == fld.lower():
                                    self.relevent_columns.append(counter)
                                    if field[1].lower() == 'objectid' or field[1].lower() == 'oid':
                                        self.id_col = counter
                                        self.object_id_in_input = True
                                    if field[1] == 'SingleLine':
                                        singleLine_field_position = counter
                                counter += 1
                        else:
                            shutil.rmtree(self.output_location,ignore_errors=True)
                            if not self.analyze_obj.started_as_feature_service:
                                self.delete_unwanted_folder()
                            no_unicode = convert_list_to_ascii(new_header_row)
                            msg_dict = dict(field=field[0], supported_fields=no_unicode)
                            msg_code = 100201
                            msg = self.ERROR_CODES[msg_code].format(**msg_dict)
                            self.addError(msg_code, msg, msg_dict)
                            

                lower_arcgis_fields = [fld.lower() for fld in self.user_arcgis_fields]

                # if 'SingleLine' in self.user_arcgis_fields:
                #     len_of_relevant = len(self.relevent_columns) - 1
                #     if "objectid" in lower_arcgis_fields:
                #         len_of_relevant -= 1
                #     if len_of_relevant > 0:
                #         # If there a sufficient amount of multiline fields,
                #         # then drop the singleLine field and use the multiline
                #         # fields
                #         self.relevent_columns.remove(singleLine_field_position)
                #         idx_to_remove = self.user_arcgis_fields.index('SingleLine')
                #         self.user_arcgis_fields.remove('SingleLine')
                #         del self.user_field_nums[idx_to_remove]
                #         # Notice that we weren't precise here, thats because
                #         # this array never actually gets used again. only its
                #         # length gets used! So, we just delete one at random.
                #         del self.relevant_user_fields[0]
                #         self.logger.log('SingleLine removed','INFO')

            else:
                first_row_values = []
                if self.analyze_obj.in_table.isLayer:
                    fields = arcpy.ListFields(self.analyze_obj.in_table.layer)
                    self.aliases = [fld.aliasName for fld in self.analyze_obj.in_table.describe.fields]
                else:
                    #fields = arcpy.ListFields(self.analyze_obj.in_table_path)
                    desc = arcpy.Describe(self.analyze_obj.in_table_path)
                    self.aliases = [fld.aliasName for fld in desc.fields]
                    fields = desc.fields
                

                first_row_values = [fld.name for fld in fields]
                self.oid_index = 0
                for fld in fields:
                    if fld.type == "OID":
                        #arcpy.AddMessage("FIELD LENGTH OF OID IS {}".format(fld.length))
                        if fld.length == 8:
                            self.sixty_four_bit_OID = True
                        # arcpy.AddMessage("precision IS {}".format(fld.precision))
                        break
                    else:
                        self.oid_index += 1
                #arcpy.AddMessage('64-bit boolean is {}'.format(self.sixty_four_bit_OID))
                header_row_lower = [f.lower() for f in first_row_values]
                for field in self.all_fields:
                    if field.lower() not in header_row_lower:
                        shutil.rmtree(self.output_location,ignore_errors=True)
                        if not self.analyze_obj.started_as_feature_service:
                            self.delete_unwanted_folder()
                        new_list = convert_list_to_ascii(first_row_values)
                        msg_dict = dict(field=field, supported_fields=new_list)
                        msg_code = 100201
                        msg = self.ERROR_CODES[msg_code].format(**msg_dict)
                        self.addError(msg_code, msg, msg_dict)
                self.correct_len = len(first_row_values)
                ### current theory: USER_ fields get lowercased SOMEWHERE SOMEHOW in ONLY the input hosted table case
                self.id_col = None
                self.object_id_in_input = False
                for tuple in self.field_info:
                    self.all_fields.append(tuple[0])
                    
                # arcpy.AddMessage("PRINTING SELF.ALL_FIELDS FOR A HOSTED INPUT {}".format(self.all_fields))
                # arcpy.AddMessage("PRINTING SELF.field_mapping {}".format(self.field_mapping))
                # arcpy.AddMessage("PRINTING header_row_lower {}".format(header_row_lower))
                for field in self.field_mapping:
                    counter = 0
                    #self.all_fields.append(field[0])
                    if field[1] != '':
                        self.relevant_user_fields.append(field[0])
                        if field[0].lower() in header_row_lower:
                            for fld in first_row_values:
                                if field[0].lower() == fld.lower():
                                    self.relevent_columns.append(counter)
                                    if field[1].lower() == 'objectid' or field[1].lower() == 'oid':
                                        self.id_col = counter
                                        self.object_id_in_input = True
                                    if field[1] == 'SingleLine':
                                        singleLine_field_position = counter
                                counter += 1
                        else:
                            shutil.rmtree(self.output_location,ignore_errors=True)
                            if not self.analyze_obj.started_as_feature_service:
                                self.delete_unwanted_folder()
                            no_unicode = convert_list_to_ascii(new_header_row)
                            msg_dict = dict(field=field[0], supported_fields=no_unicode)
                            msg_code = 100201
                            msg = self.ERROR_CODES[msg_code].format(**msg_dict)
                            self.addError(msg_code, msg, msg_dict)
                            

                lower_arcgis_fields = [fld.lower() for fld in self.user_arcgis_fields]

                # if 'SingleLine' in self.user_arcgis_fields:
                #     len_of_relevant = len(self.relevent_columns) - 1
                #     if "objectid" in lower_arcgis_fields:
                #         len_of_relevant -= 1
                #     if len_of_relevant > 0:
                #         # If there a sufficient amount of multiline fields,
                #         # then drop the singleLine field and use the multiline
                #         # fields
                #         self.relevent_columns.remove(singleLine_field_position)
                #         idx_to_remove = self.user_arcgis_fields.index('SingleLine')
                #         self.user_arcgis_fields.remove('SingleLine')
                #         del self.user_field_nums[idx_to_remove]
                #         # Notice that we weren't precise here, thats because
                #         # this array never actually gets used again. only its
                #         # length gets used! So, we just delete one at random.
                #         del self.relevant_user_fields[0]
                #         self.logger.log('SingleLine removed','INFO')
        except Exception as e:
            self.logger.log('Your header row could not be parsed {}'\
                            .format(e),'USER_ERROR')
            if self.DEBUG:
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                self.logger.log('{0} {1} {2}'.format(exc_type, fname, exc_tb.tb_lineno),'ERROR')
            # Close whatever file handles are open and return an error
            if self.analyze_obj.in_file_type == 'csv':
                self.csv_file_handle.close()
            if self.analyze_obj.in_file_type == 'xls':
                # xls_out_workbook.save(os.path.join(self.working_dir,'results.xls'))
                pass
            if self.analyze_obj.in_file_type == 'txt':
                self.f.close()

            shutil.rmtree(self.output_location,ignore_errors=True)
            if not self.analyze_obj.started_as_feature_service:
                self.delete_unwanted_folder()
            sys.exit()


        # user_fields are mapped to self.user_arcgis_fields
        # relevant columns numbers from the input data are extracted
        self.logger.log('user_fields looks like' +\
                        str(self.relevant_user_fields),'INFO')
        self.logger.log('self.user_arcgis_fields looks like' +\
                         str(self.user_arcgis_fields),'INFO')
        self.logger.log('relevant_columns looks like' +\
                         str(self.relevent_columns),'INFO')

        # open the writer file handle of each file type
        try:
            if self.out_file_type == 'csv':
                self.csv_out_file = open(os.path.join(self.working_dir,self.output_name + '.csv'),'w', encoding="utf8", newline='')
                #arcpy.AddMessage("file is here: {}".format(os.path.join(self.working_dir,self.output_name + '.csv')))
                # Open the csv writer that will write our self.output CSV
                self.csvWriter = csv.writer(self.csv_out_file)
            elif self.out_file_type == 'xls':
                self.xls_out_workbook = openpyxl.Workbook()
                self.out_worksheet = self.xls_out_workbook.active
                #self.out_worksheet = self.xls_out_workbook.create_sheet('Sheet1')
                self.xls_out_workbook.save(os.path.join(self.working_dir,self.output_name + '.xlsx'))
            else:
                self.txt_file_out = open(os.path.join(self.working_dir,self.output_name + '.txt'),'a')
        except Exception as e:
            shutil.rmtree(self.output_location,ignore_errors=True)
            if not self.analyze_obj.started_as_feature_service:
                self.delete_unwanted_folder()
            msg_code = 100194
            msg = self.ERROR_CODES[msg_code]
            self.addError(msg_code, msg)

    def postprocess(self):
        if not self.error_q.empty():
            error_message = str(self.error_q.get()[1])
            arcpy.AddWarning("There was an error in the BatchGeocode job {}, returning results thusfar".format(error_message))
            # self.logger.log(error_message,'ERROR')
            if 'category' in error_message:
                arcpy.AddError(error_message)
                raise arcpy.ExecuteError
            # So there is something severely wrong here. Let's exit gracefully.

            # Close the input and self.output files because we are done reading from
            # and writing to them
            if self.analyze_obj.in_file_type == 'csv':
                self.csv_file_handle.close()
            if self.analyze_obj.in_file_type == 'xlsx':
                # CSV files need to be closed, xls don't
                pass
            if self.analyze_obj.in_file_type == 'txt':
                self.f.close()
            if self.out_file_type == 'csv':
                self.csv_out_file.close()
                ### for debugging~
                #scratch_folder = arcpy.env.scratchFolder
                #copyfile(os.path.join(self.working_dir, self.output_name + '.csv'), os.path.join(scratch_folder, 'debug.csv'))
                if self.feature_service_output:
                    self.convert_csv_to_feature_service()
                    try:
                        pass
                        # leave the CSV for debugging...
                        #os.remove(os.path.join(self.working_dir, self.output_name + '.csv'))
                    except Exception as e:
                        self.logger.log('Could not remove intermittent files {}'.format(e), 'INFO')
                        # we're done here
                    return
            if self.out_file_type == 'xls':
                self.xls_out_workbook.save(os.path.join(self.working_dir,self.output_name + '.xlsx'))
            if self.out_file_type == 'txt':
                self.txt_file_out.close()
            # Benchmark how long the whole process took
            self.end_time = time.time()

            # Write the report file
            # try:
            #     with open(os.path.join(self.working_dir,'report.txt'), 'w') as report_file:
            #         report_file.write('Your file could not be geocoded because ' + \
            #                           str(error_message) + '\n')
            #         time_taken = self.end_time-self.start_time
            #         report_file.write('The job took {}'.format(round(time_taken,1)) + \
            #                           ' seconds to fail' + '\n')
            #         dropped_records_list = []
            #         while self.dropped_records_q.qsize() > 0:
            #             dropped_records_list.append(self.dropped_records_q.get())
            #         report_file.write('The records that could not be parsed are {}'\
            #                           .format(dropped_records_list))
            # except Exception as e:
            #     self.logger.log('Problem with writing report: {}'.format(e),'ERROR')


            # Write to self.output file
            if self.out_file_type == 'csv':
                self.output = os.path.join(self.working_dir,self.output_name + '.csv')
            elif self.out_file_type == 'xls':
                self.output = os.path.join(self.working_dir,self.output_name + '.xlsx')
            elif self.out_file_type == 'txt':
                self.output = os.path.join(self.working_dir,self.output_name + '.txt')
            # self.output.write(os.path.join(self.working_dir,'report.txt'),'report.txt')



            # Remove the files from the workspace now that they've
            # been written into the zip file
            # try:
            #     if self.out_file_type == 'csv':
            #         os.remove(os.path.join(self.working_dir,self.output_name + '.csv'))
            #     elif self.out_file_type == 'xls':
            #         os.remove(os.path.join(self.working_dir,self.output_name + '.xls'))
            #     elif self.out_file_type == 'txt':
            #         os.remove(os.path.join(self.working_dir,self.output_name + '.txt'))
            #     os.remove(os.path.join(self.working_dir,'report.txt'))
            #     if not self.analyze_obj.started_as_feature_service:
            #         self.delete_unwanted_folder()
            #     # self.logger.log('Was able to remove intermittent file','INFO')
            #
            # except Exception as e:
            #     # Not the end of the world since the only thing
            #     # returned out to the user is the zip file
            #     self.logger.log('Could not remove intermittent files {}'.format(e),'INFO')

        else:
            # The whole above if statement is if the process encounters an error
            # and needs to exit gracefully. Everything below is the case where
            # no error was thrown and we are at the end of our script
            self.logger.log('No global errors were found','INFO')

            # Close file handles as appropriate
            if self.analyze_obj.in_file_type == 'csv':
                self.csv_file_handle.close()
            if self.analyze_obj.in_file_type == 'xlsx':
                # CSV files need to be closed, xls don't
                pass
            if self.analyze_obj.in_file_type == 'txt':
                self.f.close()
            if self.out_file_type == 'csv':
                self.csv_out_file.close()
                if self.feature_service_output:
                    # start_time = time.time()
                    self.convert_csv_to_feature_service()
                    self.logger.log('{0} Matched ({1}%)'.format(self.matched,round((float(self.matched)/float(self.row_count - self.header_rows_to_skip))*100,1)), 'USER')
                    self.logger.log('{0} Unmatched ({1}%)'.format(self.unmatched, round((float(self.unmatched)/float(self.row_count - self.header_rows_to_skip))*100,1)), 'USER')
                    self.logger.log('{0} Tied ({1}%)'.format(self.tied, round((float(self.tied)/float(self.row_count - self.header_rows_to_skip))*100,1)), 'USER')
                    # end_time = time.time()

                    stats = {}
                    stats["Matched"] = '{0} Matched ({1}%)'.format(self.matched,round((float(self.matched)/float(self.row_count - self.header_rows_to_skip))*100,1))
                    stats["Unmatched"] = '{0} Unmatched ({1}%)'.format(self.unmatched, round((float(self.unmatched)/float(self.row_count - self.header_rows_to_skip))*100,1))
                    stats["Tied"] = '{0} Tied ({1}%)'.format(self.tied, round((float(self.tied)/float(self.row_count - self.header_rows_to_skip))*100,1))
                    if self.cannot_parse_records:
                        stats["Dropped records"] = str(self.cannot_parse_records)
                    self.geocoding_stats = json.dumps(stats)
                    ### do not remove the csv for debugging purposes right now~
                    # try:
                    #     os.remove(os.path.join(self.working_dir, self.output_name + '.csv'))
                    # except Exception as e:
                    #     self.logger.log('Could not remove intermittent files {}'.format(e), 'INFO')
                    #     # we're done here
                    return
            if self.out_file_type == 'xls':
                #arcpy.AddMessage('successfully saved xls')
                self.xls_out_workbook.save(os.path.join(self.working_dir,self.output_name + '.xlsx'))
            if self.out_file_type == 'txt':
               self.txt_file_out.close()

            # This list is the number of records that did not make it
            # to the self.output
            leftover_keys = list(self.input_data.keys())

            # Benchmark how long the whole process took
            self.end_time = time.time()

            self.logger.log('About to write report','INFO')

            # Write the report file
            # try:
            #     with open(os.path.join(self.working_dir,'report.txt'), 'w') as report_file:
            #         report_file.write('Number of records not written is ' + \
            #                           str(len(leftover_keys)) + '\n')
            #
            #         report_file.write('The records unwritten were: {}'.format(leftover_keys) + '\n')
            #         time_taken = self.end_time-self.start_time
            #         report_file.write('The job took {}'.format(round(time_taken,1)) + \
            #                           ' seconds to finish' + '\n')
            #         report_file.write('This job finished at a rate of ' + \
            #                           str(int((self.row_count/time_taken)*3600)) + \
            #                           ' geocodes per hour.')
            #         report_file.write('\n' + 'The mapping used was {0}'.format(self.field_mapping))
            #         dropped_records_list = []
            #         while self.dropped_records_q.qsize() > 0:
            #             dropped_records_list.append(self.dropped_records_q.get())
            #         report_file.write('\n The records that could not be parsed are {}'\
            #                           .format(dropped_records_list))
            # except Exception as e:
            #     self.logger.log('Problem with writing report: {}'.format(e),'ERROR')

            # Write to self.output file
            if self.out_file_type == 'csv':
                self.output = os.path.join(self.working_dir,self.output_name + '.csv')
                # pass
            elif self.out_file_type == 'xls':
                self.output = os.path.join(self.working_dir,self.output_name + '.xlsx')

    def execute(self):
        ''' Calls different methods to perform batch geocoding
            First does some preprocessing
            Then launches 6 threads (4 rest, 1 input parsing, 1 self.output writing)
            Finally writes the report, puts the report and self.output file in the
            zip file and closes the zip file and deletes any intermediate files'''
        # self.preprocess()
        #arcpy.AddMessage('makes it to execute')

        true_step = 100 / self.num_of_batches

        rounded_step = int(true_step)

        if rounded_step == 0:
            step = 1
        else:
            step = rounded_step
        
        #arcpy.AddMessage("step value is: {}".format(step))

        arcpy.SetProgressor("step", "Beginning to batch geocode...", 0, 100, step)
        step_total = 0


        # Run the function that creates JSON (1 thread), the function that writes
        # to results.csv (1 thread), and the function that makes REST requests
        # (4 threads) The thread that creates JSON will finish first and
        # all threads will return when there the file is done being written.
        # We use a queue of max size 4 called num_of_REST_threads to put a lock
        # on the number of REST threads that are fired at once, since the service
        # can only handle 4 threads at any given time

        #arcpy.AddMessage('TYPE IS {}'.format(self.analyze_obj.in_file_type))

        #arcpy.AddMessage("allowing for {} threads".format(self.concurrent_batches+2))
        try:
            with concurrent.futures.ThreadPoolExecutor(max_workers=self.concurrent_batches+2) as executor:
                self.logger.log('Firing off createJSON and create_results',\
                                'INFO')
                if self.analyze_obj.in_file_type == 'table':
                    # do a different subroutine so that we can use cursors on the main thread,
                    # arcpy doesn't work on sub threads
                    #arcpy.AddMessage('INSIDE TABLE')
                    if self.analyze_obj.in_table.isLayer:
                        self.cursor = arcpy.da.SearchCursor(self.analyze_obj.in_table.layer,self.analyze_obj.list_of_fields)
                    else:
                        #arcpy.AddMessage("gets to adding the cursor")
                        self.cursor = arcpy.da.SearchCursor(self.analyze_obj.in_table_path,self.analyze_obj.list_of_fields)
                    # A TOTALLY DIFFERENT SUBROUTINE

                    if self.out_file_type == 'csv':
                        results_future = executor.submit(self.create_results_csv)

                    if self.out_file_type == 'xls':
                        results_future = executor.submit(self.create_results_xls)

                    if self.out_file_type == 'txt':
                        results_future = executor.submit(self.create_results_txt)

                    ### DEBUG
                    ### non-async version for debugging purposes

                    # self.recordset_builder_fs(0)
                    # self.num_of_REST_threads.put(1)
                    # self.geocode_batch_handler(0)
                    # self.create_results_csv()

                    ### end non-async version for debugging purposes
                    ### END DEBUG

                    try:
                        time.sleep(2)
                        for i in range(self.num_of_batches):                            
                            if arcpy.env.isCancelled:
                                self.is_cancelled = True
                            if not self.error_q.empty():
                                self.error_out = True
                                # Basically this queue should be empty at all times. If not,
                                # there is a fatal error.
                                # The error_q is what the threads use to communicate to each other
                                # If one thread encounters throws an error, it logs it in the
                                # error_q and all other threads check the error_q while waiting,
                                # and if it is empty, they continue, but if it is nonempty, they
                                # all return immidately
                                self.logger.log('Something went severely wrong','INFO')
                                executor.shutdown(wait=True)
                                break
                            if not self.processed_data_q.full():
                                self.recordset_builder_fs(i)
                            if not self.processed_data_q.empty():
                                # This queue makes sure that only 4 REST threads are fired at any given
                                # time, otherwise we wait on this line. It is basically a lock.
                                self.logger.log('trying to put into rest_threads_q','INFO')
                                self.num_of_REST_threads.put(1)
                                self.logger.log('was able to put into rest_threads_q','INFO')
                                REST_future = executor.submit(self.geocode_batch_handler,i)
                                # self.logger.log(str(round((float(i)/float(self.num_of_batches))*\
                                #     100,1)) + ' percent done with' +\
                                #     ' the pending job','USER')
                                step_total = step_total + true_step
                                if step_total >= 100:
                                    step_total = 99
                                arcpy.SetProgressorLabel(str(int(step_total)) + ' percent done with the pending job')
                                arcpy.SetProgressorPosition(int(step_total))
                                    # a = REST_future.result(timeout=30)
                    except Exception as e:
                        #arcpy.AddMessage("exception is {}".format(e))
                        self.logger.log('problem in firing off threads','ERROR')


                    if self.error_q.empty():
                        self.logger.log('no errors were found in the threads','INFO')
                        executor.shutdown(wait=True)

                else:
                    #arcpy.AddMessage('firing off create_json')

                    #DEBUG
                    #self.create_JSON()
                    #END DEBUG

                    JSON_future = executor.submit(self.create_JSON)

                    if self.out_file_type == 'csv':
                        results_future = executor.submit(self.create_results_csv)

                    if self.out_file_type == 'xls':
                        results_future = executor.submit(self.create_results_xls)

                    if self.out_file_type == 'txt':
                        results_future = executor.submit(self.create_results_txt)

                    try:
                        time.sleep(2)
                        for i in range(self.num_of_batches):
                            # This queue makes sure that only 4 REST threads are fired at any given
                            # time, otherwise we wait on this line. It is basically a lock.
                            self.logger.log('trying to put into rest_threads_q','INFO')
                            self.num_of_REST_threads.put(1)
                            self.logger.log('was able to put into rest_threads_q','INFO')
                            if arcpy.env.isCancelled:
                                self.is_cancelled = True
                            if not self.error_q.empty():
                                self.error_out = True
                                # Basically this queue should be empty at all times. If not,
                                # there is a fatal error.
                                # The error_q is what the threads use to communicate to each other
                                # If one thread encounters throws an error, it logs it in the
                                # error_q and all other threads check the error_q while waiting,
                                # and if it is empty, they continue, but if it is nonempty, they
                                # all return immidately
                                self.logger.log('Something went severely wrong','INFO')
                                executor.shutdown(wait=True)
                                break
                            self.logger.log('firing off a thread','INFO')
                            #arcpy.AddMessage("JSON_future is running {}".format(JSON_future.running()))
                            #arcpy.AddMessage("Size of processed_data_q is {}".format(self.processed_data_q.qsize()))
                            #arcpy.AddMessage("SHOULD BE FALSE. Is the processed_data_q_empty?? {}".format(self.processed_data_q.empty()))
                            REST_future = executor.submit(self.geocode_batch_handler,i)

                            # DEBUG
                            #self.geocode_batch_handler(i)
                            # DEBUG END

                            # self.logger.log(str(round((float(i)/float(self.num_of_batches))*\
                            #     100,1)) + ' percent done with' +\
                            #     ' the pending job','USER')
                            step_total = step_total + true_step
                            if step_total >= 100:
                                step_total = 99
                            arcpy.SetProgressorLabel(str(int(step_total))  + ' percent done with the pending job')
                            arcpy.SetProgressorPosition(int(step_total))
                    except Exception as e:
                        #arcpy.AddMessage(e)
                        self.logger.log('problem in firing off threads','ERROR')


                    if self.error_q.empty():
                        self.logger.log('no errors were found in the threads','INFO')
                        executor.shutdown(wait=True)

                    #DEBUG
                    #self.create_results_csv()
                    #END DEBUG
                    
                    # Does this do anything?? Not sure...
                    # if arcpy.env.isCancelled:
                    #     self.is_cancelled = True

        except Exception as e:
            self.logger.log('Error in thread_pool_executor {}'.format(e),'ERROR')


        self.logger.log('All threads finished, finishing up job with 1 thread','INFO')