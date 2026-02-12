import datetime
import logging
import os
import pathlib
import shutil
import tempfile
import uuid
import copy
import warnings
from pathlib import Path
from typing import Optional, Union, TYPE_CHECKING, Set, Any

import arcpy
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.cell_range import CellRange, MultiCellRange
from openpyxl.worksheet.datavalidation import DataValidation, DataValidationList
from openpyxl.worksheet.worksheet import Worksheet

from . import helper
from . import workbook_styles
from .helper import MsgType
from .settings import *

if TYPE_CHECKING:
    from .workspace import TableMatch
    from .helper import TableWrapper, FieldMatch, FieldWrapper

logger = logging.getLogger(__name__)

__all__ = ["WorkbookGenerator", "WorkbookLoader", "WorkbookUpdater"]

# date to store on the workbook for last modified.
MAGIC_DATE = datetime.datetime(year=2000, month=1, day=1).timestamp()

N_ = helper.get_message


def is_query_valid(table: str, query: str) -> bool:
    """Validates that query can successfully be applied to table."""
    if helper.is_select_all(query):
        return True
    try:
        # Add 1=2 to ensure no features are returned to optimize the call for services
        arcpy.MakeTableView_management(
            in_table=table, out_view=f"f_{uuid.uuid4().hex}", where_clause=f"{query} AND 1=2"
        )
        return True
    except arcpy.ExecuteError as e:
        msg = "".join(e.args)
        if any(f"ERROR {i:06}" in msg for i in [358, 840]):
            # ERROR 000358: Invalid expression.
            # ERROR 000840: The value is not a SQL Expression.
            return False
        else:
            raise e


def _is_service_invalid(describe) -> bool:
    """If describe points to a service, validate it. Returns True if invalid"""
    service: str = describe.path.lower()
    if not service.startswith("http"):
        return False

    if not service.endswith("FeatureServer".lower()):
        return True

    return False


def _is_spatial_reference_problem(describe) -> bool:
    """If Feature Dataset has differing SR, could cause problems."""
    dirname = os.path.dirname(describe.catalogPath)
    desc = helper.describe_object(dirname)
    if hasattr(desc, "datasetType") and desc.dataElementType == "DEFeatureDataset":
        return desc.spatialReference.name != describe.spatialReference.name
    return False


def is_mapping_valid(source, target, msg: str = "") -> bool:
    """Warns if source cannot be loaded into target"""

    white_list = {"DEDbaseTable", "DETable", "DEShapeFile", "DEFeatureClass", "DETextFile"}
    for data_elem in (source.dataElementType, target.dataElementType):
        if data_elem not in white_list:
            arcpy.AddIDMessage(MsgType.WRN, 3798, msg, data_elem)  # "%2" is not a supported dataType.
            return False

    # Valid source and target combinations
    #
    # +--------------+--------------+-----------+
    # |    Source    |    Target    | Supported |
    # +--------------+--------------+-----------+
    # | Table        | Table        |     X     |
    # +--------------+--------------+-----------+
    # | Table        | FeatureClass |     X     |
    # +--------------+--------------+-----------+
    # | FeatureClass | FeatureClass |     X     |
    # +--------------+--------------+-----------+
    # | FeatureClass | Table        |     X     |
    # +--------------+--------------+-----------+

    for obj in (source, target):
        if _is_service_invalid(obj):
            arcpy.AddIDMessage(MsgType.WRN, 3801, msg)  # Service must be a FeatureServer service.
            return False

    if target.datasetType == "FeatureClass" and _is_spatial_reference_problem(target):
        # This could cause problems but loads can still succeed
        logger.debug(f"{target.catalogPath} has a different Spatial Reference than its Feature Dataset.")

    return True


def get_user_scripts(dlw: pathlib.Path, skip_file: str = "") -> list:
    """Loads system and user defined scripts to be used inside code block"""

    # rglob on a path that doesn't exist returns an empty iterator, so this is safe
    scripts: pathlib.Path = dlw / SCRIPTS_FOLDER
    return [s.read_text(encoding="utf-8") for s in scripts.rglob("*.py") if s.stem.lower() != skip_file.lower()]


class WorkbookWrapper:
    def __init__(self, workbook: str, on_error: str, read_only: bool = True):
        """
        Thin utility wrapper for openpyxl.Workbook

        Args:
            workbook: str - The workbook on disk.
            on_error: behavior when an error occurs
                raise - log error message and raise SystemExit
                warn - log warning message and return None
            read_only: Load workbook as read only. Default is True.
        """
        self.catalog_path: str = os.path.normpath(workbook)
        self.on_error = on_error
        self.wb = self._load_workbook(read_only=read_only)

    def _load_workbook(self, read_only: bool = True) -> openpyxl.Workbook:
        """wrapper for openpyxl.load_workbook with validation"""
        from zipfile import BadZipFile
        import io

        try:
            # In case openpyxl does not get an opportunity to safely close later (eg hard error, user cancels),
            # we do not want to place a lock on the Workbook. By reading into memory, we can avoid this entirely.
            # https://stackoverflow.com/a/43789779/10231563
            with open(self.catalog_path, "rb") as f:
                in_mem_file = io.BytesIO(f.read())
            if not read_only:
                # Suppress Openpyxl UserWarning about Data Validation when opening with read/write
                # https://docs.python.org/3/library/warnings.html#temporarily-suppressing-warnings
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore", category=UserWarning)
                    return openpyxl.load_workbook(filename=in_mem_file)
            return openpyxl.load_workbook(filename=in_mem_file, read_only=read_only, data_only=True)
        except (BadZipFile, KeyError):
            # Invalid Excel file
            self._on_error(3808, self.catalog_path)  # is not a valid Excel file.
        except Exception as e:
            # Something bad happened...
            raise e

    def _on_error(self, mess_id, arg1=None, arg2=None):
        if self.on_error == "raise":
            arcpy.AddIDMessage(MsgType.ERR, mess_id, arg1, arg2)
            raise SystemExit
        else:
            arcpy.AddIDMessage(MsgType.WRN, mess_id, arg1, arg2)

    @helper.lru_cache()
    def _sheets_lookup(self):
        """cache lookup of sheets with lower case name"""
        return {s.lower(): s for s in self.wb.sheetnames}

    def _proper_name(self, sheet_name: str) -> Optional[str]:
        """openpyxl does not perform case insensitive lookups, so we need to get the proper name ourselves"""
        return self._sheets_lookup().get(sheet_name.lower(), None)

    def get_sheet(self, sheet_name: str, raise_error: bool = True) -> Optional[Worksheet]:
        """Gets sheet from workbook if it exists"""
        proper_name = self._proper_name(sheet_name)
        if proper_name is None:
            if not raise_error:
                return
            return self._on_error(3809, sheet_name, self.catalog_path)  # Sheet "%1" does not exist in %2

        return self.wb[proper_name]

    def remove_sheet(self, sheet_name: str):
        """Removes sheet from notebook regardless of case"""
        proper_name = self._proper_name(sheet_name)
        if proper_name is None:
            self._on_error(3811, sheet_name, self.catalog_path)  # Sheet "%1" does not exist in %2
        else:
            self.wb.remove(self.wb[proper_name])

    def get_index(self, sheet_name: str) -> Optional[int]:
        proper_name = self._proper_name(sheet_name)
        if proper_name is None:
            return None
        else:
            return self.wb.get_index(self.wb[proper_name])

    @staticmethod
    def _get_header(sheet: Worksheet) -> list[str]:
        """Gets values from first row in sheet, casting to lower case string"""
        return [str(val).lower() for row in sheet.iter_rows(max_row=1, values_only=True) for val in row]

    def is_header_missing(self, sheet: Worksheet, expected: list[str]) -> bool:
        """Validates that the expected header is present in sheet"""
        actual = set(self._get_header(sheet))

        missing = [column for column in expected if column.lower() not in actual]

        if missing:
            # Column names %1 are missing from sheet %2
            self._on_error(3810, missing, f'"{sheet.title}", {self.catalog_path}')
            return True
        return False

    def sheet_to_df(
        self, sheet: Union[str, Worksheet], columns: list[str], optional_columns: list[str] = None
    ) -> pd.DataFrame:
        """Converts columns from sheet into a DataFrame"""
        import warnings
        import numpy as np

        if isinstance(sheet, str):
            sheet = self.get_sheet(sheet)

        if optional_columns is None:
            optional_columns = []

        header = self._get_header(sheet)

        # Certain spreadsheets with conditional formatting, data validation, etc are not supported by openpyxl
        # We want to swallow this warning when reading data.
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            df = pd.DataFrame.from_records(data=sheet.iter_rows(min_row=2, values_only=True), columns=header)

        # In case there are duplicate columns with the same name, we choose the first.
        # Reset names back to expected case.
        index = [header.index(col.lower()) for col in columns + optional_columns if col.lower() in header]
        df = df.iloc[:, index]

        for col in optional_columns:
            if col.lower() not in df:
                df[col] = None
        df.columns = columns + optional_columns

        # Fill all NaN with None
        # df.where(pd.notnull(df), None) does not work at pandas 1.3 - https://github.com/pandas-dev/pandas/issues/42423
        return df.replace(np.nan, None)


class WorkbookGenerator:
    HEADER = [
        SOURCE,
        DEFINITION_QUERY,
        TARGET,
        DELETE_QUERY,
        MAPPING_WORKBOOK,
        ENABLED,
        ATTACHMENTS,
        GLOBALIDS,
        TRANSFORMATIONS,
    ]

    def __init__(self, source_target: list["TableMatch"], output_folder: str, compute_stats: bool = False):
        self.mapping = source_target
        self.folder = pathlib.Path(arcpy.CreateUniqueName(WORKSPACE_FOLDER, output_folder))
        self.compute_stats = compute_stats

        self.book: openpyxl.Workbook = None

        # Running list of domain hyperlinks and field sheets
        self.domain_hyperlink: dict[tuple, str] = {}
        self.field_sheets: dict[str, Worksheet] = {}

    def create_mapping_sheet(self, match: "TableMatch"):
        """Create source-target mapping sheet"""
        source_table = match.source.describe

        header = [TARGET_FIELD, FIELD_TYPE, EXPRESSION, LOOKUP_SHEET, LOOKUP_KEYS, LOOKUP_VALUE, LOOKUP_DEFAULT]
        sheet = self.book.create_sheet(MAPPING, index=1)
        sheet.sheet_properties.tabColor = workbook_styles.TAB_GREEN
        sheet.append(header)

        target_fields = match.target.fields
        skip_fields = target_fields.default_fields()

        # I cannot explain why, but this needs to happen as late as possible...
        match.match_fields()

        # Add shape row first
        is_shape_row = self._append_shape_row(sheet, match)

        for field in match.target.fields:
            # Non-editable fields cannot be loaded into.
            if not field.editable or field.name.lower() in skip_fields:
                continue

            m = field.match
            if m:
                row = [m.target.name, helper.property_to_pretty(m.target.type), m.expression]
                if m.lookup:
                    self._add_domain_mapping(m, target_workspace=match.target.workspace)
                    row.extend([m.source.name, m.source.name, m.target.domain.name])

            else:
                row = [field.name, helper.property_to_pretty(field.type)]
            sheet.append(row)

        # Hyperlink any lookup sheets
        lookup_col = header.index(LOOKUP_SHEET) + 1
        for row in sheet.iter_rows(min_row=2, min_col=lookup_col, max_col=lookup_col):
            for cell in row:
                if cell.value is not None:
                    cell.hyperlink = f"#{cell.value}!A1"

        scripts_location = pathlib.Path(__file__).parents[1] / "dataloading"
        self._add_drop_down(source_table, sheet, scripts_location, "C3" if is_shape_row else "C2")

    @staticmethod
    def _append_shape_row(sheet: Worksheet, match: "TableMatch") -> bool:
        """Add shape row if target is not a table"""
        if match.target.shape_type == "Table":
            return False
        shape_field = match.target.fields.get(match.target.fields.shape)
        m = shape_field.match
        if match.source.shape_type == match.target.shape_type:
            row = [m.target.name, helper.property_to_pretty(m.target.type), m.expression]
        else:
            row = [m.target.name, helper.property_to_pretty(m.target.type), helper.get_shape_expression(match)]
        sheet.append(row)
        return True

    def create_info_sheet(self, subtype_code: Optional[int] = None):
        """Create sheet with instructions and subtype property"""

        sheet = self.book.create_sheet(INFO, index=0)
        sheet.sheet_properties.tabColor = workbook_styles.TAB_ORANGE
        sheet.append([MAPPING_PROPERTY, VALUE])
        sheet.append([SOURCE_ST, subtype_code])
        subtype_tbl = self._add_table("A1:B2", "SubtypeTbl", sheet)
        helper.filter_button_off(subtype_tbl)

        self._create_quick_links(sheet)
        self._create_legend(sheet)

    def _create_quick_links(self, sheet: Worksheet):
        """Generate quick links table"""
        start = idx = sheet.max_row + 1
        data = [
            (N_(280105), None),  # Quick Links
            (HELP_TEXT, HELP),
            (GLOBAL_LOOKUP_WORKBOOK, f"../{GLOBAL_LOOKUP}/{GLOBAL_LOOKUP}.xlsx"),
            (f"{SCRIPTS_FOLDER} Folder", f"../../{SCRIPTS_FOLDER}"),
        ]
        for domain_sheet in sorted(set(map(lambda x: x.split("#")[0], self.domain_hyperlink.values()))):
            data.append((f"{domain_sheet.split('Domains/')[-1].removesuffix('.xlsx')} Domains", domain_sheet))

        for row in data:
            sheet[f"A{idx}"].value, sheet[f"A{idx}"].hyperlink = row
            idx += 1

        quick = self._add_table(f"A{start}:A{idx - 1}", "QuickLinks", sheet)
        helper.filter_button_off(quick)

    def _create_legend(self, sheet: Worksheet):
        """Generate validation legend"""
        start = idx = sheet.max_row + 1

        field_name = f"!{N_(220308).replace(' ', '')}!"
        text = N_(220305)
        link3811 = "https://pro.arcgis.com/en/pro-app/latest/tool-reference/tool-errors-and-warnings/001001-010000/tool-errors-and-warnings-03801-03825-003811.htm"
        link3813 = "https://pro.arcgis.com/en/pro-app/latest/tool-reference/tool-errors-and-warnings/001001-010000/tool-errors-and-warnings-03801-03825-003813.htm"
        link3814 = "https://pro.arcgis.com/en/pro-app/latest/tool-reference/tool-errors-and-warnings/001001-010000/tool-errors-and-warnings-03801-03825-003814.htm"

        legend_data = [
            [N_(280103), None],  # Validation Legend
            [N_(280104), N_(84919)],  # Formatting, Description
            [field_name, N_(280101)],  # !FieldName!, Field Mapping Error
            [field_name, N_(280102)],  # !FieldName!, Field Mapping Warning
            # hardcoded strings because these are abbreviation of actual message, follow hyperlink to get translation
            [text, "Error 003811: Only one of expression or sheet must be specified"],
            [text, "Error 003813: LookupKeys and LookupValue are required when LookupSheet is specified"],
            [text, "Error 003814: Target field cannot be duplicated"],
        ]

        for row in legend_data:
            sheet.append(row)
            idx += 1

        # add hyperlinks
        link: str
        for i, link in enumerate((link3814, link3813, link3811), 1):
            sheet.cell(idx - i, 2).hyperlink = link

        # add formatting
        wbs = workbook_styles
        style: wbs.FormatOptions
        for i, style in enumerate(
            (
                wbs.DUPLICATE_TARGET_FIELD,
                wbs.LOOKUP_REQUIRED,
                wbs.EXPRESSION_LOOKUP,
                wbs.FIELD_WARNING,
                wbs.FIELD_ERROR,
            ),
            1,
        ):
            style.apply(sheet.cell(idx - i, 1))

        legend = self._add_table(f"A{start + 1}:B{idx - 1}", "ValidationLegend", sheet)
        helper.filter_button_off(legend)

    @staticmethod
    def _get_script_functions(scripts_location: pathlib.Path):
        """convert user scripts to function call examples
        Ex. concatenate(!FIELD1!,!FIELD2!,delimiter=" ")
        """
        scripts = get_user_scripts(scripts_location, skip_file="shape_operations")
        scripts_string = "\n\n".join(scripts)
        cf = helper.ConvertFunctions(scripts_string)
        for func in cf.function_calls():
            yield func

    def _add_drop_down(
        self, source_desc: arcpy.Describe, mapping_sheet: Worksheet, scripts_location, start_cell: str = "C2"
    ):
        """Add drop down validation to mapping sheet"""
        sheet: Worksheet = self.book.create_sheet(HIDDEN_SHEET, index=0)
        sheet.sheet_state = "hidden"

        # Create a lookup column with the !fieldName! values and script functions.
        rows = [f"!{field.name}!" for field in sorted(source_desc.fields, key=lambda x: x.name.lower())]
        rows.extend(self._get_script_functions(scripts_location))
        rows.insert(0, "FieldDropDown")
        for row in rows:
            sheet.append([row])

        # wipe out previous validation before adding
        mapping_sheet.data_validations = DataValidationList()

        valid = DataValidation(
            type="list",
            formula1=f"{HIDDEN_SHEET!r}!$A$2:$A${len(rows) + 1}",
            allow_blank=True,
            showErrorMessage=False,
        )
        mapping_sheet.add_data_validation(valid)
        valid.add(f"{start_cell}:C{MAX_ROWS}")

        self._create_field_type_combinations(sheet)

    @staticmethod
    def _create_field_type_combinations(sheet: Worksheet):
        """Creates table for source-target field type combinations"""
        from openpyxl.worksheet.table import Table
        from gdbschema.constants import esriFieldType

        header = (
            esriFieldType.esriFieldTypeSmallInteger,
            esriFieldType.esriFieldTypeInteger,
            esriFieldType.esriFieldTypeBigInteger,
            esriFieldType.esriFieldTypeSingle,
            esriFieldType.esriFieldTypeDouble,
            esriFieldType.esriFieldTypeString,
            esriFieldType.esriFieldTypeDate,
            esriFieldType.esriFieldTypeDateOnly,
            esriFieldType.esriFieldTypeTimeOnly,
            esriFieldType.esriFieldTypeTimestampOffset,
            esriFieldType.esriFieldTypeOID,
            esriFieldType.esriFieldTypeGeometry,
            esriFieldType.esriFieldTypeGUID,
            esriFieldType.esriFieldTypeGlobalID,
            esriFieldType.esriFieldTypeBlob,
            esriFieldType.esriFieldTypeRaster,
            esriFieldType.esriFieldTypeXML,
        )

        # Rows are source and columns are target.
        # 0 - error, 1 - valid, 2 - warn
        array = (
            (1, 1, 1, 1, 1, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0),
            (2, 1, 1, 1, 1, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0),
            (2, 2, 1, 2, 1, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0),
            (2, 2, 2, 1, 1, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0),
            (2, 2, 2, 2, 1, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0),
            (2, 2, 2, 2, 2, 1, 2, 2, 2, 2, 0, 0, 2, 2, 0, 0, 0),
            (0, 0, 0, 0, 0, 1, 1, 2, 2, 1, 0, 0, 0, 0, 0, 0, 0),
            (0, 0, 0, 0, 0, 1, 1, 1, 0, 1, 0, 0, 0, 0, 0, 0, 0),
            (0, 0, 0, 0, 0, 1, 0, 0, 1, 2, 0, 0, 0, 0, 0, 0, 0),
            (0, 0, 0, 0, 0, 1, 2, 2, 2, 1, 0, 0, 0, 0, 0, 0, 0),
            (2, 1, 1, 1, 1, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0),
            (0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0),
            (0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 1, 1, 0, 0, 0),
            (0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 1, 1, 0, 0, 0),
            (0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0),
            (0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0),
            (0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1),
        )

        start_row, start_col = 1, 3
        end_row, end_col = start_row + len(header), start_col + len(header)

        for i, field in enumerate(header, 1):
            sheet.cell(start_row + i, start_col, field.value)
            sheet.cell(1, start_col + i, field.value)

        for i, field_row in enumerate(array, start_row + 1):
            for j, num in enumerate(field_row, start_col + 1):
                sheet.cell(i, j, num)

        sheet.add_table(
            Table(
                displayName="FieldType",
                ref=f"{get_column_letter(start_col + 1)}{start_row}:{get_column_letter(end_col)}{end_row}",
            )
        )

    def _add_domain_mapping(self, match: "FieldMatch", target_workspace: str):
        """Adds a source-target domain mapping to the field sheet"""

        sheet = self.field_sheets[match.source.name.lower()]

        # Target mappings can already exist, so we need to just add new columns at the end.
        max_col = sheet.max_column
        domain_name = match.target.domain.name

        # If we've already did this mapping (eg the same source field going to multiple targets with the same domain)
        # we do not need to repeat.
        if domain_name in {c.value for c in next(sheet.iter_rows(max_row=1, min_col=3))}:
            return

        sheet.cell(row=1, column=max_col + 1, value=domain_name)
        sheet.cell(row=1, column=max_col + 2, value=f"{domain_name} (Description)")

        link = self.domain_hyperlink.get((target_workspace.lower(), domain_name.lower()), None)
        if link:
            sheet.cell(1, max_col + 2).hyperlink = link

        lookup = match.code_lookup()
        for source_code, *_, target_code, target_desc in sheet.iter_rows(min_row=2):
            target_code.value, target_desc.value = lookup.get(source_code.value, (None, None))

    @staticmethod
    def _add_table(ref: str, table_name: str, sheet: Worksheet):
        """Add an Excel table to sheet"""
        from openpyxl.worksheet.table import Table, TableStyleInfo

        title: str = sheet.title.strip("-")
        if title.isdigit():  # Table names cannot start with a number
            title = f"T{title}"

        table = Table(
            displayName=f"{title}_{table_name}",
            ref=ref,
            tableStyleInfo=TableStyleInfo(name="TableStyleLight1", showRowStripes=True),
        )
        sheet.add_table(table)
        sheet.append([None])

        return table

    def _rows_to_table(self, rows: list, table_name: str, sheet: Worksheet):
        """Writes rows to Excel table"""
        sheet.append([table_name])

        start = sheet.max_row + 1  # Offset for header.
        for row in rows:
            sheet.append(row)
        end = start + 1 if sheet.max_row == start else sheet.max_row  # empty table needs at least one row

        self._add_table(f"A{start}:{get_column_letter(len(rows[0]))}{end}", table_name, sheet)

    def _create_table_properties(self, table, sheet: Worksheet):
        """Writes table properties to Excel Sheet"""

        data = [("Key", "Value")]

        data.extend(
            [
                ("Name", table.name),
                ("Alias", getattr(table, "aliasName", "")),
                ("Dataset Type", table.datasetType),
                ("Subtype Field", getattr(table, "subtypeFieldName", "")),
            ]
        )

        if hasattr(table, "shapeType"):  # Spatial
            data.extend(
                [
                    ("Shape Type", table.shapeType),
                    ("Feature Type", table.featureType),
                    ("Spatial Reference", table.spatialReference.name),
                ]
            )

        self._rows_to_table(data, "Properties", sheet)

    def _create_table_subtypes(self, table, sheet: Worksheet):
        """Writes table subtypes to Excel Sheet"""

        subtypes = helper.get_subtypes(table.catalogPath)
        if not subtypes:
            return

        subtype_rows = [["Name", "Code"]]
        info_rows = [["Subtype Name", "Subtype Code", "Field Name", "Default Value", "Domain Name"]]
        for subtype_code, info in subtypes.items():
            subtype_rows.append([subtype_code, subtype_name := info["Name"]])

            for field, (default, domain) in info["FieldValues"].items():
                if default is None and domain is None:
                    continue
                info_rows.append([subtype_name, subtype_code, field, default, getattr(domain, "name", None)])

        self._rows_to_table(subtype_rows, "Subtypes", sheet)
        self._rows_to_table(info_rows, "SubtypeFieldInfos", sheet)

    def _create_schema_sheet(self, table, sheet_name: str, calc_stats: bool = True, index: int = None):
        """Create sheet detailing the schema (fields and feature counts)"""

        sheet = self.book.create_sheet(sheet_name, index)
        sheet.sheet_properties.tabColor = workbook_styles.TAB_YELLOW

        self._create_table_properties(table, sheet)
        self._create_table_fields(table, sheet, calc_stats=calc_stats & self.compute_stats)
        self._create_table_subtypes(table, sheet)

    def _create_table_fields(self, table, sheet, calc_stats: bool):
        """Writes field properties to Excel Sheet"""
        header = SCHEMA_HEADER.copy()
        if calc_stats:
            header.extend([COUNT, FILL_FACTOR])

        data = [header]

        stats = {}
        row_count = helper.get_count(table.catalogPath) if calc_stats else 0
        if calc_stats and row_count:
            # These fields are not valid inputs to Summary Stats
            # TODO: we can optimize this further by skipping non-nullable fields and then filling based on row count
            fields = [
                f.name for f in table.fields if f.type not in ("Guid", "GlobalID", "Blob", "XML", "Raster", "Geometry")
            ]
            if fields:
                logger.debug(f"\tCalculating statistics on {table.name}")
                try:
                    stats_table = arcpy.Statistics_analysis(
                        in_table=table.catalogPath,
                        out_table=f"memory/t_{uuid.uuid4().hex}",
                        statistics_fields=[(f, "COUNT") for f in fields],
                    )[0]

                    with arcpy.da.SearchCursor(stats_table, "*") as cursor:
                        stats = dict(zip(map(str.lower, cursor.fields), next(cursor)))
                    del cursor
                except arcpy.ExecuteError as e:
                    # Raise specific warning for no oids (csv)
                    if error_message := helper.get_tool_message(339):
                        # Failed to calculate statistics :
                        logger.warning(N_(86120))
                        # Input %s does not have OIDs
                        logger.warning(error_message)
                    else:
                        logger.warning("".join(e.args))

        field: arcpy.Field
        for field in table.fields:
            field_row = [
                field.name,
                field.aliasName,
                helper.property_to_pretty(field.type),
                field.length if field.type == "String" else None,
                field.domain or None,
                field.defaultValue,
                field.isNullable,
                field.editable,
            ]

            if calc_stats:
                count = stats.get(f"count_{field.name.lower()}") or 0

                field_row.extend([count, (count / row_count) if row_count else 0])

            data.append(field_row)

        self._rows_to_table(data, "Fields", sheet)

    def _create_field_sheets(self, table: "TableWrapper"):
        """Creates field sheet for fields with CodedValue domains"""

        # Reset field sheets from previous runs.
        self.field_sheets = {}

        # If subtypes actually do exist we need to generate a field sheet. This occurs when user unchecks
        # "Create Matches by Subtype".
        st_name = getattr(table.describe, "subtypeFieldName", "").lower()
        create_subtype_field_sheet = (table.subtype_code is None) and bool(st_name)

        skip_fields = table.fields.default_fields()
        for field in table.fields:
            if field.name.lower() in skip_fields:
                continue

            if field.name.lower() == st_name and create_subtype_field_sheet:
                # handle subtype fields if "Create Matches by Subtypes" unchecked
                sheet = self.book.create_sheet(self._create_valid_sheet_name(self.book, field.name))
                self.field_sheets[field.name.lower()] = sheet
                sheet.append([field.name, f"SubtypeName"])
                sheet.sheet_properties.tabColor = workbook_styles.TAB_BLUE
                subtypes = {(code, info["Name"]) for code, info in helper.get_subtypes(table.path).items()}
                for code, name in sorted(subtypes):
                    sheet.append([code, name])
                continue

            if field.domain is None or not field.domain.codedValues:
                continue

            self._create_field_sheet(field, table, create_subtype_field_sheet)

    def _create_field_sheet(self, field: "FieldWrapper", table: "TableWrapper", check_subtype_domains: bool = False):
        """Create new field sheet for Source field."""
        # For lookups to work, the name of the column containing the codes must be the field name.
        sheet = self.book.create_sheet(self._create_valid_sheet_name(self.book, field.name))
        self.field_sheets[field.name.lower()] = sheet
        sheet.append([field.name, f"{field.domain.name} (Description)"])
        sheet.sheet_properties.tabColor = workbook_styles.TAB_BLUE

        link = self.domain_hyperlink.get((table.workspace.lower(), field.domain.name.lower()), None)
        if link:
            sheet.cell(1, 2).hyperlink = link

        domains = []
        if check_subtype_domains:
            # get domains by subtype, which may be different from root, so we can list the values in the Field sheet.
            # This is for reference purposes and does not affect any automatching.
            max_col = sheet.max_column
            unique_domains = table.get_all_domains(field.name)
            for domain, subtypes in unique_domains.domains_and_subtypes:
                if not domain:
                    continue
                # skip root domain
                if domain.name.casefold() == field.domain.name.casefold():
                    continue
                sheet.cell(row=1, column=max_col + 1, value=f"{domain.name} ({','.join(map(str, subtypes))})")
                link = self.domain_hyperlink.get((table.workspace.casefold(), domain.name.casefold()), None)
                if link:
                    sheet.cell(1, max_col + 1).hyperlink = link
                domains.append(domain)
                max_col += 1

        if domains:
            sheet.cell(1, 2).value = f"{field.domain.name} (Description) (Root)"

        for code, desc in field.domain.codedValues.items():
            row = [code, desc]
            for domain in domains:
                row.extend([domain.codedValues.get(code, None)])
            sheet.append(row)

    def create_sheets(self, match: "TableMatch"):
        """Creates the source and target schema / mapping sheets"""
        source = match.source.describe
        target = match.target.describe

        self._create_schema_sheet(target, TARGET_SCHEMA, calc_stats=False)
        self._create_schema_sheet(source, SOURCE_SCHEMA)
        self._create_field_sheets(match.source)

    @staticmethod
    def _format_header(sheets: list[Worksheet]):
        """Formats the header across multiple sheets"""
        font = Font(bold=True)
        alignment = Alignment(horizontal="center")

        for sheet in sheets:
            # Bold and center the header
            for row in sheet.iter_rows(max_row=1):
                for cell in row:
                    cell.font = font
                    cell.alignment = alignment

            # Freeze first pane if we haven't set it before.
            if sheet.freeze_panes is None:
                sheet.freeze_panes = "A2"

    @staticmethod
    def _set_column_width(sheets: list[Worksheet]):
        """Sets column widths based on cell contents"""
        from gdbschema.writer.excel_helper import cell_width

        for sheet in sheets:
            # Row number of the table headers.
            header_rows = {range_boundaries(table.ref)[1] for table in sheet.tables.values()}

            width_array = [0] * sheet.max_column
            for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
                if i in header_rows:  # Add extra spacing to the values to account for filter button width.
                    row = [f"{r}..." if r is not None else None for r in row]

                for j, value in enumerate(row):
                    if value is None:
                        continue
                    width_array[j] = max(width_array[j], cell_width(str(value)))

            for i, width in enumerate(width_array, 1):
                sheet.column_dimensions[get_column_letter(i)].width = width

    @staticmethod
    def _write_metadata(book: openpyxl.Workbook, **kwargs):
        """Writes metadata to the workbook"""

        props = book.properties
        props.creator = ""

        info = arcpy.GetInstallInfo()
        props.description = f'ArcGIS Pro {info["Version"]}.{info["BuildNumber"]}'

        for k, v in kwargs.items():
            if hasattr(props, k):
                setattr(props, k, v)

    def _hyperlink(self, gdb: str, rows: tuple[tuple, ...], column: str):
        i = [c.value for c in rows[0]].index(column)  # Position of domain column
        for row in rows:
            row[i].hyperlink = self.domain_hyperlink.get((gdb, (row[i].value or "").lower()), None)

    def _hyperlink_domains(self, book: openpyxl.Workbook, source: str, target: str):
        """Hyperlinks domains at field/subtype level to domain workbook"""

        source_workspace = helper.get_workspace(source).lower()
        target_workspace = helper.get_workspace(target).lower()

        for sheet_name, workspace in [(SOURCE_SCHEMA, source_workspace), (TARGET_SCHEMA, target_workspace)]:
            if sheet_name not in book:
                continue
            sheet: Worksheet = book[sheet_name]

            if fields := sheet.tables.get(f"{sheet.title.strip('-')}_Fields"):
                self._hyperlink(workspace, sheet[fields.ref], "Domain")

            if subtypes := sheet.tables.get(f"{sheet.title.strip('-')}_SubtypeFieldInfos"):
                self._hyperlink(workspace, sheet[subtypes.ref], "Domain Name")

    @staticmethod
    def _format_range(start: int) -> str:
        letter = get_column_letter(start)
        return f"{letter}2:{letter}{MAX_ROWS}"

    def _restrict_formatting(self, key: int, others: list[int]) -> tuple:
        """creates a conditional formatting formula"""

        # By locking the column here, applying the formatting to multiple ranges will properly shift.
        source = f"${get_column_letter(key)}2"
        other_columns = ", ".join(f"NOT( ISBLANK(${get_column_letter(x)}2) )" for x in others)

        formula = f"""
=AND(
    NOT( ISBLANK({source}) ),
    OR( 
        {other_columns}
    )    
)
        """

        # The conditional formatting needs to be applied to all the other columns.
        multi_range = MultiCellRange()
        for idx in others:
            multi_range.add(CellRange(self._format_range(idx)))

        return str(multi_range), formula

    def _mapping_sheet_conditional_formatting(self, sheet: Worksheet):
        """Adds conditional formatting for data validation"""
        from openpyxl.formatting.rule import Rule
        from openpyxl.formatting.formatting import ConditionalFormattingList

        # Clear cond formatting if exists
        if bool(sheet.conditional_formatting._cf_rules):
            sheet.conditional_formatting = ConditionalFormattingList()

        # Index positions
        header = [c.value for c in sheet[1]]
        target_idx = header.index(TARGET_FIELD) + 1
        field_type_idx = header.index(FIELD_TYPE) + 1
        expression_idx = header.index(EXPRESSION) + 1
        sheet_idx = header.index(LOOKUP_SHEET) + 1
        sheet_keys_idx = header.index(LOOKUP_KEYS) + 1
        sheet_value_idx = header.index(LOOKUP_VALUE) + 1
        lookup_idx_list = [sheet_idx, sheet_keys_idx, sheet_value_idx]

        # factor in optional Default column
        try:
            lookup_idx_list.append(header.index(LOOKUP_DEFAULT) + 1)
        except ValueError:
            pass

        # Target field cannot be duplicated
        rule = Rule(
            type="duplicateValues",
            dxf=workbook_styles.DUPLICATE_TARGET_FIELD.to_differential(),
        )
        sheet.conditional_formatting.add(self._format_range(target_idx), rule)

        # expression specified, all others must be null
        range_string, formula = self._restrict_formatting(key=expression_idx, others=lookup_idx_list)
        rule = Rule(
            type="expression",
            dxf=workbook_styles.EXPRESSION_LOOKUP.to_differential(),
            formula=[formula],
            stopIfTrue=True,
        )
        sheet.conditional_formatting.add(range_string, rule)

        # If one of the sheet columns are specified, the others are required. DefaultLookup is not required.
        cells = ", ".join(f"${get_column_letter(idx)}2" for idx in [sheet_idx, sheet_keys_idx, sheet_value_idx])
        formula = f"""
=AND(
    COUNTA({cells}) > 0,
    COUNTA({cells}) < 3 )
"""

        rule = Rule(
            type="expression",
            dxf=workbook_styles.LOOKUP_REQUIRED.to_differential(),
            formula=[formula],
            stopIfTrue=True,
        )

        multi_range = MultiCellRange()
        for idx in [sheet_idx, sheet_keys_idx, sheet_value_idx]:
            multi_range.add(CellRange(self._format_range(idx)))

        sheet.conditional_formatting.add(str(multi_range), rule)

        # add field type conditional formatting to expression field
        self._expression_field_type_formatting(sheet, field_type_idx, expression_idx, error=False)
        self._expression_field_type_formatting(sheet, field_type_idx, expression_idx, error=True)

    def _expression_field_type_formatting(
        self,
        sheet: Worksheet,
        field_type: int,
        expression: int,
        error: bool,
    ):
        """Adds conditional formatting to expression column for field type"""
        from openpyxl.formatting.rule import Rule

        if error:
            style = workbook_styles.FIELD_ERROR.to_differential()
        else:
            style = workbook_styles.FIELD_WARNING.to_differential()

        formula = """
=INDEX(
    INDIRECT("FieldType"),
    MATCH(
        VLOOKUP(MID(${exp}2, 2, LEN(${exp}2)-2), INDIRECT("SourceSchema_Fields"), 3, FALSE),
        INDIRECT("FieldType[#Headers]"),
        0),
    MATCH(
        ${field}2,
        INDIRECT("FieldType[#Headers]"),
        0)
)={val}
""".strip().format(
            exp=get_column_letter(expression),
            field=get_column_letter(field_type),
            val=0 if error else 2,
        )

        sheet.conditional_formatting.add(
            self._format_range(expression),
            Rule(
                type="expression",
                dxf=style,
                formula=[formula],
                stopIfTrue=True,
            ),
        )

    @staticmethod
    def _format_hyperlinks(book: openpyxl.Workbook):
        """Formats all cells in the book with hyperlinks to be blue font with underline"""
        for sheet in book.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.hyperlink:
                        workbook_styles.HYPERLINK.apply(cell)

    def post_process(self, book: openpyxl.Workbook):
        """Cleanup the worksheets"""

        skip = {SOURCE_SCHEMA, TARGET_SCHEMA, INFO}
        self._format_header([s for s in book.worksheets if s.title not in skip])
        self._set_column_width(book.worksheets)
        self._write_metadata(book)
        self._format_hyperlinks(book)

        if MAPPING in book:
            mapping: Worksheet = book[MAPPING]
            book.active = mapping
            self._mapping_sheet_conditional_formatting(mapping)

        # If compute stats is enabled, set the numeric format on the cells
        if self.compute_stats:
            for sheet in (SOURCE_SCHEMA, TARGET_SCHEMA):
                if sheet not in book:
                    continue

                schema: Worksheet = book[sheet]
                header = [c.value for c in schema[1]]
                if COUNT not in header or FILL_FACTOR not in header:
                    continue

                # Comma format for Count, hide 0 values
                count_col = header.index(COUNT)
                for row in schema.iter_rows(min_row=2, min_col=count_col + 1, max_col=count_col + 1):
                    for cell in row:
                        cell.number_format = r"#,##0;\-0;;@"

                # Percent format for FillFactor, hide 0 values
                fill_col = header.index(FILL_FACTOR)
                for row in schema.iter_rows(min_row=2, min_col=fill_col + 1, max_col=fill_col + 1):
                    for cell in row:
                        cell.number_format = r"0%;\-0;;@"

    @staticmethod
    def _create_valid_sheet_name(workbook: openpyxl.Workbook, name: str):
        """Ensures the sheet name is valid"""
        from openpyxl.workbook.child import INVALID_TITLE_REGEX

        # Replace invalid characters
        name = INVALID_TITLE_REGEX.sub("_", name)

        # Append running number to sheet name to ensure uniqueness, still staying < 32 characters
        existing_sheets = {s.lower() for s in workbook.sheetnames}
        base = name
        i = 1
        while name[:31].lower() in existing_sheets:
            # If adding a number exceeds the length, we need to start chopping off the end of the base name.
            name = f"{base[:31 - len(str(i))]}{i}"
            i += 1

        return name[:31]

    def _create_domain_sheets(self, domains: list, book: openpyxl.Workbook):
        """Creates all the domain worksheets"""
        header = [
            "Domain Name",
            "Description",
            "Field Type",
            "Domain Type",
            "Split Policy",
            "Merge Policy",
            "# of Codes",
        ]
        toc_rows = [header]
        toc = book.active
        toc.title = "List"

        for i, domain in enumerate(domains, 1):
            row = [
                domain.name,
                domain.description or None,
                helper.property_to_pretty(domain.type),
                helper.property_to_pretty(domain.domainType),
                helper.property_to_pretty(domain.splitPolicy),
                helper.property_to_pretty(domain.mergePolicy),
                len(domain.codedValues or {}) or None,
            ]
            toc_rows.append(row)

            # Excel has limitations with sheet names and length so for ease of use, we just use incrementing
            # number as the sheet name.
            sheet = book.create_sheet(str(i))
            self._rows_to_table([["Key", "Value"], *zip(header, row)], "Properties", sheet)

            if domain.domainType == "CodedValue":
                label = "DomainCodedValue"
                records = [["Code", "Description"], *domain.codedValues.items()]
            else:
                label = "DomainRange"
                records = [["Minimum", "Maximum"], domain.range]
            self._rows_to_table(records, label, sheet)

        self._rows_to_table(toc_rows, "Domains", toc)

    @staticmethod
    def _supports_attachments(mapping: "TableMatch") -> bool:
        """Determine if MaintainAttachments can be supported based on shape type matching. The only mapping for
        tables that should default to False is Table > Point"""
        source = mapping.source.shape_type
        target = mapping.target.shape_type
        if source == "Table" and target == "Point":
            return False
        return not (source != "Table" and target != "Table" and source != target)

    def create_domain_workbooks(self, tables):
        """Creates a workbook for each target workspace with a list of all domains"""

        logger.debug("Creating domain workbooks")

        seen = set()
        folder = self.folder / DOMAINS_FOLDER

        for table in tables:
            workspace = helper.get_workspace(table.path)
            # Only need to create a new workbook if we haven't processed this workspace before
            if workspace in seen:
                continue
            seen.add(workspace)
            domains = sorted(helper.get_domains(table.path).values(), key=lambda d: d.name.lower())
            if not domains:
                continue

            logger.debug(f"\tCreating domain workbook for {workspace!r}")
            book_path = helper.create_unique_path(folder / f"{os.path.basename(workspace)}.xlsx")
            rel_path = f"../../{DOMAINS_FOLDER}/{os.path.basename(book_path)}"
            folder.mkdir(parents=True, exist_ok=True)
            book = openpyxl.Workbook()
            toc: Worksheet = book.active
            self._create_domain_sheets(domains, book)

            for i, domain in enumerate(domains, 1):
                # We need to keep a running list of workspace and domain to the sheet hyperlink
                # This way, the lookup during post-processing hyperlinking is straightforward.
                self.domain_hyperlink[(workspace.lower(), domain.name.lower())] = f"{rel_path}#{i}!A1"

                # Hyperlink TOC to numeric sheets and numeric sheets back to TOC.
                toc.cell(row=i + 2, column=1).hyperlink = f"#{i}!A1"
                book[str(i)].cell(row=1, column=1).hyperlink = f"#{toc.title}!A1"

            self._set_column_width(book.worksheets)
            self._write_metadata(book)
            self._format_hyperlinks(book)

            book.save(book_path)

    def _filter_mappings(self):
        """Removes invalid mappings"""

        for match in self.mapping:
            if not is_mapping_valid(
                match.source.describe,
                match.target.describe,
                msg=f"{match.source.base_name} -> {match.target.base_name}",
            ):
                continue

            yield match

    def create_mapping_workbook(self, match: "TableMatch"):
        """Creates the source-target workbook"""

        self.book = openpyxl.Workbook()
        self.book.active.title = INTRO
        self.book.active.sheet_properties.tabColor = workbook_styles.TAB_WHITE

        self.create_sheets(match)
        self.create_mapping_sheet(match)
        self._hyperlink_domains(self.book, source=match.source.describe.path, target=match.target.describe.path)
        self.create_info_sheet(match.source.subtype_code)
        self.post_process(self.book)

    def create_global_workbook(self, folder: Path):
        """Create GlobalLookups workbook where lookup sheets can be copied"""
        folder: Path = folder / MAPPING_FOLDER / GLOBAL_LOOKUP
        folder.mkdir(parents=True, exist_ok=True)

        book = openpyxl.Workbook()
        book.save(folder / f"{GLOBAL_LOOKUP}.xlsx")

    def finalize_data_reference(self, book: openpyxl.Workbook):
        """Finalize DataReference workbook"""

        self._format_header(book.worksheets)
        self._set_column_width(book.worksheets)
        self._format_hyperlinks(book)
        self._write_metadata(book)
        if INTRO in book:
            book.remove(book[INTRO])
        book.active = book[SOURCE_TARGET]

    def main(self):
        mappings = list(self._filter_mappings())
        if not mappings:
            return

        self.folder.mkdir(parents=True, exist_ok=True)

        mapping_book = openpyxl.Workbook()
        mapping_book.active.title = INTRO
        mapping_book.active.sheet_properties.tabColor = workbook_styles.TAB_WHITE

        reference_sheet = mapping_book.create_sheet(SOURCE_TARGET)
        reference_sheet.append(self.HEADER)

        tables = []
        for m in mappings:
            tables.extend([m.source.describe, m.target.describe])
        self.create_domain_workbooks(tables)

        # Create drop-down validation for Enabled, MaintainAttachments, PreserveGlobalids
        valid = DataValidation(type="list", formula1='"True,False"', allow_blank=True, showErrorMessage=False)
        reference_sheet.add_data_validation(valid)
        for heading in (ENABLED, ATTACHMENTS, GLOBALIDS):
            letter = get_column_letter(self.HEADER.index(heading) + 1)
            valid.add(f"{letter}2:{letter}{MAX_ROWS}")

        logger.debug("Creating mapping workbooks")
        arcpy.SetProgressor(type="STEP", message="Processing workbooks", min_range=0, max_range=len(mappings))
        for mapping in mappings:
            source = mapping.source
            target = mapping.target

            message = f"{source.name_key} -> {target.name_key}"
            logger.debug(f"\t{message}")
            arcpy.SetProgressorPosition()
            arcpy.SetProgressorLabel(message)

            if source.describe.datasetType == "Table":
                folder = "Table"
            else:
                folder = source.describe.shapeType

            folder = self.folder / MAPPING_FOLDER / f"{folder}s"
            folder.mkdir(parents=True, exist_ok=True)

            xlsx = f"{source.name_key}-{target.name_key}.xlsx".translate(str.maketrans("/", "_", r'\:*?"<>|'))
            output_file = pathlib.Path(arcpy.CreateUniqueName(xlsx, str(folder)))

            # Add row pointing from source -> target and add a relative hyperlink to the new workbook.
            reference_sheet.append(
                [
                    source.describe.catalogPath,
                    source.definition_query,
                    target.describe.catalogPath,
                    None,
                    f"{folder.parent.stem}/{folder.stem}/{output_file.name}",
                    True,
                    self._supports_attachments(mapping),
                    False,
                    None,
                ]
            )
            cell = reference_sheet.cell(row=reference_sheet.max_row, column=self.HEADER.index(MAPPING_WORKBOOK) + 1)
            cell.hyperlink = cell.value

            self.create_mapping_workbook(mapping)
            self.book.remove(self.book[INTRO])
            self.book.active = self.book[MAPPING]
            self.book.save(output_file)

        self.finalize_data_reference(mapping_book)
        workbook = self.folder / "DataReference.xlsx"
        mapping_book.save(workbook)

        self.create_global_workbook(self.folder)

        shutil.copytree(
            src=pathlib.Path(__file__).parents[1] / "dataloading" / "scripts",
            dst=self.folder / SCRIPTS_FOLDER,
            ignore=shutil.ignore_patterns("__pycache__"),
        )


class WorkbookLoader:
    MAPPING_HEADER = [TARGET_FIELD, EXPRESSION, LOOKUP_SHEET, LOOKUP_KEYS, LOOKUP_VALUE, LOOKUP_DEFAULT]

    def __init__(self, workbook: str, preview: str = None):
        self.workbook = workbook
        self.preview = preview

        self.field_map = {}

        self.source_describe = None
        self.target_describe = None
        self.mapping_msg = None
        self._tempdir = None
        self._preview_gdb = None
        self._attachments = None
        self._globalids = None
        self._transformations = None

        self.code_block = []

        self.row = None

    @staticmethod
    def _create_code_block(function_name: str, lookup: dict, default=None):
        return f"""
def {function_name}(*fields):
    lookup = {lookup}
    return lookup.get(fields, {default!r})"""

    def _sheet_mapping(
        self, workbook: WorkbookWrapper, target_field: str, sheet_name: str, keys: str, value: str, default=None
    ):
        if keys is None or value is None:
            arcpy.AddIDMessage(MsgType.WRN, 3813, self.mapping_msg, self.row)
            return

        if (sheet := workbook.get_sheet(sheet_name)) is None:
            return

        # Keys and value represent column names in the referenced sheet.
        fields = [x.strip() for x in str(keys).split(",")]  # save for creating lookup with correct field names
        keys = [str(k).casefold() for k in fields]
        columns = keys + [value]
        if workbook.is_header_missing(sheet, expected=columns):
            return

        if len(keys) != len(set(keys)):
            # Problem with mapping workbook: %1, Row %2, Duplicate LookupKeys were provided.
            arcpy.AddIDMessage(MsgType.WRN, 3876, self.mapping_msg, self.row)
            return

        # Because the lookup is based on the source data, the keys need to be source fields.
        source_fields = {f.name.lower(): f for f in self.source_describe.fields}
        stringify = []
        for key in keys:
            if (source_f := source_fields.get(key.casefold(), None)) is None:
                # Problem with mapping workbook: %1, Row %2, Key must be a source field to be used in mapping.
                arcpy.AddIDMessage(MsgType.WRN, 3816, self.mapping_msg, self.row)
                return
            # cast string fields to str
            if source_f.type == "String":
                stringify.append(key.casefold())

        header = workbook._get_header(sheet)
        if value in columns[:-1]:
            # the value column (which is a domain name) is a duplicate of a key column (which are fields). This is
            # allowed but we need to rename the value column to build lookup properly.
            header_rindex = len(header) - 1 - header[::-1].index(value)  # get right index in header
            value = header[header_rindex] = columns[-1] = uuid.uuid4().hex

        df = pd.DataFrame.from_records(data=sheet.iter_rows(min_row=2, values_only=True), columns=header)
        index = [header.index(col.casefold()) for col in columns if col.casefold() in header]
        df = df.iloc[:, index]

        import numpy as np

        df = df.replace(np.nan, None)

        # Cast all string column values to str. Target values do not need to be cast since Pro handles casting during
        # insert.
        for k in stringify:
            df[k] = [None if x is None else str(x) for x in df[k].values]

        # Even though mapping a source value to a target value of None is valid, we want to filter them out here
        # because the python dictionary lookup created in the code block uses .get() which by default returns None
        # TODO: what if there are duplicate keys with different target values? do we throw a warning here?
        df = df[df[value.lower()].notnull()]

        # Because keys is a list, our dictionary keys will always be a tuple, even if len(keys) == 1.
        lookup = dict(zip(map(tuple, df[[k.lower() for k in keys]].values), df[value].values))

        # We need to create a python function that will be in the code block and called as the expression.
        function_name = f"{sheet_name}_{target_field}"
        self.code_block.append(self._create_code_block(function_name, lookup, default))

        return "{}({})".format(function_name, ",".join(f"!{field}!" for field in fields))

    @staticmethod
    def _expression_mapping(expression: str):
        # TODO: Extract things that look like python fields and verify they exist.
        return expression

    @staticmethod
    def get_global_workbook(workbook: str) -> Optional["WorkbookWrapper"]:
        """find GlobalLookup workbook if exists"""
        global_path = pathlib.Path(workbook).parents[1] / GLOBAL_LOOKUP / f"{GLOBAL_LOOKUP}.xlsx"
        if global_path.exists():
            return WorkbookWrapper(workbook=global_path.as_posix(), on_error="warning")

    def _validate_single_workbook(self, workbook: WorkbookWrapper, for_update: bool = False) -> Optional[pd.DataFrame]:
        """Validates a single workbook"""
        if workbook.wb is None:
            return

        # try getting old Mapping tab as well
        mapping_name = "Mapping" if "Mapping" in workbook.wb else MAPPING
        sheet = workbook.get_sheet(mapping_name)
        if sheet is None:
            return

        if workbook.is_header_missing(sheet, expected=self.MAPPING_HEADER[:-1]):
            return

        df = workbook.sheet_to_df(sheet, columns=self.MAPPING_HEADER[:-1], optional_columns=[LOOKUP_DEFAULT])
        if not for_update:
            # We only need to process where target field is specified and at least 1 of the 2 options is present
            keep_rows = df[self.MAPPING_HEADER[1:]].notnull().any(axis=1)
            df = df[df[TARGET_FIELD].notnull() & keep_rows]

        # Multiple target fields is ambiguous. Ignore multiple nulls.
        duplicates = df.loc[df[TARGET_FIELD].str.lower().duplicated(), TARGET_FIELD]
        duplicates = duplicates.loc[duplicates.notnull()]
        if not duplicates.empty:
            for field in duplicates.values:
                arcpy.AddIDMessage(MsgType.WRN, 3814, self.mapping_msg, field)  # Target field %2 cannot be duplicated.
            return

        return df

    def create_lookup(self, mapping_workbook: str) -> Optional[int]:
        workbook = WorkbookWrapper(workbook=mapping_workbook, on_error="warning")
        global_workbook = self.get_global_workbook(mapping_workbook)
        mapping_df = self._validate_single_workbook(workbook)
        if mapping_df is None:
            return 1

        target_fields: dict[str, arcpy.Field] = {f.name.lower(): f for f in self.target_describe.fields}
        for row, target_field, expression, sheet, keys, value, default in mapping_df[self.MAPPING_HEADER].itertuples():
            # Row is offset by 2 because of zero indexing + header column
            self.row = row + 2

            # If lookup mode is used, expression must be null
            count = sum(x is not None for x in (sheet, keys, value))
            if expression and count:
                # Only one of expression or sheet must be specified.
                arcpy.AddIDMessage(MsgType.WRN, 3811, self.mapping_msg, self.row)
                continue

            target_field_object = target_fields.get(target_field.lower(), None)
            if target_field_object is None or target_field_object.type in {"OID"}:
                # Target field does not exist or is not supported.
                arcpy.AddIDMessage(MsgType.WRN, 3812, self.mapping_msg, self.row)
                continue

            code = None
            if expression is not None:
                code = self._expression_mapping(expression)
            elif sheet is not None:
                sheet = str(sheet).strip()
                value = str(value).strip().casefold()
                # check for global sheet reference
                if helper.is_global_sheet(sheet, global_workbook):
                    sheet = sheet[1:-1]
                    lookup_workbook = global_workbook
                else:
                    lookup_workbook = workbook

                code = self._sheet_mapping(
                    workbook=lookup_workbook,
                    target_field=target_field,
                    sheet_name=sheet,
                    keys=keys,
                    value=value,
                    default=default,
                )

            if code is not None:
                # Code here might not be a string, but when we are passing it to calculateFields, it needs to be.
                self.field_map[target_field] = str(code)

    def _delete_target_features(self, target_desc: arcpy.Describe, sql_query: str):
        """Deletes target features as needed"""
        if not sql_query:
            return

        count = int(arcpy.GetCount_management(target_desc.catalogPath)[0])
        if count < 1:
            return

        if helper.is_select_all(sql_query):
            layer_count = count
            layer = target_desc.catalogPath
        else:
            layer, layer_count, *_ = arcpy.SelectLayerByAttribute_management(
                in_layer_or_view=target_desc.catalogPath, selection_type="NEW_SELECTION", where_clause=sql_query
            )
            layer_count = int(layer_count)
            if layer_count < 1:
                return

        # Truncate is faster, so we use it if possible. If target has attachments table we cannot use Truncate.
        cannot_truncate = not helper.is_service(target_desc) and target_desc.relationshipClassNames
        truncate = layer_count == count and not cannot_truncate
        delete_rows = self._delete_service_rows_API if helper.is_service(target_desc) else arcpy.DeleteRows_management
        self.layer_count = layer_count

        try:
            func = arcpy.TruncateTable_management if truncate else delete_rows
            logger.debug(f"Deleting {layer_count:,}/{count:,} rows on {target_desc.catalogPath}")
            func(layer)
        except arcpy.ExecuteError as e:
            if truncate:
                # There are likely numerous ExecuteError possibilities, so retry with DeleteRows.
                logger.debug("".join(e.args))
                logger.debug(f"\tTrying again with {delete_rows.__name__}")
                delete_rows(layer)
            else:
                raise e

    def _delete_service_rows_API(self, layer):
        """using the python api to delete feature removes a series of queries and creates an optimized delete payload"""
        from arcgis.features import FeatureLayer
        from arcgis.gis import GIS
        import arcpy

        gis = GIS("pro")
        d = arcpy.Describe(layer)  # cannot lru_cache a Layer object
        fl = FeatureLayer(d.catalogPath, gis=gis)
        with arcpy.da.SearchCursor(layer, ["OID@"]) as curs:
            oids = list(curs)
        del curs

        chunk_size = 500
        arcpy.SetProgressor(
            type="STEP",
            message=N_(86616, alias=d.aliasname, progress=f"0/{self.layer_count}"),
            min_range=0,
            max_range=self.layer_count // chunk_size,
        )
        for i in range(0, len(oids), chunk_size):
            progress = self.layer_count if i + chunk_size > self.layer_count else i + chunk_size
            arcpy.SetProgressorLabel(N_(86616, alias=d.aliasname, progress=f"{progress}/{self.layer_count}"))
            fl.edit_features(deletes=",".join([str(oid[0]) for oid in oids[i : i + chunk_size]]))
            arcpy.SetProgressorPosition()

    def process(self, mapping: str, def_query: str):
        """Parse the mapping workbook and call ETL"""
        from .etl import ETL

        source = self.source_describe.catalogPath
        if def_query:
            source = arcpy.SelectLayerByAttribute_management(
                in_layer_or_view=source, selection_type="NEW_SELECTION", where_clause=def_query
            )[0]
        if int(arcpy.GetCount_management(source)[0]) < 1:
            value = f"{self.source_describe.name}{f' WHERE {def_query}' if def_query else ''}"
            arcpy.AddIDMessage(MsgType.WRN, 3797, self.mapping_msg, value)  # No features in %2.
            return

        # Create lookup only modifies class attributes, so we reset in between runs.
        self.field_map = {}
        self.code_block = get_user_scripts(pathlib.Path(self.workbook).parent)
        ret = self.create_lookup(mapping)

        if not self.field_map:
            if ret != 1:
                arcpy.AddIDMessage(MsgType.WRN, 3815, self.mapping_msg)  # No fields to map.
            return

        with arcpy.EnvManager(maintainAttachments=self._attachments, preserveGlobalIds=self._globalids):
            etl = ETL(
                source=source,
                target=self.target_describe.catalogPath,
                mapping=self.field_map,
                code_block="\n\n".join(self.code_block),
                preview=self.preview,
                tempdir=self._tempdir,
                preview_gdb=self._preview_gdb,
                transformations=self._transformations,
            )
            result = etl.main()

            return result

    def _validate_mapping_workbook(self, for_update: False) -> pd.DataFrame:
        """Validates mapping workbook and gets dataframe of records"""
        wb = WorkbookWrapper(self.workbook, on_error="raise")
        sheet = wb.get_sheet(SOURCE_TARGET)

        header = [SOURCE, DEFINITION_QUERY, TARGET, DELETE_QUERY, MAPPING_WORKBOOK]
        optional = [ATTACHMENTS, GLOBALIDS, TRANSFORMATIONS]
        wb.is_header_missing(sheet, expected=header)
        df = wb.sheet_to_df(sheet, columns=header, optional_columns=[ENABLED] + optional)
        # Attachments, Globalids are optional columns so set defaults
        df[ATTACHMENTS].replace([None], value=True, inplace=True)
        df[GLOBALIDS].replace([None], value=False, inplace=True)

        if for_update:
            # We don't care about enabled setting during update, process every source and target pair
            return df[header + optional]
        # Because enabled is optional, we treat True and None as the same (ie, process this row)
        return df.loc[df[ENABLED].isin({True, None}), header + optional]

    def validate_mapping_values(self, for_update: bool = False):
        df = self._validate_mapping_workbook(for_update)

        for (
            row,
            source,
            def_query,
            target,
            delete_query,
            mapping,
            attachments,
            globalids,
            transformations,
        ) in df.itertuples():
            # Row is offset by 2 because of zero indexing + header column
            self.row = row + 2

            # All of these values are required
            if any(x is None for x in (source, target, mapping)):
                arcpy.AddIDMessage(MsgType.WRN, 3804, self.row)  # A required value is missing.
                continue

            mapping = os.path.abspath(os.path.join(os.path.dirname(self.workbook), mapping))

            msg = f"{os.path.basename(source)} -> {os.path.basename(target)}"
            logger.debug(msg)

            skip = False
            for file, col in [(source, SOURCE), (target, TARGET), (mapping, MAPPING_WORKBOOK)]:
                # os.path.exists is faster check for Excel workbook than arcpy.Exists
                if os.path.exists(file):
                    continue
                if not helper.does_exist(file):
                    skip = True
                    arcpy.AddIDMessage(MsgType.WRN, 3803, self.row, file)  # Does not exist or is not supported.

            if skip:
                continue

            logger.debug("\tDescribing source...")
            source_describe = helper.describe_object(source)
            logger.debug("\tDescribing target...")
            target_describe = helper.describe_object(target)
            if not for_update:
                # don't check this if we are only doing an update to Data Loading workspace
                if not is_mapping_valid(source_describe, target_describe, msg=msg):
                    continue

                if def_query and not is_query_valid(source, def_query):
                    arcpy.AddIDMessage(MsgType.WRN, 3802, self.row, DEFINITION_QUERY)  # DefinitionQuery is invalid.
                    continue

                if delete_query and not is_query_valid(target, delete_query):
                    arcpy.AddIDMessage(MsgType.WRN, 3802, self.row, DELETE_QUERY)  # DeleteQuery is invalid.
                    continue

            yield source_describe, def_query, target_describe, delete_query, mapping, attachments, globalids, transformations

    def main(self):
        rows = list(self.validate_mapping_values())
        if not rows:
            return

        if self.preview:
            self._preview_gdb = helper.create_preview_gdb(self.preview)

        else:
            # Remove data in target based on delete query
            for (
                source_describe,
                def_query,
                target_describe,
                delete_query,
                mapping,
                attachments,
                globalids,
                transformations,
            ) in rows:
                if delete_query:
                    logger.debug(f"Removing data from {target_describe.name}...")
                self._delete_target_features(target_describe, delete_query)

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as temp_dir:
            self._tempdir = temp_dir

            # Load data to target
            arcpy.SetProgressor(type="STEP", message="Loading data", min_range=0, max_range=len(rows))
            results = set()
            for (
                source_describe,
                def_query,
                target_describe,
                delete_query,
                mapping,
                attachments,
                globalids,
                transformations,
            ) in rows:
                self.mapping_msg = f"{source_describe.name} -> {target_describe.name}"
                logger.debug(self.mapping_msg)
                arcpy.SetProgressorLabel(self.mapping_msg)

                self._attachments = attachments
                self._globalids = globalids
                self._transformations = transformations
                self.source_describe = source_describe
                self.target_describe = target_describe
                if ret := self.process(mapping, def_query):
                    results.add(ret)

                arcpy.SetProgressorPosition()

        return list(results)


class WorkbookUpdater:
    def __init__(self, workbook: str):
        self.workbook = pathlib.Path(workbook)

        self.source_describe = None
        self.target_describe = None
        self.source_subtype = None

        self.wg = WorkbookGenerator([], str(self.workbook.parent))
        # reset wg folder to correct folder
        self.wg.folder = self.workbook.parent

        self._shape_inserted = False

    def _copy_dlw(self):
        """copy data loading workspace to new folder so we can update without locks"""
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        new = helper.create_unique_path(self.wg.folder.parent / f"{self.wg.folder.stem}-{timestamp}")

        # copy folders
        for folder in (MAPPING_FOLDER, DOMAINS_FOLDER, SCRIPTS_FOLDER):
            if (path := self.wg.folder / folder).exists():
                self._copy_folders_with_excel(path, new / folder)

        # copy Data Ref workbook
        shutil.copy(self.workbook, new / self.workbook.name)

        self.wg.folder = new
        self.workbook = new / self.workbook.name

    @staticmethod
    def _copy_folders_with_excel(old: Path, new: Path) -> None:
        """Copy folders. We don't care about excel lock files, swallow those errors"""
        try:
            shutil.copytree(str(old), str(new))
        except shutil.Error as e:
            errors = e.args[0]
            for error in errors:
                src, dst, msg = error
                src_p = pathlib.Path(src)
                if src_p.stem.startswith("~$"):
                    pass
                else:
                    raise

    def update_domain_workbooks(self, rows):
        """Regenerate domain workbooks. This will overwrite any old domain workbooks."""
        folder = self.wg.folder / DOMAINS_FOLDER

        tables = []
        for source_describe, _, target_describe, *_ in rows:
            tables.extend([source_describe, target_describe])

        # Delete existing domain workbooks
        for table in tables:
            workspace = helper.get_workspace(table.path)
            wb_path = folder / f"{os.path.basename(workspace)}.xlsx"
            if wb_path.exists():
                wb_path.unlink()

        # Recreate domain workbooks
        self.wg.create_domain_workbooks(tables)

    def update_scripts(self):
        """Give fresh copy of base.py and new copy of user.py if does not exist"""
        scripts: pathlib.Path = self.wg.folder / SCRIPTS_FOLDER
        src: pathlib.Path = pathlib.Path(__file__).parents[1] / "dataloading" / "scripts"

        if not scripts.exists():
            shutil.copytree(src=src, dst=scripts, ignore=shutil.ignore_patterns("__pycache__"))
        else:
            shutil.copy(src / "base.py", scripts / "base.py")
            if not (scripts / "user.py").exists():
                shutil.copy(src / "user.py", scripts / "user.py")

    def add_global_workbook(self):
        global_workbook: pathlib.Path = self.wg.folder / MAPPING_FOLDER / GLOBAL_LOOKUP / f"{GLOBAL_LOOKUP}.xlsx"
        if global_workbook.exists():
            return
        self.wg.create_global_workbook(self.wg.folder / MAPPING_FOLDER)

    @staticmethod
    def _existing_values(df: pd.DataFrame, field: str) -> Set:
        return set(df[field].str.casefold())

    def insert_shape_row(self, sheet: Worksheet) -> str:
        """If shape row is missing we need to add it"""

        if not (target_name := getattr(self.target_describe, "shapeFieldName", "")):
            return ""

        existing = set()
        for val, *_ in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
            existing.add(self._lower(val))
        if target_name.casefold() in existing:
            return ""

        sheet.insert_rows(2)
        row = [target_name, "Geometry"]
        if getattr(self.source_describe, "shapeType", "") == self.target_describe.shapeType:
            row.append(f"!{self.source_describe.shapeFieldName}!")

        for idx, val in enumerate(row, 1):
            sheet.cell(2, idx).value = val
        return f"\n\t\t{N_(86615, value=f'{TARGET_FIELD}: {target_name} - Geometry')}"  # Adding missing

    @staticmethod
    def update_lookup_hyperlinks(sheet: Worksheet, parent_wb: WorkbookWrapper):
        """Hyperlink any existing local or global lookup sheets"""
        header = parent_wb._get_header(sheet)
        lookup_col = header.index(LOOKUP_SHEET.casefold()) + 1
        global_workbook = WorkbookLoader.get_global_workbook(parent_wb.catalog_path)
        for row in sheet.iter_rows(min_row=2, min_col=lookup_col, max_col=lookup_col):
            for cell in row:
                if cell.value is not None:
                    if helper.is_global_sheet(cell.value, global_workbook):
                        sheet = cell.value.strip("[] ")
                        if sheet.casefold() in global_workbook._sheets_lookup():
                            cell.hyperlink = f"../{GLOBAL_LOOKUP}/{GLOBAL_LOOKUP}.xlsx#{sheet}!A1"
                        else:
                            cell.hyperlink = None
                            workbook_styles.NON_HYPERLINK.apply(cell)
                    else:
                        if cell.value.casefold() in parent_wb._sheets_lookup():
                            cell.hyperlink = f"#{cell.value}!A1"
                        else:
                            cell.hyperlink = None
                            workbook_styles.NON_HYPERLINK.apply(cell)

    def update_mapping_sheet(self, mapping_sheet: Worksheet, workbook: WorkbookWrapper) -> None:
        self.update_target_fields(mapping_sheet)
        self.update_columns(mapping_sheet, workbook)
        self.update_lookup_hyperlinks(mapping_sheet, workbook)

    def update_target_fields(self, sheet: Worksheet):
        """Add missing fields and update existing in the Mapping worksheet"""

        target_fields_lookup = {
            f.name: helper.property_to_pretty(f.type)
            for f in self.target_describe.fields
            if f.editable and f.type not in {"OID"}
        }
        source_fields_lookup = {f.name.casefold(): f.name for f in self.source_describe.fields}
        message = self.insert_shape_row(sheet)
        # track whether shape was inserted to control drop down validation
        self._shape_inserted = bool(message)

        message += self.update_first_two_cols(
            sheet, target_fields_lookup, TARGET_FIELD, "FieldType", source_fields_lookup
        )
        if message:
            logger.info(f"\t{N_(86613, message=message.rstrip())}")  # Updating Mapping sheet

    @staticmethod
    def update_columns(sheet: Worksheet, parent_wb: WorkbookWrapper):
        """Add default lookup column if needed"""
        header = parent_wb._get_header(sheet)
        if LOOKUP_DEFAULT.casefold() not in header:
            sheet[f"{get_column_letter(sheet.max_column + 1)}1"].value = LOOKUP_DEFAULT

    def update_schema_subtype_sheets(self, workbook: WorkbookWrapper):
        """Regenerate schema tabs and delete subtype tabs"""

        # check fill factor on existing source schema
        calc_stats = False
        source_schema_name = "SourceSchema" if "SourceSchema" in workbook.wb else SOURCE_SCHEMA
        source_sheet = workbook.get_sheet(source_schema_name)
        if source_sheet:
            cols = workbook._get_header(source_sheet)
            if FILL_FACTOR.lower() in cols and COUNT.lower() in cols:
                calc_stats = True
        self.wg.compute_stats = calc_stats

        # drop old sheets
        for sheet_name in (
            TARGET_SCHEMA,
            TARGET_SUBTYPES,
            SOURCE_SCHEMA,
            SOURCE_SUBTYPES,
            HIDDEN_SHEET,
            "TargetSchema",
            "SourceSchema",
        ):
            if sheet_name in workbook.wb:
                workbook.remove_sheet(sheet_name)

        # create new schema and subtype sheets to the right of the mapping sheet
        mapping_sheet = workbook.wb[MAPPING]
        idx = workbook.wb.get_index(mapping_sheet) + 1
        self.wg.book = workbook.wb
        self.wg._create_schema_sheet(self.target_describe, TARGET_SCHEMA, calc_stats=False, index=idx)
        idx += 1
        self.wg._create_schema_sheet(self.source_describe, SOURCE_SCHEMA, calc_stats=calc_stats, index=idx)

        # Add drop down validation again.
        self.wg._add_drop_down(
            self.source_describe, mapping_sheet, self.wg.folder, "C3" if self._shape_inserted else "C2"
        )
        self.wg._hyperlink_domains(workbook.wb, source=self.source_describe.path, target=self.target_describe.path)

    def _validate_info_worksheet(self, workbook: WorkbookWrapper) -> Optional[pd.DataFrame]:
        """Validates a single workbook"""
        if workbook.wb is None:
            return

        sheet = workbook.get_sheet(INFO, raise_error=False)
        if sheet is None:
            return

        if workbook.is_header_missing(sheet, expected=[MAPPING_PROPERTY, VALUE]):
            return

        df = workbook.sheet_to_df(sheet, columns=[MAPPING_PROPERTY, VALUE])
        if SOURCE_ST.casefold() not in self._existing_values(df, MAPPING_PROPERTY):
            logger.debug(f"\t{INFO} sheet does not have a {SOURCE_ST} property.")
            return

        return df

    def _get_subtype_from_excel_name(self, workbook_path: str) -> Optional[str]:
        """Extract subtype name from excel file name. Makes assumption the Table/FC name is more likely to have
        underscores than subtype name. No real way for us to differentiate if they both have them."""
        path = pathlib.Path(workbook_path)
        if "_" not in path.stem:
            # there is no Subtype
            return None

        st_name_candidate = path.stem.split("-")[0].split("_")[-1].lower()
        st_name_lookup = {
            (info["Name"] or "").lower(): code
            for code, info in helper.get_subtypes(self.source_describe.catalogPath).items()
        }
        if st_name_candidate not in st_name_lookup:
            logger.debug(f"\tCannot determine Source subtype in {path.name}. Using domains from root.")
            return None
        return st_name_lookup[st_name_candidate]

    def _get_source_subtype(self, workbook: WorkbookWrapper) -> Optional[int]:
        """extract the Source subtype from the workbook"""

        # if subtypes don't exist then we don't care about the workbook setting
        subtypes = helper.get_subtypes(self.source_describe.catalogPath)
        if not subtypes:
            return

        info_df = self._validate_info_worksheet(workbook)

        if info_df is not None:
            # return Source Subtype value from Info workbook
            subtype = info_df.set_index(MAPPING_PROPERTY).loc[SOURCE_ST, VALUE]
            if subtype is not None and subtype not in subtypes.keys():
                # in case invalid subtype is provided, we still want to try to extract subtype from file name
                logger.debug(f"{INFO} sheet contains invalid subtype in {SOURCE_ST} property.")
            else:
                return subtype

        return self._get_subtype_from_excel_name(workbook.catalog_path)

    @staticmethod
    def _lower(thing: Any):
        """if string, lowercase it"""
        return thing.casefold() if isinstance(thing, str) else thing

    def update_first_two_cols(
        self, sheet: Worksheet, lookup: dict, key_name: str = None, value_name: str = None, source_lookup: dict = None
    ) -> str:
        """
        Update the first two columns in a worksheet using a dictionary lookup. Ignores the first row. Case of string
        values in lookup does not matter.

        Args:
            sheet (Worksheet): The worksheet to be updated
            lookup (dict): The dictionary used to update the first two columns of the worksheet. Values in the worksheet
             but not in the lookup will be nulled. Values in the lookup but not in the worksheet will be added.
            key_name (str): Name of Column A to be used for logging changes
            value_name (str): Name of Column B to be used for logging changes
            source_lookup (dict): When values are added from first lookup, determine if exact match exists in this
            lookup and add !expression! to third column
        Returns:
            message (str): Formatted string listing all changes made.
        """

        message = ""
        lower_lookup = {self._lower(k): v for k, v in lookup.items()}

        # update existing values
        for cell_a, cell_b in sheet.iter_rows(min_row=2, max_col=2, values_only=False):
            cell_a_val = self._lower(cell_a.value)
            # null row if it doesn't exist in lookup
            if cell_a_val not in lower_lookup:
                cell_a_orig = copy.deepcopy(cell_a.value)
                cell_a.value = None
                cell_b.value = None
                if cell_a_orig is not None:
                    message += f"\n\t\t{N_(86611, column=key_name, value=cell_a_orig)}"
            # update Col B if necessary
            elif self._lower(cell_b.value) != self._lower(lower_lookup.get(cell_a_val, None)):
                cell_b_orig = copy.deepcopy(cell_b.value)
                cell_b.value = lower_lookup.get(cell_a_val, None)
                message += (
                    f"\n\t\t{N_(86612, column=value_name, value=f'{cell_a.value} - {cell_b_orig} -> {cell_b.value}')}"
                )

        # add missing key value pairs
        existing = ()
        for col in sheet.iter_cols(min_row=2, max_col=1, values_only=True):
            existing = [self._lower(x) for x in col]
            break
        for key, value in lookup.items():
            if (lower_key := self._lower(key)) not in existing:
                new_row = [key, value]
                if source_lookup and lower_key in source_lookup:
                    new_row.append(f"!{source_lookup[lower_key]}!")
                sheet.append(new_row)
                message += f"\n\t\t{N_(86615, value=f'{key_name}: {key} - {value}')}"  # Adding missing

        return message

    @staticmethod
    def _is_valid_subtype_sheet(sheet: Worksheet):
        """validate column headers are as expected"""
        if WorkbookUpdater._lower(sheet["B1"].value) == "subtypename":
            return True
        else:
            logger.debug(f"\tInvalid column headers found in subtype sheet {sheet.title}")
            return False

    def update_field_sheets(self, workbook: WorkbookWrapper):
        """Update any existing Source Field Sheets and add new sheets if needed."""

        # determine if this is subtype worksheet
        st_name = getattr(self.source_describe, "subtypeFieldName", "").casefold()

        # get Source table info
        self.source_subtype = self._get_source_subtype(workbook)
        table = helper.TableWrapper(self.source_describe.catalogPath, self.source_subtype)

        # Get Field Sheets lookup
        # For Field Sheets - the sheet name may differ from field name. Use value from A1 cell.
        non_field_sheets = [x.casefold() for x in (TARGET_SCHEMA, SOURCE_SCHEMA, MAPPING, INFO)]
        field_sheets = {
            (workbook.wb[name]["A1"].value or "").casefold(): workbook.wb[name]
            for name in workbook.wb.get_sheet_names()
            if name.casefold() not in non_field_sheets
        }

        skip_fields = table.fields.default_fields()
        for field in table.fields:
            if field.name.casefold() in skip_fields:
                continue

            if field.domain is None or not field.domain.codedValues:
                # checking if domain exists also filters out subtype field
                continue

            if field.name.casefold() not in field_sheets:
                # Create new sheet
                logger.debug(f"\tCreating new field sheet - {field.name}")
                self.wg._create_field_sheet(field, table)
                continue

            sheet = field_sheets[field.name.casefold()]
            domain_header = f"{field.domain.name} (Description){' (Root)' if st_name in field_sheets else ''}"
            if sheet.cell(1, 2).value in {
                f"{field.domain.name} (Description)",
                f"{field.domain.name} (Description) (Root)",
            }:
                domain_header = sheet.cell(1, 2).value
            if workbook.is_header_missing(sheet, [field.name, domain_header]):
                continue

            # update header
            sheet.cell(1, 1).value = field.name
            sheet.cell(1, 2).value = domain_header
            link = self.wg.domain_hyperlink.get((table.workspace.casefold(), field.domain.name.casefold()), None)
            sheet.cell(1, 2).hyperlink = link

            # update coded value pairs
            message = self.update_first_two_cols(sheet, field.domain.codedValues, "Domain Code", "Domain Description")
            if message:
                logger.info(f"\t{N_(86614, sheet=sheet.title, message=message.rstrip())}")  # Update field sheet

        # if subtype sheet exists run special update
        if st_name in field_sheets and self._is_valid_subtype_sheet(field_sheets[st_name]):
            subtypes = {code: info["Name"] for code, info in helper.get_subtypes(table.path).items()}
            message = self.update_first_two_cols(field_sheets[st_name], subtypes, "Subtype Code", "Subtype Name")
            if message:
                logger.debug(f"\tUpdating subtype sheet - {field_sheets[st_name].title}\n" + message.rstrip())

        # TODO: we could check if code is None and subtypes exist to determine whether to create a Subtype sheet if it
        #  doesn't exist

    def recreate_info_sheet(self, workbook: WorkbookWrapper):
        """remove -info- sheet if exists and recreate it"""
        if INFO in workbook.wb:
            workbook.remove_sheet(INFO)
        self.wg.create_info_sheet(self.source_subtype)

    def process(self, mapping_workbook):
        logger.info(N_(86610, mapping_workbook=mapping_workbook))

        workbook = WorkbookWrapper(workbook=mapping_workbook, on_error="warn", read_only=False)
        wl = WorkbookLoader("")
        mapping_df = wl._validate_single_workbook(workbook=workbook, for_update=True)
        if mapping_df is None:
            return

        mapping_name = "Mapping" if "Mapping" in workbook.wb else MAPPING
        mapping_sheet = workbook.get_sheet(mapping_name)
        mapping_sheet.title = MAPPING

        # run updates
        self.update_mapping_sheet(mapping_sheet, workbook)
        self.update_schema_subtype_sheets(workbook)
        self.update_field_sheets(workbook)
        self.recreate_info_sheet(workbook)
        self.wg.post_process(workbook.wb)

        workbook.wb.save(mapping_workbook)

    def refresh_data_ref(self, workbook: WorkbookWrapper, rows: list):
        """drop data path tab and add source and target full paths"""
        datapath_sheet = workbook.get_sheet(DATA_PATH, raise_error=False)
        if datapath_sheet is not None:
            workbook.remove_sheet(DATA_PATH)
        reference_sheet = workbook.get_sheet(SOURCE_TARGET)

        for idx, (source_describe, _, target_describe, *_) in enumerate(rows, 2):
            cell = reference_sheet.cell(row=idx, column=1)
            cell.value = source_describe.catalogPath
            cell.hyperlink = None
            cell.font = None
            cell = reference_sheet.cell(row=idx, column=3)
            cell.value = target_describe.catalogPath
            cell.hyperlink = None
            cell.font = None

    def update_env_columns(self, workbook: WorkbookWrapper):
        """add attachments, globalids columns in data reference sheet"""
        reference_sheet = workbook.get_sheet(SOURCE_TARGET)
        header = workbook._get_header(reference_sheet)

        new_columns = []
        col_number = reference_sheet.max_column
        for column, value in ((ATTACHMENTS, True), (GLOBALIDS, False)):
            if column.casefold() in header:
                continue
            col_number += 1
            reference_sheet.cell(row=1, column=col_number, value=column)
            new_columns.append(get_column_letter(col_number))

            for row in reference_sheet.iter_rows(min_row=2, min_col=col_number, max_col=col_number):
                row[0].value = value

        if new_columns:
            # Create drop-down validation for Maintain Attachments
            valid = DataValidation(type="list", formula1='"True,False"', allow_blank=True, showErrorMessage=False)
            reference_sheet.add_data_validation(valid)
            valid.add(f"{new_columns[0]}2:{new_columns[-1]}{MAX_ROWS}")

        if TRANSFORMATIONS.casefold() not in header:
            reference_sheet[f"{get_column_letter(col_number+1)}1"].value = TRANSFORMATIONS

    def main(self):
        # copy to fresh folder
        self._copy_dlw()

        # read in DataReference.xlsx to determine what to process
        wl = WorkbookLoader(str(self.workbook))
        rows = list(wl.validate_mapping_values(for_update=True))
        if not rows:
            return

        # update DataReference
        workbook = WorkbookWrapper(str(self.workbook), on_error="warn", read_only=False)
        self.refresh_data_ref(workbook, rows)
        self.update_env_columns(workbook)
        self.wg.finalize_data_reference(workbook.wb)
        workbook.wb.save(str(self.workbook))
        # delete and recreate domains workbooks
        self.update_domain_workbooks(rows)
        # refresh Scripts folder
        self.update_scripts()
        # add global lookup workbook
        self.add_global_workbook()

        # process each mapping workbook
        arcpy.SetProgressor(type="STEP", message="Updating sheets", min_range=0, max_range=len(rows))
        for source_describe, def_query, target_describe, delete_query, mapping, *_ in rows:
            msg = f"{source_describe.name} -> {target_describe.name}"
            logger.debug(msg)
            arcpy.SetProgressorLabel(msg)

            self.source_describe = source_describe
            self.target_describe = target_describe
            self.process(mapping)

            arcpy.SetProgressorPosition()
