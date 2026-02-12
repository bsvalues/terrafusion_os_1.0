"""
Tool name: Table To Excel
Source: TableToExcel.py
Author: ESRI

Convert an table to a MS Excel spreadsheet.
"""
from __future__ import print_function, unicode_literals, absolute_import
import arcpy
import os
import re
import sys
import xlwt
import datetime
import ConversionUtils

# openpyxl imports
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import openpyxl.styles
import openpyxl.cell


class clsField(object):
    """ Class to hold properties and behavior of the output fields """

    @property
    def alias(self):
        return self._field.aliasName

    @property
    def name(self):
        return self._field.name

    @property
    def domain(self):
        return self._field.domain

    @property
    def type(self):
        return self._field.type

    @property
    def length(self):
        return self._field.length

    def __init__(self, f, i, subtypes, cvdomains):
        """ Create the object from a describe field object """
        self._field = f
        self.subtype_field = ''
        self.domain_desc = {}
        self.subtype_desc = {}
        self.index = i

        # Get coded value domain info from field
        if f.domain:
            for cvd in cvdomains:
                if cvd.name.casefold() == f.domain.casefold():
                    self.domain_desc = {0: cvd.codedValues}
                    break

        # Get coded value domain info from subtype
        for st_key, st_val in subtypes.items():
            if st_val['SubtypeField'].casefold() == f.name.casefold():
                self.subtype_desc[st_key] = st_val['Name']
                self.subtype_field = st_val['SubtypeField'].casefold()
            for field_name, (default, domain) in st_val['FieldValues'].items():
                if domain and field_name.casefold() == f.name.casefold():
                    self.domain_desc[st_key] = domain.codedValues
                    self.subtype_field = st_val['SubtypeField'].casefold()
                    break



    def __repr__(self):
        """ Nice representation for debugging  """
        return '<clsfield object name={}, alias={}, domain_desc={}>'.format(
            self.name, self.alias, self.domain_desc)

    def updateValue(self, row, fields, is_xlsx=False):
        """ Update value based on domain description """

        value = row[self.index]
        if is_xlsx:
            value = ooxml_value(value)

        if self.type == "TimestampOffset" and (value is not None):
            value = value.isoformat(timespec='milliseconds')

        # updates based on gdb types and domains
        if self.subtype_field:
            subtype_val = row[fields.index(self.subtype_field)]
        else:
            subtype_val = 0

        if self.subtype_desc:
            # pass if value is <Null>
            try:
                value = self.subtype_desc[row[self.index]]
            except:
                pass

        if self.domain_desc:
            try:
                value = self.domain_desc[subtype_val][row[self.index]]
            except:
                pass  # not all subtypes will have domain

        # Return the validated value
        return value


ESC_REGEX = "_x(?=([0-9A-Fa-f]{4}_))"

def ooxml_value(value):
    """ update individual values before adding to the output excel file """

    if (type(value) is str) and ("_x" in value) :
        #as per OOXML spec _xHHHH_ is an escape sequence
        # https://stackoverflow.com/questions/67614622/writing-to-excel-using-openpyxl-results-in-a-different-character-symbol
        return re.sub(ESC_REGEX, "_x005F_x", value)
    else:
        return value

# create openpyxl style for first row
STYLE0 = openpyxl.styles.NamedStyle(name="STYLE0")
STYLE0.font = openpyxl.styles.Font(bold=True, size=12)
STYLE0.border = openpyxl.styles.Border(
    bottom=openpyxl.styles.Side(style='medium', color="000000"))
STYLE0.alignment = openpyxl.styles.Alignment(horizontal="left")
STYLE0.fill = openpyxl.styles.PatternFill(start_color='e6e6e6',
                                          end_color='e6e6e6',
                                          fill_type='solid')


def WriteOnlyCellEx(ws, value):
    cell = openpyxl.cell.WriteOnlyCell(ws=ws, value=ooxml_value(value))
    cell.style = STYLE0
    return cell


def get_field_defs(in_table, use_domain_desc):
    """ returns nice field definition """
    desc = arcpy.Describe(in_table)

    subtypes = {}
    cvdomains = {}
    if use_domain_desc:
        try:
            subtypes = arcpy.da.ListSubtypes(in_table)
        except:
            pass
        try:
            ws = arcpy.Describe(in_table).workspace.catalogpath
            cvdomains = [i for i in arcpy.da.ListDomains(ws) if (i.domainType == 'CodedValue')]
        except:
            pass

    fields = []
    for i, field in enumerate([f for f in desc.fields
                               if f.type in ["Date", "Double", "Guid",
                                             "Integer", "OID", "Single",
                                             "SmallInteger", "String",
                                             "GlobalID", "BigInteger",
                                             "TimeOnly", "DateOnly",
                                             "TimestampOffset"]]):
        fields.append(clsField(field, i, subtypes, cvdomains))

    return fields


def validate_sheet_name(sheet_name):
    """ Validate sheet name to excel limitations
         - 31 character length
         - there characters not allowed : \\ / ? * [ ]
    """
    import re
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]

    # Replace invalid sheet character names with an underscore
    r = re.compile(r'[:\\\/?*\[\]]')
    sheet_name = r.sub("_", sheet_name)

    return sheet_name


def add_error(id, s=None):
    """ Return errors """

    arcpy.AddIDMessage("ERROR", id, s if s else None)
    if __name__ == '__main__':
        sys.exit(1)
    else:
        raise arcpy.ExecuteError(arcpy.GetIDMessage(id))


def table_to_excel(in_table, output_file, use_field_alias=False,
                   use_domain_desc=False):
    """ Writes a table to an Excel file """

    in_table_list = ConversionUtils.SplitMultiInputs(in_table)

    if os.path.isfile(output_file):
        if not arcpy.env.overwriteOutput:
            add_error(258, output_file)
        else:
            os.remove(output_file)

    output_dir = os.path.dirname(output_file)
    if not os.path.isdir(output_dir):
        arcpy.AddIDMessage("ERROR", 792, output_dir)
        return

    unique_table_names = set()
    def generate_unique_name(table_name):
        nonlocal unique_table_names

        unique_name = arcpy.Describe(table_name).name

        index = 1
        while( unique_name in unique_table_names):
            unique_name = table_name + " (" + str(index) + ")"
            index += 1

        unique_table_names.add(unique_name)
        return unique_name

    if output_file.lower().endswith("xlsx"):
        # use openpyxl to write generate output xlsx file

        workbook = Workbook(write_only=True)

        for input_table in in_table_list:

            fields = get_field_defs(input_table, use_domain_desc)

            worksheet = workbook.create_sheet()
            worksheet.title = validate_sheet_name(generate_unique_name(input_table))

            # set the column names with style
            worksheet.append(
                [WriteOnlyCellEx(worksheet,
                                 (i.alias if use_field_alias else i.name))
                 for i in fields])

            # set column width
            for index, field in enumerate(fields):
                continue  # skip, not supported with WriteOnlyCell
                if field.type == 'String':
                    worksheet.column_dimensions[col_letter].width = min(50, field.length)
                else:
                    worksheet.column_dimensions[col_letter].width = 16

            # Loop through input rows
            field_names = tuple(i.name.casefold() for i in fields)
            with arcpy.da.SearchCursor(input_table, field_names) as cursor:
                for row in cursor:
                    # convert to list which allows item assignment
                    rowUpdated = list(row)

                    for col_index, value in enumerate(row):
                        value = fields[col_index].updateValue(row, field_names, True)
                        rowUpdated[col_index] = value

                    worksheet.append(rowUpdated)

        # save workbook out to file
        workbook.save(output_file)
        del(workbook)

    else:
        # use xlwt to write generate output xls file

        workbook = xlwt.Workbook()

        for input_table in in_table_list:

            fields = get_field_defs(input_table, use_domain_desc)

             # Error if table exceeds the 65535 rows limit of the .xls file format
            if (int(arcpy.GetCount_management(input_table)[0]) > 65535 ):
                add_error(1531)

            # Input table exceeds the 256 columns limit of the .xls file format.
            elif (len(fields) > 255):
                add_error(1530)

            worksheet = workbook.add_sheet(validate_sheet_name(generate_unique_name(input_table)))

            # Add first (header) row
            header_style = xlwt.easyxf(
                "font: bold on; align: horiz center; pattern: pattern solid, fore-colour 0x16;")
            for index, field in enumerate(fields):
                worksheet.write(0, index,
                                field.alias if use_field_alias else field.name,
                                header_style)
                if field.type == 'String':
                    worksheet.col(index).width = min(50, field.length) * 256
                else:
                    worksheet.col(index).width = 16 * 256

            # Freeze panes
            worksheet.set_panes_frozen(True)
            worksheet.set_horz_split_pos(1)
            worksheet.set_remove_splits(True)

            # Set cell format/styles for data types
            styleDefault = xlwt.XFStyle()

            styleDateOnly = xlwt.XFStyle()
            styleDateOnly.num_format_str = 'YYYY-MM-DD'

            styleTimeOnly = xlwt.XFStyle()
            styleTimeOnly.num_format_str = 'h:mm:ss'

            styleDateTime = xlwt.XFStyle()
            styleDateTime.num_format_str = 'YYYY-MM-DD h:mm:ss'

            styleInt = xlwt.XFStyle()
            styleInt.num_format_str = '0'

            # Loop through input records
            field_names = tuple(i.name.casefold() for i in fields)
            with arcpy.da.SearchCursor(input_table, field_names) as cursor:
                row_index = 1
                for row in cursor:
                    for col_index, value in enumerate(row):
                        value = fields[col_index].updateValue(row, field_names)

                        if isinstance(value, datetime.datetime):

                            if (value.hour == 0) and (value.minute == 0):
                                style = styleDateOnly
                            elif (value.year == 1899) and (value.month == 12) and (
                                value.day == 30):
                                style = styleTimeOnly
                                value = (value - datetime.datetime(1899, 12, 30, 0,
                                                                   0,
                                                                   0)).total_seconds() / 86400.0
                            else:
                                style = styleDateTime

                        elif isinstance(value, int):
                            style = styleInt

                        elif isinstance(value, datetime.time):
                            style = styleTimeOnly

                        elif isinstance(value, datetime.date):
                            style = styleDateOnly

                        else:
                            style = styleDefault

                        # write to the cell
                        worksheet.write(row_index, col_index, value, style)
                    row_index += 1

        # save workbook out to file
        workbook.save(output_file)
        del(workbook)

        # fix the WriteAccess record on XLS files
        #  as per Section 2.4.249 in MS-XLS spec
        import struct
        with open(output_file, mode='rb') as f:
            arr = bytearray(f.read())
            update = b'\x0f', b'\x00', b'\x00', b'N', b'o', b'n', b'e'
            struct.pack_into('s'*7, arr, 552, *update)

        with open(output_file, mode='wb') as f:
            f.write(arr)


if __name__ == "__main__":
    table_to_excel(arcpy.GetParameterAsText(0),
                   arcpy.GetParameterAsText(1),
                   arcpy.GetParameter(2),
                   arcpy.GetParameter(3))
