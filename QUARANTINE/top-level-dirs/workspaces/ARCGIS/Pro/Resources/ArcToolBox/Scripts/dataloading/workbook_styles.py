import dataclasses

from openpyxl.cell import Cell
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.styles.colors import Color
from openpyxl.styles.differential import DifferentialStyle


@dataclasses.dataclass
class FormatOptions:
    """Collection of formatting options"""

    font: Font = None
    fill: PatternFill = None
    border: Border = None

    def to_differential(self) -> DifferentialStyle:
        """Creates DifferentialStyle to be used in conditional formatting"""
        return DifferentialStyle(
            font=self.font,
            fill=self.fill,
            border=self.border,
        )

    def apply(self, cell: Cell):
        """Applies formatting to cell"""
        if self.font:
            cell.font = self.font
        if self.fill:
            cell.fill = self.fill
        if self.border:
            cell.border = self.border


TAB_GREEN = Color("FF90D4B8")
TAB_ORANGE = Color("00FFCC99")
TAB_YELLOW = Color("FFFAF5A8")
TAB_BLUE = Color("FFA7E1F3")
TAB_WHITE = Color("FFFFFFFF")

white_bold = Font(color=Color("00FFFFFF"), bold=True)
black_bold = Font(color=Color("00000000"), bold=True)

HYPERLINK = FormatOptions(
    font=Font(underline="single", color=Color("0000FF")),
)

NON_HYPERLINK = FormatOptions(font=Font(underline=None, color=Color("000000")))

DUPLICATE_TARGET_FIELD = FormatOptions(
    font=white_bold,
    fill=PatternFill(bgColor=Color("FF2B83BA"), fgColor=Color("FF2B83BA"), fill_type="solid"),  # dark blue
)

EXPRESSION_LOOKUP = FormatOptions(
    font=black_bold,
    fill=PatternFill(bgColor=Color("FFABD9E9"), fgColor=Color("FFABD9E9"), fill_type="solid"),  # light blue
)

LOOKUP_REQUIRED = FormatOptions(
    font=black_bold,
    border=Border(left=(side := Side(style="thick", color=Color("FFFF0000"))), right=side, top=side, bottom=side),
)  # thick bright red

FIELD_ERROR = FormatOptions(
    font=white_bold,
    fill=PatternFill(bgColor=Color("FFD7191C"), fgColor=Color("FFD7191C"), fill_type="solid"),  # deep red
)

FIELD_WARNING = FormatOptions(
    font=black_bold,
    fill=PatternFill(bgColor=Color("FFFDAE61"), fgColor=Color("FFFDAE61"), fill_type="solid"),  # light orange
)

del side
