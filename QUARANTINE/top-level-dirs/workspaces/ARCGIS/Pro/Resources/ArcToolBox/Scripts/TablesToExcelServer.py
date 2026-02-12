"""
Tool name: TablesToExcelServer
Source: TablesToExcelServer.py
Author: ESRI

This function takes multiple tables and writes them to an output excel file.
Each table will become a worksheet in the excel file.
The function returns the excel file path.

The first parameter is a json string with this syntax:

  {"tables" : [ <table1>, <table2>,...] }

where each table object is:

  { "path": <table path>", "targetName": <targetName>, "where": <where clause> , "includeGeometry" : true/false, "xColumnName": <xname>, "yColumnName" : <yname> }

The path property is required. The others are optional.
The targetName will be used as the worksheet name. Otherwise the name of the table will be used.
"""

import json
import arcpy
import os
import sys
from openpyxl import Workbook


def add_error(id, s=None):
    """ Return errors """

    arcpy.AddIDMessage("ERROR", id, s if s else None)
    if __name__ == '__main__':
        sys.exit(1)
    else:
        raise arcpy.ExecuteError(arcpy.GetIDMessage(id))


def validate_sheet_name(sheet_name):
    """ Validate sheet name to excel limitations
         - 31 character length
         - there characters not allowed : \ / ? * [ ]
    """
    import re
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]

    # Replace invalid sheet character names with an underscore
    r = re.compile(r'[:\\\/?*\[\]]')
    sheet_name = r.sub("_", sheet_name)

    return sheet_name

def generate_unique_name(unique_names, name):

    unique_name = name

    index = 1
    while(unique_name in unique_names):
        unique_name = name + " (" + str(index) + ")"
        index += 1

    unique_names.add(unique_name)
    return unique_name

def get_fields(table_desc, sub_fields, include_geometry):
    """"Returns a list of field names, and the index of the X and Y fields in the list
        If table is point feature class, the last two will be SHAPE@X, SHAPE@Y
    """

    include_shape = include_geometry and table_desc.datasetType == 'FeatureClass' and table_desc.shapeType == 'Point'

    types_valid_excel = ["Date", "Double", "Guid",
                         "Integer", "OID", "Single",
                         "SmallInteger", "String",
                         "GlobalID", "BigInteger",
                         "TimeOnly", "DateOnly"]

    eval_sub_fields = sub_fields is not None and len(sub_fields) > 0

    shape_field = None
    fields = []
    for f in table_desc.fields:
        if include_shape and f.type == "Geometry":
            shape_field = f
        elif f.type in types_valid_excel:
            if eval_sub_fields:
                if f.name in sub_fields:
                    fields.append(f.name)
            else:
                fields.append(f.name)

    x_index = -1
    y_index = -1
    if shape_field is not None:
        # these field names have special meaning for arcpy: "include the point's x/y coord in the search cursor"
        fields.append("SHAPE@X")
        fields.append("SHAPE@Y")
        x_index = fields.index("SHAPE@X")
        y_index = fields.index("SHAPE@Y")

    return fields, x_index, y_index


def table_to_sheet(table_spec, sr, sheet):

    table_path = table_spec["path"]
    where_clause = table_spec.get("where", "")
    include_geometry = table_spec.get("includeGeometry", True)
    x_column_name = table_spec.get("xColumnName", "x")
    y_column_name = table_spec.get("yColumnName", "y")
    sub_fields = table_spec.get("subfields")

    if int(arcpy.GetCount_management(table_path)[0]) > 1048576:
        add_error(1531)   # Input table exceeds the 65535 rows limit of the .xlsx file format.

    table_desc = arcpy.Describe(table_path)

    field_names, x_index, y_index = get_fields(table_desc, sub_fields, include_geometry)

    # note: excel row and columns 'indices' start counting from 1 (ugh)

    # write first excel row with field names
    for field_index, field_name in enumerate(field_names):
        if field_index == x_index:
            sheet.cell(row=1, column=field_index + 1, value=x_column_name)
        elif field_index == y_index:
            sheet.cell(row=1, column=field_index + 1, value=y_column_name)
        else:
            sheet.cell(row=1, column=field_index + 1, value=field_name)

    # write the rows with the values
    with arcpy.da.SearchCursor(table_path, field_names, where_clause, sr) as cursor:        
        for row in cursor:
            row_as_list = list(row)
            sheet.append(row_as_list)
    
def tables_to_excel(tables, target_srid, excel_file):
    if os.path.isfile(excel_file):
        if not arcpy.env.overwriteOutput:
            add_error(258, excel_file)
        else:
            os.remove(excel_file)
            
    output_dir = os.path.dirname(excel_file)
    if not os.path.isdir(output_dir):
        arcpy.AddIDMessage("ERROR", 792, output_dir)
        return
        
    sr = None
    if target_srid != "":
        sr = arcpy.SpatialReference(int(target_srid))

    wb = Workbook()
    default_sheet = wb.active

    spec = json.loads(tables)

    sheet_names = set()

    for table_spec in spec["tables"]:
        table_path = table_spec["path"]
        target_name = table_spec.get("targetName", os.path.basename(table_path))
        sheet_name = validate_sheet_name(target_name)
        sheet_name = generate_unique_name(sheet_names, sheet_name)
        sheet = wb.create_sheet(sheet_name)
        table_to_sheet(table_spec, sr, sheet)

    wb.remove(default_sheet)
    wb.save(excel_file)
    arcpy.SetParameterAsText(2, excel_file)


def main():
    tables_spec = arcpy.GetParameterAsText(0)
    target_srid = arcpy.GetParameterAsText(1)
    excel_file = arcpy.GetParameterAsText(2)
    tables_to_excel(tables_spec, target_srid, excel_file)

    # table1 = { "path": r"F:\FastLocalData\fgdb\redlandsFewFeats.gdb\RedlandsFewFeats\ParcelsFew", "where": "", "includeGeometry" : False }
    # table2 = { "path": r"F:\FastLocalData\fgdb\redlandsFewFeats.gdb\RedlandsFewFeats\TreesFew",   "where": "OBJECTID >= 38" , "includeGeometry" : True, "xColumnName": "x_1"  }
    # tables = { "tables" : [table1, table2] }
    # param1 = json.dumps(tables)
    # target_srid = "4326"
    # excel_file = r"F:\FastLocalData\temp\foo.xlsx"

    # table0 = { "path": r"F:\FastLocalData\fgdb\USA_64_Test.gdb\BigIntegerTesting\Cities64", "where": "CAPITAL = 'Y'", "subfields": ["CITY_NAME","STATE_NAME","CAPITAL"]}
    # table1 = { "path": r"F:\FastLocalData\fgdb\USA_64_Test.gdb\BigIntegerTesting\Cities64", "where": "", "includeGeometry" : False }
    # table2 = { "path": r"F:\FastLocalData\fgdb\USA_64_Test.gdb\PointNewDateFieldTypes", "includeGeometry" : True, "xColumnName": "x_1"  }
    # tables = { "tables" : [table0, table1, table2] }
    # param1 = json.dumps(tables)
    # target_srid = "4326"
    # excel_file = r"F:\FastLocalData\temp\fieldTypes.xlsx"
    #
    # tables_to_excel(param1, target_srid, excel_file)

if __name__ == "__main__":
    main()


