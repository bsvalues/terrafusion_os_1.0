# -*- coding: utf-8 -*-

"""
------------------------------------------------------------------------------
BatchImport.py
------------------------------------------------------------------------------
requirements: ArcGIS 2.5, Python 3.6
author: ArcGIS Solutions for Intelligence
contact: intelsolutions@esri.com
company: Esri
------------------------------------------------------------------------------
* 01/16/2018 - elinz - original writeup
* 03/13/2018 - elinz - added _add_to_map function to add converted dataset
                       to the active map
* 03/16/2018 - elinz - modified _add_to_map function to dynamically build
                       group lyrx file path
* 03/23/2018 - elinz - fixed issue where KML converter failed if output
                       raster folder existed
* 03/29/2018 - elinz - changed folder suffix where raster files are extracted
                     - implemented arcpy.ValidateTableName function
                     - added support for .lyr files; added _get_layer_file
                       function
                     - added _layer_time_stamp property and added layer_time_stamp
                       parameter to _add_to_map function
* 05/31/2018 - elinz - added import support for excel worksheets and text files
* 06/26/2018 - elinz - modified ExcelWorkbookDataConverter class to import Excel
                       worksheet into scratch gdb table using "Excel To Table"
                       tool and using that scratch table as an input source
                       into "XY Table To Point" tool; achieve better results
                       using this methodology then "feeding" Excel worksheet
                       directly into "XY Table To Point" tool
* 09/13/2018 - elinz - modified to allow user to "override" KML layer symbology.
* 10/18/2018 - elinz - modified code to handle cases where a map is not available.
* 10/31/2018 - elinz - added import support for GeoJSON files.
* 04/15/2019 - elinz - added support for adding spatial index
* 04/18/2019 - elinz - added support to filter input data
* 07/16/2019 - elinz - added import support for GPX files.
* 2019-09-06 - mfunk - update lib references and module name for Pro integration
* 2019-09-17 - mfunk - move GDAL import into GeoJSONDataConverter
* 2019-09-17 - mfunk - ID messaging updates
* 2019-10-14 - mfunk - move xlrd, fnmatch imports into method
* 2020-01-13 - mfunk - module name change for 'intel'
* 2020-08-03 - elinz - add new parameter for kml ground overlay
* 2020-12-03 - jjones - reorganized into intel Subfolders, fixed relative imports
* 2020-12-29 - elinz - modified find_source_files method to handle files
* 2021-01-19 - elinz - added support for mgrs, usng, and lat/long in single field
* 2021-01-28 - elinz - Fixed issue with filter parameter value raising error
*                      when pattern doesn't include file extension; fixed issue
*                      where files where not found when filter pattern didn't
*                      include file extension
* 2021-08-05 - elinz - Updated to use openpyxl library to access Excel .xlsx files
*                      because xlrd library has removed support for .xlsx files
*                      at release 2.0.1 of library
* 2022-11-01 - mfunk - Intel 3257- Replace FeatureClassToFeatureClass with 
*                      ExportFeatures
------------------------------------------------------------------------------
"""

import os
import sys
import traceback
import datetime
import arcpy
from intel.enumerations import ConvertCoordNotation as ccn
from intel.utilities import DEBUG
import xlrd
import openpyxl

class DataConverterBase(object):
    """ Data Converter Base Class. """

    _layer_time_stamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Valid coordinate fields and priority order for data formats 
    # that don't have "built-in" spatial info
    _latitude_fields = [
        'latitude', 
        'lat', 
        'y_coordinate', 
        'y_coord', 
        'y'
        ]
    _longitude_fields = [
        'longitude', 
        'long', 
        'lon', 
        'x_coordinate', 
        'x_coord', 
        'x'
        ]
    _mgrs_fields = [
        'mgrs', 
        'mgrs_coordinate',
        'mgrs_coordinates', 
        'mgrs_coord',
        'mgrs_coords'
        ]
    _usng_fields = [
        'usng', 
        'usng_coordinate',
        'usng_coordinates', 
        'usng_coord',
        'usng_coords'
        ] 
    _latitude_longitude_fields = [
        'latitude_longitude', 
        'lat_long', 
        'lat_lon', 
        'xy_coordinate', 
        'xy_coordinates', 
        'xy_coord', 
        'xy_coords', 
        'xy'
        ]

    _source_file_extensions = [
        '.kml', 
        '.kmz', 
        '.shp', 
        '.xlsx', 
        '.xls', 
        '.csv', 
        '.txt', 
        '.tab', 
        '.geojson', 
        '.gpx'
        ]

    _verbose_messaging = False

    @property
    def layer_time_stamp(self):
        return self._layer_time_stamp

    @layer_time_stamp.setter
    def layer_time_stamp(self, value):
        self._layer_time_stamp = value

    @property
    def _valid_latitude_fields(self):
        return self._latitude_fields

    @property
    def _valid_longitude_fields(self):
        return self._longitude_fields

    @property
    def valid_source_file_extensions(self):
        return self._source_file_extensions

    @property
    def verbose_messaging(self):
        return self._verbose_messaging

    @verbose_messaging.setter
    def verbose_messaging(self, value):
        self._verbose_messaging = value

    @staticmethod
    def normalize_data_filter(data_filter):
        """ Returns 'cleaned' data filter """

        # Remove leading and trailing spaces
        cleaned_filter = [i.strip() for i in data_filter]

        # Remove trailing periods
        cleaned_filter = [i.rstrip('.') for i in cleaned_filter]

        return cleaned_filter

    @staticmethod
    def validate_data_filter_extensions(data_filter):
        """ Checks data_filter items for unsupported file extensions """

        valid = True
        msg = None

        if data_filter is None:
            return valid, msg

        invalid_extensions = []

        data_filter = DataConverterBase.normalize_data_filter(data_filter)

        for pattern in data_filter:
            pattern_name, pattern_ext = os.path.splitext(pattern)
            # Validate only if pattern has a file extension - Issue #2443
            if len(pattern_ext) > 0:
                if pattern_ext.lower() not in DataConverterBase._source_file_extensions:
                    valid = False
                    invalid_extensions.append(pattern_ext)

        if not valid:
            err_msg = arcpy.GetIDMessage(190175)
            msg = err_msg.format(', '.join(invalid_extensions), ', '.join(DataConverterBase._source_file_extensions))

        return valid, msg

    @staticmethod
    def _create_unique_name_in_workspace(base_name, workspace):
        """ Creates a unique name in the specified workspace by appending a number
        to the base name. The number is increased until the name is unique. """

        unique_name = base_name
        i = 1
        while arcpy.Exists(os.path.join(workspace, unique_name)):
            unique_name = '{}_{}'.format(base_name, i)
            i = i + 1
        else:
            return unique_name

    @staticmethod
    def _create_unique_name_in_list(base_name, name_list):
        """ Creates a unique name in the specified list by appending a number
        to the base name. The number is increased until the name is unique. """

        unique_name = base_name
        test_name_list = [name.lower() for name in name_list]
        i = 1
        while unique_name.lower() in test_name_list:
            unique_name = '{}_{}'.format(base_name, i)
            i = i + 1
        else:
            return unique_name

    @staticmethod
    def _add_group_layer(group_layer_name, output_map):

        if output_map is None:
            return None

        map_group_layer = None
        add_group_layer_position = 'TOP'

        try:

            # Check if group layer already exists in the map
            layers = output_map.listLayers()
            if layers:
                for layer in layers:
                    if layer.name == group_layer_name:
                        map_group_layer = layer

            # If group layer doesn't exist in map, create a layer file and add to map
            if map_group_layer is None:

                # Empty group layer CIM represented as string
                cim_json_str = (r'{"type" : "CIMLayerDocument", "version" : "2.4.0", "build" : 19948, '
                    r'"layers" : ["CIMPATH=map/new_group_layer.xml"], "layerDefinitions" : [{"type" : "CIMGroupLayer", '
                    r'"name" : "New Group Layer", "uRI" : "CIMPATH=map/new_group_layer.xml", '
                    r'"sourceModifiedTime" : {"type" : "TimeInstant"}, "metadataURI" : "CIMPATH=Metadata/f58db033472b3500f9cdbb94a58ac5ba.xml", '
                    r'"useSourceMetadata" : true, "description" : "New Group Layer", "layerElevation" : {"type" : "CIMLayerElevationSurface", '
                    r'"mapElevationID" : "{4A509A8E-9F0F-4839-B584-96190DF51B04}"}, '
                    r'"layerType" : "Operational", "showLegends" : true, "visibility" : true, "displayCacheType" : "Permanent", '
                    r'"maxDisplayCacheAge" : 5, "showPopups" : true, "serviceLayerID" : -1, "refreshRate" : -1, "refreshRateUnit" : "esriTimeUnitsSeconds"}], '
                    r'"binaryReferences" : [{"type" : "CIMBinaryReference", "uRI" : "CIMPATH=Metadata/f58db033472b3500f9cdbb94a58ac5ba.xml", '
                    r'"data" : "<?xml version=\"1.0\"?>\r\n<metadata xml:lang=\"en\"><Esri><CreaDate>20191001</CreaDate>'
                    r'<CreaTime>16284000</CreaTime><ArcGISFormat>1.0</ArcGISFormat><SyncOnce>TRUE</SyncOnce></Esri>'
                    r'<dataIdInfo><idCitation><resTitle>New Group Layer</resTitle></idCitation><idAbs>New Group Layer</idAbs><idCredit></idCredit><idPurp></idPurp>'
                    r'<resConst><Consts><useLimit></useLimit></Consts></resConst></dataIdInfo></metadata>\r\n"}], '
                    r'"elevationSurfaces" : [{"type" : "CIMMapElevationSurface", "elevationMode" : "BaseGlobeSurface", "name" : "Ground", "verticalExaggeration" : 1, '
                    r'"mapElevationID" : "{4A509A8E-9F0F-4839-B584-96190DF51B04}", "color" : {"type" : "CIMRGBColor", "values" : [255, 255, 255, 100]}, '
                    r'"surfaceTINShadingMode" : "Smooth", "visibility" : true, "expanded" : true}]}')

                # Write group layer CIM to layer file on disk
                group_layer_path = os.path.join(arcpy.env.scratchFolder, 'BatchImportData_Group_Layer.lyrx')
                if os.path.exists(group_layer_path):
                    os.remove(group_layer_path)
                f = open(group_layer_path, 'w')
                f.write(cim_json_str)
                f.close()

                if not os.path.exists(group_layer_path):
                    # arcpy.AddMessage('ERROR: file {} does not exist.'.format(group_layer_path))
                    arcpy.AddMessage(arcpy.GetIDMessage(190176).format(group_layer_path))
                else:
                    group_layer = arcpy.mp.LayerFile(group_layer_path).listLayers()[0]
                    group_layer.name = group_layer_name
                    output_map.addLayer(group_layer, add_group_layer_position)
                    map_group_layer = output_map.listLayers(group_layer.name)[0]

        except Exception:
            tb = sys.exc_info()[2]
            tbinfo = traceback.format_tb(tb)[0]
            pymsg = '{}\n{}\n{}'.format(tbinfo,
                                        str(sys.exc_info()[1]),
                                        arcpy.GetMessages(2))
            arcpy.AddError(pymsg)

        finally:
            # Delete temp layer file
            try:
                arcpy.Delete_management(group_layer_path)
            except Exception:
                pass

            return map_group_layer

    @staticmethod
    def _add_to_map(input_dataset_path, output_map, output_layer_name=None, symbology_layer=None,
                    layer_time_stamp=None):

        if output_map is None:
            return

        add_layer_position = 'AUTO_ARRANGE'
        makelayer_results = None
        applysym_results = None
        display_layer = None

        # map_group_layer = DataConverterBase._add_group_layer('Batch Import {}'.format(layer_time_stamp), output_map)
        map_group_layer = DataConverterBase._add_group_layer(arcpy.GetIDMessage(190177).format(layer_time_stamp), output_map)

        # Unless explict layer name is given, set layer name the same as input dataset
        if output_layer_name is None:
            output_layer_name = os.path.basename(input_dataset_path)

        desc = arcpy.Describe(input_dataset_path)

        if hasattr(desc, "dataType"):
            input_data_type = desc.dataType
        else:
            return

        # Create the layer for the input dataset and apply symbology if appropriate for input data type
        if input_data_type == 'FeatureClass':
            makelayer_results = arcpy.MakeFeatureLayer_management(
                                                    input_dataset_path,
                                                    output_layer_name).getOutput(0)

            if symbology_layer:
                applysym_results = arcpy.ApplySymbologyFromLayer_management(
                                                         makelayer_results,
                                                         symbology_layer).getOutput(0)

        if input_data_type == 'MosaicDataset':
            makelayer_results = arcpy.MakeMosaicLayer_management(
                                                    input_dataset_path,
                                                    output_layer_name).getOutput(0)

        if applysym_results:
            display_layer = applysym_results
        elif makelayer_results:
            display_layer = makelayer_results

        if display_layer:
            output_map.addLayerToGroup(map_group_layer, display_layer, add_layer_position)

    @staticmethod
    def _get_layer_file(base_name, workspace):
        """ Returns layer file with path for the specified base_name in the
        workspace if it exists. """

        lyr_file = None
        lyr_file_extensions = ['.lyrx', '.lyr']
        
        # Perform testing on inputs
        if not base_name:
            return lyr_file
        if not workspace:
            return lyr_file
        if not os.path.exists(workspace):
            return lyr_file
        
        # Convert 'workspace' parameter to directory in case
        # it was was passed in as a file
        if os.path.isfile(workspace):
            workspace = os.path.dirname(workspace)
        
        # Remove file extension from 'base_name' parameter
        base_name = os.path.splitext(base_name)[0]
        
        # Determine if a layer file exists
        for lyr_file_extension in lyr_file_extensions:
            f = os.path.join(workspace, '{}{}'.format(base_name, lyr_file_extension))
            if os.path.exists(f):
                lyr_file = f
                break
        
        return lyr_file
    
    @staticmethod
    def _get_latitude_field(table):
        lat_field = None
        fields = arcpy.ListFields(table)
        for field in fields:
            if field.name.lower() in DataConverterBase._latitude_fields:
                lat_field = field
                break
        return lat_field
    
    @staticmethod
    def _get_latitude_field_name(table):
        lat_field = DataConverterBase._get_latitude_field(table)
        if lat_field:
            return lat_field.name
        else:
            return None
    
    @staticmethod
    def _get_longitude_field(table):
        long_field = None
        fields = arcpy.ListFields(table)
        for field in fields:
            if field.name.lower() in DataConverterBase._longitude_fields:
                long_field = field
                break
        return long_field
    
    @staticmethod
    def _get_longitude_field_name(table):
        long_field = DataConverterBase._get_longitude_field(table)
        if long_field:
            return long_field.name
        else:
            return None

    @staticmethod
    def _get_mgrs_field(table):
        mgrs_field = None
        fields = arcpy.ListFields(table)
        for field in fields:
            if field.name.lower() in DataConverterBase._mgrs_fields:
                mgrs_field = field
                break
        return mgrs_field

    @staticmethod
    def _get_mgrs_field_name(table):
        mgrs_field = DataConverterBase._get_mgrs_field(table)
        if mgrs_field:
            return mgrs_field.name
        else:
            return None

    @staticmethod
    def _get_usng_field(table):
        usng_field = None
        fields = arcpy.ListFields(table)
        for field in fields:
            if field.name.lower() in DataConverterBase._usng_fields:
                usng_field = field
                break
        return usng_field

    @staticmethod
    def _get_usng_field_name(table):
        usng_field = DataConverterBase._get_usng_field(table)
        if usng_field:
            return usng_field.name
        else:
            return None

    @staticmethod
    def _get_latitude_longitude_field(table):
        lat_long_field = None
        fields = arcpy.ListFields(table)
        for field in fields:
            if field.name.lower() in DataConverterBase._latitude_longitude_fields:
                lat_long_field = field
                break
        return lat_long_field

    @staticmethod
    def _get_latitude_longitude_field_name(table):
        lat_long_field = DataConverterBase._get_latitude_longitude_field(table)
        if lat_long_field:
            return lat_long_field.name
        else:
            return None

    @staticmethod
    def _get_field_names(table):
        fields = arcpy.ListFields(table)
        field_names = [field.name for field in fields]
        return field_names

    @staticmethod
    def _get_convert_coordinate_params(table):
        """ Returns dictionary of coordinate related parameters 
        for Convert Coordinate Notation gp tool. """

        convert_coord = {}

        # Get all valid coordinate fields in the table
        lat_fld_name = DataConverterBase._get_latitude_field_name(table)
        long_fld_name = DataConverterBase._get_longitude_field_name(table)
        latlong_fld_name = DataConverterBase._get_latitude_longitude_field_name(table)
        mgrs_fld_name = DataConverterBase._get_mgrs_field_name(table)
        usng_fld_name = DataConverterBase._get_usng_field_name(table)

        # Determine which coordinate field(s) to use for import.
        # Precedence is: latitude and longitude in separate fields,
        # latitude and longitude in a single field, MGRS, and then USNG
        if lat_fld_name and long_fld_name:
            # latitude and longitude in separate fields
            convert_coord[ccn.INPUT_COORD_FORMAT.value] = ccn.INPUT_COORD_FORMAT_DD_2.value
            convert_coord[ccn.X_FIELD.value] = long_fld_name
            convert_coord[ccn.Y_FIELD.value] = lat_fld_name
            convert_coord[ccn.OUTPUT_COORD_FORMAT.value] = ccn.OUTPUT_COORD_FORMAT_DD_2.value
            if DEBUG:
                arcpy.AddMessage(f'convert_coord: {convert_coord}')
            return convert_coord
        elif latlong_fld_name:
            # latitude and longitude in a single field
            convert_coord[ccn.INPUT_COORD_FORMAT.value] = ccn.INPUT_COORD_FORMAT_DD_1.value
            convert_coord[ccn.X_FIELD.value] = latlong_fld_name
            convert_coord[ccn.Y_FIELD.value] = None
            convert_coord[ccn.OUTPUT_COORD_FORMAT.value] = ccn.OUTPUT_COORD_FORMAT_DD_2.value
            if DEBUG:
                arcpy.AddMessage(f'convert_coord: {convert_coord}')
            return convert_coord
        elif mgrs_fld_name:
            # MGRS
            convert_coord[ccn.INPUT_COORD_FORMAT.value] = ccn.INPUT_COORD_FORMAT_MGRS.value
            convert_coord[ccn.X_FIELD.value] = mgrs_fld_name
            convert_coord[ccn.Y_FIELD.value] = None
            convert_coord[ccn.OUTPUT_COORD_FORMAT.value] = ccn.OUTPUT_COORD_FORMAT_DD_2.value
            if DEBUG:
                arcpy.AddMessage(f'convert_coord: {convert_coord}')
            return convert_coord
        elif usng_fld_name:
            # USNG
            convert_coord[ccn.INPUT_COORD_FORMAT.value] = ccn.INPUT_COORD_FORMAT_USNG.value
            convert_coord[ccn.X_FIELD.value] = usng_fld_name
            convert_coord[ccn.Y_FIELD.value] = None
            convert_coord[ccn.OUTPUT_COORD_FORMAT.value] = ccn.OUTPUT_COORD_FORMAT_DD_2.value
            if DEBUG:
                arcpy.AddMessage(f'convert_coord: {convert_coord}')
            return convert_coord
        else:
            # There are no valid coordinate fields in the table
            if DEBUG:
                arcpy.AddMessage(f'convert_coord: None')
            return None

        return

    @staticmethod
    def _get_dataset_geometry_type_suffix_name(esri_geometry_type):
        """ Returns geometry suffix to append to dataset name. """

        # NOTE: input parameter value corresponds to geometry type parameter values of 'Create Feature Class' gp tool

        suffix = None
        if esri_geometry_type.upper() == 'POINT':
            suffix = 'POINTS'
        elif esri_geometry_type.upper() == 'MULTIPOINT':
            suffix = 'MULTIPOINT'
        elif esri_geometry_type.upper() == 'POLYGON':
            suffix = 'POLYGONS'
        elif esri_geometry_type.upper() == 'POLYLINE':
            suffix = 'POLYLINES'
        elif esri_geometry_type.upper() == 'MULTIPATCH':
            suffix = 'MULTIPATCH'

        return suffix

    @staticmethod
    def _replicate_patterns_to_include_extensions(patterns):
        """ Replicate each pattern to include valid file extensions """
        replicated_patterns = []

        if not patterns:
            return patterns
        
        for pattern in patterns:
            pattern_name, pattern_ext = os.path.splitext(pattern)
            if len(pattern_ext) > 0:
                # Pattern does not include file extension so 
                # add unaltered original pattern to return list 
                replicated_patterns.append(pattern)
            else:
                # Pattern does not include file extension so
                # replicate pattern for each valid file extension
                # and append file extension
                new_patterns = [i.replace('.', f'{pattern_name}.').lower() for i in DataConverterBase._source_file_extensions]
                replicated_patterns.extend(new_patterns)

        return replicated_patterns

    @staticmethod
    def find_source_files(data_paths, recursive=True, patterns=None):
        """ Create list of files to import. """
        import fnmatch
        source_files = []

        if patterns is None:
            # Create default file extension patterns if necessary
            updated_patterns = [i.replace('.', '*.').lower() for i in DataConverterBase._source_file_extensions]
        else:
            # Update patterns to include file extensions
            updated_patterns = DataConverterBase._replicate_patterns_to_include_extensions(patterns)

        for data_path in data_paths:
            if os.path.isfile(data_path):
                source_files.append(data_path)
            else:
                for dir_path, dir_names, file_names in os.walk(data_path):
                    for file_name in file_names:
                        # Issue 806: skip temp files
                        if os.path.basename(file_name)[:1] in ["~", "$"]:
                            continue

                        file_path = os.path.join(dir_path, file_name)

                        for pattern in updated_patterns:
                            if fnmatch.fnmatch(file_path, pattern):
                                source_files.append(file_path)

                    if not recursive:
                        break

        return source_files
    
    def _add_spatial_index(self, features):
        """ Add spatial index to a shapefile, file geodatabase, or enterprise geodatabase feature class. """

        try:
            if arcpy.Exists(features):
                if self._verbose_messaging:
                    arcpy.AddMessage(arcpy.GetIDMessage(190178).format(features))
                arcpy.AddSpatialIndex_management(features)
                if self._verbose_messaging:
                    arcpy.AddMessage(arcpy.GetMessages())
        except Exception:
            tb = sys.exc_info()[2]
            tbinfo = traceback.format_tb(tb)[0]
            pymsg = '{}\n{}\n{}\n{}'.format(self.__class__.__name__,
                                            tbinfo,
                                            str(sys.exc_info()[1]),
                                            arcpy.GetMessages(2))
            arcpy.AddError(pymsg)


class KMLDataConverter(DataConverterBase):
    """ Convert an input KML/KMZ file to a geodatabase feature class. """

    def __init__(self, input_file, output_gdb, output_map, include_groundoverlay):
        self._input_file = input_file
        self._output_gdb = output_gdb
        self._output_map = output_map
        self._include_groundoverlay = include_groundoverlay

    @property
    def input_file(self):
        return self._input_file

    @input_file.setter
    def input_file(self, value):
        self._input_file = value

    @property
    def output_gdb(self):
        return self._output_gdb

    @output_gdb.setter
    def output_gdb(self, value):
        self._output_gdb = value

    @property
    def output_map(self):
        return self._output_map

    @output_map.setter
    def output_map(self, value):
        self._output_map = value

    @property
    def include_groundoverlay(self):
        return self._include_groundoverlay

    @include_groundoverlay.setter
    def include_groundoverlay(self, value):
        self._include_groundoverlay = value

    def convert(self):
        """ Convert an input KML/KMZ file to a geodatabase feature class. """

        feature_classes = None
        fc = None
        mosaic_datasets = None
        md = None

        # Store current Add Outputs to Map parameter
        initial_add_output_to_map = arcpy.env.addOutputsToMap
        arcpy.env.addOutputsToMap = False

        # Generate folder name/path to temporarily store KML/KMZ output
        temp_output_folder = '{}_{}'.format(os.path.join(arcpy.env.scratchFolder, 'KMLDataConverter'),
                                            '{0:%Y%m%d%H%M%S}'.format(datetime.datetime.now()))

        try:

            # Root folder where KML/KMZ Ground Overlay rasters will be stored
            target_groundoverlay_folder = '{}{}'.format(
                os.path.join(os.path.dirname(self._output_gdb), os.path.splitext(self._output_gdb)[0]), '_Rasters')
            
            # Create the temporary folder if it does not exist
            if not os.path.exists(temp_output_folder):
                os.makedirs(temp_output_folder)
            
            # Convert the KML/KMZ file to feature class in temporary workspace
            temp_file_gdb = os.path.join(temp_output_folder,
                                         os.path.basename(os.path.splitext(self._input_file)[0]) + '.gdb')
            if self._verbose_messaging:
                    arcpy.AddMessage(arcpy.GetIDMessage(190179).format(temp_file_gdb))
            arcpy.KMLToLayer_conversion(self._input_file, temp_output_folder, '#', self._include_groundoverlay)
            if self._verbose_messaging:
                arcpy.AddMessage(arcpy.GetMessages())
            
            # Remove the layer that was just added to the map; have to do this because of a bug
            # in the KMLToLayer tool where it doesn't obey the env.addOutputsToMap property
            if self._output_map:
                layer_name = os.path.basename(os.path.splitext(temp_file_gdb)[0])
                lyr_list = self._output_map.listLayers()
                for lyr in lyr_list:
                    if lyr.supports("NAME"):
                        if lyr.name == layer_name:
                            if self._verbose_messaging:
                                arcpy.AddMessage(arcpy.GetIDMessage(190180).format(layer_name))
                            self._output_map.removeLayer(lyr)

            # Get the layer file created from KMLToLayer gp tool
            layer_file = self._get_layer_file(
                os.path.basename(os.path.splitext(temp_file_gdb)[0]), os.path.dirname(temp_file_gdb))

            # For every Featureclass inside, copy it to the target geodatabase and use the name from the temporary fGDB
            arcpy.env.workspace = temp_file_gdb
            feature_classes = arcpy.ListFeatureClasses('*', '', 'Placemarks')
            if feature_classes:
                for fc in feature_classes:

                    fc_copy = os.path.join(temp_file_gdb, 'Placemarks', fc)
                    target_dataset_name = temp_file_gdb[temp_file_gdb.rfind(os.sep)+1:-4] + '_' + fc
                    target_dataset_name = arcpy.ValidateTableName(target_dataset_name, self._output_gdb)
                    target_dataset_name = self._create_unique_name_in_workspace(target_dataset_name, self._output_gdb)
                    if self._verbose_messaging:
                        arcpy.AddMessage(arcpy.GetIDMessage(190181).format(fc_copy,
                                                                           os.path.join(self._output_gdb,
                                                                                        target_dataset_name)))
                    arcpy.conversion.ExportFeatures(fc_copy, os.path.join(self._output_gdb, target_dataset_name))
                    if self._verbose_messaging:
                        arcpy.AddMessage(arcpy.GetMessages())

                    # Add spatial index
                    # self._add_spatial_index(os.path.join(self._output_gdb, target_dataset_name))

                    layer_name = '{} {}'.format(target_dataset_name, self._layer_time_stamp)

                    # Check if symbology is being overridden (i.e. user has created a lyrx/lyr file at same location
                    # of input file KML (File Name_Points, File Name_Polygons, or File Name_Polylines), otherwise
                    # use the layer file generated by the KML to Layer gp tool.
                    override_layer_file = self._get_layer_file(
                        os.path.basename(os.path.splitext(self._input_file)[0]) + '_' + fc,
                        os.path.dirname(self._input_file))

                    symbology_source = None
                    if override_layer_file is not None:
                        symbology_source = override_layer_file
                    else:
                        # Extract the layer from the group layer in the layer file associated with the input dataset
                        kml_sym_layers = arcpy.mp.LayerFile(layer_file).listLayers()
                        kml_layer_prefix = kml_sym_layers[0]
                        kml_layer_prefix = arcpy.ValidateTableName(kml_layer_prefix, self._output_gdb)
                        for lyr in kml_sym_layers[1:]:
                            kml_layer_name = '{}_{}'.format(kml_layer_prefix, lyr.name)
                            if kml_layer_name in target_dataset_name:
                                symbology_source = lyr

                    self._add_to_map(os.path.join(self._output_gdb, target_dataset_name),
                                     self._output_map, layer_name, symbology_source, self._layer_time_stamp)
            
            # For every Mosaic Dataset that exists in the temporary geodatabase, copy to the target geodatabase
            arcpy.env.workspace = temp_file_gdb
            mosaic_datasets = arcpy.ListDatasets('*', 'Mosaic')
            if mosaic_datasets:
                # Create the root folder to store KML Ground Overlayer rasters
                if not os.path.exists(target_groundoverlay_folder):
                    os.makedirs(target_groundoverlay_folder)
                    
                for md in mosaic_datasets:
                    
                    # Get list of all datasets in output gdb and output raster
                    # folder and determine unique target dataset name across
                    # all lists
                    target_md_name = temp_file_gdb[temp_file_gdb.rfind(os.sep)+1:-4] + '_' + md
                    target_md_name = arcpy.ValidateTableName(target_md_name, self._output_gdb)
                    arcpy.env.workspace = self._output_gdb
                    name_list = set(arcpy.ListDatasets('*') + os.listdir(target_groundoverlay_folder))
                    target_md_name = self._create_unique_name_in_list(target_md_name, name_list)
                    target_md_path = os.path.join(self._output_gdb, target_md_name)
                    
                    # Set env workspace back to another workspace otherwise
                    # when we add layer to map it will fail
                    arcpy.env.workspace = temp_file_gdb
                    
                    # Copy mosaic dataset to target geodatabase
                    md_copy = os.path.join(temp_file_gdb, md)
                    if self._verbose_messaging:
                        arcpy.AddMessage(arcpy.GetIDMessage(190182).format(md_copy, target_md_path))
                    arcpy.Copy_management(md_copy, target_md_path)
                    if self._verbose_messaging:
                        arcpy.AddMessage(arcpy.GetMessages())

                    # Add spatial index
                    # self._add_spatial_index(target_md_path)

                    # Copy the GroundOverlay folder to the same location as target geodatabase
                    raster_source_folder_path = os.path.join(temp_output_folder, os.path.basename(temp_file_gdb).replace('.gdb', '.grd'))
                    raster_target_folder_path = os.path.join(target_groundoverlay_folder, f'{target_md_name}.grd')
                    if self._verbose_messaging:
                        arcpy.AddMessage(arcpy.GetIDMessage(190183).format(raster_source_folder_path,
                                                                           raster_target_folder_path))
                    arcpy.Copy_management(raster_source_folder_path, raster_target_folder_path)
                    if self._verbose_messaging:
                        arcpy.AddMessage(arcpy.GetMessages())
                    
                    # Update raster paths in mosaic dataset to reference final raster location
                    paths = '"{}" "{}"'.format(raster_source_folder_path, raster_target_folder_path)
                    if self._verbose_messaging:
                        arcpy.AddMessage(arcpy.GetIDMessage(190184))
                    arcpy.RepairMosaicDatasetPaths_management(target_md_path, paths)
                    if self._verbose_messaging:
                        arcpy.AddMessage(arcpy.GetMessages())
                    
                    # Add mosaic dataset to the map
                    layer_name = '{} {}'.format(target_md_name, self._layer_time_stamp)
                    self._add_to_map(target_md_path, self._output_map, layer_name, None, self._layer_time_stamp)
                    
        except Exception:
            tb = sys.exc_info()[2]
            tbinfo = traceback.format_tb(tb)[0]
            pymsg = '{}\n{}\n{}\n{}'.format(self.__class__.__name__,
                                            tbinfo,
                                            str(sys.exc_info()[1]),
                                            arcpy.GetMessages(2))
            arcpy.AddError(pymsg)

        finally:
            # Reset environment setting
            arcpy.env.addOutputsToMap = initial_add_output_to_map
            
            # Delete objects
            try:
                for v in [feature_classes, fc, mosaic_datasets, md]:
                    del v
            except:
                pass

            # Delete temporary folder
            arcpy.Delete_management(temp_output_folder)


class ShapefileDataConverter(DataConverterBase):
    """ Convert an input shapefile to a geodatabase feature class """

    def __init__(self, input_file, output_gdb, output_map):
        self._input_file = input_file
        self._output_gdb = output_gdb
        self._output_map = output_map

    @property
    def input_file(self):
        return self._input_file

    @input_file.setter
    def input_file(self, value):
        self._input_file = value

    @property
    def output_gdb(self):
        return self._output_gdb

    @output_gdb.setter
    def output_gdb(self, value):
        self._output_gdb = value

    @property
    def output_map(self):
        return self._output_map

    @output_map.setter
    def output_map(self, value):
        self._output_map = value

    def convert(self):
        """ Convert an input shapefile to a geodatabase feature class """

        try:
            target_dataset_name = os.path.basename(os.path.splitext(self._input_file)[0])
            target_dataset_name = arcpy.ValidateTableName(target_dataset_name, self._output_gdb)
            target_dataset_name = self._create_unique_name_in_workspace(target_dataset_name, self._output_gdb)
            if self._verbose_messaging:
                arcpy.AddMessage(arcpy.GetIDMessage(190185).format(self._input_file,
                                                                   os.path.join(self._output_gdb,
                                                                                target_dataset_name)))
            arcpy.conversion.ExportFeatures(self._input_file, os.path.join(self._output_gdb, target_dataset_name))
            if self._verbose_messaging:
                arcpy.AddMessage(arcpy.GetMessages())

            # Add spatial index
            # self._add_spatial_index(os.path.join(self._output_gdb, target_dataset_name))

            layer_name = '{} {}'.format(target_dataset_name, self._layer_time_stamp)
            symbology_source = self._get_layer_file(os.path.basename(os.path.splitext(self._input_file)[0]),
                                                    os.path.dirname(self._input_file))
            self._add_to_map(os.path.join(self._output_gdb, target_dataset_name),
                             self._output_map, layer_name, symbology_source, self._layer_time_stamp)

        except Exception:
            tb = sys.exc_info()[2]
            tbinfo = traceback.format_tb(tb)[0]
            pymsg = '{}\n{}\n{}\n{}'.format(self.__class__.__name__,
                                            tbinfo,
                                            str(sys.exc_info()[1]),
                                            arcpy.GetMessages(2))
            arcpy.AddError(pymsg)


class TextFileDataConverter(DataConverterBase):
    """ Convert an input text file to a geodatabase feature class """

    def __init__(self, input_file, output_gdb, output_map):
        self._input_file = input_file
        self._output_gdb = output_gdb
        self._output_map = output_map

    @property
    def input_file(self):
        return self._input_file

    @input_file.setter
    def input_file(self, value):
        self._input_file = value

    @property
    def output_gdb(self):
        return self._output_gdb

    @output_gdb.setter
    def output_gdb(self, value):
        self._output_gdb = value

    @property
    def output_map(self):
        return self._output_map

    @output_map.setter
    def output_map(self, value):
        self._output_map = value

    def convert(self):
        """ Convert an input text file to a geodatabase feature class """

        try:
            target_dataset_name = os.path.basename(os.path.splitext(self._input_file)[0])
            target_dataset_name = arcpy.ValidateTableName(target_dataset_name, self._output_gdb)
            target_dataset_name = self._create_unique_name_in_workspace(target_dataset_name, self._output_gdb)
            target_dataset_path = os.path.join(self._output_gdb, target_dataset_name)
            arcpy.AddMessage(arcpy.GetIDMessage(190186).format(self._input_file,
                                                               os.path.join(self._output_gdb,
                                                                            target_dataset_name)))
            
            # Convert text file to point feature class
            convert_coord = self._get_convert_coordinate_params(self._input_file)
            
            if convert_coord:
                arcpy.ConvertCoordinateNotation_management(self._input_file, 
                                                           target_dataset_path,
                                                           convert_coord[ccn.X_FIELD.value],
                                                           convert_coord[ccn.Y_FIELD.value],
                                                           convert_coord[ccn.INPUT_COORD_FORMAT.value],
                                                           output_coordinate_format=convert_coord[ccn.OUTPUT_COORD_FORMAT.value],
                                                           exclude_invalid_records='INCLUDE_INVALID')
                if self._verbose_messaging:
                    arcpy.AddMessage(arcpy.GetMessages())

                # Add spatial index
                # self._add_spatial_index(target_dataset_path)

                layer_name = '{} {}'.format(target_dataset_name, self._layer_time_stamp)
                symbology_source = self._get_layer_file(os.path.basename(os.path.splitext(self._input_file)[0]),
                                                        os.path.dirname(self._input_file))
                self._add_to_map(target_dataset_path, self._output_map, layer_name, symbology_source,
                                self._layer_time_stamp)
            else:
                # Missing coordinate field(s) warning
                arcpy.AddWarning(arcpy.GetIDMessage(190239).format(self._input_file))
                
        except Exception:
            tb = sys.exc_info()[2]
            tbinfo = traceback.format_tb(tb)[0]
            pymsg = '{}\n{}\n{}\n{}'.format(self.__class__.__name__,
                                            tbinfo,
                                            str(sys.exc_info()[1]),
                                            arcpy.GetMessages(2))
            arcpy.AddError(pymsg)


class ExcelWorkbookDataConverter(DataConverterBase):
    """ Convert Excel worksheets contained in an Excel workbook to geodatabase feature classes. """

    def __init__(self, input_file, output_gdb, output_map):
        self._input_file = input_file
        self._output_gdb = output_gdb
        self._output_map = output_map

    @property
    def input_file(self):
        return self._input_file

    @input_file.setter
    def input_file(self, value):
        self._input_file = value

    @property
    def output_gdb(self):
        return self._output_gdb

    @output_gdb.setter
    def output_gdb(self, value):
        self._output_gdb = value

    @property
    def output_map(self):
        return self._output_map

    @output_map.setter
    def output_map(self, value):
        self._output_map = value

    def _get_sheet_names(self):
        """ Return a list of sheet names in the workbook """
        try:
            if os.path.splitext(self._input_file)[1].lower() == ".xlsx":
                workbook = openpyxl.load_workbook(self._input_file, read_only=True, data_only=True, keep_links=False)
                return workbook.sheetnames
            else:
                workbook = xlrd.open_workbook(self._input_file)
                return [sheet.name for sheet in workbook.sheets()]
        
        except Exception:
            tb = sys.exc_info()[2]
            tbinfo = traceback.format_tb(tb)[0]
            pymsg = '{}\n{}\n{}\n{}'.format(self.__class__.__name__,
                                            tbinfo,
                                            str(sys.exc_info()[1]),
                                            arcpy.GetMessages(2))
            arcpy.AddError(pymsg)

    def _has_sheet_data(self, sheet_name):
        """ Return boolean indicating if worksheet has data """
        
        has_data = False

        try:
            if os.path.splitext(self._input_file)[1].lower() == ".xlsx":
                workbook = openpyxl.load_workbook(self._input_file, read_only=True, data_only=True, keep_links=False)
                worksheet = workbook[sheet_name]
                
                for row in worksheet.values:
                    for cell_value in row:
                        if cell_value is not None:
                            has_data = True
                            break
                    if has_data:
                        break

            else:
                workbook = xlrd.open_workbook(self._input_file)
                worksheet = workbook.sheet_by_name(sheet_name)

                if worksheet.nrows > 0:
                    has_data = True
    
        except Exception:
            tb = sys.exc_info()[2]
            tbinfo = traceback.format_tb(tb)[0]
            pymsg = '{}\n{}\n{}\n{}'.format(self.__class__.__name__,
                                            tbinfo,
                                            str(sys.exc_info()[1]),
                                            arcpy.GetMessages(2))
            arcpy.AddError(pymsg)
            
        return has_data

    def convert(self):
        """ Convert Excel worksheets contained in an Excel workbook to geodatabase feature classes. """

        initial_workspace = arcpy.env.workspace
        scratch_tables = []

        try:
            scratch_gdb_path = arcpy.env.scratchGDB
            worksheet_names = self._get_sheet_names()
            
            for worksheet_name in worksheet_names:
                source_dataset_path = os.path.join(self._input_file, worksheet_name)
                arcpy.AddMessage(arcpy.GetIDMessage(190188).format(source_dataset_path))
                
                if not self._has_sheet_data(worksheet_name):
                    arcpy.AddWarning(arcpy.GetIDMessage(190238).format(source_dataset_path))
                    continue

                # Import source Excel worksheet to temporary scratch geodatabase table
                scratch_path = arcpy.CreateScratchName('', '', 'Dataset', scratch_gdb_path)
                scratch_tables.append(scratch_path)
                if self._verbose_messaging:
                    arcpy.AddMessage(arcpy.GetIDMessage(190190).format(source_dataset_path, scratch_path))
                arcpy.ExcelToTable_conversion(self._input_file, scratch_path, worksheet_name)
                if self._verbose_messaging:
                    arcpy.AddMessage(arcpy.GetMessages())

                # Determine unique output feature class name
                target_dataset_name = '{}_{}'.format(os.path.splitext(os.path.basename(self._input_file))[0],
                                                     worksheet_name)
                target_dataset_name = arcpy.ValidateTableName(target_dataset_name, self._output_gdb)
                target_dataset_name = self._create_unique_name_in_workspace(target_dataset_name, self._output_gdb)
                target_dataset_path = os.path.join(self._output_gdb, target_dataset_name)

                # Convert Excel worksheet to point feature class
                if self._verbose_messaging:
                    arcpy.AddMessage(arcpy.GetIDMessage(190231).format(scratch_path, target_dataset_path))
                
                convert_coord = self._get_convert_coordinate_params(scratch_path)

                if convert_coord:
                    arcpy.ConvertCoordinateNotation_management(scratch_path, 
                                                            target_dataset_path,
                                                            convert_coord[ccn.X_FIELD.value],
                                                            convert_coord[ccn.Y_FIELD.value],
                                                            convert_coord[ccn.INPUT_COORD_FORMAT.value],
                                                            output_coordinate_format=convert_coord[ccn.OUTPUT_COORD_FORMAT.value],
                                                            exclude_invalid_records='INCLUDE_INVALID')
                    if self._verbose_messaging:
                        arcpy.AddMessage(arcpy.GetMessages())

                    # Add spatial index
                    # self._add_spatial_index(target_dataset_path)

                    # Add layer to map with symbology
                    layer_name = '{} {}'.format(target_dataset_name, self._layer_time_stamp)
                    symbology_base_name = '{}_{}'.format(os.path.splitext(os.path.basename(self._input_file))[0],
                                                        worksheet_name)
                    symbology_source = self._get_layer_file(symbology_base_name, os.path.dirname(self._input_file))
                    self._add_to_map(target_dataset_path, self._output_map, layer_name, symbology_source,
                                    self._layer_time_stamp)
                else:
                    # Missing coordinate field(s) warning
                    arcpy.AddWarning(arcpy.GetIDMessage(190239).format(source_dataset_path))

        except Exception:
            tb = sys.exc_info()[2]
            tbinfo = traceback.format_tb(tb)[0]
            pymsg = '{}\n{}\n{}\n{}'.format(self.__class__.__name__,
                                            tbinfo,
                                            str(sys.exc_info()[1]),
                                            arcpy.GetMessages(2))
            arcpy.AddError(pymsg)

        finally:
            # Reset environment setting
            arcpy.env.workspace = initial_workspace
            
            # Delete scratch tables
            if len(scratch_tables) > 0:
                for scratch_table in scratch_tables:
                    if arcpy.Exists(scratch_table):
                        if self._verbose_messaging:
                            arcpy.AddMessage(arcpy.GetIDMessage(190232).format(scratch_table))
                        arcpy.Delete_management(scratch_table)


class GeoJSONDataConverter(DataConverterBase):
    """ Convert GeoJSON file to geodatabase feature classes. GeoJSON can support multiple
         geometry types within a single file. Each geometry type within the file is converted
         to a separate geodatabase feature class. """

    def __init__(self, input_file, output_gdb, output_map):
        self._input_file = input_file
        self._output_gdb = output_gdb
        self._output_map = output_map

    @property
    def input_file(self):
        return self._input_file

    @input_file.setter
    def input_file(self, value):
        self._input_file = value

    @property
    def output_gdb(self):
        return self._output_gdb

    @output_gdb.setter
    def output_gdb(self, value):
        self._output_gdb = value

    @property
    def output_map(self):
        return self._output_map

    @output_map.setter
    def output_map(self, value):
        self._output_map = value

    def _get_geometry_types(self):
        """ Return list of unique native geometry types from input source """
        
        # According to GeoJSON specification, GeoJSON files have the following geometry types:
        # "Point", "MultiPoint", "LineString", "MultiLineString", "Polygon", "MultiPolygon",
        # and "GeometryCollection".
        from osgeo import ogr  # Part of GDAL library installed by ArcGIS Pro

        # Use GDAL/ODR module to read GeoJSON file.
        driver = ogr.GetDriverByName('GeoJSON')
        data_source = driver.Open(self._input_file, 0)
        layer = data_source.GetLayer()
        geometry_types = []

        for feature in layer:
            geometry = feature.GetGeometryRef()
            geometry_type_name = geometry.GetGeometryName().upper()
            if geometry_type_name not in geometry_types:
                geometry_types.append(geometry_type_name)

        if len(geometry_types) == 0:
            return None
        else:
            return geometry_types

    def _get_esri_geometry_types(self, source_geometry_types=None):
        """ Return list of unique geometry types from input source mapped to Esri geometry types """

        # GeoJSON geometry type to gp tool mapping:
        #   Point                       -> POINT
        #   MultiPoint                  -> MULTIPOINT
        #   LineString, MultiLineString -> POLYLINE
        #   Polygon, MultiPolygon       -> POLYGON
        #   GeometryCollection   NOTE: this mapping is unknown at this time; need a sample dataset.

        if not source_geometry_types:
            source_geometry_types = self._get_geometry_types()

        esri_geometry_types = []

        for source_geometry_type in source_geometry_types:

            esri_geometry_type = None

            if source_geometry_type == 'POINT':
                esri_geometry_type = 'POINT'
            elif source_geometry_type == 'MULTIPOINT':
                esri_geometry_type = 'MULTIPOINT'
            elif source_geometry_type == 'LINESTRING' or source_geometry_type == 'MULTILINESTRING':
                esri_geometry_type = 'POLYLINE'
            elif source_geometry_type == 'POLYGON' or source_geometry_type == 'MULTIPOLYGON':
                esri_geometry_type = 'POLYGON'

            if esri_geometry_type:
                if esri_geometry_type not in esri_geometry_types:
                    esri_geometry_types.append(esri_geometry_type)

        if len(esri_geometry_types) == 0:
            return None
        else:
            return esri_geometry_types

    def convert(self):
        """ Convert GeoJSON file to geodatabase feature classes """

        try:
            # GeoJSON files can store multiple geometry types, so we need to determine what geometry
            # types the source GeoJSON file has so we know which "geometry_type" parameter value to
            # use on JSON to Features gp tool.
            source_geometry_types = self._get_geometry_types()

            if 'GEOMETRYCOLLECTION' in source_geometry_types:
                arcpy.AddWarning(arcpy.GetIDMessage(190233))

            # Return the equivalent Esri geometry types
            esri_geometry_types = self._get_esri_geometry_types(source_geometry_types)

            if esri_geometry_types:

                # Loop through each geometry type and execute the JSON to Features gp tool
                for esri_geometry_type in esri_geometry_types:
                    dataset_geometry_suffix = self._get_dataset_geometry_type_suffix_name(esri_geometry_type).title()
                    target_dataset_name = '{}_{}'.format(
                        os.path.basename(os.path.splitext(self._input_file)[0]), dataset_geometry_suffix)
                    symbology_source_basename = target_dataset_name
                    target_dataset_name = arcpy.ValidateTableName(target_dataset_name, self._output_gdb)
                    target_dataset_name = self._create_unique_name_in_workspace(target_dataset_name, self._output_gdb)
                    arcpy.AddMessage(arcpy.GetIDMessage(190234).format(self._input_file, 
                                                                       os.path.join(self._output_gdb, 
                                                                                    target_dataset_name)))
                    arcpy.JSONToFeatures_conversion(self._input_file,
                                                    os.path.join(self._output_gdb, target_dataset_name),
                                                    esri_geometry_type)
                    if self._verbose_messaging:
                        arcpy.AddMessage(arcpy.GetMessages())

                    # Add spatial index
                    # self._add_spatial_index(os.path.join(self._output_gdb, target_dataset_name))

                    layer_name = '{} {}'.format(target_dataset_name, self._layer_time_stamp)
                    symbology_source = self._get_layer_file(symbology_source_basename,
                                                            os.path.dirname(self._input_file))
                    self._add_to_map(os.path.join(self._output_gdb, target_dataset_name), self._output_map, layer_name,
                                     symbology_source, self._layer_time_stamp)

            else:
                arcpy.AddWarning(arcpy.GetIDMessage(190235))

        except Exception:
            tb = sys.exc_info()[2]
            tbinfo = traceback.format_tb(tb)[0]
            pymsg = '{}\n{}\n{}\n{}'.format(self.__class__.__name__,
                                            tbinfo,
                                            str(sys.exc_info()[1]),
                                            arcpy.GetMessages(2))
            arcpy.AddError(pymsg)


class GPXDataConverter(DataConverterBase):
    """ Convert an input GPX file to a geodatabase feature class """

    def __init__(self, input_file, output_gdb, output_map):
        self._input_file = input_file
        self._output_gdb = output_gdb
        self._output_map = output_map

    @property
    def input_file(self):
        return self._input_file

    @input_file.setter
    def input_file(self, value):
        self._input_file = value

    @property
    def output_gdb(self):
        return self._output_gdb

    @output_gdb.setter
    def output_gdb(self, value):
        self._output_gdb = value

    @property
    def output_map(self):
        return self._output_map

    @output_map.setter
    def output_map(self, value):
        self._output_map = value

    def convert(self):
        """ Convert an input GPX file to a geodatabase feature class """

        try:
            target_dataset_name = os.path.basename(os.path.splitext(self._input_file)[0])
            target_dataset_name = arcpy.ValidateTableName(target_dataset_name, self._output_gdb)
            target_dataset_name = self._create_unique_name_in_workspace(target_dataset_name, self._output_gdb)
            output_dataset = os.path.join(self._output_gdb, target_dataset_name)
            arcpy.AddMessage(arcpy.GetIDMessage(190236).format(self._input_file, output_dataset))
            arcpy.GPXtoFeatures_conversion(self._input_file, output_dataset)
            if self._verbose_messaging:
                    arcpy.AddMessage(arcpy.GetMessages())

            # Add spatial index
            # self._add_spatial_index(output_dataset)

            layer_name = '{} {}'.format(target_dataset_name, self._layer_time_stamp)
            symbology_source = self._get_layer_file(os.path.basename(os.path.splitext(self._input_file)[0]),
                                                    os.path.dirname(self._input_file))
            self._add_to_map(output_dataset, self._output_map, layer_name, symbology_source, self._layer_time_stamp)

        except Exception:
            tb = sys.exc_info()[2]
            tbinfo = traceback.format_tb(tb)[0]
            pymsg = '{}\n{}\n{}\n{}'.format(self.__class__.__name__,
                                            tbinfo,
                                            str(sys.exc_info()[1]),
                                            arcpy.GetMessages(2))
            arcpy.AddError(pymsg)
