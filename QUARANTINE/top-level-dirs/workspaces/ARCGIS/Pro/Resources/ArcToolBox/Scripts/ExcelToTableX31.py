"""
Tool name: Excel To Table 31
Source: ExcelToTableX31.py
Author: ESRI

Convert a Microsoft Excel xlsx file to an gdb, dbf or INFO table.

ExcelToTable31 vs ExcelToTableX31
 - ExcelToTable31.py uses xlrd to convert either xls and xlsx files
 - ExcelToTableX31.py uses openpyxl to convert xlsx files

"""

import arcpy
import os
import sys
import datetime
import re
import locale
import openpyxl
import arcgisscripting
fieldHelper = arcgisscripting._sharing._TextFileFieldHelper()


class Field(object):
    """ Class to hold properties and behavior of the output fields """

    def __init__(self, name, workspace, is_gdb=True, fields=[], col_num=None):
        """ Validate name of field based on output table format as well
              as existing/used field names
        """
        self.ftype = None
        self.length = None
        self.is_gdb = is_gdb
        self.col_num = col_num
        self.original_name = name   # if name else None

        if name and type(name) is str:
            name = openpyxl.utils.escape.unescape(name)

        self.alias = name if name else None

        name = arcpy.ValidateFieldName("" if name is None else name, workspace)
        name = arcpy.ValidateFieldName(name, "in_memory")

        if not name:
            try:
                name = 'COL_' + openpyxl.utils.get_column_letter(col_num)
            except IndexError:
                name = 'COL_'

        if not is_gdb:
            # dbf can't take non-alpha character to begin field name
            if not name[0].isalpha():
                name = 'f' + name
            name = name[:10]

        names = [f.name.lower() for f in fields] + ['objectid']

        # generate a unique name (using "_1", "_2", etc...)
        i = 0
        orig_name = name
        while names.count(name.lower()) != 0:
            i += 1
            name = "{0}_{1}".format(orig_name[:8], i)

        self.name = name

    def __repr__(self):
        """ Nice repr for debugging. """
        return f'<clsfield.name={self.name}, alias={self.alias}, ftype={self.ftype}, length={self.length}>'

    def validate_value(self, value):
        """ Validate the value against the output field type """
        # Non-gdb output do not support None/Null (except with Date)
        if not self.is_gdb:
            if value in [None, '', "#N/A"]:
                if self.ftype == 'Text':
                    value = ""
                elif self.ftype == 'Date':
                    value = None
                else:
                    value = 0
        else: # gdb
            if value == "#N/A":
                value = None  #value of #N/A is always None
            elif self.ftype in ['Double', 'Date', 'Long'] and value in ['', ]:
                # Cannot set '' into a numeric or date field in a gdb
                value = None
            elif self.ftype == "Text":
                if value == None: # equiv with xlrd
                    value = ""
                elif isinstance(value, (datetime.date, datetime.time, datetime.datetime)):
                    value = str(value)
                elif isinstance(value, str):
                    # deal with xlsx escape charcters
                    value = openpyxl.utils.escape.unescape(value)
        return value

def get_sheet_names(in_excel):
    """ Returns a list of sheet names for the selected excel file.
          This function is used in the script tool's Validation
    """
    wb = openpyxl.load_workbook(in_excel, read_only=True, data_only=True, keep_links=False)
    return wb.sheetnames

def open_excel_table(in_excel, sheet_name):
    """ Open the excel file, return sheet and workbook. """

    try:
        workbook = openpyxl.load_workbook(in_excel, read_only=True, data_only=True, keep_links=False)
        worksheet = workbook[sheet_name]

        # Read-only mode relies on applications and libraries that created the
        # file providing correct information about the worksheets, specifically
        # the used part of it, known as the dimensions. Some applications set
        # this incorrectly. Resetting will correct this.
        if worksheet.max_column == worksheet.max_row == 1:
            worksheet.reset_dimensions()
        return worksheet, workbook

    except Exception as err:
        arcpy.AddError(err)
        raise

def validate_fields(in_excel, out_table, sheet_name, rows=100):
    """ Validates field names, eliminating duplicate names and invalid
        characters. This is only used in the script tool's Validation.
    """
    # Code works, but open_workbook can be quite slow (20 sec), don't want
    #  to pay this cost in updateParameter as it's called in tool dialog each
    #  time a parameter value is changed.
    return []

    # workbook = open_workbook(in_excel)
    # sheet, workbook = open_excel_table(in_excel, sheet_name)
    # out_path, out_table_name = os.path.split(out_table)
    # is_gdb = not arcpy.Describe(out_path).workspaceType == "FileSystem"
    # return [f.name for f in gen_out_fields(workbook,
    #                                        sheet,
    #                                        out_path,
    #                                        is_gdb,
    #                                        rows,
    #                                        check_for_MGRS)]
    # Note: check_for_MGRS is only true it this is called from ExcelAnalyze
    #            from within the server context

def gen_out_fields(workbook, sheet, out_path, is_gdb, eval_count=None,
                   check_for_MGRS=False, cell_range=None, field_names_row=1):
    """ Generate the list of output field names based on inputs """

    if cell_range is None:
        cell_range = gen_cell_range(None, sheet)

    if sheet.max_row == 0:
        return []

    # Set the Field types based on values
    out_fields = []

    fieldmap = {'str': 'Text', 'float': 'Double', 'datetime': 'Date',
                'time': 'Text', 'int': 'Long', 'bool': 'Short',
                'NoneType': "none", 5: 'MGRS', 6: 'USNG'}

    out_fields = []

    # Generate the list of output fields
    # need loop since can be independent of  data range
    for row in sheet.iter_rows(max(field_names_row, 1),
                               max(field_names_row, 1),
                               cell_range.get("left"),
                               cell_range.get("right")):

        for i, value in enumerate([j.value for j in row]):
            name = "" if field_names_row == 0 else value
            out_fields.append(Field(name, out_path, is_gdb, out_fields,
                                    i + (cell_range.get("left") or 1)))

    # Get list of all types in column
    column_info = {}
    rowid = None
    for row_counter, row in enumerate(
            sheet.iter_rows(cell_range.get("top"), cell_range.get("bottom"),
                            cell_range['left'], cell_range['right'])):

        if eval_count is not None:
            if row_counter >= eval_count:
                break

        if row_counter == 0:
            if (cell_range['right'] is not None) and (cell_range['left'] is not None):
                for i in range(cell_range['right'] - cell_range['left']+1):
                    column_info[i] = {'cell_types': set(), 'is_lat_long': set(), 'maxlen': 0}
            else:
                for i in range(len(row)):
                    column_info[i] = {'cell_types': set(), 'is_lat_long': set(), 'maxlen': 0}

        if field_names_row is not None:
            # not all cells have row property, so loop until find one
            for cell in row:
                rowid = getattr(cell, 'row', None)
                if rowid:
                    break

        for i, cell in enumerate(row):
            value = cell.value

            if rowid == field_names_row:
                field_names_row = None # set to none, so stop looking for it
                break # break out of the cells loop

            data_type = value.__class__.__name__

            # check if int should be stored in float field
            if data_type == "int":
                if value > 2.1e+9 or value < -2.1e+9:
                    data_type = 'float'

            # treat #N/A same as None
            elif value == "#N/A":
                data_type = "NoneType"

            # track all the column cell's data types
            column_info[i]['cell_types'].add(data_type)

            if data_type == 'str':
                column_info[i]["maxlen"] = max(column_info[i]["maxlen"], len(value))
                if check_for_MGRS:
                    column_info[i]["is_lat_long"].add(fieldHelper.IsLatLong(value))

    # now look at all values in the column_info array
    for i, field_type in column_info.items():
        cell_types = set(field_type['cell_types']) - {'NoneType'}
        maxlen = None

        # if mix of float and integer, go with float
        if cell_types == {'float', 'int'}:
            cell_types = {'float'}

        if len(cell_types) == 1:
            # if all cells are same type use that
            ftype = fieldmap[list(cell_types)[0]]
        else:
            # if all cells are mix of types use text
            ftype = 'Text'

        # if text in a gdb, set field's length based on longest value
        if ftype == 'Text':
            maxlen = max(field_type['maxlen'] + 10, 255)

            if all(field_type['is_lat_long']) and check_for_MGRS:
                maxlen = 16    # 15 + 1 characters
                check_for_MGRS = False  # stop checking for MGRS
                if 'USNG' in out_fields[i].name.upper():
                    ftype = 'USNG'
                else:
                    ftype = 'MGRS'

        out_fields[i].ftype = ftype
        out_fields[i].length = maxlen if ftype in ["Text", "USNG", "MGRS"] else None

    return out_fields

def col_letter_to_int(s):
    """ converts column header to an integer
        A=0, Z=26, AA=27, CV=100, ...

        Not case sensitive, a=A=0, b=B=1
    """
    s = list(s.upper())
    s.reverse()
    val2 = ((ord(s[2]) - 64) * 26 * 26) if len(s) > 2 else 0
    val1 = ((ord(s[1]) - 64) * 26) if len(s) > 1 else 0
    val0 = ord(s[0]) - 64

    # all all of them then substract 0 (index base)
    return val2 + val1 + val0 - 1

def unpack_cell_range(cell_range):
    """ insure cell_range is valid and turn it into
        left, right, top, bottom

        So B10:E50 becomes left: 2, right: 5, top: 10, bottom: 50
    """

    if re.match("^[A-Za-z]{1,3}[0-9]{1,7}:[A-Za-z]{1,3}[0-9]{1,7}$",
                  cell_range) is None:
        raise AssertionError

    # pull the 4 components out of the cell_range
    left = col_letter_to_int(re.search("^[A-Za-z]{1,3}", cell_range)[0])
    right = col_letter_to_int(re.search("(?<=\:)[A-Za-z]{1,3}", cell_range)[0])

    top = int(re.search("[0-9]{1,7}(?=:)", cell_range)[0])
    bottom = int(re.search("[0-9]{1,7}$", cell_range)[0])

    # make sure ranges are valid
    assert left <= right
    assert top <= bottom

    return left +1, right +1, top, bottom

def gen_cell_range(cell_range, sheet):
    """ create a dict of the left, right, top, bottom of the cell_range """

    if cell_range in [None, '', "#"]:
        cell_range = {'left': None,
                      'right': None,
                      'top': None,
                      'bottom': None}
    else:
        try:
            left, right, top, bottom = unpack_cell_range(cell_range)
        except Exception as e:

            arcpy.AddIDMessage("ERROR", 2926, "cell_range", str(cell_range))
            if __name__ == '__main__':
                sys.exit(1)
            else:
                msg = arcpy.GetIDMessage(2926).replace("%1", "cell_range"
                                             ).replace("%2", str(cell_range))
                raise arcpy.ExecuteError(msg)

        # turn cell_range into a dictionary for easier use later
        cell_range = {'left': left,
                      'right': min(right, sheet.max_column),
                      'top': top,
                      'bottom': min(bottom, sheet.max_row) }

        # throw error if the right or bottom of cell_range is outside the data
        if ((cell_range['left'] > cell_range['right']) or
             cell_range['top'] > cell_range['bottom'] or
             cell_range['top'] < 0):

            # the cell_range is below or to the left of the data range
            arcpy.AddIDMessage("ERROR", 3301)
            if __name__ == '__main__':
                sys.exit(1)
            else:
                raise arcpy.ExecuteError(arcpy.GetIDMessage(3301))

    return cell_range

def excel_to_table(in_excel, out_table, sheet_name=None, field_names_row=1,
                   cell_range=None, eval_count=None):
    """
    Convert an excel sheet to a gdb table, dbf, or info table

    :param in_excel: path to an xls or xlsx file
    :param out_table: path to gdb table or dbf to be created
    :param sheet_name: sheet from input file to be converted - if blank sheet[0] is used
    :param cell_range: cell range to convert (A5:B10) - if blank use all
    :param eval_count: number of input rows to eval to determine field data type - if blank use all
    :param field_names_row: row for field names
    :return: None
    """
    if sheet_name in [None, '', '#']:
        sheet_name = get_sheet_names(in_excel)[0]

    if arcpy.Exists(out_table):
        if not arcpy.env.overwriteOutput:
            arcpy.AddIDMessage("ERROR", 258, out_table)
            if __name__ == '__main__':
                print('ERROR ' + arcpy.GetIDMessage(258))
                sys.exit(1)
            else:
                raise arcpy.ExecuteError(arcpy.GetIDMessage(258))

    sheet, workbook = open_excel_table(in_excel, sheet_name)
    cell_range = gen_cell_range(cell_range, sheet)
    out_path, out_table_name = os.path.split(out_table)

    if out_path.lower() in ["in_memory", "memory"]:
        out_path_ws_type = "Memory"
    else:
        out_path_ws_type = getattr(arcpy.Describe(out_path), "workspaceType", "?")

    is_gdb = out_path_ws_type != "FileSystem"

    # Note: If the check_for_MGRS flag wo out_path_ws_type == "FileSysuld be set to True, then we would
    # get some field type that we cannot create. Only set this to True to
    # get for evaluating fields, such as MGRS (text) fields
    out_fields = gen_out_fields(workbook, sheet, out_path, is_gdb, eval_count,
                                False, cell_range, field_names_row)

    # file extension for out_table
    ext = os.path.splitext(out_table_name)[1].lower()

    if ext == ".csv":
        coll_sep = "," if locale.localeconv()['decimal_point'] == "." else";"
        with open(out_table, 'w', encoding="utf-8") as _outf:
            _outf.write("\ufeff" + str(coll_sep).join([i.name for i in out_fields]) + "\n")
    else:
        # For performance reasons, add the fields to an in_memory table
        tmp_table = os.path.join('in_memory', 'tmp_exceltotable_template')
        if arcpy.Exists(tmp_table):
            arcpy.Delete_management(tmp_table)

        arcpy.CreateTable_management(*os.path.split(tmp_table))

        # Add the fields that were validates using the previous function
        arcpy.AddFields_management(tmp_table,
                                   [(f.name, f.ftype, f.alias, f.length)
                                               for f in out_fields])

        if out_path_ws_type == "RemoteDatabase":
            try:
                # pluck 'table' from 'db.owner.table'
                out_table_name = arcpy.ParseTableName(out_table_name, out_path).split(", ")[2]
            except:
                pass

        if out_path_ws_type == "FileSystem":
            if not ext in [".dbf"]:
                out_table_name += ".dbf"
        else:
            out_table_name = arcpy.ValidateTableName(out_table_name, out_path)


        # Now create the actual output table
        out_table = arcpy.CreateTable_management(out_path,
                                                 out_table_name,
                                                 template=tmp_table)[0]

        arcpy.Delete_management(tmp_table)

        # Output info table has trouble with OBJECTID field
        if not is_gdb and ext != '.dbf':
            arcpy.DeleteField_management(out_table, 'OBJECTID')

    arcpy.SetParameterAsText(1, out_table)

    # If the sheet has no rows, warn and exit
    #if sheet.max_row < 2:
    #    arcpy.AddIDMessage('WARNING', 117)
    #    return

    # Loop through each row and insert values into the output table
    with arcpy.da.InsertCursor(out_table, [f.name for f in out_fields]) as cursor:
        field_names_row_id = field_names_row - (cell_range['top'] or 1)
        blank_rows = []

        for rowid, rowvals in enumerate(sheet.iter_rows(cell_range['top'],
                                                        cell_range['bottom'],
                                                        cell_range['left'],
                                                        cell_range['right'],
                                                        values_only=True)):

            if rowid == field_names_row_id:
                continue

            row = list(rowvals)

            for colid, field in enumerate(out_fields):
                try:
                    row[colid] = field.validate_value(row[colid], )
                except IndexError:
                    row.append(field.validate_value(None))

            # collect blank rows
            if not any(row):
                blank_rows.append(row)
                continue

            # if blank rows were not at end, then add them to the output
            if blank_rows:
                for row2 in blank_rows:
                    oid = cursor.insertRow(row2)
                blank_rows = []

            # add row to the output table
            oid = cursor.insertRow(row)
